# Builds "RX Comparison - All", "RX Comparison +ve", and "MFP Drugs - RX" per-prescription analysis sheets.

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def _build_kinray_price_table(kinray_df, max_month):
    """
    Build a price lookup table: NDC_norm x Month -> unit price.
    Uses ffill (backward search) then bfill (forward search) per NDC
    to replicate find_kinray_price_by_month() fallback logic:
      1. Same month -> use latest price
      2. Search backwards (ffill)
      3. Search forwards (bfill)
      4. Return 0 if never found
    """
    kdf = kinray_df.copy()

    # Normalize NDC
    kdf['NDC_norm'] = (kdf['NDC #'].astype(str)
        .str.replace(r'\D', '', regex=True).str.zfill(11))

    # Use parser-calculated unit price: Invoice $ / Ship Qty.
    kdf['__UnitPrice__'] = pd.to_numeric(kdf.get('__UnitPrice__'), errors='coerce')

    # Parse date and derive month
    kdf['DATE'] = pd.to_datetime(kdf['DATE'], errors='coerce')
    kdf = kdf.dropna(subset=['DATE', '__UnitPrice__'])
    kdf['Month'] = kdf['DATE'].dt.to_period('M')

    # Take latest price per NDC per month
    price_table = (kdf.sort_values('DATE')
        .groupby(['NDC_norm', 'Month'], as_index=False)
        .agg(__unit_price__=('__UnitPrice__', 'last'))
    )

    if price_table.empty:
        return price_table

    # Build full month grid from earliest Kinray month to max fill date month
    min_month = price_table['Month'].min()
    all_months = pd.period_range(min_month, max_month, freq='M')
    all_ndcs = price_table['NDC_norm'].unique()

    full_grid = pd.MultiIndex.from_product(
        [all_ndcs, all_months],
        names=['NDC_norm', 'Month']
    ).to_frame(index=False)

    # Merge known prices onto full grid
    price_filled = full_grid.merge(price_table, on=['NDC_norm', 'Month'], how='left')

    # ffill = backward search, bfill = forward search
    try:
        price_filled = (price_filled
            .sort_values(['NDC_norm', 'Month'])
            .groupby('NDC_norm', group_keys=False)
            .apply(lambda g: g.assign(
                __unit_price__=g['__unit_price__'].ffill().bfill()
            ), include_groups=False)
            .reset_index(drop=True)
        )
    except TypeError:
        # pandas < 2.2 doesn't support include_groups
        price_filled = (price_filled
            .sort_values(['NDC_norm', 'Month'])
            .groupby('NDC_norm', group_keys=False)
            .apply(lambda g: g.assign(
                __unit_price__=g['__unit_price__'].ffill().bfill()
            ))
            .reset_index(drop=True)
        )

    return price_filled[['NDC_norm', 'Month', '__unit_price__']]


def add_rx_unit_compare_sheet_exact(
    wb,
    log_df,
    kinray_df,
    sheet_name: str = "RX Comparison - All"
):
    """
    Output columns (exact order):
      Rx, NDC, Drug Name, Fill date, Qty filled, Package billed,
      Kinray Unit Price, Ins paid, Unit Ins paid, Difference

    ✅ Shows ONLY rows where Difference < 0 (underpaid RXs)
    ✅ Sorted by Fill Date descending (latest first)
    """

    df = log_df.copy()
    #print(df.head())
    if '* SDRA Amt' in df.columns and 'SDRA Amt' not in df.columns:
        df.rename(columns={'* SDRA Amt': 'SDRA Amt'}, inplace=True)
    elif 'SDRA' in df.columns and 'SDRA Amt' not in df.columns:
        df.rename(columns={'SDRA': 'SDRA Amt'}, inplace=True)
    if 'Copay' in df.columns and 'COPAY' not in df.columns:
        df.rename(columns={'Copay': 'COPAY'}, inplace=True)

    # --- Normalize numeric columns ---
    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'Plan 1 BIN',
              'Plan 2 BIN', 'SDRA Amt', 'COPAY']:
        df[c] = pd.to_numeric(df.get(c, 0), errors='coerce').fillna(0)

    # Normalize NDC
    df['NDC #'] = (df['NDC #'].astype(str)
                   .str.replace('-', '', regex=False)
                   .str.replace(r'\D', '', regex=True)
                   .str.zfill(11))

    # Detect Fill Date column first (needed for price lookup)
    date_candidates = ['Fill Date', 'Date',
                       'Rx Date', 'Dispense Date', 'Service Date']
    fill_date_col = next((c for c in date_candidates if c in df.columns), None)
    if fill_date_col:
        df['Fill Date'] = pd.to_datetime(df[fill_date_col], errors='coerce')
    else:
        df['Fill Date'] = pd.NaT

    # Vectorized price lookup with backward/forward fallback
    df['NDC_norm'] = df['NDC #'].astype(str).str.zfill(11)
    df['Month'] = df['Fill Date'].dt.to_period('M')
    _max_month = df['Month'].max()
    _price_table = _build_kinray_price_table(kinray_df, _max_month)
    if not _price_table.empty:
        df = df.merge(_price_table, on=['NDC_norm', 'Month'], how='left')
        df['Kinray Unit Price'] = df['__unit_price__'].fillna(0)
        df.drop(columns=['NDC_norm', 'Month', '__unit_price__'], inplace=True, errors='ignore')
    else:
        df['Kinray Unit Price'] = 0
        df.drop(columns=['NDC_norm', 'Month'], inplace=True, errors='ignore')

    # --- Winning insurance paid ---
    df['Ins paid'] = np.where(
        df['Ins Paid Plan 1'].fillna(0) >= df['Ins Paid Plan 2'].fillna(0),
        df['Ins Paid Plan 1'].fillna(0),
        df['Ins Paid Plan 2'].fillna(0)
    )

    # --- Package billed ---
    df['Package billed'] = np.where(
        df['Drug Pkg Size'] > 0,
        df['Qty Filled'] / df['Drug Pkg Size'],
        np.nan
    )

    df['Kinray final Price'] = np.where(
        (df['Drug Pkg Size'] > 0) & (df['Kinray Unit Price'] > 0),
        (df['Kinray Unit Price']/df['Drug Pkg Size']) * df['Qty Filled'],
        0.0
    )

    # --- Unit insurance paid (per package logic) ---
    df['Unit Ins paid'] = np.where(
        df['Package billed'] > 0,
        df['Ins paid'] / df['Package billed'],
        np.nan
    )

    # Total paid (Insurance + SDRA + Copay)
    df['Total Ins paid'] = df['Ins paid'] + df['SDRA Amt'] + df['COPAY']

    # Difference = Total Ins paid - Kinray final Price
    # If Kinray Unit Price is 0, force Difference = 0
    df['Difference'] = np.where(
        df['Kinray Unit Price'] > 0,
        df['Total Ins paid'] - df['Kinray final Price'],
        0.0
    )

    # Drop rows where Difference is positive or 0
    # df = df[df['Difference'] > 0]

    # Map Rx column
    rx_col = 'Rx #' if 'Rx #' in df.columns else (
        'Rx' if 'Rx' in df.columns else None)
    df['RX'] = df[rx_col] if rx_col else pd.NA
    df['NDC'] = df['NDC #']
    df['Drug Name'] = df['Drug Name']
    df['Pkg Size'] = df['Drug Pkg Size']
    df['BIN'] = df['Winning_BIN']
    df['Processor'] = df['Processor']
    df['PCN'] = df['Winning PCN']
    df['Group'] = df['Winning Group']
    df['Fill Date'] = df['Fill Date'].dt.date
    df['Pkgs Billed to Insurance'] = df['Package billed']
    df['Kinray Price (Pkgs Billed × Unit Price)'] = df['Kinray final Price']
    df['Total Ins Paid for Pkgs Billed = (Ins Paid + SDRA + COPAY)'] = df['Total Ins paid']
    if 'Drug Type' not in df.columns:
        df['Drug Type'] = 'Unclassified'
    out_cols = [
        'RX', 'Fill Date', 'NDC', 'Drug Name', 'Pkg Size', 'Drug Type',
        'Pkgs Billed to Insurance',
        'Total Ins Paid for Pkgs Billed = (Ins Paid + SDRA + COPAY)',
        'Kinray Price (Pkgs Billed × Unit Price)',
        'Difference', 'BIN', 'Processor',
        'PCN', 'Group'
    ]

    # Filter, then sort by Fill Date DESCENDING (latest first)
    out = df.loc[:, out_cols].copy()
    # latest first, then largest diff
    out = out.sort_values('Drug Name', ascending=True)
    _num_cols = out.select_dtypes(include='number').columns
    out[_num_cols] = out[_num_cols].round(2)

    # Compute summary stats (before sheet creation so out is fully built)
    _total_rx_analyzed = len(out)
    _kinray_price_col  = 'Kinray Price (Pkgs Billed × Unit Price)'
    _rx_with_price = int((out[_kinray_price_col] > 0).sum()) if _kinray_price_col in out.columns else 0
    _rx_no_price   = _total_rx_analyzed - _rx_with_price
    _diff_col = out['Difference'] if 'Difference' in out.columns else pd.Series([], dtype=float)
    _diff_num  = pd.to_numeric(_diff_col, errors='coerce').fillna(0)
    _overpaid_count  = int((_diff_num > 0).sum())
    _underpaid_count = int((_diff_num < 0).sum())
    _overpaid_amt    = float(_diff_num[_diff_num > 0].sum())
    _underpaid_amt   = float(_diff_num[_diff_num < 0].sum())
    if _kinray_price_col in out.columns:
        _priced_mask = out[_kinray_price_col] > 0
        _profit_loss = float(_diff_num[_priced_mask.values].sum())
    else:
        _profit_loss = 0.0

    # # If no underpaid rows, create placeholder sheet
    # if out.empty:
    #     if sheet_name in wb.sheetnames:
    #         del wb[sheet_name]
    #     ws = wb.create_sheet(title=sheet_name)
    #     ws['A1'] = "No underpaid RXs found (Difference ≥ 0)."
    #     return

    # --- Create Sheet ---
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(out_cols))
    t = ws.cell(row=1, column=1, value="RX Comparision Analysis (All RXs)")
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 32   # title row

    # Write table
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
            else:
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

    # ✅ Wrap specific headers
    for cell_ref in ["E2", "F2", "G2", "H2","I2"]:
        ws[cell_ref].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

    widths = {
        'RX': 9, 'Fill Date': 12, 'NDC': 14, 'Drug Name': 45, 'Pkg Size': 8,
        'Drug Type': 15,
        'Pkgs Billed to Insurance': 12, 'Kinray Price (Pkgs Billed × Unit Price)': 18,
        'Total Ins Paid for Pkgs Billed = (Ins Paid + SDRA + COPAY)': 24,
        'Difference': 14, 'BIN': 8, 'Processor': 15, 'PCN': 12, 'Group': 12
    }
    ws.row_dimensions[2].height = 55   # header row
    ws.sheet_format.defaultRowHeight = 15
    ws.sheet_format.customHeight = False  # let Excel auto-size data rows

    for i, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    last_data_row = ws.max_row
    total_row = last_data_row + 1

    label_cell = ws.cell(row=total_row, column=out_cols.index('Drug Name') + 1, value="TOTALS")
    label_cell.font = Font(bold=True, size=12)
    label_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{last_data_row}"

    for col_name in ['Kinray Price (Pkgs Billed × Unit Price)', 'Total Ins Paid for Pkgs Billed = (Ins Paid + SDRA + COPAY)', 'Difference']:
        try:
            idx = out_cols.index(col_name) + 1
            col_letter = get_column_letter(idx)
            tcell = ws.cell(row=total_row, column=idx)
            tcell.value = f"=SUBTOTAL(109,{col_letter}3:{col_letter}{last_data_row})"
            tcell.number_format = '"$"#,##0.00'
            tcell.font = Font(bold=True, size=12)
            tcell.alignment = Alignment(horizontal='center', vertical='center')
        except ValueError:
            pass

    # Ensure Excel recalculates when opening
    ws.parent.calculation.fullCalcOnLoad = True

    # Freeze panes
    ws.freeze_panes = "A3"
    diff_col_letter = get_column_letter(out_cols.index("Difference") + 1)
    data_range = f"{diff_col_letter}3:{diff_col_letter}{ws.max_row}"

    # 🔴 Negative values = RED FILL
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='lessThan', formula=['0'],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )

    # 🟢 Positive values = GREEN FILL
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    # Set page orientation to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── RX Comparison Summary box (placed 2 cols right of data) ──
    _S = len(out_cols) + 2  # column P when out_cols has 14 cols
    _money = '"$"#,##0.00'
    _count = '#,##0'

    def _sum_cell(ws, row, col, value, fill_hex, font_color="1A202C",
                  bold=False, num_fmt=None, label=False):
        c = ws.cell(row=row, column=col, value=value)
        c.fill      = PatternFill("solid", fgColor=fill_hex)
        c.font      = Font(bold=bold, color=font_color, size=10)
        c.alignment = Alignment(
            horizontal='left' if label else 'center',
            vertical='center', wrap_text=True)
        if num_fmt:
            c.number_format = num_fmt
        return c

    # Header
    ws.merge_cells(start_row=1, start_column=_S, end_row=1, end_column=_S + 2)
    hdr = ws.cell(row=1, column=_S, value="RX COMPARISON SUMMARY")
    hdr.fill      = PatternFill("solid", fgColor="0F4C81")
    hdr.font      = Font(bold=True, color="FFFFFF", size=11)
    hdr.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Section 1 — RX counts
    _rows_s1 = [
        ("Total RX Analyzed",      _total_rx_analyzed, _count, "E8F2FF", "185FA5"),
        ("Kinray Price Available",  _rx_with_price,     _count, "EAF3DE", "375623"),
        ("No Kinray Price",         _rx_no_price,       _count, "FAEEDA", "854F0B"),
    ]
    _r = 2
    for label, val, fmt, fill, color in _rows_s1:
        _sum_cell(ws, _r, _S,     label, fill, color, label=True)
        _sum_cell(ws, _r, _S + 1, val,   fill, color, num_fmt=fmt)
        ws.merge_cells(start_row=_r, start_column=_S + 1,
                       end_row=_r, end_column=_S + 2)
        ws.row_dimensions[_r].height = 20
        _r += 1

    # Divider
    ws.merge_cells(start_row=_r, start_column=_S, end_row=_r, end_column=_S + 2)
    ws.cell(row=_r, column=_S).fill = PatternFill("solid", fgColor="C3D9F5")
    ws.row_dimensions[_r].height = 4
    _r += 1

    # Section 2 — Overpaid / Underpaid
    _sum_cell(ws, _r, _S,     "Overpaid (Ins > Kinray)", "EAF3DE", "375623", bold=True, label=True)
    _sum_cell(ws, _r, _S + 1, _overpaid_count, "EAF3DE", "375623", num_fmt=_count)
    _sum_cell(ws, _r, _S + 2, _overpaid_amt,   "EAF3DE", "375623", num_fmt=_money)
    ws.row_dimensions[_r].height = 22
    _r += 1

    _sum_cell(ws, _r, _S,     "Underpaid (Ins < Kinray)", "FCEBEB", "A32D2D", bold=True, label=True)
    _sum_cell(ws, _r, _S + 1, _underpaid_count, "FCEBEB", "A32D2D", num_fmt=_count)
    _sum_cell(ws, _r, _S + 2, _underpaid_amt,   "FCEBEB", "A32D2D", num_fmt=_money)
    ws.row_dimensions[_r].height = 22
    _r += 1

    # Divider
    ws.merge_cells(start_row=_r, start_column=_S, end_row=_r, end_column=_S + 2)
    ws.cell(row=_r, column=_S).fill = PatternFill("solid", fgColor="C3D9F5")
    ws.row_dimensions[_r].height = 4
    _r += 1

    # Section 3 — Profit / Loss
    ws.merge_cells(start_row=_r, start_column=_S, end_row=_r, end_column=_S + 2)
    pl_fill  = "EAF3DE" if _profit_loss >= 0 else "FCEBEB"
    pl_color = "375623" if _profit_loss >= 0 else "A32D2D"
    pl_cell  = ws.cell(row=_r, column=_S,
                       value=f"Profit / Loss of {_rx_with_price:,} RX")
    pl_cell.fill      = PatternFill("solid", fgColor=pl_fill)
    pl_cell.font      = Font(bold=True, color=pl_color, size=11)
    pl_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[_r].height = 22
    _r += 1

    ws.merge_cells(start_row=_r, start_column=_S, end_row=_r, end_column=_S + 2)
    pl_val              = ws.cell(row=_r, column=_S, value=_profit_loss)
    pl_val.fill         = PatternFill("solid", fgColor=pl_fill)
    pl_val.font         = Font(bold=True, color=pl_color, size=14)
    pl_val.alignment    = Alignment(horizontal='center', vertical='center')
    pl_val.number_format = _money
    ws.row_dimensions[_r].height = 28

    # Column widths for summary columns
    ws.column_dimensions[get_column_letter(_S)].width     = 22  # label
    ws.column_dimensions[get_column_letter(_S + 1)].width = 12  # count
    ws.column_dimensions[get_column_letter(_S + 2)].width = 14  # amount


def add_mfp_drugs_sheet(
    wb,
    log_df,
    kinray_df,
    sheet_name: str = "MFP Drugs - RX"
):
    """
    Per-RX MFP analysis sheet.

    Rule used:
    - Any RX row with SDRA Amt != 0 is considered an MFP RX.
    """
    df = log_df.copy()

    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'SDRA Amt', 'COPAY']:
        df[c] = pd.to_numeric(df.get(c, 0), errors='coerce').fillna(0)
    df['Ins Paid Total'] = pd.to_numeric(
        df.get('Ins Paid Total', pd.Series(0, index=df.index)),
        errors='coerce'
    ).fillna(0)

    df['NDC #'] = (df['NDC #'].astype(str)
                   .str.replace('-', '', regex=False)
                   .str.replace(r'\D', '', regex=True)
                   .str.zfill(11))

    date_candidates = ['Fill Date', 'Date', 'Rx Date', 'Dispense Date', 'Service Date']
    fill_date_col = next((c for c in date_candidates if c in df.columns), None)
    if fill_date_col:
        df['Fill Date'] = pd.to_datetime(df[fill_date_col], errors='coerce')
    else:
        df['Fill Date'] = pd.NaT

    df = df[df['SDRA Amt'].fillna(0) != 0].copy()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    if df.empty:
        ws['A1'] = "No MFP RXs found (SDRA Amt is 0/blank for all rows)."
        ws['A1'].font = Font(size=14, bold=True)
        return

    # Vectorized price lookup with backward/forward fallback
    df['NDC_norm'] = df['NDC #'].astype(str).str.zfill(11)
    df['Month'] = df['Fill Date'].dt.to_period('M')
    _max_month = df['Month'].max()
    _price_table = _build_kinray_price_table(kinray_df, _max_month)
    if not _price_table.empty:
        df = df.merge(_price_table, on=['NDC_norm', 'Month'], how='left')
        df['Kinray Unit Price'] = df['__unit_price__'].fillna(0)
        df.drop(columns=['NDC_norm', 'Month', '__unit_price__'], inplace=True, errors='ignore')
    else:
        df['Kinray Unit Price'] = 0
        df.drop(columns=['NDC_norm', 'Month'], inplace=True, errors='ignore')

    df['Package billed'] = np.where(
        df['Drug Pkg Size'] > 0,
        df['Qty Filled'] / df['Drug Pkg Size'],
        np.nan
    )
    df['Kinray Final Price'] = np.where(
        (df['Drug Pkg Size'] > 0) & (df['Kinray Unit Price'] > 0),
        (df['Kinray Unit Price'] / df['Drug Pkg Size']) * df['Qty Filled'],
        0.0
    )

    df['Pkgs Billed to Insurance'] = df['Package billed']
    df['Kinray Cost (Pkgs × Unit Price)'] = df['Kinray Final Price']
    df['Total Ins Paid For Pkgs Billed'] = df['Ins Paid Total']
    df['Total = (SDRA + Ins Paid Total)'] = df['SDRA Amt'] + df['Ins Paid Total']
    df['Difference'] = np.where(
        df['Kinray Final Price'] > 0,
        df['Total = (SDRA + Ins Paid Total)'] - df['Kinray Cost (Pkgs × Unit Price)'],
        0.0
    )

    rx_col = 'Rx #' if 'Rx #' in df.columns else ('Rx' if 'Rx' in df.columns else None)
    df['RX'] = df[rx_col] if rx_col else pd.NA
    df['NDC'] = df['NDC #']
    df['Drug Name'] = df.get('Drug Name', '')
    df['Pkg Size'] = df.get('Drug Pkg Size', 0)
    df['BIN'] = df.get('Winning_BIN', '')
    df['PCN'] = df.get('Winning PCN', '')
    df['Group'] = df.get('Winning Group', '')

    out_cols = [
        'RX', 'Fill Date', 'NDC', 'Drug Name', 'Pkg Size',
        'Pkgs Billed to Insurance',
        'Total Ins Paid For Pkgs Billed', 'SDRA Amt', 'Total = (SDRA + Ins Paid Total)',
        'Kinray Cost (Pkgs × Unit Price)',
        'Difference', 'BIN', 'PCN', 'Group'
    ]
    out = df.loc[:, out_cols].copy().sort_values(['Drug Name', 'Fill Date'], ascending=[True, False])
    out['Fill Date'] = pd.to_datetime(out['Fill Date'], errors='coerce').dt.date
    _num_cols = out.select_dtypes(include='number').columns
    out[_num_cols] = out[_num_cols].round(2)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_cols))
    title = ws.cell(row=1, column=1, value="MFP DRUGS")
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 26

    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    widths = {
        'RX': 9, 'Fill Date': 12, 'NDC': 14, 'Drug Name': 40, 'Pkg Size': 8,
        'Pkgs Billed to Insurance': 12, 'Kinray Cost (Pkgs × Unit Price)': 14,
        'Total Ins Paid For Pkgs Billed': 22, 'SDRA Amt': 10,
        'Total = (SDRA + Ins Paid Total)': 14, 'Difference': 12,
        'BIN': 9, 'PCN': 12, 'Group': 12
    }
    ws.row_dimensions[2].height = 45
    for i, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    last_data_row = ws.max_row
    total_row = last_data_row + 1
    label_col = max(1, out_cols.index('Difference'))
    ws.cell(row=total_row, column=label_col, value='Totals').font = Font(bold=True)
    ws.cell(row=total_row, column=label_col).alignment = Alignment(horizontal='right', vertical='center')

    for name in ['Kinray Cost (Pkgs × Unit Price)', 'Total Ins Paid For Pkgs Billed', 'SDRA Amt',
                 'Total = (SDRA + Ins Paid Total)', 'Difference']:
        if name not in out_cols:
            continue
        idx = out_cols.index(name) + 1
        col_letter = get_column_letter(idx)
        tcell = ws.cell(row=total_row, column=idx)
        tcell.value = f"=SUBTOTAL(109,{col_letter}3:{col_letter}{last_data_row})"
        tcell.font = Font(bold=True)
        tcell.number_format = '"$"#,##0.00'
        tcell.alignment = Alignment(horizontal='center', vertical='center')
        tcell.border = thin

    # Add table for borders and row stripes
    tab = Table(displayName="TableMFPDrugs",
                ref=f"A2:{get_column_letter(len(out_cols))}{last_data_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight2", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
    ws.add_table(tab)
    ws.auto_filter.ref = None  # Table manages filter

    ws.sheet_format.defaultRowHeight = 20
    ws.sheet_format.customHeight = True

    # Freeze rows 1-2 - set AFTER table to avoid conflict
    ws.freeze_panes = "A3"

    diff_col_letter = get_column_letter(out_cols.index('Difference') + 1)
    data_range = f"{diff_col_letter}3:{diff_col_letter}{last_data_row}"
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
