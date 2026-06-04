# Builds "BIN to Processor" (including unmapped BINs) sheet.

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side


def create_bin_to_processor_sheet(wb, rx_compare_source, bin_to_proc,
                                  dropped_status_counts=None,
                                  total_csv_rows=None):
    # ===== Create/replace "BIN to Processor" sheet =====
    title_sheet = "BIN to Processor"
    if title_sheet in wb.sheetnames:
        del wb[title_sheet]
    ws2 = wb.create_sheet(title_sheet)

    # Title
    ws2.insert_rows(1)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    tcell = ws2.cell(
        row=1, column=1, value="BIN Numbers Billed (from Custom Log)")
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal='center', vertical='center')

    def find_fill_date_column(df):
        """Return the first column name that looks like a Fill Date."""
        candidates = ['Fill Date']
        # exact match first
        for c in candidates:
            if c in df.columns:
                return c
        # loose match (case-insensitive contains "fill" & "date")
        for c in df.columns:
            cl = str(c).strip().lower()
            if "date" in cl and ("fill" in cl or "filled" in cl):
                return c
        return None

    # --- Build BIN -> Processor counts from the UNFILTERED custom log ---
    src_df = rx_compare_source.copy()  # unfiltered copy created earlier
    # Use the UNFILTERED log for totals so processor filters don't shrink the counts

    # <- "rows" (count rows), "qty" (sum Qty Filled), or "unique_rx" (distinct Rx #)
    COUNT_MODE = "rows"

    def build_rx_counts(src_df, mode="rows"):
        # Normalize BIN; include NaN/blank -> '000000'
        bins = (src_df['Winning_BIN']
                .astype('string')
                .fillna('')                       # keep empties
                .str.replace(r'\D', '', regex=True)
                .str.zfill(6))                    # '' -> '000000'

        df = src_df.copy().assign(__BIN=bins)

        if mode == "rows":
            out = (df.groupby('__BIN', as_index=False)
                   .size()
                   .rename(columns={'__BIN': 'BIN', 'size': 'Total Rx'}))
            label = 'Total Rx'
        elif mode == "qty":
            out = (df.groupby('__BIN', as_index=False)['Qty Filled']
                   .sum()
                   .rename(columns={'__BIN': 'BIN', 'Qty Filled': 'Total Qty'}))
            label = 'Total Qty'
        else:  # unique_rx
            out = (df.groupby('__BIN', as_index=False)['Rx #']
                   .nunique()
                   .rename(columns={'__BIN': 'BIN', 'Rx #': 'Total Rx'}))
            label = 'Total Rx'
        return out, label

    rx_counts_df, RX_LABEL = build_rx_counts(src_df, COUNT_MODE)

    bin_proc_df = (rx_counts_df[['BIN']].copy()
                   .assign(Processor=lambda d: d['BIN'].map(bin_to_proc))
                   # keep 000000
                   .assign(Processor=lambda d: d['Processor'].fillna('Unmapped'))
                   .merge(rx_counts_df, on='BIN', how='left')
                   .sort_values(['Processor', 'BIN'])
                   .reset_index(drop=True))

    # Write headers with the dynamic label in C
    headers = ["BIN", "Processor", RX_LABEL]
    for cidx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=2, column=cidx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
    # header_row = 2
    # start_data_row = header_row + 1

    # Write A:C
    for r, (bin_, proc, total) in enumerate(
            bin_proc_df[['BIN', 'Processor', RX_LABEL]].itertuples(index=False, name=None), start=3):
        ws2.cell(row=r, column=1, value=str(bin_))
        ws2.cell(row=r, column=2, value=str(proc))
        ws2.cell(row=r, column=3, value=int(total))

    # Optional grand total row to sanity-check equals src_df.shape[0] when COUNT_MODE=="rows"
    gt_row = ws2.max_row + 1
    ws2.cell(row=gt_row, column=2, value="Grand Total").font = Font(bold=True)
    ws2.cell(row=gt_row, column=3,
             value=f"=SUM(C3:C{gt_row-1})").font = Font(bold=True)

    # === Dropped RX summary ===
    if dropped_status_counts:
        # Blank separator row
        gap_row = ws2.max_row + 2

        # Section header
        ws2.merge_cells(start_row=gap_row, start_column=1,
                        end_row=gap_row, end_column=3)
        hdr = ws2.cell(row=gap_row, column=1,
                       value="Dropped RX (excluded from report)")
        hdr.font = Font(bold=True, size=12, color="FF0000")
        hdr.alignment = Alignment(horizontal='left', vertical='center')

        r = gap_row + 1
        total_dropped = 0
        for status, count in sorted(dropped_status_counts.items()):
            ws2.cell(row=r, column=1, value=status).font = Font(bold=False)
            ws2.cell(row=r, column=2, value=count).font = Font(bold=False)
            total_dropped += count
            r += 1

        # Total dropped row
        ws2.cell(row=r, column=1, value="Total Dropped").font = Font(bold=True)
        ws2.cell(row=r, column=2, value=total_dropped).font = Font(bold=True)
        r += 1

        # Total CSV rows
        if total_csv_rows:
            ws2.cell(row=r, column=1, value="Total CSV Rows").font = Font(bold=True)
            ws2.cell(row=r, column=2, value=total_csv_rows).font = Font(bold=True)

    # Widths / filter
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 10
    ws2.freeze_panes = "A3"

    # # Optional: bottom TOTAL row (helps you QA against expected 7,100 etc.)
    # end_row = ws2.max_row
    # total_row = end_row + 1
    # ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    # # Sum of column C
    # ws2.cell(row=total_row, column=3,
    #         value=f"=SUM(C{start_data_row}:C{end_row})").font = Font(bold=True)

    src_norm = src_df.copy()
    src_norm['__BIN'] = (src_norm['Winning_BIN'].astype('string')
                         .fillna('')
                         .str.replace(r'\D', '', regex=True)
                         .str.zfill(6))

    unmapped_rows = src_norm[src_norm['__BIN'] == '000000'].copy()
    fill_col = find_fill_date_column(unmapped_rows)

    ws2.merge_cells('F1:M1')
    title = ws2.cell(row=1, column=6, value="Unmapped BIN Numbers (000000)")
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.font = Font(bold=True, size=14)
    ws2['F2'] = "BIN"
    ws2['G2'] = "RX #"
    ws2['H2'] = "Drug Name"
    ws2['I2'] = "Fill Date"
    ws2['J2'] = "Plan 1 Name"
    ws2['K2'] = "Ins Paid Plan 1"
    ws2['L2'] = "Plan 2 Name"
    ws2['M2'] = "Ins Paid Plan 2"

    # Style + widths
    for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        head = ws2[f'{col}2']
        head.font = Font(bold=True, color="000000")
        head.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[col].width = 18
    ws2.column_dimensions['H'].width = 40
    ws2.column_dimensions['I'].width = 14
    ws2.column_dimensions['J'].width = 25
    ws2.column_dimensions['K'].width = 18
    ws2.column_dimensions['L'].width = 25
    ws2.column_dimensions['M'].width = 18

    # Coerce date (for pretty output); safe even if mixed types
    if fill_col:
        try:
            unmapped_rows[fill_col] = pd.to_datetime(
                unmapped_rows[fill_col], errors='coerce')
        except Exception:
            pass

    # Normalize numeric columns safely
    for pay_col in ['Ins Paid Plan 1', 'Ins Paid Plan 2']:
        if pay_col in unmapped_rows.columns:
            unmapped_rows[pay_col] = pd.to_numeric(
                unmapped_rows[pay_col], errors='coerce').fillna(0).round(2)

    # Write ALL rows (no set()/groupby dedupe): F=BIN, G=RX #, H=Drug Name, I=Fill Date
    start_row_unmapped = 3
    optional_unmapped_cols = [
        (9, fill_col),
        (10, 'Plan 1 Name'),
        (11, 'Ins Paid Plan 1'),
        (12, 'Plan 2 Name'),
        (13, 'Ins Paid Plan 2'),
    ]
    optional_unmapped_cols = [
        (col_idx, col_name) for col_idx, col_name in optional_unmapped_cols
        if col_name and col_name in unmapped_rows.columns
    ]
    cols = ['__BIN', 'Rx #', 'Drug Name'] + [col_name for _, col_name in optional_unmapped_cols]

    for r_idx, row in enumerate(
            unmapped_rows[cols].itertuples(index=False, name=None),
            start=start_row_unmapped):
        # F -> BIN (000000)
        ws2.cell(row=r_idx, column=6, value=row[0])
        ws2.cell(row=r_idx, column=7, value=str(row[1]))          # G -> RX #
        ws2.cell(row=r_idx, column=8, value=row[2])               # H -> Drug Name
        for (col_idx, col_name), v in zip(optional_unmapped_cols, row[3:]):
            # format Timestamp nicely
            if col_name == fill_col and hasattr(v, "strftime"):
                v = v.strftime('%Y-%m-%d')
            ws2.cell(row=r_idx, column=col_idx, value=v)

    # ---- Formatting
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 10
    ws2.freeze_panes = 'A3'  # keep title+headers fixed
    ws2.auto_filter.ref = "F2:M2"
