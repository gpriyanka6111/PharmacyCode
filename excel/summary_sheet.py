# Builds the redesigned "Summary" sheet with 5 sections.

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Colors
BRAND_BLUE  = "185FA5"
FILL_BLUE   = "E6F1FB"
FILL_GREEN  = "EAF3DE"
FILL_AMBER  = "FAEEDA"
FILL_RED    = "FCEBEB"
FILL_PURPLE = "EEEDFE"
FILL_GRAY   = "F1EFE8"
FILL_WHITE  = "FFFFFF"
TEXT_GREEN  = "0F6E56"
TEXT_AMBER  = "854F0B"
TEXT_RED    = "A32D2D"
TEXT_BLUE   = "185FA5"
TEXT_PURPLE = "3C3489"
TEXT_DARK   = "1A202C"
TEXT_MUTED  = "64748B"


def filled_cell(ws, row, col, value, fill_hex, text_hex, bold=False, size=11, align='center'):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=fill_hex)
    c.font = Font(bold=bold, color=text_hex, size=size)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    return c


def section_header(ws, row, col_start, col_end, label):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=label.upper())
    c.fill = PatternFill("solid", fgColor=FILL_GRAY)
    c.font = Font(bold=True, color=TEXT_BLUE, size=10)
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = Border(bottom=Side(style='thin'))


def add_summary_sheet(
    wb,
    pharmacy_name,
    date_range,
    processors=None,
    final_data=None,
    log_df=None,
    kinray_df=None,
    kinray_total=None,
    rx_compare_df=None,
    needs_df=None,
    dno_df=None,
):
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "0F4C81"

    money_fmt = '"$"#,##0.00'
    count_fmt = '#,##0'

    # Default column widths (overridden at end for brand table / processor table)
    ws.column_dimensions['A'].width = 22
    for _cl in 'BCDEFGH':
        ws.column_dimensions[_cl].width = 18

    # ──────────────────────────────────────────────────────────────
    # Row 1: Title
    # ──────────────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    c = ws.cell(row=1, column=1, value="Summary")
    c.fill = PatternFill("solid", fgColor="0F4C81")
    c.font = Font(bold=True, color="FFFFFF", size=18)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    # Row 2: Subtitle
    ws.merge_cells('A2:H2')
    c = ws.cell(row=2, column=1, value=f"Summary of {pharmacy_name or ''} for {date_range or ''}")
    c.font = Font(color=TEXT_MUTED, size=11)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Row 3: blank + freeze
    ws.row_dimensions[3].height = 8
    ws.freeze_panes = 'A3'

    # ══════════════════════════════════════════════════════════════
    # SECTION 1 — Pharmacy Overview (rows 4–8)
    # ══════════════════════════════════════════════════════════════
    total_rx      = 0
    insurance_paid = 0.0
    kinray_bill    = 0.0

    if log_df is not None:
        try:
            total_rx = len(log_df)
            if 'Ins Paid Total' in log_df.columns:
                insurance_paid = float(
                    pd.to_numeric(log_df['Ins Paid Total'], errors='coerce').fillna(0).sum())
            elif 'Winning_Paid' in log_df.columns:
                insurance_paid = float(
                    pd.to_numeric(log_df['Winning_Paid'], errors='coerce').fillna(0).sum())
            elif 'Ins Paid Plan 1' in log_df.columns:
                p1 = pd.to_numeric(log_df['Ins Paid Plan 1'], errors='coerce').fillna(0)
                p2 = pd.to_numeric(
                    log_df['Ins Paid Plan 2'] if 'Ins Paid Plan 2' in log_df.columns
                    else pd.Series(0, index=log_df.index),
                    errors='coerce'
                ).fillna(0)
                insurance_paid = float((p1 + p2).sum())
        except Exception:
            pass

    # Use pre-computed total passed from pipeline — most reliable source
    if kinray_total is not None:
        kinray_bill = float(kinray_total)
    elif kinray_df is not None:
        try:
            _inv_col = next((c for c in kinray_df.columns
                            if 'invoice' in c.lower() and '$' in c), None)
            if _inv_col:
                _vals = pd.to_numeric(kinray_df[_inv_col], errors='coerce').fillna(0)
                kinray_bill = float(_vals[_vals > 0].sum())
        except Exception:
            kinray_bill = 0.0

    net_profit = insurance_paid - kinray_bill

    section_header(ws, 4, 1, 8, "Pharmacy Overview")
    ws.row_dimensions[4].height = 22

    kpi_defs = [
        ("Total RX Processed",       total_rx,       FILL_BLUE,  TEXT_BLUE,  count_fmt),
        ("Insurance Paid (BestRx)",  insurance_paid, FILL_GREEN, TEXT_GREEN, money_fmt),
        ("Kinray Total Bill",         kinray_bill,    FILL_AMBER, TEXT_AMBER, money_fmt),
        ("Net (Paid − Purchased)", net_profit,
         FILL_GREEN if net_profit >= 0 else FILL_RED,
         TEXT_GREEN if net_profit >= 0 else TEXT_RED,
         money_fmt),
    ]
    for i, (label, value, fill, text, fmt) in enumerate(kpi_defs):
        cs, ce = i * 2 + 1, i * 2 + 2
        ws.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=ce)
        filled_cell(ws, 5, cs, label, fill, TEXT_MUTED, bold=False, size=9)
        ws.merge_cells(start_row=6, start_column=cs, end_row=6, end_column=ce)
        vc = filled_cell(ws, 6, cs, value, fill, text, bold=True, size=14)
        vc.number_format = fmt
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[7].height = 4
    ws.row_dimensions[8].height = 4

    # Row 9: blank separator
    ws.row_dimensions[9].height = 10

    # ══════════════════════════════════════════════════════════════
    # SECTION 2 — Order Analysis (rows 10–13)
    # ══════════════════════════════════════════════════════════════
    def _count_sheet_rows(sheet_name, data_start=3):
        if sheet_name not in wb.sheetnames:
            return 0
        ws_s = wb[sheet_name]
        return sum(
            1 for r in range(data_start, ws_s.max_row + 1)
            if ws_s.cell(row=r, column=1).value not in (None, '')
        )

    needs_order_count    = len(needs_df) if needs_df is not None else _count_sheet_rows("Needs to be ordered - All")
    do_not_order_count   = len(dno_df)   if dno_df   is not None else _count_sheet_rows("Do Not Order - ALL")

    section_header(ws, 10, 1, 8, "Order Analysis")
    ws.row_dimensions[10].height = 22

    order_cards = [
        ("Needs to be Ordered", needs_order_count,     FILL_AMBER,  TEXT_AMBER),
        ("Do Not Order",        do_not_order_count,    FILL_RED,    TEXT_RED),
    ]
    for i, (label, value, fill, text) in enumerate(order_cards):
        cs, ce = i * 2 + 1, i * 2 + 2
        ws.merge_cells(start_row=11, start_column=cs, end_row=11, end_column=ce)
        filled_cell(ws, 11, cs, label, fill, TEXT_MUTED, size=9)
        ws.merge_cells(start_row=12, start_column=cs, end_row=12, end_column=ce)
        vc = filled_cell(ws, 12, cs, f"{value:,} Drugs", fill, text, bold=True, size=14)
        vc.number_format = '@'  # text format since we're using string
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 28
    ws.row_dimensions[13].height = 4

    # Row 14: blank separator
    ws.row_dimensions[14].height = 10

    # ══════════════════════════════════════════════════════════════
    # SECTION 3 — RX Comparison Analysis (rows 15–24)
    # ══════════════════════════════════════════════════════════════
    total_rx_analyzed = 0
    rx_with_price     = 0
    rx_no_price       = 0
    rx_overpaid       = 0
    rx_underpaid      = 0
    profit_loss       = 0.0

    if rx_compare_df is not None:
        try:
            total_rx_analyzed = len(rx_compare_df)
            if 'Kinray Unit Price' in rx_compare_df.columns:
                price_s = pd.to_numeric(rx_compare_df['Kinray Unit Price'], errors='coerce').fillna(0)
                rx_with_price = int((price_s > 0).sum())
                rx_no_price   = total_rx_analyzed - rx_with_price
                if 'Difference' in rx_compare_df.columns:
                    mask_p = price_s > 0
                    diffs  = pd.to_numeric(rx_compare_df.loc[mask_p, 'Difference'], errors='coerce').fillna(0)
                    rx_overpaid  = int((diffs > 0).sum())
                    rx_underpaid = int((diffs < 0).sum())
                    profit_loss  = float(diffs.sum())
        except Exception:
            pass
    elif "RX Comparison - All" in wb.sheetnames:
        try:
            ws_rx = wb["RX Comparison - All"]
            rx_hdrs = [ws_rx.cell(row=2, column=c).value for c in range(1, ws_rx.max_column + 1)]
            _price_candidates = [
                'Kinray Final Price',
                'Kinray Price (Pkgs Billed × Unit Price)',
                'Kinray Unit Price',
            ]
            _price_key = next((h for h in _price_candidates if h in rx_hdrs), None)
            _price_idx = rx_hdrs.index(_price_key) + 1 if _price_key else None
            _diff_idx  = rx_hdrs.index('Difference') + 1 if 'Difference' in rx_hdrs else None
            if _price_idx and _diff_idx:
                for r in range(3, ws_rx.max_row + 1):
                    if ws_rx.cell(row=r, column=1).value in (None, ''):
                        continue
                    total_rx_analyzed += 1
                    try:
                        pv = float(ws_rx.cell(row=r, column=_price_idx).value or 0)
                        dv = float(ws_rx.cell(row=r, column=_diff_idx).value or 0)
                    except (TypeError, ValueError):
                        continue
                    if pv > 0:
                        rx_with_price += 1
                        rx_overpaid  += 1 if dv > 0 else 0
                        rx_underpaid += 1 if dv < 0 else 0
                        profit_loss  += dv
                rx_no_price = total_rx_analyzed - rx_with_price
        except Exception:
            pass

    section_header(ws, 15, 1, 8, "RX Comparison Analysis")
    ws.row_dimensions[15].height = 22

    rx_kpi_boxes = [
        ("Total RX Analyzed",      total_rx_analyzed, FILL_BLUE,  TEXT_BLUE,  count_fmt),
        ("Kinray Price Available",  rx_with_price,     FILL_GREEN, TEXT_GREEN, count_fmt),
        ("No Kinray Price",         rx_no_price,       FILL_AMBER, TEXT_AMBER, count_fmt),
    ]
    for i, (label, value, fill, text, fmt) in enumerate(rx_kpi_boxes):
        cs, ce = i * 2 + 1, i * 2 + 2
        ws.merge_cells(start_row=16, start_column=cs, end_row=16, end_column=ce)
        filled_cell(ws, 16, cs, label, fill, TEXT_MUTED, size=9)
        ws.merge_cells(start_row=17, start_column=cs, end_row=17, end_column=ce)
        vc = filled_cell(ws, 17, cs, f"{value:,} RX", fill, text, bold=True, size=14)
        vc.number_format = '@'
    ws.row_dimensions[16].height = 18
    ws.row_dimensions[17].height = 28

    # Row 18: divider
    for _c in range(1, 9):
        ws.cell(row=18, column=_c).border = Border(bottom=Side(style='thin'))
    ws.row_dimensions[18].height = 6

    # Row 19: "Out of X RX..." label
    ws.merge_cells('A19:H19')
    lc = ws.cell(row=19, column=1, value=f"Out of {rx_with_price:,} RX with Kinray price:")
    lc.font = Font(color=TEXT_MUTED, size=10)
    lc.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[19].height = 16

    # Rows 20–21: overpaid / underpaid (4 cols each)
    op_cards = [
        ("Overpaid (Insurance > Kinray)",  rx_overpaid,  FILL_GREEN, TEXT_GREEN),
        ("Underpaid (Insurance < Kinray)", rx_underpaid, FILL_RED,   TEXT_RED),
    ]
    for i, (label, value, fill, text) in enumerate(op_cards):
        cs, ce = i * 4 + 1, i * 4 + 4
        ws.merge_cells(start_row=20, start_column=cs, end_row=20, end_column=ce)
        filled_cell(ws, 20, cs, label, fill, TEXT_MUTED, size=9)
        ws.merge_cells(start_row=21, start_column=cs, end_row=21, end_column=ce)
        vc = filled_cell(ws, 21, cs, f"{value:,} RX", fill, text, bold=True, size=14)
        vc.number_format = '@'
    ws.row_dimensions[20].height = 18
    ws.row_dimensions[21].height = 28

    # Row 22: profit / loss banner
    pl_text  = (f"Profit / Loss of {rx_with_price:,} RX "
                f"(Kinray price available) = ${profit_loss:,.2f}")
    pl_color = TEXT_GREEN if profit_loss >= 0 else TEXT_RED
    ws.merge_cells('A22:H22')
    plc = ws.cell(row=22, column=1, value=pl_text)
    plc.fill      = PatternFill("solid", fgColor=FILL_BLUE)
    plc.font      = Font(bold=True, color=pl_color, size=12)
    plc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[22].height = 24

    ws.row_dimensions[23].height = 4
    ws.row_dimensions[24].height = 4

    # Row 25: blank separator
    ws.row_dimensions[25].height = 10

    # ══════════════════════════════════════════════════════════════
    # SECTION 4 — Purchased But Never Billed (rows 26–35+)
    # ══════════════════════════════════════════════════════════════
    never_billed_brand_df   = pd.DataFrame()
    never_billed_generic_df = pd.DataFrame()
    brand_count   = 0
    generic_count = 0

    if kinray_df is not None and log_df is not None:
        try:
            kdf = kinray_df.copy()
            kdf['NDC_norm'] = (
                kdf['NDC #'].astype(str)
                .str.replace(r'\D', '', regex=True)
                .str.zfill(11)
            )
            billed_ndcs = set(
                log_df['NDC #'].astype(str)
                .str.replace(r'\D', '', regex=True)
                .str.zfill(11)
                .unique()
            )

            _type_col  = 'TYPE'  if 'TYPE'  in kdf.columns else None
            _price_col = 'PRICE' if 'PRICE' in kdf.columns else (
                         'Invoice $' if 'Invoice $' in kdf.columns else None)
            _ship_col  = 'Shipped' if 'Shipped' in kdf.columns else None

            if _price_col and _ship_col:
                agg_kwargs = {
                    'QtyPurchased': (_ship_col,  'sum'),
                    'TotalCost':    (_price_col, 'sum'),
                }
                if _type_col:
                    agg_kwargs['TYPE'] = (_type_col, 'first')
                kdf_agg = kdf.groupby('NDC_norm', as_index=False).agg(**agg_kwargs)

                never_mask = ~kdf_agg['NDC_norm'].isin(billed_ndcs)

                # Filter out invalid NDCs (all zeros = null/blank NDC)
                valid_ndc_mask = kdf_agg['NDC_norm'].str.replace('0', '').str.len() > 0

                if 'TYPE' in kdf_agg.columns:
                    type_s       = kdf_agg['TYPE'].astype(str).str.upper()
                    brand_mask   = type_s.str.contains(r'BRAND|^B$|^BR$',     na=False, regex=True)
                    generic_mask = type_s.str.contains(r'GENERIC|^G$|^GEN$',  na=False, regex=True)
                else:
                    brand_mask   = pd.Series(False, index=kdf_agg.index)
                    generic_mask = pd.Series(False, index=kdf_agg.index)

                never_billed_brand_df   = kdf_agg[brand_mask   & never_mask & valid_ndc_mask].copy()
                never_billed_generic_df = kdf_agg[generic_mask & never_mask & valid_ndc_mask].copy()

                # Enrich brand table with drug names and pkg sizes where possible
                drug_name_map = {}
                # First try log_df Drug Name
                if 'Drug Name' in log_df.columns:
                    drug_name_map = dict(zip(
                        log_df['NDC #'].astype(str).str.replace(r'\D', '', regex=True).str.zfill(11),
                        log_df['Drug Name'].astype(str)
                    ))
                # Also build from Kinray Description column as fallback
                kinray_name_map = {}
                _desc_col = next((c for c in kdf.columns if c.strip().lower() == 'description'), None)
                if _desc_col:
                    kinray_name_map = dict(zip(
                        kdf['NDC_norm'].astype(str),
                        kdf[_desc_col].astype(str)
                    ))
                pkg_size_map = {}
                if final_data is not None and 'NDC #' in final_data.columns and 'Package Size' in final_data.columns:
                    pkg_size_map = dict(zip(
                        final_data['NDC #'].astype(str).str.replace(r'\D', '', regex=True).str.zfill(11),
                        pd.to_numeric(final_data['Package Size'], errors='coerce')
                    ))
                # Also build from Kinray Size column as fallback
                kinray_size_map = {}
                _size_col = next((c for c in kdf.columns if c.strip().lower() == 'size'), None)
                if _size_col:
                    kinray_size_map = dict(zip(
                        kdf['NDC_norm'].astype(str),
                        pd.to_numeric(kdf[_size_col], errors='coerce')
                    ))

                never_billed_brand_df['Drug Name'] = (
                    never_billed_brand_df['NDC_norm']
                    .map(drug_name_map)
                    .fillna(never_billed_brand_df['NDC_norm'].map(kinray_name_map))
                    .fillna('')
                )
                never_billed_brand_df['Pkg Size'] = (
                    never_billed_brand_df['NDC_norm']
                    .map(pkg_size_map)
                    .fillna(never_billed_brand_df['NDC_norm'].map(kinray_size_map))
                )
                never_billed_brand_df['NDC'] = never_billed_brand_df['NDC_norm']

                brand_count   = len(never_billed_brand_df)
                generic_count = len(never_billed_generic_df)
        except Exception as e:
            print(f"[Summary] Section 4 computation failed: {e}")

    section_header(ws, 26, 1, 8, "Purchased But Never Billed")
    ws.row_dimensions[26].height = 22

    nb_cards = [
        ("Brand Drugs — Never Billed",   brand_count,   FILL_PURPLE, TEXT_PURPLE),
        ("Generic Drugs — Never Billed", generic_count, FILL_GREEN,  TEXT_GREEN),
    ]
    for i, (label, value, fill, text) in enumerate(nb_cards):
        cs, ce = i * 4 + 1, i * 4 + 4
        ws.merge_cells(start_row=27, start_column=cs, end_row=27, end_column=ce)
        filled_cell(ws, 27, cs, label, fill, TEXT_MUTED, size=9)
        ws.merge_cells(start_row=28, start_column=cs, end_row=28, end_column=ce)
        vc = filled_cell(ws, 28, cs, value, fill, text, bold=True, size=14)
        vc.number_format = count_fmt
    ws.row_dimensions[27].height = 18
    ws.row_dimensions[28].height = 28

    # Row 29: blank
    ws.row_dimensions[29].height = 8

    # Row 30: sub-header for brand detail table
    ws.merge_cells('A30:H30')
    sh_c = ws.cell(row=30, column=1,
                   value="Brand Drugs Purchased But Never Billed — Detail")
    sh_c.fill      = PatternFill("solid", fgColor=FILL_PURPLE)
    sh_c.font      = Font(bold=True, color=TEXT_PURPLE, size=10)
    sh_c.alignment = Alignment(horizontal='left', vertical='center')
    sh_c.border    = Border(bottom=Side(style='thin'))
    ws.row_dimensions[30].height = 18

    # Row 31: brand table column headers
    brand_table_cols = ['Drug Name', 'NDC', 'Pkg Size', 'Qty Purchased', 'Total Cost', 'Type']
    for ci, col_name in enumerate(brand_table_cols, start=1):
        c = ws.cell(row=31, column=ci, value=col_name)
        c.fill      = PatternFill("solid", fgColor=FILL_PURPLE)
        c.font      = Font(bold=True, color=TEXT_PURPLE, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[31].height = 18

    # Brand drug data rows
    current_row    = 32
    alt_row_fills  = ["FFFFFF", "F8F8F8"]
    _type_in_brand = 'TYPE' in never_billed_brand_df.columns

    if not never_billed_brand_df.empty:
        for row_idx, (_, row) in enumerate(never_billed_brand_df.iterrows()):
            fill_hex    = alt_row_fills[row_idx % 2]
            drug_name   = str(row.get('Drug Name', ''))
            ndc         = str(row.get('NDC', row.get('NDC_norm', '')))
            pkg_size    = row.get('Pkg Size', '')
            qty_purch   = float(row.get('QtyPurchased', 0) or 0)
            total_cost  = float(row.get('TotalCost', 0) or 0)
            drug_type   = str(row.get('TYPE', '')) if _type_in_brand else ''

            row_vals = [drug_name, ndc, pkg_size, qty_purch, total_cost, drug_type]
            for ci, val in enumerate(row_vals, start=1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.fill      = PatternFill("solid", fgColor=fill_hex)
                c.font      = Font(size=9)
                c.alignment = Alignment(
                    horizontal='left' if ci == 1 else 'center',
                    vertical='center',
                    wrap_text=(ci == 1)
                )
                if ci == 4:
                    c.number_format = '#,##0'
                elif ci == 5:
                    c.number_format = money_fmt
            ws.row_dimensions[current_row].height = 15
            current_row += 1
    else:
        ws.cell(row=current_row, column=1,
                value="No brand drugs purchased without billing found")
        ws.row_dimensions[current_row].height = 15
        current_row += 1

    # Brand table specific column widths (applied after general widths)
    _brand_col_widths = {1: 45, 2: 16, 3: 10, 4: 12, 5: 14, 6: 10}
    for ci, w in _brand_col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Blank row after section 4
    ws.row_dimensions[current_row].height = 10
    current_row += 1

    # ══════════════════════════════════════════════════════════════
    # SECTION 5 — Processor Breakdown (flipped: processors as rows)
    # ══════════════════════════════════════════════════════════════
    if processors is None and final_data is not None:
        _procs_found = set()
        for _col in final_data.columns:
            for _sfx in ('_T', '_Pur', '_Net'):
                if _col.endswith(_sfx):
                    _procs_found.add(_col[:-len(_sfx)])
        processors = sorted(_procs_found)
        if 'ALL_PBM' in processors:
            processors = ['ALL_PBM'] + sorted(p for p in processors if p != 'ALL_PBM')

    section_header(ws, current_row, 1, 5, "Processor Breakdown")
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # Table header row
    proc_headers = [
        'Processor',
        'Insurance $ (BestRx)',
        '100% Purchased (Kinray)',
        'Net (Paid−Purchased)',
        'Needs to Order Est. ($)',
    ]
    _thick_b = Border(bottom=Side(style='thick'))
    _thick_t = Border(top=Side(style='thick'))
    for ci, h in enumerate(proc_headers, start=1):
        c = ws.cell(row=current_row, column=ci, value=h)
        c.fill      = PatternFill("solid", fgColor="0F4C81")
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _thick_b
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    totals = [0.0, 0.0, 0.0, 0.0]
    alt_proc_fills = ["FFFFFF", "F0F6FF"]

    for p_idx, proc in enumerate(processors or []):
        fill_hex = "0F4C81" if proc == 'ALL_PBM' else alt_proc_fills[p_idx % 2]
        font_color = "FFFFFF" if proc == 'ALL_PBM' else TEXT_DARK
        is_allpbm = proc == 'ALL_PBM'
        ins = pur = net = needs_est = 0.0
        if final_data is not None:
            try:
                t_col   = f'{proc}_T'
                pur_col = f'{proc}_Pur'
                net_col = f'{proc}_Net'
                d_col   = f'{proc}_D'
                if t_col   in final_data.columns:
                    ins = float(pd.to_numeric(final_data[t_col],   errors='coerce').fillna(0).sum())
                if pur_col in final_data.columns:
                    pur = float(pd.to_numeric(final_data[pur_col], errors='coerce').fillna(0).sum())
                if net_col in final_data.columns:
                    net = float(pd.to_numeric(final_data[net_col], errors='coerce').fillna(0).sum())
                if d_col in final_data.columns and 'Kinray_UPrice' in final_data.columns:
                    d_s   = pd.to_numeric(final_data[d_col],           errors='coerce').fillna(0)
                    kp    = pd.to_numeric(final_data['Kinray_UPrice'], errors='coerce').fillna(0)
                    needs_est = float((d_s.clip(upper=0).abs() * kp).sum())
            except Exception:
                pass

        row_vals = [proc, ins, pur, net, needs_est]
        for ci, val in enumerate(row_vals, start=1):
            c = ws.cell(row=current_row, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=fill_hex)
            c.font      = Font(size=10, bold=is_allpbm, color=font_color)
            c.alignment = Alignment(
                horizontal='left' if ci == 1 else 'center',
                vertical='center'
            )
            if ci > 1:
                c.number_format = money_fmt
        ws.row_dimensions[current_row].height = 18
        # Exclude ALL_PBM from totals — it's the aggregate of all processors
        if proc != 'ALL_PBM':
            totals[0] += ins
            totals[1] += pur
            totals[2] += net
            totals[3] += needs_est
        current_row += 1

    # Total row
    total_row_vals = ['TOTAL'] + totals
    for ci, val in enumerate(total_row_vals, start=1):
        c = ws.cell(row=current_row, column=ci, value=val)
        c.fill      = PatternFill("solid", fgColor=FILL_BLUE)
        c.font      = Font(bold=True, color=TEXT_BLUE, size=10)
        c.alignment = Alignment(
            horizontal='left' if ci == 1 else 'center',
            vertical='center'
        )
        c.border = _thick_t
        if ci > 1:
            c.number_format = money_fmt
    ws.row_dimensions[current_row].height = 22

    # Final column widths — processor table overrides B-E to 22
    ws.column_dimensions['A'].width = 45   # brand table drug name
    for _cl in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[_cl].width = 22
    for _cl in ['F', 'G', 'H']:
        ws.column_dimensions[_cl].width = 18
