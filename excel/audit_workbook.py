# Generates the separate per-processor audit workbook via generate_master_audit_workbook().

import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.helpers import unblock_file


def generate_master_audit_workbook(final_df, pharmacy_name, date_range, output_dir):
    """
    ONE workbook with ALL insurances (including ALL_PBM).

    For each insurance/processor PR, creates:
      - Sheet: "{PR} - Top 100 Packages"
      - Sheet: "{PR} - Top 100 Ins Paid"

        Columns (both sheets):
            Rank | Insurance | NDC # | Drug Name | Package Size |
            Qty Billed | Packages Billed | Total Purchased | Difference | Actual $ Paid | Amount to be Paid to Insurance (If Audit)
    """
    os.makedirs(output_dir, exist_ok=True)

    # ---------- Helper: extract processor from column name ----------
    def _proc_from_col(col: str):
        """Extract processor/insurance name from *_Q/_P/_D/_T/_Pur/_Net columns."""
        suffixes = ('_Q', '_P', '_D', '_T', '_Pur', '_Net')
        for sfx in suffixes:
            if col.endswith(sfx):
                return col[:-len(sfx)]
        return None

    # ---- Discover processors from columns (INCLUDING ALL_PBM) ----
    processors = set()
    for c in final_df.columns:
        p = _proc_from_col(c)
        if p:
            processors.add(p)

    processors = sorted(processors)
    if not processors:
        print("[audit] No processor metrics found. Skipping audit sheets.")
        return None

    # Safe filename & sheet names
    safe_pharmacy = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', str(pharmacy_name)).strip()
    safe_range    = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', str(date_range)).strip()
    filename      = f"{safe_pharmacy}_Audit_{safe_range}.xlsx"
    filepath      = os.path.join(output_dir, filename)

    def safe_sheet_name(name: str) -> str:
        r"""
        Clean a string to be a valid Excel sheet name:
        - remove invalid characters: \ / ? * [ ] :
        - truncate to max 31 characters
        """
        cleaned = re.sub(r'[\\/*?:\[\]]', '_', str(name))
        return cleaned[:31]

    # ---------- Shared styling ----------
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color="A9A9A9"),
        right=Side(style='thin', color="A9A9A9"),
        top=Side(style='thin', color="A9A9A9"),
        bottom=Side(style='thin', color="A9A9A9")
    )

    def _apply_header(ws, insurance_name, last_col_index):
        """Row 1: Pharmacy — Date Range — INSURANCE."""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_index)
        cell = ws.cell(row=1, column=1)
        cell.value = f"{pharmacy_name} — {date_range} — {insurance_name}"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=20, bold=True)
        ws.row_dimensions[1].height = 30

    def _build_top_df(df_proc, pr, sort_col, top_n=None):
        """Build the Top N dataframe for one processor, one metric (P or T)."""
        col_q = f"{pr}_Q"
        col_p = f"{pr}_P"
        col_d = f"{pr}_D"
        col_t = f"{pr}_T"

        df = df_proc.copy()

        # Ensure numeric
        for c in [col_q, col_p, col_d, col_t, 'Total Purchased']:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

        if sort_col not in df.columns:
            return pd.DataFrame(columns=[
                "Rank", "Insurance", "NDC #", "Drug Name", "Package Size",
                "Qty Billed", "Packages Billed", "Total Purchased",
                "Difference", "Actual $ Paid", "Amount to be Paid to Insurance (If Audit)",
            ])

        df = df[df[sort_col] > 0].copy()
        if df.empty:
            return pd.DataFrame(columns=[
                "Rank", "Insurance", "NDC #", "Drug Name", "Package Size",
                "Qty Billed", "Packages Billed", "Total Purchased",
                "Difference", "Actual $ Paid", "Amount to be Paid to Insurance (If Audit)",
            ])

        # 🔑 reset index so values line up row-by-row
        df = df.sort_values(sort_col, ascending=False)
        if top_n is not None:
            df = df.head(top_n)
        df = df.reset_index(drop=True)
        out = pd.DataFrame()
        out["Rank"]            = range(1, len(df) + 1)
        out["Insurance"]       = pr
        out["NDC #"]           = df.get("NDC #", "")
        out["Drug Name"]       = df.get("Drug Name", "")
        out["Package Size"]    = df.get("Package Size", "")
        out["Qty Billed"]      = df.get(col_q, 0)
        out["Packages Billed"] = df.get(col_p, 0)
        out["Total Purchased"] = df.get("Total Purchased", 0)
        out["Difference"]      = df.get(col_d, 0)
        out["Actual $ Paid"]   = df.get(col_t, 0)
        out["Amount to be Paid to Insurance (If Audit)"] = 0

        # Clean <NA> → proper numbers / blanks
        num_cols  = ["Rank", "Qty Billed", "Packages Billed",
                 "Total Purchased", "Difference", "Actual $ Paid", "Amount to be Paid to Insurance (If Audit)"]
        text_cols = ["Insurance", "NDC #", "Drug Name", "Package Size"]

        #Rename Qty Billed to Qty Billed to {pr}
        out.rename(columns={"Qty Billed": f"Qty Billed to {pr}"}, inplace=True)
        #Rename Packages Billed to Packages Billed to {pr}
        out.rename(columns={"Packages Billed": f"Packages Billed to {pr}"}, inplace=True)
        #Rename Total Purchased to Total Qty Purchased
        out.rename(columns={"Total Purchased": "Total Qty Purchased"}, inplace=True)
        #Rename Difference to Qty Difference for {pr}
        out.rename(columns={"Difference": f"Qty Difference for {pr}"}, inplace=True)
        #Rename Actual $ Paid to Actual $ Paid by {pr}
        out.rename(columns={"Actual $ Paid": f"Actual $ Paid by {pr} (BestRX)"}, inplace=True)
        for c in num_cols:
            if c in out.columns:
                out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)

        for c in text_cols:
            if c in out.columns:
                out[c] = out[c].astype(str).where(out[c].notna(), "")

        return out

    #def _write_table(ws, df, subtitle=None):
    def _write_table(ws, df, subtitle=None):
        """Write df to ws starting row 3, with optional subtitle on row 2."""
        if df is None or df.empty:
            # Even if empty, give a subtitle (merged row 2) if requested
            if subtitle:
                empty_last_col = len(df.columns) if (df is not None and hasattr(df, 'columns') and len(df.columns) > 0) else 11
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=empty_last_col)
                c2 = ws.cell(row=2, column=1, value=subtitle)
                c2.font = Font(size=15, bold=True)
                c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=3, column=1, value="No data for this criteria.")
            return

        # Make sure there is no pandas <NA> for openpyxl
        df = df.copy()
        df = df.astype(object).where(pd.notna(df), None)

        cols = list(df.columns)
        last_col = len(cols)

        # ---- Row 2: merged subtitle (Rank .. Actual $ Paid) ----
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
            c2 = ws.cell(row=2, column=1, value=subtitle)
            c2.font = Font(size=15, bold=True)   # 🔹 font size 15 & bold
            c2.alignment = Alignment(horizontal='center',
                                     vertical='center',
                                     wrap_text=True)
        ws.row_dimensions[2].height = 24

        # ---- Row 3: headers ----
        for c_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=3, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True   # 🔹 wrap text for header row
            )
            cell.border = thin_border

        # Row 3 height = 35
        ws.row_dimensions[3].height = 65

        # ---- Data from row 4 (positionally) ----
        for r_idx, row_vals in enumerate(df.itertuples(index=False, name=None), start=4):
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                # Default alignment: left for text, right for numbers
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border

        # ---- Column K formula: per-row audit estimate (negative only) ----
        # =IFERROR(IF((ROUND((J4/G4)*I4,1))<0,(ROUND((J4/G4)*I4,1)),0),0)
        if last_col >= 11:
            for r_idx in range(4, ws.max_row + 1):
                k_cell = ws.cell(row=r_idx, column=11)
                k_cell.value = f"=IFERROR(IF((ROUND((J{r_idx}/G{r_idx})*I{r_idx},1))<0,(ROUND((J{r_idx}/G{r_idx})*I{r_idx},1)),0),0)"
                k_cell.border = thin_border

        # ---- Column-specific formatting: E..K ----
        # E: Package Size
        # F: Qty Billed
        # G: Packages Billed
        # H: Total Purchased
        # I: Difference
        # J: Actual $ Paid
        # K: Amount to be Paid to Insurance (If Audit)
        first_data_row = 4
        last_data_row = ws.max_row

        accounting_fmt = '$#,##0.00;[Red]-$#,##0.00'

        for row in ws.iter_rows(min_row=first_data_row, max_row=last_data_row,
                                min_col=5, max_col=min(11, last_col)):
            for cell in row:
                # Center alignment & 2 decimal format
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=(cell.column == 11)
                )
                if isinstance(cell.value, (int, float)):
                    if cell.column in (10, 11):
                        cell.number_format = accounting_fmt
                    else:
                        cell.number_format = "0.00"

        if last_col >= 11:
            for r_idx in range(first_data_row, last_data_row + 1):
                ws.cell(row=r_idx, column=11).number_format = accounting_fmt

        if last_col >= 10:
            for r_idx in range(first_data_row, last_data_row + 1):
                ws.cell(row=r_idx, column=10).number_format = accounting_fmt

        # ---- Conditional formatting: highlight negatives in E..K ----
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        rng_neg = f"E{first_data_row}:{get_column_letter(min(11, last_col))}{last_data_row}"
        ws.conditional_formatting.add(
            rng_neg,
            CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False, fill=red_fill, font=red_font)
        )

        # ---- Auto-sum row at bottom (F..K) ----
        total_row = last_data_row + 1
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=total_row, column=5).border = thin_border
        ws.cell(row=total_row, column=5).fill = total_fill

        for c_idx in range(6, min(11, last_col) + 1):
            col_letter = get_column_letter(c_idx)
            tcell = ws.cell(row=total_row, column=c_idx)
            tcell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            tcell.font = Font(bold=True)
            tcell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(c_idx == 11))
            if c_idx in (10, 11):
                tcell.number_format = accounting_fmt
            else:
                tcell.number_format = "0.00"
            tcell.border = thin_border
            tcell.fill = total_fill

        # ---- Bottom quick metrics ----
        if last_col >= 11:
            audited_count_row = total_row + 1
            exposure_row = total_row + 2

            ws.cell(row=audited_count_row, column=10, value="Audited Drugs Count").font = Font(bold=True)
            ws.cell(row=audited_count_row, column=10).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=audited_count_row, column=10).border = thin_border
            ws.cell(row=audited_count_row, column=11, value=f"=COUNTIF(K{first_data_row}:K{last_data_row},\"<0\")")
            ws.cell(row=audited_count_row, column=11).font = Font(bold=True)
            ws.cell(row=audited_count_row, column=11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=audited_count_row, column=11).border = thin_border

            ws.cell(row=exposure_row, column=10, value="Total Audit Exposure").font = Font(bold=True)
            ws.cell(row=exposure_row, column=10).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=exposure_row, column=10).border = thin_border
            ws.cell(row=exposure_row, column=11, value=f"=SUMIF(K{first_data_row}:K{last_data_row},\"<0\",K{first_data_row}:K{last_data_row})")
            ws.cell(row=exposure_row, column=11).font = Font(bold=True)
            ws.cell(row=exposure_row, column=11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=exposure_row, column=11).number_format = accounting_fmt
            ws.cell(row=exposure_row, column=11).border = thin_border

        # ---- Auto column widths ----
        for col_idx in range(1, last_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=col_idx, max_col=col_idx):
                val = row[0].value
                if val is None:
                    continue
                txt = str(val)
                if len(txt) > max_len:
                    max_len = len(txt)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        ws.column_dimensions['A'].width = 5   # Rank
        ws.column_dimensions['B'].width = 12  # Insurance
        ws.column_dimensions['C'].width = 12  # NDC #
        ws.column_dimensions['D'].width = 35  # Drug Name
        ws.column_dimensions['E'].width = 8  # Package Size
        ws.column_dimensions['F'].width = 12  # Qty Billed
        ws.column_dimensions['G'].width = 12  # Packages Billed
        ws.column_dimensions['H'].width = 12  # Total Purchased
        ws.column_dimensions['I'].width = 12  # Difference
        ws.column_dimensions['J'].width = 15  # Actual $ Paid
        if last_col >= 11:
            ws.column_dimensions['K'].width = 15  # Amount to be Paid to Insurance (If Audit)
        # ---- Freeze panes at row 4 (row 1–3 fixed) ----
        ws.freeze_panes = "E4"

        # ---- Enable filter on header row (row 3) ----
        if last_col >= 11:
            ws.auto_filter.ref = f"A3:K{last_data_row}"
        else:
            ws.auto_filter.ref = f"A3:{get_column_letter(last_col)}{last_data_row}"
         # ---- 📄 Page setup: Landscape + fit all columns on one page width ----
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # or "landscape"
        ws.page_setup.fitToWidth = 1   # fit all columns to one page wide
        ws.page_setup.fitToHeight = 0  # allow multiple pages down
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Optional: tiny margins for better fit
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        # ✅ Print debug info
        #print(f"[audit] Sheet '{ws.title}' written with {len(df)} rows and {last_col} columns (landscape, fit-to-width).")


    # ---------- ONE workbook for all processors ----------
    wb = Workbook()
    wb.remove(wb.active)   # we'll create sheets ourselves

    for pr in processors:
        col_q = f"{pr}_Q"
        col_p = f"{pr}_P"
        col_t = f"{pr}_T"

        # skip if we have neither Q nor T
        if col_q not in final_df.columns and col_t not in final_df.columns:
            continue

        df_proc = final_df.copy()

        top_packages = _build_top_df(df_proc, pr, col_q)
        top_paid     = _build_top_df(df_proc, pr, col_t)

        # If both empty, skip
        if top_packages.empty and top_paid.empty:
            continue

        # # ---- Sheet 1: Top 100 by Packages ----
        # sheet_name_pkg = safe_sheet_name(f"{pr} - Top 100 Qty Billed")
        # ws1 = wb.create_sheet(title=sheet_name_pkg)
        # last_col_index_1 = len(top_packages.columns) if not top_packages.empty else 10
        # _apply_header(ws1, pr, last_col_index_1)
        # _write_table(ws1, top_packages,
        #              subtitle=f"Top 100 drugs for {pr} by Qty Billed")

        # Top 100 by Ins Paid ----
        sheet_name_paid = safe_sheet_name(f"{pr}")
        ws2 = wb.create_sheet(title=sheet_name_paid)
        last_col_index_2 = len(top_paid.columns) if not top_paid.empty else 10
        _apply_header(ws2, pr, last_col_index_2)
        _write_table(ws2, top_paid,
                     subtitle=f"Overall {pr} Overview")

    # Save single master workbook
    wb.save(filepath)
    unblock_file(filepath, "audit warn")

    #print(f"[audit] Master audit workbook created: {filepath}")
    return filepath
