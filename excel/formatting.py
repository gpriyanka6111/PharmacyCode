# Shared Excel formatting helpers: apply_common_sheet_settings, add_autosum_by_processors, set_print_area_excluding, get_column_index, adjust_specific_columns, discover_processors_from_df.

import re

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


def get_column_index(ws, header_name, header_row=None):
    """Find a column by header text. Returns 1-based index or None."""
    rows = [header_row] if header_row else [2, 3]
    for r in rows:
        try:
            for cell in ws[r]:
                if cell.value == header_name:
                    return cell.col_idx
        except Exception:
            continue
    return None


def adjust_specific_columns(ws, col_letters, width=12):
    for col_letter in col_letters:
        if col_letter:
            ws.column_dimensions[col_letter].width = width


def discover_processors_from_df(final_df):
    """
    Derive processor prefixes from columns like <Processor>_{Q,P,D,T,Pur,Net,Diff$}.
    Use this if you don't already have a processors list.
    """
    suffixes = ('_Q', '_P', '_D', '_T', '_Pur', '_Net', '_Net')
    procs = set()
    for c in final_df.columns:
        for s in suffixes:
            if c.endswith(s):
                procs.add(c[:-len(s)])
                break
    return sorted(procs)


def add_autosum_by_processors(ws, processors, start_row, end_row, header_row=3):
    """
    Write SUM formulas for each processor's *_T, *_Pur, *_Net into the row after end_row,
    and format totals in currency.
    """
    if not processors:
        return

    total_row = end_row + 1
    # Write "Totals" label if blank
    if ws.cell(row=total_row, column=1).value in (None, ""):
        ws.cell(row=total_row, column=1, value="Totals").font = Font(bold=True)

    for pr in processors:
        for suf in ("_T", "_Pur", "_Net"):
            hdr = f"{pr}{suf}"
            col_idx = get_column_index(ws, hdr, header_row=header_row)
            if not col_idx:
                continue

            col_letter = get_column_letter(col_idx)
            cell = ws.cell(
                row=total_row,
                column=col_idx,
                value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            )
            # Format in currency and bold
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Optional: highlight the total row visually
    for c in range(1, ws.max_column + 1):
        ws.cell(row=total_row, column=c).border = Border(
            top=Side(style="thick"))


def apply_common_sheet_settings(
    wb,
    pharmacy_name: str,
    date_range: str,
    processors: list[str] | None = None,
    header_row_main: int = 3,
):
    """
    Apply titles, orientations, print settings to known sheets.
    If processors are provided (or discoverable), run autosum & width tweaks on 'Processed Data'.
    """
    RULES = {
        "Processed Data": {
            "title": "{pharmacy} ({range})",
            "font_size": 35,
            "min_row_for_height": header_row_main + 1,  # data starts after header
            "orientation": "landscape",
            "header_row": header_row_main,
        },
        "Needs to be Ordered": {
            "title": "{pharmacy} ({range}) - NTO CVS",
            "font_size": 25,
            "min_row_for_height": 2,
            "orientation": "landscape",
        },
        "Missing Items": {
            "title": "{pharmacy} ({range}) - Missing items, To be updated in master file",
            "font_size": 15,
            "min_row_for_height": 2,
            "orientation": "landscape",
        },
        "Do Not Order CVS": {
            "title": "{pharmacy} ({range}) - DNO CVS",
            "font_size": 25,
            "min_row_for_height": 2,
            "orientation": "landscape",
        },
        "Needs to be ordered - All": {
            "title": "{pharmacy} ({range}) - Need to Order - ALL",
            "font_size": 25,
            "min_row_for_height": 3,
            "orientation": "landscape",
        },
        "Do Not Order - ALL": {
            "title": "{pharmacy} ({range}) - Do Not Order",
            "font_size": 25,
            "min_row_for_height": 3,
            "orientation": "portrait",
        },
        "Never Ordered - Check": {
            "title": "{pharmacy} ({range}) - Never Ordered Package - Check",
            "font_size": 25,
            "min_row_for_height": 3,
            "orientation": "landscape",
        },
        "Never Ordered  - Check": {  # tolerate double-space variant
            "title": "{pharmacy} ({range}) - Never Ordered Package - Check",
            "font_size": 25,
            "min_row_for_height": 3,
            "orientation": "landscape",
        },
    }

    # Apply per-sheet render + common print settings
    for sheet in wb.worksheets:
        name = sheet.title
        cfg = RULES.get(name)
        if cfg:
            # Title in A1
            t = cfg["title"].format(pharmacy=pharmacy_name, range=date_range)
            c = sheet.cell(row=1, column=1)
            c.value = t
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=cfg["font_size"], bold=True)

            # Row heights
            min_row = cfg.get("min_row_for_height", 2)
            for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20

            # Orientation
            sheet.page_setup.orientation = cfg.get("orientation", "landscape")

        # Common print config
        sheet.print_title_rows = "1:2"
        sheet.oddFooter.left.text = "Page &P of &N"
        sheet.oddFooter.left.size = 8
        sheet.oddFooter.left.font = "Arial,Bold"
        sheet.page_margins = PageMargins(
            left=0, right=0, top=0, bottom=0, header=0, footer=0.1)
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_options.horizontalCentered = True
        sheet.print_options.verticalCentered = True
        sheet.print_options.gridLines = True

        # Extra on main sheet: autosum + width for processor T/Pur/Diff$
        if name == "Processed Data":
            header_row = RULES[name]["header_row"]
            start_row = header_row + 1
            end_row = sheet.max_row
            if end_row >= start_row:
                # find processors if not provided
                procs = processors
                if not procs:
                    # try to discover from headers on the sheet
                    # (if you still have the DataFrame, prefer discover_processors_from_df(final))
                    # Here, we scan row `header_row` for names that look like <Proc>_<suffix>
                    import re
                    suffixes = ("_T", "_Pur", "_Net")
                    procs = set()
                    for cell in sheet[header_row]:
                        val = str(cell.value or "")
                        for s in suffixes:
                            if val.endswith(s):
                                procs.add(val[:-len(s)])
                    procs = sorted(procs)

                # autosum
                add_autosum_by_processors(
                    sheet, procs, start_row, end_row, header_row=header_row)

                # width tweaks for *_T, *_Pur, *_Net
                cols_to_adjust = []
                for pr in procs:
                    for suf in ("_T", "_Pur", "_Net"):
                        idx = get_column_index(
                            sheet, f"{pr}{suf}", header_row=header_row)
                        if idx:
                            cols_to_adjust.append(get_column_letter(idx))
                if cols_to_adjust:
                    adjust_specific_columns(sheet, cols_to_adjust, width=12)


def set_print_area_excluding(wb, sheet_name, headers_to_exclude):
    """
    Define the Excel print area for a sheet, excluding specific columns by header name.
    This does NOT hide them in the workbook — it only removes them from the print page.
    """
    ws = wb[sheet_name]
    header_row = 1  # change if your headers are on another row

    # Map each column index to its header text
    headers = {
        c: (ws.cell(row=header_row, column=c).value or '').strip()
        for c in range(1, ws.max_column + 1)
    }

    # Identify which columns to skip
    exclude_cols = {
        c for c, name in headers.items()
        if name.strip().lower() in {h.lower() for h in headers_to_exclude}
    }

    # Build contiguous column ranges to print
    ranges = []
    start = None
    for c in range(1, ws.max_column + 1):
        if c in exclude_cols:
            if start is not None:
                ranges.append((start, c - 1))
                start = None
        else:
            if start is None:
                start = c
    if start is not None:
        ranges.append((start, ws.max_column))

    # Convert to A1 ranges spanning all rows
    last_row = ws.max_row
    a1_ranges = [
        f"{get_column_letter(a)}1:{get_column_letter(b)}{last_row}"
        for a, b in ranges if a <= b
    ]

    # ✅ Assign the print area (multiple blocks allowed)
    ws.print_area = a1_ranges


def set_print_area_excluding_headers(ws, header_row=2, exclude_headers=()):
    # Build a case-insensitive set of headers to exclude
    bad = {str(h).strip().casefold() for h in exclude_headers}

    # Read headers on header_row
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        headers.append(((v if v is not None else "").strip(), c))

    # Keep columns not in exclude list
    include_cols = [c for (name, c) in headers if name.casefold() not in bad]
    if not include_cols:
        return  # nothing to include

    # Build contiguous column blocks -> "A2:C{max_row},E2:G{max_row},..."
    maxr = ws.max_row
    blocks = []
    run_start = run_prev = include_cols[0]
    for c in include_cols[1:] + [None]:
        if c is None or c != run_prev + 1:
            # close current run
            left = get_column_letter(run_start)
            right = get_column_letter(run_prev)
            blocks.append(f"{left}{header_row}:{right}{maxr}")
            # start new
            if c is not None:
                run_start = c
        run_prev = c if c is not None else run_prev

    # Assign comma-separated print area
    ws.print_area = ",".join(blocks)
