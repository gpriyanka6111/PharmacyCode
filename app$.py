
<<<<<<< HEAD
=======
License:
This script is the intellectual property of Priyanka Gulgari.
Unauthorized copying, distribution, modification, or use of this code, via any medium, is strictly prohibited without prior written permission from the author.

Contact:
For permissions or inquiries, please contact priyankagulgari@gmail.com .

==================================================
"""
# Standard library imports
import csv
import json
import mimetypes
import os
import re
import shutil
import smtplib
import sys
import tempfile
import threading
import urllib.parse
import webbrowser
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from glob import glob
from urllib.parse import quote
from uuid import uuid4

# Third-party imports
import numpy as np
import pandas as pd
from flask import (Flask, current_app, jsonify, make_response, redirect,
                   render_template, request, send_file, send_from_directory,
                   url_for)
from flaskwebgui import FlaskUI
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break, PageBreak
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename
import tkinter as tk
from tkinter import filedialog

from config import APP_DISPLAY_NAME, BASE_DIR, UPLOAD_FOLDER, PROCESSED_FOLDER
from utils.helpers import resource_path, get_screen_dimensions, unblock_file
from processing.all_pbm_parser import _load_all_pbm_csv
from processing.log_parser import _filter_custom_log_transmitted_paid_ins
from processing.vendor_parser import parse_vendor_files
from excel.formatting import (get_column_index, discover_processors_from_df,
                               apply_common_sheet_settings, set_print_area_excluding_headers)
from excel.order_sheets import add_max_difference_sheet, min_difference_sheet
from excel.support_sheets import create_never_ordered_check_sheet, create_bin_to_processor_sheet
from excel.rx_comparison_sheets import (add_rx_unit_compare_sheet_exact,
                                        add_rx_unit_compare_sheet_exact_pos,
                                        add_mfp_drugs_sheet)
from excel.refill_sheets import add_zero_refills_sheet
from excel.summary_sheet import add_summary_sheet
from excel.audit_workbook import generate_master_audit_workbook
from excel.processed_data_sheet import build_processed_data_sheet
from processing.pipeline import process_custom_log_data

# Initialize Tkinter window for screen dimensions
screen_width, screen_height = get_screen_dimensions()


app = Flask(__name__, template_folder='templates')
app.static_folder = 'static'

# window = webview.create_window('Pharmacy Data Processing Application',app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(PROCESSED_FOLDER):
    os.makedirs(PROCESSED_FOLDER)


from routes.main import bp as main_bp
app.register_blueprint(main_bp)


@app.errorhandler(404)
def not_found(_e):
    return render_template('error.html',
        title='Page Not Found',
        message='The page you are looking for does not exist.',
        hint='Use the button below to go back to the home page.',
        back_url='/',
        back_label='Go to Home'
    ), 404


@app.errorhandler(500)
def server_error(_e):
    return render_template('error.html',
        title='Something Went Wrong',
        message='An unexpected error occurred.',
        hint='Please restart the app and try again. If the problem persists, re-upload your CSVs.',
        back_url='/',
        back_label='Go to Home'
    ), 500
