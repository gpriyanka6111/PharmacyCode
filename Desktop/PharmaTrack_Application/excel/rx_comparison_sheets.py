# Builds "RX Comparison - All", "RX Comparison +ve", and "MFP Drugs - RX" per-prescription analysis sheets.

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from processing.kinray_pricing import find_kinray_price_by_month


def add_rx_unit_compare_sheet_exact(
    wb,
    log_df,
    kinray_df,
    sheet_name: str = "RX Comparison - All"
):
    """
    Output columns (exact order):
      Rx, NDC, Drug Name, Fill date, Qty filled, Package billed,
      Kinray Unit Price, Ins paid, Unit Ins paid, Difference

    ✅ Shows ONLY rows where Difference < 0 (underpaid RXs)
    ✅ Sorted by Fill Date descending (latest first)
    """

    df = log_df.copy()
    #print(df.head())
    if '* SDRA Amt' in df.columns and 'SDRA Amt' not in df.columns:
        df.rename(columns={'* SDRA Amt': 'SDRA Amt'}, inplace=True)
    elif 'SDRA' in df.columns and 'SDRA Amt' not in df.columns:
        df.rename(columns={'SDRA': 'SDRA Amt'}, inplace=True)
    if 'Copay' in df.columns and 'COPAY' not in df.columns:
        df.rename(columns={'Copay': 'COPAY'}, inplace=True)

    # --- Normalize numeric columns ---
    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'Plan 1 BIN',
              'Plan 2 BIN', 'SDRA Amt', 'COPAY']:
        df[c] = pd.to_numeric(df.get(c, 0), errors='coerce').fillna(0)

    # Normalize NDC
    df['NDC #'] = (df['NDC #'].astype(str)
                   .str.replace('-', '', regex=False)
                   .str.replace(r'\D', '', regex=True)
                   .str.zfill(11))

    # Detect Fill Date column first (needed for price lookup)
    date_candidates = ['Fill Date', 'Date',
                       'Rx Date', 'Dispense Date', 'Service Date']
    fill_date_col = next((c for c in date_candidates if c in df.columns), None)
    if fill_date_col:
        df['Fill Date'] = pd.to_datetime(df[fill_date_col], errors='coerce')
    else:
        df['Fill Date'] = pd.NaT

    # Apply month-based Kinray price lookup
    df['Kinray Unit Price'] = df.apply(
        lambda row: find_kinray_price_by_month(row['NDC #'], row['Fill Date'], kinray_df),
        axis=1
    )

    # --- Winning insurance paid ---
    df['Ins paid'] = np.where(
        df['Ins Paid Plan 1'].fillna(0) >= df['Ins Paid Plan 2'].fillna(0),
        df['Ins Paid Plan 1'].fillna(0),
        df['Ins Paid Plan 2'].fillna(0)
    )

    # --- Package billed ---
    df['Package billed'] = np.where(
        df['Drug Pkg Size'] > 0,
        df['Qty Filled'] / df['Drug Pkg Size'],
        np.nan
    )

    df['Kinray final Price'] = np.where(
        (df['Drug Pkg Size'] > 0) & (df['Kinray Unit Price'] > 0),
        (df['Kinray Unit Price']/df['Drug Pkg Size']) * df['Qty Filled'],
        0.0
    )

    # --- Unit insurance paid (per package logic) ---
    df['Unit Ins paid'] = np.where(
        df['Package billed'] > 0,
        df['Ins paid'] / df['Package billed'],
        np.nan
    )

    # Total paid (Insurance + SDRA + Copay)
    df['Total Ins paid'] = df['Ins paid'] + df['SDRA Amt'] + df['COPAY']

    # Difference = Total Ins paid - Kinray final Price
    # If Kinray Unit Price is 0, force Difference = 0
    df['Difference'] = np.where(
        df['Kinray Unit Price'] > 0,
        df['Total Ins paid'] - df['Kinray final Price'],
        0.0
    )

    # Drop rows where Difference is positive or 0
    # df = df[df['Difference'] > 0]

    # Map Rx column
    rx_col = 'Rx #' if 'Rx #' in df.columns else (
        'Rx' if 'Rx' in df.columns else None)
    df['RX'] = df[rx_col] if rx_col else pd.NA
    df['NDC'] = df['NDC #']
    df['Drug Name'] = df['Drug Name']
    df['Pkg Size'] = df['Drug Pkg Size']
    df['Qty Filled'] = df['Qty Filled']
    df['BIN'] = df['Winning_BIN']
    df['Processor'] = df['Processor']
    df['PCN'] = df['Winning PCN']
    df['Group'] = df['Winning Group']
    df['Fill Date'] = df['Fill Date']
    df['Kinray Final Price'] = df['Kinray final Price']
    df['Ins Paid'] = df['Ins paid']
    df['SDRA Amt'] = df['SDRA Amt']
    df['COPAY'] = df['COPAY']
    df['Total = (Ins Paid + SDRA + COPAY)'] = df['Total Ins paid']
    df['Package Billed'] = df['Package billed']
    out_cols = [
        'RX', 'Fill Date', 'NDC', 'Drug Name', 'Pkg Size', 'Qty Filled',
        'Package Billed', 'Kinray Final Price', 'Ins Paid', 'SDRA Amt', 'COPAY', 'Total = (Ins Paid + SDRA + COPAY)', 'Difference', 'BIN', 'Processor',
        'PCN', 'Group'
    ]

    # Filter, then sort by Fill Date DESCENDING (latest first)
    out = df.loc[:, out_cols].copy()
    # latest first, then largest diff
    out = out.sort_values('Drug Name', ascending=True)

    # # If no underpaid rows, create placeholder sheet
    # if out.empty:
    #     if sheet_name in wb.sheetnames:
    #         del wb[sheet_name]
    #     ws = wb.create_sheet(title=sheet_name)
    #     ws['A1'] = "No underpaid RXs found (Difference ≥ 0)."
    #     return

    # --- Create Sheet ---
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(out_cols))
    t = ws.cell(row=1, column=1, value="RX Comparision Analysis (All RXs)")
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 26

    # Write table
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
            else:
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

    # ✅ Wrap specific headers
    for cell_ref in ["E2", "F2", "G2", "L2"]:
        ws[cell_ref].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

    # Borders & formatting
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(out_cols)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 9, 'NDC': 14, 'Drug Name': 45, 'Pkg Size': 8, 'Fill Date': 12,
        'Qty Filled': 8, 'Package Billed': 9, 'Kinray Final Price': 16,
        'Ins Paid': 14, 'SDRA Amt': 12, 'COPAY': 10, 'Total = (Ins Paid + SDRA + COPAY)': 24,
        'Difference': 14, 'BIN': 8,
        'PCN': 12, 'Group': 12, 'Processor': 15
    }
    ws.row_dimensions[2].height = 50

    for i, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    # Number formats
    for r in range(3, ws.max_row + 1):
        for name in ['Kinray Final Price', 'Ins Paid', 'SDRA Amt', 'COPAY', 'Total = (Ins Paid + SDRA + COPAY)', 'Difference']:
            idx = out_cols.index(name) + 1
            ws.cell(row=r, column=idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=out_cols.index(
            'Qty Filled') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index(
            'Package Billed') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index(
            'Fill Date') + 1).number_format = 'yyyy-mm-dd'

    diff_idx = out_cols.index('Difference') + 1
    last_data_row = ws.max_row
    total_row = last_data_row + 1

    # Label cell (optional)
    label_col = diff_idx - 1
    label_cell = ws.cell(row=total_row, column=label_col,
                         value="Total Difference")
    label_cell.font = Font(bold=True, size=12)
    label_cell.alignment = Alignment(horizontal='center', vertical='center')
    drug_idx = out_cols.index('Drug Name') + 1
    diff_idx = out_cols.index('Difference') + 1
    left_idx = min(drug_idx, diff_idx)
    right_idx = max(drug_idx, diff_idx)

    left_col = get_column_letter(left_idx)
    right_col = get_column_letter(right_idx)

    # Apply number format to Total Difference cell
    total_diff_cell = ws.cell(row=total_row, column=diff_idx)
    total_diff_cell.number_format = '"$"#,##0.00'
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{last_data_row}"

    # AutoSum cell
    sum_col_letter = get_column_letter(diff_idx)
    total_cell = ws.cell(row=total_row, column=diff_idx)
    # total_cell.value = f"=SUM({sum_col_letter}3:{sum_col_letter}{last_data_row})"
    total_cell.value = f"=SUBTOTAL(109,{sum_col_letter}3:{sum_col_letter}{last_data_row})"
    total_cell.number_format = '"$"#,##0.00'  # ✅ Currency format

    total_cell.font = Font(bold=True, size=12)
    total_cell.number_format = 'General'
    total_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ensure Excel recalculates when opening
    ws.parent.calculation.fullCalcOnLoad = True
    # Freeze panes
    ws.freeze_panes = "A3"
    diff_col_letter = get_column_letter(out_cols.index("Difference") + 1)
    data_range = f"{diff_col_letter}3:{diff_col_letter}{ws.max_row}"

    # 🔴 Negative values = RED FILL
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='lessThan', formula=['0'],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )

    # 🟢 Positive values = GREEN FILL
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    # Set page orientation to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_rx_unit_compare_sheet_exact_pos(
    wb,
    log_df,
    kinray_df,
    sheet_name: str = "RX Comparison +ve"
):
    """
    Output columns (exact order):
      Rx, NDC, Drug Name, Fill date, Qty filled, Package billed,
      Kinray Unit Price, Ins paid, Unit Ins paid, Difference

    ✅ Shows ONLY rows where Difference < 0 (underpaid RXs)
    ✅ Sorted by Fill Date descending (latest first)
    """
    df = log_df.copy()
    # print(df.head())
    if '* SDRA Amt' in df.columns and 'SDRA Amt' not in df.columns:
        df.rename(columns={'* SDRA Amt': 'SDRA Amt'}, inplace=True)
    elif 'SDRA' in df.columns and 'SDRA Amt' not in df.columns:
        df.rename(columns={'SDRA': 'SDRA Amt'}, inplace=True)
    if 'Copay' in df.columns and 'COPAY' not in df.columns:
        df.rename(columns={'Copay': 'COPAY'}, inplace=True)

    # --- Normalize numeric columns ---
    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'Plan 1 BIN',
              'Plan 2 BIN', 'SDRA Amt', 'COPAY']:
        df[c] = pd.to_numeric(df.get(c, 0), errors='coerce').fillna(0)

    # Normalize NDC
    df['NDC #'] = (df['NDC #'].astype(str)
                   .str.replace('-', '', regex=False)
                   .str.replace(r'\D', '', regex=True)
                   .str.zfill(11))

    # Detect Fill Date column first (needed for price lookup)
    date_candidates = ['Fill Date', 'Date',
                       'Rx Date', 'Dispense Date', 'Service Date']
    fill_date_col = next((c for c in date_candidates if c in df.columns), None)
    if fill_date_col:
        df['Fill Date'] = pd.to_datetime(df[fill_date_col], errors='coerce')
    else:
        df['Fill Date'] = pd.NaT

    # Apply month-based Kinray price lookup
    df['Kinray Unit Price'] = df.apply(
        lambda row: find_kinray_price_by_month(row['NDC #'], row['Fill Date'], kinray_df),
        axis=1
    )

    # --- Winning insurance paid ---
    df['Ins paid'] = np.where(
        df['Ins Paid Plan 1'].fillna(0) >= df['Ins Paid Plan 2'].fillna(0),
        df['Ins Paid Plan 1'].fillna(0),
        df['Ins Paid Plan 2'].fillna(0)
    )

    # --- Package billed ---
    df['Package billed'] = np.where(
        df['Drug Pkg Size'] > 0,
        df['Qty Filled'] / df['Drug Pkg Size'],
        np.nan
    )

    df['Kinray final Price'] = np.where(
        (df['Drug Pkg Size'] > 0) & (df['Kinray Unit Price'] > 0),
        (df['Kinray Unit Price']/df['Drug Pkg Size']) * df['Qty Filled'],
        0.0
    )

    # --- Unit insurance paid (per package logic) ---
    df['Unit Ins paid'] = np.where(
        df['Package billed'] > 0,
        df['Ins paid'] / df['Package billed'],
        np.nan
    )

    # Total paid (Insurance + SDRA + Copay)
    df['Total Ins paid'] = df['Ins paid'] + df['SDRA Amt'] + df['COPAY']

    # Difference = Total Ins paid - Kinray final Price
    # If Kinray Unit Price is 0, force Difference = 0
    df['Difference'] = np.where(
        df['Kinray Unit Price'] > 0,
        df['Total Ins paid'] - df['Kinray final Price'],
        0.0
    )

    # Drop rows where Difference is negative or 0
    df = df[df['Difference'] > 0]

    # Map Rx column
    rx_col = 'Rx #' if 'Rx #' in df.columns else (
        'Rx' if 'Rx' in df.columns else None)
    df['RX'] = df[rx_col] if rx_col else pd.NA
    df['NDC'] = df['NDC #']
    df['Drug Name'] = df['Drug Name']
    df['Pkg Size'] = df['Drug Pkg Size']
    df['Qty Filled'] = df['Qty Filled']
    df['BIN'] = df['Winning_BIN']
    df['Processor'] = df['Processor']
    df['PCN'] = df['Winning PCN']
    df['Group'] = df['Winning Group']
    df['Fill Date'] = df['Fill Date']
    df['Kinray Final Price'] = df['Kinray final Price']
    df['Ins Paid'] = df['Ins paid']
    df['SDRA Amt'] = df['SDRA Amt']
    df['COPAY'] = df['COPAY']
    df['Total = (Ins Paid + SDRA + COPAY)'] = df['Total Ins paid']
    df['Package Billed'] = df['Package billed']
    out_cols = [
        'RX', 'Fill Date', 'NDC', 'Drug Name', 'Pkg Size', 'Qty Filled',
        'Package Billed', 'Kinray Final Price', 'Ins Paid', 'SDRA Amt', 'COPAY', 'Total = (Ins Paid + SDRA + COPAY)', 'Difference', 'BIN', 'Processor',
        'PCN', 'Group'
    ]

    # Filter, then sort by Fill Date DESCENDING (latest first)
    out = df.loc[:, out_cols].copy()
    # latest first, then largest diff
    out = out.sort_values('Drug Name', ascending=True)

    # # If no underpaid rows, create placeholder sheet
    # if out.empty:
    #     if sheet_name in wb.sheetnames:
    #         del wb[sheet_name]
    #     ws = wb.create_sheet(title=sheet_name)
    #     ws['A1'] = "No underpaid RXs found (Difference ≥ 0)."
    #     return

    # --- Create Sheet ---
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(out_cols))
    t = ws.cell(row=1, column=1, value="RX Comparision +ve Analysis (All RXs)")
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 26

    # Write table
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
            else:
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

    # ✅ Wrap specific headers
    for cell_ref in ["E2", "F2", "G2", "L2"]:
        ws[cell_ref].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

    # Borders & formatting
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(out_cols)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 9, 'NDC': 14, 'Drug Name': 45, 'Pkg Size': 8, 'Fill Date': 12,
        'Qty Filled': 8, 'Package Billed': 9, 'Kinray Final Price': 16,
        'Ins Paid': 14, 'SDRA Amt': 12, 'COPAY': 10, 'Total = (Ins Paid + SDRA + COPAY)': 24,
        'Difference': 14, 'BIN': 8,
        'PCN': 12, 'Group': 12, 'Processor': 15
    }
    ws.row_dimensions[2].height = 50

    for i, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    # Number formats
    for r in range(3, ws.max_row + 1):
        for name in ['Kinray Final Price', 'Ins Paid', 'SDRA Amt', 'COPAY', 'Total = (Ins Paid + SDRA + COPAY)', 'Difference']:
            idx = out_cols.index(name) + 1
            ws.cell(row=r, column=idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=out_cols.index(
            'Qty Filled') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index(
            'Package Billed') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index(
            'Fill Date') + 1).number_format = 'yyyy-mm-dd'

    diff_idx = out_cols.index('Difference') + 1
    last_data_row = ws.max_row
    total_row = last_data_row + 1

    # Label cell (optional)
    label_col = diff_idx - 1
    label_cell = ws.cell(row=total_row, column=label_col,
                         value="Total Difference")
    label_cell.font = Font(bold=True, size=12)
    label_cell.alignment = Alignment(horizontal='center', vertical='center')
    drug_idx = out_cols.index('Drug Name') + 1
    diff_idx = out_cols.index('Difference') + 1
    left_idx = min(drug_idx, diff_idx)
    right_idx = max(drug_idx, diff_idx)

    left_col = get_column_letter(left_idx)
    right_col = get_column_letter(right_idx)

    # Apply number format to Total Difference cell
    total_diff_cell = ws.cell(row=total_row, column=diff_idx)
    total_diff_cell.number_format = '"$"#,##0.00'
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{last_data_row}"

    # AutoSum cell
    sum_col_letter = get_column_letter(diff_idx)
    total_cell = ws.cell(row=total_row, column=diff_idx)
    # total_cell.value = f"=SUM({sum_col_letter}3:{sum_col_letter}{last_data_row})"
    total_cell.value = f"=SUBTOTAL(109,{sum_col_letter}3:{sum_col_letter}{last_data_row})"
    total_cell.number_format = '"$"#,##0.00'  # ✅ Currency format

    total_cell.font = Font(bold=True, size=12)
    total_cell.number_format = 'General'
    total_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ensure Excel recalculates when opening
    ws.parent.calculation.fullCalcOnLoad = True
    # Freeze panes
    ws.freeze_panes = "A3"

    # Set page orientation to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_mfp_drugs_sheet(
    wb,
    log_df,
    kinray_df,
    sheet_name: str = "MFP Drugs - RX"
):
    """
    Per-RX MFP analysis sheet.

    Rule used:
    - Any RX row with SDRA Amt != 0 is considered an MFP RX.
    """
    df = log_df.copy()

    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'SDRA Amt', 'COPAY']:
        df[c] = pd.to_numeric(df.get(c, 0), errors='coerce').fillna(0)

    df['NDC #'] = (df['NDC #'].astype(str)
                   .str.replace('-', '', regex=False)
                   .str.replace(r'\D', '', regex=True)
                   .str.zfill(11))

    date_candidates = ['Fill Date', 'Date', 'Rx Date', 'Dispense Date', 'Service Date']
    fill_date_col = next((c for c in date_candidates if c in df.columns), None)
    if fill_date_col:
        df['Fill Date'] = pd.to_datetime(df[fill_date_col], errors='coerce')
    else:
        df['Fill Date'] = pd.NaT

    df = df[df['SDRA Amt'].fillna(0) != 0].copy()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    if df.empty:
        ws['A1'] = "No MFP RXs found (SDRA Amt is 0/blank for all rows)."
        ws['A1'].font = Font(size=14, bold=True)
        return

    df['Kinray Unit Price'] = df.apply(
        lambda row: find_kinray_price_by_month(row['NDC #'], row['Fill Date'], kinray_df),
        axis=1
    )

    df['Package billed'] = np.where(
        df['Drug Pkg Size'] > 0,
        df['Qty Filled'] / df['Drug Pkg Size'],
        np.nan
    )
    df['Kinray Final Price'] = np.where(
        (df['Drug Pkg Size'] > 0) & (df['Kinray Unit Price'] > 0),
        (df['Kinray Unit Price'] / df['Drug Pkg Size']) * df['Qty Filled'],
        0.0
    )

    df['Winning Ins Paid'] = np.where(
        df['Ins Paid Plan 1'].fillna(0) >= df['Ins Paid Plan 2'].fillna(0),
        df['Ins Paid Plan 1'].fillna(0),
        df['Ins Paid Plan 2'].fillna(0)
    )
    df['Total Collected'] = df['Winning Ins Paid'] + df['SDRA Amt'] + df['COPAY']
    df['Difference'] = df['Total Collected'] - df['Kinray Final Price']

    rx_col = 'Rx #' if 'Rx #' in df.columns else ('Rx' if 'Rx' in df.columns else None)
    df['RX'] = df[rx_col] if rx_col else pd.NA
    df['NDC'] = df['NDC #']
    df['Drug Name'] = df.get('Drug Name', '')
    df['Pkg Size'] = df.get('Drug Pkg Size', 0)
    df['Qty Filled'] = df.get('Qty Filled', 0)
    df['BIN'] = df.get('Winning_BIN', '')
    df['Processor'] = df.get('Processor', '')
    df['PCN'] = df.get('Winning PCN', '')
    df['Group'] = df.get('Winning Group', '')

    out_cols = [
        'RX', 'Fill Date', 'NDC', 'Drug Name', 'Pkg Size', 'Qty Filled',
        'Package billed', 'Kinray Unit Price', 'Kinray Final Price',
        'Ins Paid Plan 1', 'Ins Paid Plan 2', 'SDRA Amt', 'COPAY',
        'Winning Ins Paid', 'Total Collected', 'Difference',
        'BIN', 'Processor', 'PCN', 'Group'
    ]
    out = df.loc[:, out_cols].copy().sort_values(['Drug Name', 'Fill Date'], ascending=[True, False])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_cols))
    title = ws.cell(row=1, column=1, value="MFP DRUGS")
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 26

    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(out_cols)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 9, 'Fill Date': 12, 'NDC': 14, 'Drug Name': 40, 'Pkg Size': 8, 'Qty Filled': 10,
        'Package billed': 12, 'Kinray Unit Price': 14, 'Kinray Final Price': 16,
        'Ins Paid Plan 1': 14, 'Ins Paid Plan 2': 14, 'SDRA Amt': 12, 'COPAY': 10,
        'Winning Ins Paid': 14, 'Total Collected': 15, 'Difference': 13,
        'BIN': 9, 'Processor': 15, 'PCN': 12, 'Group': 12
    }
    ws.row_dimensions[2].height = 45
    for i, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    currency_cols = {
        'Kinray Unit Price', 'Kinray Final Price',
        'Ins Paid Plan 1', 'Ins Paid Plan 2', 'SDRA Amt', 'COPAY',
        'Winning Ins Paid', 'Total Collected', 'Difference'
    }
    for r in range(3, ws.max_row + 1):
        for name in out_cols:
            idx = out_cols.index(name) + 1
            if name in currency_cols:
                ws.cell(row=r, column=idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=out_cols.index('Qty Filled') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index('Package billed') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index('Fill Date') + 1).number_format = 'yyyy-mm-dd'

    last_data_row = ws.max_row
    total_row = last_data_row + 1
    label_col = max(1, out_cols.index('Difference'))
    ws.cell(row=total_row, column=label_col, value='Totals').font = Font(bold=True)
    ws.cell(row=total_row, column=label_col).alignment = Alignment(horizontal='right', vertical='center')

    for name in ['Kinray Final Price', 'Ins Paid Plan 1', 'Ins Paid Plan 2', 'SDRA Amt', 'COPAY',
                 'Winning Ins Paid', 'Total Collected', 'Difference']:
        idx = out_cols.index(name) + 1
        col_letter = get_column_letter(idx)
        tcell = ws.cell(row=total_row, column=idx)
        tcell.value = f"=SUBTOTAL(109,{col_letter}3:{col_letter}{last_data_row})"
        tcell.font = Font(bold=True)
        tcell.number_format = '"$"#,##0.00'
        tcell.alignment = Alignment(horizontal='center', vertical='center')
        tcell.border = thin

    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{last_data_row}"
    ws.freeze_panes = "A3"

    diff_col_letter = get_column_letter(out_cols.index('Difference') + 1)
    data_range = f"{diff_col_letter}3:{diff_col_letter}{last_data_row}"
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
