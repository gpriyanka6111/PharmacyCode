# Builds "Never Ordered - Check" and "BIN to Processor" (including unmapped BINs) sheets.

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def create_never_ordered_check_sheet(wb, final_data):
    """
    Create 'Never Ordered - Check' sheet:
      • Rows where Total Purchased == 0
      • AND billed to insurance (any *_Q > 0 OR *_P > 0 OR *_T > 0)
      • Shows base cols + all insurer Q/P/T columns (excluding ALL_PBM_*).
    """

    df = final_data.copy()

    # Identify insurance bands
    # q_cols = [c for c in df.columns if c.endswith('_Q')]
    p_cols = [c for c in df.columns if c == 'ALL_PBM_P']
    # t_cols = [c for c in df.columns if c.endswith('_T') and c != 'ALL_PBM_T']

    # Ensure required base columns exist
    for base in ['Drug Name', 'NDC #', 'Package Size', 'Total Purchased']:
        if base not in df.columns:
            df[base] = 0 if base != 'Drug Name' else pd.NA

    # Coerce numeric for logic
    def _to_num(cols):
        if not cols:
            return
        df.loc[:, cols] = df[cols].apply(
            pd.to_numeric, errors='coerce').fillna(0)

    # _to_num(q_cols)
    _to_num(p_cols)
    # _to_num(t_cols)
    df['Total Purchased'] = pd.to_numeric(
        df['Total Purchased'], errors='coerce').fillna(0)

    # "Billed to insurance" mask (any positive in Q/P/T)
    billed_mask = pd.Series(False, index=df.index)
    # if q_cols: billed_mask |= df[q_cols].gt(0).any(axis=1)
    if p_cols:
        billed_mask |= df[p_cols].gt(0).any(axis=1)
    # if t_cols: billed_mask |= df[t_cols].gt(0).any(axis=1)

    mask = (df['Total Purchased'] == 0) & billed_mask

    # Build output
    display_columns = (['Drug Name', 'NDC #', 'Package Size', 'Total Purchased']
                       + p_cols)
    out = df.loc[mask, display_columns].copy()
    out.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)

    # Create/replace sheet
    title = "Never Ordered - Check"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title=title)

    if out.empty:
        ws['A1'] = "No rows with Total Purchased = 0 that were billed to insurance."
        return

    out = out.sort_values('Drug Name')

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(out.columns))
    cell = ws.cell(row=1, column=1)
    cell.value = "Never Ordered - Check (Billed to Insurance)"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=20, bold=True)
    ws.row_dimensions[1].height = 30

    # Write table (headers at row 2, data from row 3)
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 2:
                # Header formatting
                cell.font = Font(bold=True, size=12)
                # Rotate insurance columns (Q/P/T) and Total Purchased
                hdr = out.columns[c_idx - 1]
                rotate = hdr in (p_cols + ['Total Purchased', 'Pkg Size'])
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           text_rotation=(90 if rotate else 0), wrap_text=True)
            else:
                # Body formatting
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=False)
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')
                cell.font = Font(size=12)

    # Header thick border
    thick = Border(left=Side(style='thick'), right=Side(style='thick'),
                   top=Side(style='thick'), bottom=Side(style='thick'))
    for col_idx in range(1, len(out.columns) + 1):
        ws.cell(row=2, column=col_idx).border = thick

    # Thin borders for body
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(out.columns)):
        for cell in row:
            cell.border = thin

    # Thick edge borders for key columns
    def apply_column_border(sheet, col_idx):
        col_letter = get_column_letter(col_idx)
        for r in range(2, sheet.max_row + 1):
            c = sheet[f"{col_letter}{r}"]
            c.border = Border(left=thick.left, right=thick.right,
                              top=c.border.top, bottom=c.border.bottom)

    edge_cols = ['Drug Name', 'NDC #', 'Pkg Size', 'Total Purchased']
    for name in edge_cols:
        if name in out.columns:
            apply_column_border(ws, out.columns.get_loc(name) + 1)

    # Column widths
    widths = {
        'Drug Name': 70,
        'NDC #': 15,
        'Pkg Size': 10,
        'Total Purchased': 12
    }
    for idx, col_name in enumerate(out.columns, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = widths.get(col_name, 8)

    # Rotate headers already done; set header row height
    ws.row_dimensions[2].height = 80

    # # Currency for *_T columns
    # for tcol in t_cols:
    #     if tcol in out.columns:
    #         cidx = out.columns.get_loc(tcol) + 1
    #         for r in range(3, ws.max_row + 1):
    #             ws.cell(row=r, column=cidx).number_format = '"$"#,##0.00'

    # Freeze panes
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_columns))}{ws.max_row}"


def create_bin_to_processor_sheet(wb, rx_compare_source, bin_to_proc):
    # ===== Create/replace "BIN to Processor" sheet =====
    title_sheet = "BIN to Processor"
    if title_sheet in wb.sheetnames:
        del wb[title_sheet]
    ws2 = wb.create_sheet(title_sheet)

    # Title
    ws2.insert_rows(1)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    tcell = ws2.cell(
        row=1, column=1, value="BIN Numbers Billed (from Custom Log)")
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal='center', vertical='center')

    def find_fill_date_column(df):
        """Return the first column name that looks like a Fill Date."""
        candidates = ['Fill Date']
        # exact match first
        for c in candidates:
            if c in df.columns:
                return c
        # loose match (case-insensitive contains "fill" & "date")
        for c in df.columns:
            cl = str(c).strip().lower()
            if "date" in cl and ("fill" in cl or "filled" in cl):
                return c
        return None

    # --- Build BIN → Processor counts from the UNFILTERED custom log ---
    src_df = rx_compare_source.copy()  # unfiltered copy created earlier
    # Use the UNFILTERED log for totals so processor filters don't shrink the counts

    # <- "rows" (count rows), "qty" (sum Qty Filled), or "unique_rx" (distinct Rx #)
    COUNT_MODE = "rows"

    def build_rx_counts(src_df, mode="rows"):
        # Normalize BIN; include NaN/blank → '000000'
        bins = (src_df['Winning_BIN']
                .astype('string')
                .fillna('')                       # keep empties
                .str.replace(r'\D', '', regex=True)
                .str.zfill(6))                    # '' -> '000000'

        df = src_df.copy().assign(__BIN=bins)

        if mode == "rows":
            out = (df.groupby('__BIN', as_index=False)
                   .size()
                   .rename(columns={'__BIN': 'BIN', 'size': 'Total Rx'}))
            label = 'Total Rx'
        elif mode == "qty":
            out = (df.groupby('__BIN', as_index=False)['Qty Filled']
                   .sum()
                   .rename(columns={'__BIN': 'BIN', 'Qty Filled': 'Total Qty'}))
            label = 'Total Qty'
        else:  # unique_rx
            out = (df.groupby('__BIN', as_index=False)['Rx #']
                   .nunique()
                   .rename(columns={'__BIN': 'BIN', 'Rx #': 'Total Rx'}))
            label = 'Total Rx'
        return out, label

    rx_counts_df, RX_LABEL = build_rx_counts(src_df, COUNT_MODE)

    bin_proc_df = (rx_counts_df[['BIN']].copy()
                   .assign(Processor=lambda d: d['BIN'].map(bin_to_proc))
                   # keep 000000
                   .assign(Processor=lambda d: d['Processor'].fillna('Unmapped'))
                   .merge(rx_counts_df, on='BIN', how='left')
                   .sort_values(['Processor', 'BIN'])
                   .reset_index(drop=True))

    # Write headers with the dynamic label in C
    headers = ["BIN", "Processor", RX_LABEL]
    for cidx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=2, column=cidx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
    # header_row = 2
    # start_data_row = header_row + 1

    # Write A:C
    for r, (bin_, proc, total) in enumerate(
            bin_proc_df[['BIN', 'Processor', RX_LABEL]].itertuples(index=False, name=None), start=3):
        ws2.cell(row=r, column=1, value=str(bin_))
        ws2.cell(row=r, column=2, value=str(proc))
        ws2.cell(row=r, column=3, value=int(total))

    # Optional grand total row to sanity-check equals src_df.shape[0] when COUNT_MODE=="rows"
    gt_row = ws2.max_row + 1
    ws2.cell(row=gt_row, column=2, value="Grand Total").font = Font(bold=True)
    ws2.cell(row=gt_row, column=3,
             value=f"=SUM(C3:C{gt_row-1})").font = Font(bold=True)

    # Widths / filter
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 10
    ws2.auto_filter.ref = f"A2:C{ws2.max_row}"
    ws2.freeze_panes = "A3"

    # # Optional: bottom TOTAL row (helps you QA against expected 7,100 etc.)
    # end_row = ws2.max_row
    # total_row = end_row + 1
    # ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    # # Sum of column C
    # ws2.cell(row=total_row, column=3,
    #         value=f"=SUM(C{start_data_row}:C{end_row})").font = Font(bold=True)

    src_norm = src_df.copy()
    src_norm['__BIN'] = (src_norm['Winning_BIN'].astype('string')
                         .fillna('')
                         .str.replace(r'\D', '', regex=True)
                         .str.zfill(6))

    unmapped_rows = src_norm[src_norm['__BIN'] == '000000'].copy()
    fill_col = find_fill_date_column(unmapped_rows)

    ws2.merge_cells('F1:H1')
    title = ws2.cell(row=1, column=6, value="Unmapped BIN Numbers (000000)")
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.font = Font(bold=True, size=14)
    ws2['F2'] = "BIN"
    ws2['G2'] = "RX #"
    if fill_col:
        ws2['H2'] = "Fill Date"

    # Style + widths
    for col in ['F', 'G'] + (['H'] if fill_col else []):
        head = ws2[f'{col}2']
        head.font = Font(bold=True, color="000000")
        head.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[col].width = 18
    if fill_col:
        ws2.column_dimensions['H'].width = 14

    # Coerce date (for pretty output); safe even if mixed types
    if fill_col:
        try:
            unmapped_rows[fill_col] = pd.to_datetime(
                unmapped_rows[fill_col], errors='coerce')
        except Exception:
            pass

   # Write ALL rows (no set()/groupby dedupe): F=BIN, G=RX #, H=Fill Date
    start_row_unmapped = 3
    cols = ['__BIN', 'Rx #'] + ([fill_col] if fill_col else [])
    for r_idx, row in enumerate(unmapped_rows[cols].itertuples(index=False, name=None),
                                start=start_row_unmapped):
        # F -> BIN (000000)
        ws2.cell(row=r_idx, column=6, value=row[0])
        ws2.cell(row=r_idx, column=7, value=str(row[1]))          # G -> RX #
        if fill_col:
            v = row[2]
            # format Timestamp nicely
            if hasattr(v, "strftime"):
                v = v.strftime('%Y-%m-%d')
            # H -> Fill Date
            ws2.cell(row=r_idx, column=8, value=v)

    # Filter across A..H if H exists; else A..G
    last_col_letter = 'H' if fill_col else 'G'
    ws2.auto_filter.ref = f"A2:{last_col_letter}{ws2.max_row}"

    # ---- Formatting
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 10
    ws2.freeze_panes = 'A3'  # keep title+headers fixed
    ws2.auto_filter.ref = f"A2:B{ws2.max_row}"  # filter on BIN/Processor only
