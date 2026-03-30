# Builds "Refills 0 - Call Doctor" and "Missed Refill - Revenue Recovery" sheets from the custom log.

import re

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def add_zero_refills_sheet(
    wb,
    log_df,
    sheet_name: str = "Refills 0 - Call Doctor"
):
    df = log_df.copy()

    def pick_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    rx_col = pick_col(['Rx #', 'Rx', 'RX'])
    refill_col = pick_col(['Refills Left', 'Refills Remaining'])
    fill_col = pick_col(['Fill Date', 'Dispense DateTime', 'Date', 'Rx Date', 'Service Date'])
    days_col = pick_col(['Days', 'Days Supply'])

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    if not rx_col or not refill_col:
        ws['A1'] = "Missing required columns for zero-refill list (Rx and Refills Left)."
        return

    df['__RX__'] = df[rx_col].astype(str).str.strip()
    df = df[df['__RX__'].ne('')].copy()

    if fill_col:
        df['__FillDate__'] = pd.to_datetime(df[fill_col], errors='coerce')
    else:
        df['__FillDate__'] = pd.NaT

    df['__RefillsLeft__'] = pd.to_numeric(df[refill_col], errors='coerce').fillna(0)
    if days_col:
        df['__Days__'] = pd.to_numeric(df[days_col], errors='coerce').fillna(0)
    else:
        df['__Days__'] = 0

    df = df.sort_values(['__RX__', '__FillDate__'])
    latest = df.drop_duplicates(subset=['__RX__'], keep='last').copy()

    out = latest[latest['__RefillsLeft__'] <= 0].copy()
    if out.empty:
        ws['A1'] = "No RX found with Refills Left = 0."
        return

    patient_col = pick_col(['Patient Name'])
    dob_col = pick_col(['Patient DOB'])
    phone_col = pick_col(['Patient Cell Phone', 'Patient Phone', 'Patient Work Phone'])
    drug_col = pick_col(['Drug Name'])
    ndc_col = pick_col(['Drug NDC', 'NDC #'])
    prescriber_col = pick_col(['Prescriber Name'])
    status_col = pick_col(['Rx Status', 'Status'])
    workflow_col = pick_col(['Workflow Status'])

    out['Expected Next Fill Date'] = out['__FillDate__'] + pd.to_timedelta(out['__Days__'], unit='D')

    display = pd.DataFrame({
        'RX': out['__RX__'],
        'Patient Name': out[patient_col] if patient_col else pd.NA,
        'Patient DOB': out[dob_col] if dob_col else pd.NA,
        'Phone': out[phone_col] if phone_col else pd.NA,
        'Drug Name': out[drug_col] if drug_col else pd.NA,
        'NDC': out[ndc_col] if ndc_col else pd.NA,
        'Last Fill Date': out['__FillDate__'],
        'Days Supply': out['__Days__'],
        'Refills Left': out['__RefillsLeft__'],
        'Expected Next Fill Date': out['Expected Next Fill Date'],
        'Prescriber': out[prescriber_col] if prescriber_col else pd.NA,
        'Rx Status': out[status_col] if status_col else pd.NA,
        'Workflow Status': out[workflow_col] if workflow_col else pd.NA,
        'Doctor Called?': '',
        'Call Date': pd.NaT,
        'Outcome': '',
        'New RX Received?': '',
    })

    display = display.sort_values(['Expected Next Fill Date', 'Last Fill Date'], ascending=[True, False])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display.columns))
    title = ws.cell(row=1, column=1, value="Refills Left = 0 (Call Doctor List)")
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42

    for r_idx, row in enumerate(dataframe_to_rows(display, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(display.columns)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 10,
        'Patient Name': 28,
        'Patient DOB': 12,
        'Phone': 14,
        'Drug Name': 40,
        'NDC': 14,
        'Last Fill Date': 13,
        'Days Supply': 10,
        'Refills Left': 10,
        'Expected Next Fill Date': 16,
        'Prescriber': 26,
        'Rx Status': 14,
        'Workflow Status': 18,
        'Doctor Called?': 12,
        'Call Date': 12,
        'Outcome': 22,
        'New RX Received?': 14,
    }
    for i, name in enumerate(display.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)

    date_cols = [
        display.columns.get_loc('Patient DOB') + 1,
        display.columns.get_loc('Last Fill Date') + 1,
        display.columns.get_loc('Expected Next Fill Date') + 1,
        display.columns.get_loc('Call Date') + 1,
    ]
    for r in range(3, ws.max_row + 1):
        for dc in date_cols:
            ws.cell(row=r, column=dc).number_format = 'mm-dd-yyyy'

    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = 'A3'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

'''
def add_missed_refill_revenue_sheet(
    wb,
    log_df,
    sheet_name: str = "Missed Refill - Revenue Recovery",
    grace_days: int = 7
):
    df = log_df.copy()

    def pick_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    rx_col = pick_col(['Rx #', 'Rx', 'RX'])
    fill_col = pick_col(['Fill Date', 'Dispense DateTime', 'Date', 'Rx Date', 'Service Date'])
    days_col = pick_col(['Days', 'Days Supply'])
    refills_left_col = pick_col(['Refills Left', 'Refills Remaining'])

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    if not rx_col or not fill_col or not days_col or not refills_left_col:
        ws['A1'] = "Missing required columns for missed refill analysis (Rx, Fill Date, Days, Refills Left)."
        return

    df['__RX__'] = df[rx_col].astype(str).str.strip()
    df = df[df['__RX__'].ne('')].copy()

    df['__FillDate__'] = pd.to_datetime(df[fill_col], errors='coerce')
    df['__Days__'] = pd.to_numeric(df[days_col], errors='coerce').fillna(0)
    df['__RefillsLeft__'] = pd.to_numeric(df[refills_left_col], errors='coerce').fillna(0)

    total_col = pick_col(['Total', 'Ins Paid Total'])
    if total_col:
        df['__TotalCollected__'] = pd.to_numeric(df[total_col], errors='coerce').fillna(0)
    else:
        ins1 = pd.to_numeric(df.get('Ins Paid Plan 1', 0), errors='coerce').fillna(0)
        ins2 = pd.to_numeric(df.get('Ins Paid Plan 2', 0), errors='coerce').fillna(0)
        copay = pd.to_numeric(df.get('Copay', df.get('COPAY', 0)), errors='coerce').fillna(0)
        sdra = pd.to_numeric(df.get('* SDRA Amt', df.get('SDRA Amt', df.get('SDRA', 0))), errors='coerce').fillna(0)
        df['__TotalCollected__'] = np.maximum(ins1, ins2) + copay + sdra

    patient_col = pick_col(['Patient Name'])
    patient_id_col = pick_col(['Patient ID'])
    dob_col = pick_col(['Patient DOB'])
    phone_col = pick_col(['Patient Cell Phone', 'Patient Phone', 'Patient Work Phone'])
    drug_col = pick_col(['Drug Name'])
    ndc_col = pick_col(['Drug NDC', 'NDC #'])
    drug_group_col = pick_col(['Drug Group'])
    prescriber_col = pick_col(['Prescriber Name'])
    status_col = pick_col(['Rx Status', 'Status'])
    workflow_col = pick_col(['Workflow Status'])
    bin_col = pick_col(['Plan 1 BIN'])
    processor_col = pick_col(['Ins Group', 'Processor'])

    def norm_text(s):
        if pd.isna(s):
            return ''
        return re.sub(r'[^A-Z0-9]+', '', str(s).upper().strip())

    patient_base = (
        df[patient_id_col].map(norm_text)
        if patient_id_col else
        (df[patient_col].map(norm_text) if patient_col else pd.Series('', index=df.index))
    )
    if dob_col:
        dob_norm = pd.to_datetime(df[dob_col], errors='coerce').dt.strftime('%Y%m%d').fillna('')
    else:
        dob_norm = pd.Series('', index=df.index)

    drug_group_base = df[drug_group_col].map(norm_text) if drug_group_col else pd.Series('', index=df.index)
    drug_name_base = df[drug_col].map(norm_text) if drug_col else pd.Series('', index=df.index)
    ndc_base = (df[ndc_col].astype(str)
                .str.replace(r'\D', '', regex=True)
                .str.zfill(11)) if ndc_col else pd.Series('', index=df.index)

    uom_col = pick_col(['Unit of Measure'])
    uom_base = df[uom_col].map(norm_text) if uom_col else pd.Series('', index=df.index)

    def infer_form_token(name_val, uom_val):
        if uom_val:
            if uom_val in {'ML', 'EACH', 'TABLET', 'CAPSULE', 'GRAM'}:
                return uom_val
        n = '' if pd.isna(name_val) else str(name_val).upper()
        checks = [
            ('PEN', 'PEN'),
            ('INJECTION', 'INJECTION'),
            ('SYRINGE', 'SYRINGE'),
            ('TABLET', 'TABLET'),
            ('TAB', 'TABLET'),
            ('CAPSULE', 'CAPSULE'),
            ('CAP ', 'CAPSULE'),
            ('AEROSOL', 'INHALER'),
            ('INHAL', 'INHALER'),
            ('STRIP', 'STRIP'),
            ('SOLUTION', 'SOLUTION'),
            ('SUSPENSION', 'SUSPENSION'),
            ('PATCH', 'PATCH'),
            ('CREAM', 'CREAM'),
            ('OINTMENT', 'OINTMENT'),
            ('DROPS', 'DROPS'),
        ]
        for key, val in checks:
            if key in n:
                return val
        return ''

    THERAPY_ALIAS_MAP = {
        'APIXABAN': ['ELIQUIS', 'APIXABAN'],
        'RIVAROXABAN': ['XARELTO', 'RIVAROXABAN'],
        'DABIGATRAN': ['PRADAXA', 'DABIGATRAN'],
        'EDOXABAN': ['SAVAYSA', 'EDOXABAN'],
        'WARFARIN': ['COUMADIN', 'JANTOVEN', 'WARFARIN'],

        'ATORVASTATIN': ['LIPITOR', 'ATORVASTATIN'],
        'ROSUVASTATIN': ['CRESTOR', 'ROSUVASTATIN'],
        'SIMVASTATIN': ['ZOCOR', 'SIMVASTATIN'],
        'PRAVASTATIN': ['PRAVACHOL', 'PRAVASTATIN'],
        'PITAVASTATIN': ['LIVALO', 'PITAVASTATIN'],

        'LEVOTHYROXINE': ['SYNTHROID', 'LEVOXYL', 'UNITHROID', 'LEVOTHYROXINE'],
        'METFORMIN': ['GLUCOPHAGE', 'METFORMIN'],
        'EMPAGLIFLOZIN': ['JARDIANCE', 'EMPAGLIFLOZIN'],
        'DAPAGLIFLOZIN': ['FARXIGA', 'DAPAGLIFLOZIN'],
        'CANAGLIFLOZIN': ['INVOKANA', 'CANAGLIFLOZIN'],
        'ERTUGLIFLOZIN': ['STEGLATRO', 'ERTUGLIFLOZIN'],
        'SITAGLIPTIN': ['JANUVIA', 'SITAGLIPTIN'],
        'SITAGLIPTIN_METFORMIN': ['JANUMET', 'SITAGLIPTIN/METFORMIN'],
        'SAXAGLIPTIN': ['ONGLYZA', 'SAXAGLIPTIN'],
        'LINAGLIPTIN': ['TRADJENTA', 'LINAGLIPTIN'],
        'PIOGLITAZONE': ['ACTOS', 'PIOGLITAZONE'],
        'GLIMEPIRIDE': ['AMARYL', 'GLIMEPIRIDE'],
        'GLIPIZIDE': ['GLUCOTROL', 'GLIPIZIDE'],
        'GLYBURIDE': ['MICRONASE', 'DIABETA', 'GLYNASE', 'GLYBURIDE'],

        'SEMAGLUTIDE_INJ': ['OZEMPIC', 'WEGOVY', 'SEMAGLUTIDE'],
        'SEMAGLUTIDE_ORAL': ['RYBELSUS'],
        'DULAGLUTIDE': ['TRULICITY', 'DULAGLUTIDE'],
        'LIRAGLUTIDE': ['VICTOZA', 'SAXENDA', 'LIRAGLUTIDE'],
        'TIRZEPATIDE': ['MOUNJARO', 'ZEPBOUND', 'TIRZEPATIDE'],
        'EXENATIDE': ['BYETTA', 'EXENATIDE'],
        'EXENATIDE_ER': ['BYDUREON', 'BYDUREON BCISE', 'EXENATIDE ER'],

        'INSULIN_GLARGINE': ['LANTUS', 'BASAGLAR', 'INSULIN GLARGINE'],
        'INSULIN_GLARGINE_U300': ['TOUJEO'],
        'INSULIN_DETEMIR': ['LEVEMIR', 'INSULIN DETEMIR'],
        'INSULIN_DEGLUDEC': ['TRESIBA', 'INSULIN DEGLUDEC'],
        'INSULIN_LISPRO': ['HUMALOG', 'ADMELOG', 'INSULIN LISPRO'],
        'INSULIN_ASPART': ['NOVOLOG', 'FIASP', 'INSULIN ASPART'],
        'INSULIN_GLULISINE': ['APIDRA', 'INSULIN GLULISINE'],

        'BUDESONIDE_FORMOTEROL': ['SYMBICORT', 'BUDESONIDE/FORMOTEROL'],
        'FLUTICASONE_SALMETEROL': ['ADVAIR', 'WIXELA', 'FLUTICASONE/SALMETEROL'],
        'FLUTICASONE_VILANTEROL': ['BREO', 'FLUTICASONE/VILANTEROL'],
        'FLUTICASONE_UMECLIDINIUM_VILANTEROL': ['TRELEGY', 'FLUTICASONE/UMECLIDINIUM/VILANTEROL'],
        'TIOTROPIUM': ['SPIRIVA', 'TIOTROPIUM'],
        'UMECLIDINIUM': ['INCRUSE', 'UMECLIDINIUM'],
        'ALBUTEROL': ['VENTOLIN', 'PROAIR', 'PROVENTIL', 'ALBUTEROL'],

        'CYCLOSPORINE_OPHTHALMIC': ['RESTASIS', 'CYCLOSPORINE'],
        'LIFITEGRAST_OPHTHALMIC': ['XIIDRA', 'LIFITEGRAST'],
        'RIMEGEPANT': ['NURTEC', 'RIMEGEPANT'],

        'DARUNAVIR_COBICISTAT_EMTRICITABINE_TAF': ['SYMTUZA'],
    }

    def canonical_therapy(name_val, group_val):
        g = '' if pd.isna(group_val) else str(group_val).upper()
        n = '' if pd.isna(name_val) else str(name_val).upper()
        hay = f"{g} {n}".strip()
        for canon, aliases in THERAPY_ALIAS_MAP.items():
            for alias in aliases:
                if alias in hay:
                    return canon
        return ''

    canonical_base = pd.Series(
        [canonical_therapy(nv, gv) for nv, gv in zip(df[drug_col] if drug_col else pd.Series('', index=df.index),
                                                     df[drug_group_col] if drug_group_col else pd.Series('', index=df.index))],
        index=df.index,
        dtype='object'
    )

    therapy_token = canonical_base.where(canonical_base.ne(''), drug_group_base)
    therapy_token = therapy_token.where(therapy_token.ne(''), drug_name_base)
    therapy_token = therapy_token.where(therapy_token.ne(''), ndc_base)
    form_token = pd.Series(
        [infer_form_token(nv, uv) for nv, uv in zip(df[drug_col] if drug_col else pd.Series('', index=df.index), uom_base)],
        index=df.index,
        dtype='object'
    )

    df['__PatientKey__'] = patient_base + np.where(dob_norm.ne(''), '_' + dob_norm, '')
    df['__TherapyToken__'] = therapy_token
    df['__FormToken__'] = form_token
    df['__TherapyKey__'] = np.where(
        df['__TherapyToken__'].ne(''),
        df['__PatientKey__'] + '|' + df['__TherapyToken__'],
        ''
    )

    df = df.sort_values(['__RX__', '__FillDate__'])
    latest = df.drop_duplicates(subset=['__RX__'], keep='last').copy()

    latest['Expected Refill Date'] = latest['__FillDate__'] + pd.to_timedelta(latest['__Days__'], unit='D')
    latest['Expected Refill Date + Grace'] = latest['Expected Refill Date'] + pd.to_timedelta(grace_days, unit='D')
    today = pd.Timestamp.today().normalize()

    missed = latest[
        (latest['__RefillsLeft__'] > 0)
        & latest['__FillDate__'].notna()
        & (latest['__Days__'] > 0)
        & (latest['Expected Refill Date + Grace'] < today)
    ].copy()

    if missed.empty:
        ws['A1'] = "No missed refills found for the selected grace window."
        return

    missed['Days Overdue'] = (today - missed['Expected Refill Date']).dt.days.clip(lower=0)

    coverage_rows = []
    therapy_groups = latest.groupby('__TherapyKey__', dropna=False)
    for idx, row in missed.iterrows():
        tkey = row.get('__TherapyKey__', '')
        current_fill = row['__FillDate__']
        coverage_rx = ''
        coverage_fill = pd.NaT
        covered = False
        if tkey and tkey in therapy_groups.groups:
            group_rows = therapy_groups.get_group(tkey)
            cands = group_rows[
                (group_rows['__RX__'] != row['__RX__'])
                & group_rows['__FillDate__'].notna()
                & (group_rows['__FillDate__'] > current_fill)
                & (group_rows['__FillDate__'] <= today)
            ]
            row_form = row.get('__FormToken__', '')
            if row_form:
                cands = cands[(cands.get('__FormToken__', '') == row_form) | (cands.get('__FormToken__', '') == '')]
            if not cands.empty:
                hit = cands.sort_values('__FillDate__').iloc[-1]
                coverage_rx = hit['__RX__']
                coverage_fill = hit['__FillDate__']
                covered = True

        rx_status_val = str(row[status_col]).strip().lower() if status_col else ''
        workflow_val = str(row[workflow_col]).strip().lower() if workflow_col else ''
        excluded_terms = ['void', 'reversed', 'cancel', 'transferred', 'deleted']
        excluded = any(t in rx_status_val for t in excluded_terms) or any(t in workflow_val for t in excluded_terms)

        if excluded:
            final_action = 'Exclude'
        elif covered:
            final_action = 'Covered by Other Rx'
        else:
            final_action = 'Open - Missed'

        coverage_rows.append((idx, covered, coverage_rx, coverage_fill, final_action))

    coverage_df = pd.DataFrame(
        coverage_rows,
        columns=['__idx__', '__Covered__', '__CoverRx__', '__CoverFill__', '__FinalAction__']
    ).set_index('__idx__')
    missed = missed.join(coverage_df, how='left')
    missed['__Covered__'] = missed['__Covered__'].fillna(False)
    missed['Est Recoverable $'] = np.where(
        missed['__FinalAction__'].eq('Open - Missed'),
        missed['__TotalCollected__'],
        0
    )

    out = pd.DataFrame({
        'RX': missed['__RX__'],
        'Patient Name': missed[patient_col] if patient_col else pd.NA,
        'Patient DOB': missed[dob_col] if dob_col else pd.NA,
        'Phone': missed[phone_col] if phone_col else pd.NA,
        'Drug Name': missed[drug_col] if drug_col else pd.NA,
        'NDC': missed[ndc_col] if ndc_col else pd.NA,
        'Therapy Key': missed['__TherapyKey__'],
        'Last Fill Date': missed['__FillDate__'],
        'Days Supply': missed['__Days__'],
        'Refills Left': missed['__RefillsLeft__'],
        'Expected Refill Date': missed['Expected Refill Date'],
        'Days Overdue': missed['Days Overdue'],
        'Covered by Other Rx': np.where(missed['__Covered__'], 'Yes', 'No'),
        'Covering Rx #': missed['__CoverRx__'],
        'Covering Fill Date': missed['__CoverFill__'],
        'Final Action': missed['__FinalAction__'],
        'Est Recoverable $': missed['Est Recoverable $'],
        'Rx Status': missed[status_col] if status_col else pd.NA,
        'Workflow Status': missed[workflow_col] if workflow_col else pd.NA,
        'Processor': missed[processor_col] if processor_col else pd.NA,
        'BIN': missed[bin_col] if bin_col else pd.NA,
        'Prescriber': missed[prescriber_col] if prescriber_col else pd.NA,
    })

    out = out.sort_values(['Final Action', 'Days Overdue', 'Est Recoverable $'], ascending=[True, False, False])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out.columns))
    title = ws.cell(row=1, column=1, value=f"Missed Refill - Revenue Recovery (Grace: {grace_days} days)")
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42

    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(out.columns)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 10,
        'Patient Name': 28,
        'Patient DOB': 12,
        'Phone': 14,
        'Drug Name': 42,
        'NDC': 14,
        'Therapy Key': 28,
        'Last Fill Date': 13,
        'Days Supply': 10,
        'Refills Left': 10,
        'Expected Refill Date': 15,
        'Days Overdue': 11,
        'Covered by Other Rx': 15,
        'Covering Rx #': 12,
        'Covering Fill Date': 14,
        'Final Action': 18,
        'Est Recoverable $': 16,
        'Rx Status': 16,
        'Workflow Status': 18,
        'Processor': 12,
        'BIN': 9,
        'Prescriber': 28,
    }

    for i, name in enumerate(out.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)

    money_col = out.columns.get_loc('Est Recoverable $') + 1
    date_cols = [
        out.columns.get_loc('Patient DOB') + 1,
        out.columns.get_loc('Last Fill Date') + 1,
        out.columns.get_loc('Expected Refill Date') + 1,
        out.columns.get_loc('Covering Fill Date') + 1,
    ]
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=money_col).number_format = '"$"#,##0.00'
        for dc in date_cols:
            ws.cell(row=r, column=dc).number_format = 'mm-dd-yyyy'

    total_row = ws.max_row + 1
    label_col = out.columns.get_loc('Days Overdue') + 1
    ws.cell(row=total_row, column=label_col, value='Total Potential Recovery').font = Font(bold=True)
    total_cell = ws.cell(
        row=total_row,
        column=money_col,
        value=f"=SUBTOTAL(109,{get_column_letter(money_col)}3:{get_column_letter(money_col)}{ws.max_row})"
    )
    total_cell.font = Font(bold=True)
    total_cell.number_format = '"$"#,##0.00'

    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row - 1}"
    ws.freeze_panes = 'A3'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

'''
