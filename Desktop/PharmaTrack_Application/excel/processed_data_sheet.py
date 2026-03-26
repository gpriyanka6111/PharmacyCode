# Builds and formats the main "Processed Data" sheet: title, column bands, borders, conditional formatting, freeze panes.

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.page import PageMargins

from excel.support_sheets import create_bin_to_processor_sheet


def build_processed_data_sheet(wb, ws, final, desired_columns, processors,
                                pharmacy_name, date_range,
                                rx_compare_source, bin_to_proc):
    header_row = 3

    # Merge the first row and set the pharmacy name and date range in the center
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(desired_columns))
    cell = ws.cell(row=1, column=1)
    cell.value = f"{pharmacy_name} ({date_range})"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=35, bold=True)
    ws.row_dimensions[1].height = 60

    # Move the data down by one row
    ws.insert_rows(2)
    # Move the data down by one row
    ws.insert_rows(3)

    # Explicitly set the headers in the second row
    for col_num, header in enumerate(desired_columns, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=False)
        cell.font = Font(bold=False, size=15)

    # Dynamically calculate the start and end columns for each merged cell

    def get_column_index(ws, header_name, header_row=None):
        """Find a column by header text. Returns 1-based index or None."""
        rows = [header_row] if header_row else [2, 3]
        for r in rows:
            try:
                for cell in ws[r]:
                    if cell.value == header_name:
                        return cell.col_idx
            except Exception:
                continue
        return None

    for col_name in ['NDC #', 'Drug Name', 'Package Size']:
        idx = get_column_index(ws, col_name)
        if not idx:
            continue

        header_cell = ws.cell(row=header_row, column=idx)

        if col_name == 'Package Size':
            # Header rotated 90°, centered
            header_cell.alignment = Alignment(
                horizontal='center',
                vertical='bottom',
                text_rotation=90,
                wrap_text=False
            )
        else:
            # Normal center/center for NDC # and Drug Name
            header_cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=False
            )

    if "Total Purchased" in desired_columns:
        total_purchased_col = desired_columns.index("Total Purchased") + 1
        total_purchased_indices = [total_purchased_col]
    else:
        raise ValueError("'Total Purchased' not found in desired_columns")

    # >>> NEW: group headers for Insurance $ Paid, $$ Purchased (Kinray), Net
    def _band_bounds_from_suffix(sfx):
        cols = [get_column_index(ws, f"{pr}_{sfx}") for pr in processors]
        cols = [c for c in cols if c]  # drop None
        if not cols:
            return None, None
        return min(cols), max(cols)

    def _ranges_intersect(a, b):
        a_min_col, a_min_row, a_max_col, a_max_row = range_boundaries(str(a))
        b_min_col, b_min_row, b_max_col, b_max_row = range_boundaries(str(b))
        return not (a_max_col < b_min_col or b_max_col < a_min_col or
                    a_max_row < b_min_row or b_max_row < a_min_row)

    def _merge_band(row, start_col, end_col, title):
        """
        Safely put `title` across start_col..end_col on `row`.
        - write value BEFORE merge (avoids MergedCell read-only)
        - unmerge any overlapping prior merges
        - style the anchor (top-left) cell
        """
        if not (start_col and end_col and start_col <= end_col):
            return

        # If it's a single column, just set value & style—no merge needed.
        if start_col == end_col:
            anchor = ws.cell(row=row, column=start_col)
            anchor.value = title
            anchor.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            anchor.font = Font(bold=True)
            return

        target_ref = f"{ws.cell(row=row, column=start_col).coordinate}:{ws.cell(row=row, column=end_col).coordinate}"

        # Unmerge any existing merged ranges that overlap our target span
        for mr in list(ws.merged_cells.ranges):
            if _ranges_intersect(mr, target_ref):
                ws.unmerge_cells(str(mr))

        # Write BEFORE merging
        anchor = ws.cell(row=row, column=start_col)
        anchor.value = title

        # Merge & style the anchor
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=end_col)
        anchor.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        anchor.font = Font(bold=True)

    paid_s, paid_e = _band_bounds_from_suffix('T')
    pur_s,  pur_e = _band_bounds_from_suffix('Pur')
    net_s,  net_e = _band_bounds_from_suffix('Net')
    qt_s, qt_e = _band_bounds_from_suffix('Q')
    pk_b_s, pk_b_e = _band_bounds_from_suffix('P')
    pk_d_s, pk_d_e = _band_bounds_from_suffix('D')

    _merge_band(2, qt_s,  qt_e,  "Quantity Billed = BestRX")
    _merge_band(2, pk_b_s, pk_b_e,
                "Package Size Billed = Quantity Billed(BestRx) ÷ Package Size")
    _merge_band(2, pk_d_s, pk_d_e,
                "Package Size Difference = Total Packages Purchased(Vendors) − Package Size Billed(BestRx)")
    _merge_band(2, paid_s, paid_e, "Actual $ Paid by Insurance = BestRX")
    _merge_band(2, pur_s,  pur_e,
                "Actual $ Purchased (Kinray Unit Price × Packages Billed To Ins)")
    _merge_band(2, net_s,  net_e,
                "Net(Profit/Loss)$ = Actual $ Paid(BestRx) − Actual $ Purchased(Kinray)")

    # ✅ Enable wrap text for Drug Name column (column B)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=False)

    # Set the desired column widths
    column_widths = {
        'A': 15,  # NDC
        'B': 45,  # Drug Name
    }

    # Setting up width for the other columns
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    # Set widths for dynamic columns
    for col_num in range(4, len(desired_columns) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 7

    # Set the height for the first row
    ws.row_dimensions[1].height = 35

    # Set header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D0CECE",
                              end_color="D0CECE", fill_type="solid")
    # Set border style
    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(
        style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    # Enable text wrapping for the second row
    for cell in ws[3]:
        if cell.col_idx > 3:
            cell.alignment = Alignment(
                text_rotation=90, horizontal='center', wrap_text=False)
            cell.font = Font(bold=False, size=14, name='Calibri')
            cell.fill = header_fill
            cell.border = thin_border
        else:
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=False)
            cell.fill = header_fill
            cell.border = thin_border

    # Text rotation for Package Size column header
    # text rotation is not happening for Package Size column

    pkg_size_col_idx = get_column_index(ws, 'Package Size')
    if pkg_size_col_idx:
        pkg_cell = ws.cell(row=3, column=pkg_size_col_idx)
        pkg_cell.alignment = Alignment(
            text_rotation=90,
            horizontal='center',
            vertical='center',
            wrap_text=False
        )
        pkg_cell.font = Font(bold=False, size=14, name='Calibri')

    ws.row_dimensions[3].height = 100
    # Freeze the first row
    # ws.freeze_panes = 'A4'
    # freeze panes till kinray unit price column
    ws.freeze_panes = get_column_letter(total_purchased_col + 2) + '4'

    red_fill = PatternFill(start_color="FFC7CE",
                           end_color="FFC7CE", fill_type="solid")   # soft red
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # soft green
    row_fill_soft = PatternFill(
        start_color="E8E6FF", end_color="E8E6FF", fill_type="solid")  # subtle lavender-gray

    data_first_row = header_row + 1
    data_last_row = ws.max_row

    # Map header text -> column index from the worksheet
    header_map = {ws.cell(row=header_row, column=c).value: c for c in range(
        1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value}

    # All *_D columns present in the sheet
    diff_cols = [h for h in header_map.keys() if isinstance(h, str)
                 and h.endswith('_D')]

    # Apply per-column red/green CF to *_D columns
    for h in diff_cols:
        col_idx = header_map[h]
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}{data_first_row}:{col_letter}{data_last_row}"

        # red < 0
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=[
                       "0"], stopIfTrue=False, fill=red_fill)
        )
        # green > 0
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=[
                       "0"], stopIfTrue=False, fill=green_fill)
        )

    # Whole-row soft highlight if ANY *_D in that row is negative
    if diff_cols:
        first_idx = min(header_map[h] for h in diff_cols)
        last_idx = max(header_map[h] for h in diff_cols)
        first_letter = get_column_letter(first_idx)
        last_letter = get_column_letter(last_idx)

        # IMPORTANT: the formula must be relative to the TOP row of the CF range
        row_range = f"A{data_first_row}:{get_column_letter(ws.max_column)}{data_last_row}"
        formula = f'COUNTIF(${first_letter}{data_first_row}:${last_letter}{data_first_row},"<0")>0'
        ws.conditional_formatting.add(
            row_range,
            FormulaRule(formula=[formula],
                        stopIfTrue=False, fill=row_fill_soft)
        )

    # AutoFilter over the full data region (row 3 headers)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{data_last_row}"

    # Center align all data
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set the first two columns to left alignment
    for row in ws.iter_rows(min_row=4):
        row[0].alignment = Alignment(horizontal='left',  vertical='center')
        row[1].alignment = Alignment(horizontal='left',  vertical='center')
        row[2].alignment = Alignment(horizontal='center', vertical='center')

    def apply_thick_border(ws, start_col, end_col, start_row, end_row):
        # Apply the thick border to the top row
        for col_num in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.border = Border(
                top=thick_border.top,
                left=cell.border.left,
                right=cell.border.right,
                bottom=cell.border.bottom
            )

        # Apply the thick border to the bottom row
        for col_num in range(start_col, end_col + 1):
            cell = ws.cell(row=end_row, column=col_num)
            cell.border = Border(
                bottom=thick_border.bottom,
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top
            )

        # Apply the thick border to the left column
        for row_num in range(start_row, end_row + 1):
            cell = ws.cell(row=row_num, column=start_col)
            cell.border = Border(
                left=thick_border.left,
                top=cell.border.top,
                right=cell.border.right,
                bottom=cell.border.bottom
            )

        # Apply the thick border to the right column
        for row_num in range(start_row, end_row + 1):
            cell = ws.cell(row=row_num, column=end_col)
            cell.border = Border(
                right=thick_border.right,
                top=cell.border.top,
                left=cell.border.left,
                bottom=cell.border.bottom
            )
    start_row = 1
    end_row = ws.max_row

    def style_sheet(ws):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    def get_column_indices(ws, column_names):
        indices = []
        for col in ws[3]:
            if col.value in column_names:
                indices.append(col.col_idx)
        return indices
    total_purchased_col = get_column_index(ws, 'Total Purchased')
    if total_purchased_col is None:
        raise ValueError("Header 'Total Purchased' not found in the worksheet")

    # Apply thick border for specific column ranges
    quantity_billed_indices = get_column_indices(
        ws, [f'{pr}_Q' for pr in processors])
    package_size_billed_indices = get_column_indices(
        ws, [f'{pr}_P' for pr in processors])
    package_size_difference_indices = get_column_indices(
        ws, [f'{pr}_D' for pr in processors])

    total_purchased_indices = [total_purchased_col]

    def apply_thick_border_to_groups(ws, column_groups, start_row, end_row):
        for group in column_groups:
            if group:
                start_col = group[0]
                end_col = group[-1]
                apply_thick_border(ws, start_col, end_col, start_row, end_row)

    thin_border = Border(left=Side(style='thin', color="A9A9A9"), right=Side(style='thin', color="A9A9A9"), top=Side(
        style='thin', color="A9A9A9"), bottom=Side(style='thin', color="A9A9A9"))

    for cell in ws[3]:
        cell.border = thin_border

    # Set up styles
    cell_fill_red = PatternFill(
        start_color="F88379", end_color="F88379", fill_type="solid")
    row_fill_blue = PatternFill(
        start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Grouping column indices
    paid_indices = get_column_indices(ws, [f'{pr}_T' for pr in processors])
    pur_indices = get_column_indices(ws, [f'{pr}_Pur' for pr in processors])
    net_indices = get_column_indices(ws, [f'{pr}_Net' for pr in processors])
    Kinray_UPrice_index = get_column_index(ws, 'Kinray_UPrice')
    all_pbmm = get_column_index(ws, 'ALL_PBM_Pur')
    all_pbmd = get_column_index(ws, 'ALL_PBM_D')
    all_pbmt = get_column_index(ws, 'ALL_PBM_T')
    all_pbmq = get_column_index(ws, 'ALL_PBM_Q')
    all_pbmn = get_column_index(ws, 'ALL_PBM_Net')
    all_pbmp = get_column_index(ws, 'ALL_PBM_P')
    raw_groups = [
        quantity_billed_indices,
        package_size_billed_indices,
        package_size_difference_indices,
        paid_indices,
        pur_indices,
        net_indices,
        total_purchased_indices,
        Kinray_UPrice_index,
        all_pbmm,
        all_pbmd,
        all_pbmt,
        all_pbmq,
        all_pbmn,
        all_pbmp
    ]

    # Normalize groups so each entry is a sequence (list). Some items like
    # Kinray_UPrice_index are single ints; the helper expects indexable groups
    # (group[0], group[-1]). Convert ints -> [int], None -> [].
    column_groups = []
    for g in raw_groups:
        if isinstance(g, (list, tuple)):
            column_groups.append(list(g))
        elif isinstance(g, int):
            column_groups.append([g])
        else:
            column_groups.append([])

    apply_thick_border_to_groups(ws, column_groups, start_row, end_row)
    apply_thick_border(ws, start_col=1, end_col=1,
                       start_row=start_row, end_row=end_row)
    apply_thick_border(ws, start_col=2, end_col=2,
                       start_row=start_row, end_row=end_row)
    apply_thick_border(ws, start_col=3, end_col=3,
                       start_row=start_row, end_row=end_row)
    apply_thick_border(ws, start_col=4, end_col=4,
                       start_row=start_row, end_row=end_row)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0, right=0, top=0, bottom=0, header=0, footer=0)

    # Set the title of the active worksheet
    ws.title = "Processed Data"
    # ws.protection.sheet = True
    create_bin_to_processor_sheet(wb, rx_compare_source, bin_to_proc)
