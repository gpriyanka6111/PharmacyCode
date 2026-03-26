# Builds "Needs to be ordered - All" (add_max_difference_sheet) and "Do Not Order - ALL" (min_difference_sheet) sheets.

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.pagebreak import PageBreak

from excel.formatting import set_print_area_excluding_headers


def min_difference_sheet(wb, final_data, insurance_paths=None):
    """
    Build 'Do Not Order - ALL' sheet:
      - Identify all *_D columns (excluding ALL_PBM_D).
      - Keep only rows where ALL *_D values are strictly > 0 (i.e., no deficits anywhere).
      - 'Min Positive' = row-wise minimum across the *_D columns (strictly positive).
      - Display ORIGINAL *_D values (positives remain visible; we don't zero anything for display).

    Parameters
    ----------
    wb : openpyxl.Workbook
    final_data : pandas.DataFrame
        Must contain columns: 'NDC #', 'Drug Name', 'Package Size', and multiple '*_D' columns.
    insurance_paths : unused; kept only for signature parity.
    """
    df = final_data.copy()

    # --- 1) Identify difference columns (exclude ALL_PBM_D)
    difference_columns = [
        c for c in df.columns if c.endswith('_D') and c != 'ALL_PBM_D']

    # Create/replace target sheet
    sheet_title = "Do Not Order - ALL"
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    if not difference_columns:
        ws['A1'] = "No difference columns (*_D) found."
        return

    # Ensure required base columns exist
    for base in ['NDC #', 'Drug Name', 'Package Size']:
        if base not in df.columns:
            df[base] = 0 if base != 'Drug Name' else pd.NA

    # --- 2) Coerce *_D to numeric (for logic), but keep original values for display
    dnum = df[difference_columns].apply(
        pd.to_numeric, errors='coerce').fillna(0)

    # Logic: "Do Not Order" rows are those with ALL positives (>0) across *_D
    # (Mirror of your previous: negatives→0 then min>0; equivalently, (dnum > 0).all(axis=1))
    min_positive = dnum.min(axis=1)
    mask = (dnum > 0).all(axis=1) & (min_positive > 0)

    if not mask.any():
        ws['A1'] = "No rows qualify: no items with all positive differences."
        return

    # --- 3) Build display frame (use ORIGINAL *_D values)
    out = df.loc[mask, ['NDC #', 'Drug Name',
                        'Package Size'] + difference_columns].copy()
    out['Do Not Order'] = min_positive.loc[mask]
    out['Paper Work'] = " "
    out.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)

    display_columns = ['NDC #', 'Drug Name', 'Pkg Size'] + \
        difference_columns + ['Do Not Order', 'Paper Work']
    out = out[display_columns].sort_values('Drug Name')

    # --- 4) Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(display_columns))
    title_cell = ws.cell(row=1, column=1, value="Do not order")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)
    ws.row_dimensions[1].height = 30

    # --- 5) Write table (headers at row 2; data from row 3)
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                # header
                cell.font = Font(bold=True, size=12)
                align = Alignment(horizontal='center', vertical='center')
                if display_columns[c_idx - 1] in difference_columns + ['Pkg Size', 'Do Not Order']:
                    align = Alignment(horizontal='center',
                                      vertical='bottom', text_rotation=90)
                cell.alignment = align
            else:
                # body
                if display_columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=False)
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')
                cell.font = Font(size=12)

    # --- 6) Column widths
    ws.column_dimensions['A'].width = 15   # NDC #
    ws.column_dimensions['B'].width = 50   # Drug Name
    ws.column_dimensions['C'].width = 7    # Pkg Size

    for col_name in difference_columns:
        if col_name in display_columns:
            idx = display_columns.index(col_name) + 1
            ws.column_dimensions[get_column_letter(idx)].width = 7
    for col_name in ['Do Not Order', 'Pkg Size']:
        if col_name in display_columns:
            idx = display_columns.index(col_name) + 1
            ws.column_dimensions[get_column_letter(idx)].width = 8
    # wrap Paper work column text
    if 'Paper Work' in display_columns:
        idx = display_columns.index('Paper Work') + 1
        ws.column_dimensions[get_column_letter(idx)].width = 10
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    # --- 7) Borders and header height
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    thick = Border(left=Side(style='thick'), right=Side(style='thick'),
                   top=Side(style='thick'), bottom=Side(style='thick'))

    ws.row_dimensions[2].height = 80  # header row height

    # Thick border for header row
    for c in range(1, len(display_columns) + 1):
        ws.cell(row=2, column=c).border = thick

    # Thin borders for data cells
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(display_columns)):
        for cell in row:
            cell.border = thin

    # Thick edge borders for key columns
    def apply_column_border(ws_, col_idx):
        col_letter = get_column_letter(col_idx)
        for r in range(2, ws_.max_row + 1):
            c = ws_[f"{col_letter}{r}"]
            c.border = Border(left=thick.left, right=thick.right,
                              top=c.border.top, bottom=c.border.bottom)

    for key in ['NDC #', 'Drug Name', 'Pkg Size', 'Do Not Order', 'Paper Work']:
        if key in display_columns:
            apply_column_border(ws, display_columns.index(key) + 1)

    # --- 8) Freeze panes
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_columns))}{ws.max_row}"


def add_max_difference_sheet(wb, final_data, insurance_paths=None):
    """
    Create a 'Needs to be ordered - All' worksheet showing:
      - All *_D columns (package differences per processor) with original values (positives kept visible)
      - 'To Order' computed ONLY from negatives: max package deficit across all *_D per row
      - PRICE chosen as Kinray_UPrice (if > 0), else first non-zero vendor *_PRICE, else 0
      - 'Total Order Price' = To Order * PRICE

    Parameters
    ----------
    wb : openpyxl.Workbook
    final_data : pandas.DataFrame
        Must contain: 'NDC #', 'Drug Name', 'Package Size', optional 'Kinray_UPrice', vendor '*_PRICE' cols, and '*_D' cols.
    insurance_paths : any (unused; kept for signature compatibility)
    """
    df = final_data.copy()

    # 1) Identify difference columns (exclude ALL_PBM_D)
    difference_columns = [
        c for c in df.columns if c.endswith('_D') and c != 'ALL_PBM_D']

    # Create/replace the target sheet early
    sheet_title = "Needs to be ordered - All"
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    if not difference_columns:
        ws['A1'] = "No difference columns (*_D) found."
        return

    # Make sure base columns exist
    for base in ['NDC #', 'Drug Name', 'Package Size', 'Kinray_UPrice']:
        if base not in df.columns:
            df[base] = 0 if base != 'Drug Name' else pd.NA

    # 2) Coerce numerics
    df[difference_columns] = df[difference_columns].apply(
        pd.to_numeric, errors='coerce').fillna(0)
    df['Kinray_UPrice'] = pd.to_numeric(
        df['Kinray_UPrice'], errors='coerce').fillna(0)

    # 3) Choose PRICE per row
    # Prefer Kinray_UPrice (>0), else first non-zero vendor *_PRICE, else 0
    vendor_price_cols = [c for c in df.columns if c.endswith('_PRICE')]
    if vendor_price_cols:
        vendor_prices = df[vendor_price_cols].apply(
            pd.to_numeric, errors='coerce').replace(0, pd.NA)
        # bring the first non-null from the row to the leftmost position, then pick first col
        first_nonzero_vendor_price = vendor_prices.bfill(
            axis=1).iloc[:, 0].fillna(0)
    else:
        first_nonzero_vendor_price = pd.Series(0, index=df.index)

    df['PRICE'] = np.where(df['Kinray_UPrice'] > 0,
                           df['Kinray_UPrice'], first_nonzero_vendor_price)
    df['PRICE'] = pd.to_numeric(df['PRICE'], errors='coerce').fillna(0)

    # 4) Compute "To Order" using ONLY negatives (positives remain visible in the table)
    neg_for_logic = df[difference_columns].clip(
        upper=0)  # keep negatives, zero out positives
    # any negative deficit across insurers
    needs_mask = neg_for_logic.lt(0).any(axis=1)

    if not needs_mask.any():
        # No ordering needed, but keep the sheet useful
        ws['A1'] = "No rows require ordering (no negative deficits in *_D columns)."
        return

    needs = df.loc[needs_mask].copy()
    needs['To Order'] = neg_for_logic.loc[needs_mask].min(
        axis=1).abs()  # most negative (largest deficit), abs->packages
    needs['Pkg Size'] = needs['Package Size']
    needs['Paper Work'] = " "
    needs['Total Order Price'] = needs['To Order'] * needs['PRICE']
    needs['Total Order Price'] = pd.to_numeric(
        needs['Total Order Price'], errors='coerce').fillna(0)

    # 5) Build display frame (keep ORIGINAL *_D values = positives visible)
    display_columns = (
        ['NDC #', 'Drug Name', 'Pkg Size'] +
        difference_columns +
        ['To Order', 'Paper Work', 'PRICE', 'Total Order Price']
    )
    needs = needs[display_columns].sort_values('Drug Name')

    # 6) Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(display_columns))
    title_cell = ws.cell(
        row=1, column=1, value="Needs to be Ordered (Max Package Deficit Across Insurances)")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 80  # header
    # 7) Write the table (headers at row 2, data from row 3)
    for r_idx, row in enumerate(dataframe_to_rows(needs, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                # header style
                cell.font = Font(bold=True, size=12)
                align = Alignment(horizontal='center', vertical='center')
                if display_columns[c_idx - 1] in difference_columns:
                    align = Alignment(horizontal='center',
                                      vertical='center', text_rotation=90)
                cell.alignment = align
            else:
                # body style
                if display_columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=False)
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

    # 8) Column widths
    ws.column_dimensions['A'].width = 15   # NDC #
    ws.column_dimensions['B'].width = 50   # Drug Name
    ws.column_dimensions['C'].width = 7    # Pkg Size
    # Rotate text in 'Pkg Size' column (C)
    ws["C2"].alignment = Alignment(
        horizontal='center', vertical='center', text_rotation=90)

    for col_name in difference_columns:
        if col_name in display_columns:
            idx = display_columns.index(col_name) + 1
            ws.column_dimensions[get_column_letter(idx)].width = 8
    if 'To Order' in display_columns:
        ws.column_dimensions[get_column_letter(
            display_columns.index('To Order') + 1)].width = 12
    # PAPER WORK column
    if 'Paper Work' in display_columns:
        idx = display_columns.index('Paper Work') + 1
        ws.column_dimensions[get_column_letter(idx)].width = 12
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    if 'PRICE' in display_columns:
        ws.column_dimensions[get_column_letter(
            display_columns.index('PRICE') + 1)].width = 12
    # TOTAL ORDER PRICE column
    if 'Total Order Price' in display_columns:
        idx = display_columns.index('Total Order Price') + 1
        ws.column_dimensions[get_column_letter(idx)].width = 16
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    # 9) Borders and header height
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    thick = Border(left=Side(style='thick'), right=Side(style='thick'),
                   top=Side(style='thick'), bottom=Side(style='thick'))

    for c in range(1, len(display_columns) + 1):
        ws.cell(row=2, column=c).border = thick
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(display_columns)):
        for cell in row:
            cell.border = thin

    # Edge thick borders for key columns (visual groups)
    edge_cols = ['NDC #', 'Drug Name', 'Pkg Size', 'PRICE',
                 'To Order', 'Total Order Price', 'Paper Work']
    for name in edge_cols:
        if name in display_columns:
            idx = display_columns.index(name) + 1
            col_letter = get_column_letter(idx)
            for r in range(2, ws.max_row + 1):
                c = ws[f"{col_letter}{r}"]
                c.border = Border(
                    left=thick.left if c.column == idx else c.border.left,
                    right=thick.right if c.column == idx else c.border.right,
                    top=c.border.top, bottom=c.border.bottom
                )
    # --- page setup BEFORE summary is fine (heights, breaks, etc.) ---
    ws.print_title_rows = "2:2"
    ws.row_breaks = PageBreak()
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[2].height = 80
    ws.freeze_panes = "A3"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- numeric formats for data columns ---
    price_idx = display_columns.index('PRICE') + 1
    total_idx = display_columns.index('Total Order Price') + 1
    to_order_idx = display_columns.index('To Order') + 1
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=price_idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=total_idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=to_order_idx).number_format = '0'

    # --- build the two summary columns (label + amount) ---
    def _find_header_col(ws, header_text, header_row=2):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            if (v if v is not None else "") == header_text:
                return c
        return None

    top_total_hdr_col = _find_header_col(ws, "Total Order Price", header_row=2)
    if top_total_hdr_col is None:
        top_total_hdr_col = display_columns.index("Total Order Price") + 1

    summary_label_col = top_total_hdr_col + 1
    summary_value_col = top_total_hdr_col + 2

    hdr_cell = ws.cell(row=2, column=summary_label_col,
                       value="Insurance-wise Order Estimate ($)")
    hdr_cell.font = Font(bold=True, size=12)
    hdr_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    val_hdr_cell = ws.cell(row=2, column=summary_value_col, value="Amount")
    val_hdr_cell.font = Font(bold=True, size=12)
    val_hdr_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    # Lock the last data row BEFORE we start writing summary data
    last_data_row = ws.max_row
    price_col_letter = get_column_letter(price_idx)

    r = 3
    for diff_col in difference_columns:
        ws.cell(row=r, column=summary_label_col, value=diff_col).alignment = Alignment(
            horizontal="left", vertical="center")
        diff_idx = display_columns.index(diff_col) + 1
        diff_letter = get_column_letter(diff_idx)
        formula = (
            f"=SUMPRODUCT((-{diff_letter}3:{diff_letter}{last_data_row})"
            f"*({diff_letter}3:{diff_letter}{last_data_row}<0),"
            f"{price_col_letter}3:{price_col_letter}{last_data_row})"
        )
        vcell = ws.cell(row=r, column=summary_value_col, value=formula)
        vcell.number_format = '"$"#,##0.00'
        vcell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    # === Formatting for summary columns ===
    for col_idx in (summary_label_col, summary_value_col):
        col_letter = get_column_letter(col_idx)

        # Set fixed width
        ws.column_dimensions[col_letter].width = 20

        # Center + wrap for ALL cells in these columns
        for r in range(2, ws.max_row + 1):      # row 2 = header, rest = values
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    for c in (summary_label_col, summary_value_col):
        ws.cell(row=2, column=c).border = thick
    for rr in range(3, ws.max_row + 1):
        ws.cell(row=rr, column=summary_label_col).border = thin
        ws.cell(row=rr, column=summary_value_col).border = thin

    # === Grand Total footer OUTSIDE sort/filter range ===
    total_col_idx = display_columns.index('Total Order Price') + 1
    total_col_letter = get_column_letter(total_col_idx)
    footer_row = last_data_row + 2

    ws.cell(row=footer_row, column=total_col_idx - 1,
            value="Grand Total").font = Font(bold=True)
    footer_total_cell = ws.cell(
        row=footer_row, column=total_col_idx,
        value=f"=SUBTOTAL(109,{total_col_letter}3:{total_col_letter}{last_data_row})"
    )
    footer_total_cell.font = Font(bold=True)
    footer_total_cell.number_format = '"$"#,##0.00'

    # strong top border across the table width (not across summary columns)
    for c in range(1, len(display_columns) + 1):
        ws.cell(row=footer_row, column=c).border = Border(
            top=Side(style='thick'))

    # === Add conditional formatting for negative values (light grey for B&W printing) ===
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey

    # Apply to all *_D columns
    for diff_col in difference_columns:
        diff_idx = display_columns.index(diff_col) + 1
        diff_letter = get_column_letter(diff_idx)
        data_range = f"{diff_letter}3:{diff_letter}{last_data_row}"

        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=False, fill=grey_fill)
        )

    # === NOW set print area to exclude the two summary columns ===
    set_print_area_excluding_headers(
        ws, header_row=2,
        exclude_headers=["Insurance-wise Order Estimate ($)", "Amount"]
    )

    # === Filter range limited to data only (keeps footer fixed) ===
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_columns))}{last_data_row}"
