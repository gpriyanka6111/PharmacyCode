# Main orchestrator: process_custom_log_data() — loads inputs, merges all data, drives Excel sheet builders.

import os
import re
import shutil

import numpy as np
import pandas as pd
from flask import current_app
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from processing.log_parser import _filter_custom_log_transmitted_paid_ins
from processing.all_pbm_parser import _load_all_pbm_csv
from processing.vendor_parser import parse_vendor_files
from utils.helpers import unblock_file
from excel.formatting import discover_processors_from_df, apply_common_sheet_settings
from excel.order_sheets import add_max_difference_sheet, min_difference_sheet
from excel.support_sheets import create_never_ordered_check_sheet
from excel.rx_comparison_sheets import (add_rx_unit_compare_sheet_exact,
                                        add_rx_unit_compare_sheet_exact_pos,
                                        add_mfp_drugs_sheet)
from excel.summary_sheet import add_summary_sheet
from excel.audit_workbook import generate_master_audit_workbook
from excel.processed_data_sheet import build_processed_data_sheet


def process_custom_log_data(custom_log_path, bin_master_path, vendor_paths, pharmacy_name, date_range, all_pbm_path,
                            selected_processors=None, selected_sheets=None, vendor_count=None,
                            job_dir=None, user_audit_dir=None):
    # ===== Load =====
    bin_df = pd.read_csv(bin_master_path, dtype=str)
    log_df = pd.read_csv(custom_log_path, dtype=str)

    # Normalize incoming headers
    log_df.columns = [str(c).strip() for c in log_df.columns]

    # Keep only insurance-adjudicated rows early (ignore cash/other states)
    log_df, _status_col, _kept_rows, _dropped_rows = _filter_custom_log_transmitted_paid_ins(log_df)

    # Normalize "Drug NDC" -> "NDC #", if needed
    for c in list(log_df.columns):
        if c.strip().lower() in ('drug ndc', 'drug ndc#', 'drug ndc #'):
            log_df.rename(columns={c: 'NDC #'}, inplace=True)
            break

    # Normalize SDRA naming to one canonical column
    if '* SDRA Amt' in log_df.columns and 'SDRA Amt' not in log_df.columns:
        log_df.rename(columns={'* SDRA Amt': 'SDRA Amt'}, inplace=True)
    elif 'SDRA' in log_df.columns and 'SDRA Amt' not in log_df.columns:
        log_df.rename(columns={'SDRA': 'SDRA Amt'}, inplace=True)

    # Normalize Copay naming/case to one canonical column
    if 'Copay' in log_df.columns and 'COPAY' not in log_df.columns:
        log_df.rename(columns={'Copay': 'COPAY'}, inplace=True)

    # ===== Validate required columns =====
    need_log = [
        'Rx #', 'NDC #', 'Drug Name',
        'Plan 1 BIN', 'Plan 1 PCN', 'Plan 1 Group #',
        'Plan 2 BIN', 'Plan 2 PCN', 'Plan 2 Group #',
        'Ins Paid Plan 1', 'Ins Paid Plan 2',
        'Qty Filled', 'Drug Pkg Size',
        'SDRA Amt', 'COPAY'
    ]
    missing = [c for c in need_log if c not in log_df.columns]
    if missing:
        raise ValueError(
            f"Custom Log missing required column(s): {', '.join(missing)}")

    # User requirement: treat the custom log as a 20-column mandatory input
    if len(log_df.columns) < 20:
        raise ValueError(
            f"Custom Log must contain at least 20 columns. Found {len(log_df.columns)} columns.")

    if 'BIN' not in bin_df.columns or 'Processor' not in bin_df.columns:
        raise ValueError(
            "BIN master must contain 'BIN' and 'Processor' columns.")

    # ===== Normalize =====
    # BIN master
    bin_df['BIN'] = (bin_df['BIN'].astype(str)
                     .str.replace(r'\D', '', regex=True)
                     .str.zfill(6))

    bin_df['Processor'] = bin_df['Processor'].astype(str).str.strip()
    # >>> Build the BIN -> Processor map ONCE (used below for row-level filter)
    bin_to_proc = dict(zip(bin_df['BIN'], bin_df['Processor']))

    # Custom log
    for col in ['Plan 1 BIN', 'Plan 2 BIN']:
        log_df[col] = (log_df[col].astype(str)
                                  .str.replace(r'\D', '', regex=True)
                                  .str.zfill(6))

    for col in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'SDRA Amt', 'COPAY']:
        log_df[col] = pd.to_numeric(log_df[col], errors='coerce').fillna(0)

    log_df['NDC #'] = (log_df['NDC #'].astype(str)
                       .str.replace('-', '', regex=False)
                       .str.strip()
                       .str.zfill(11))

    log_df['Drug Name'] = (
        log_df['Drug Name']
        .astype(str)
        .str.strip()
        .str.replace(r'\*+$', '', regex=True)  # remove only trailing ****
    )

    # ===== Choose winning BIN per row =====
    log_df['Winning_BIN'] = log_df.apply(
        lambda r: r['Plan 1 BIN'] if r['Ins Paid Plan 1'] >= r['Ins Paid Plan 2'] else r['Plan 2 BIN'],
        axis=1
    ).str.zfill(6)

    # capture the winning insurance dollars only (this becomes Processor_T later)
    log_df['Winning_Paid'] = np.where(
        log_df['Winning_BIN'] == log_df['Plan 1 BIN'],
        log_df['Ins Paid Plan 1'],
        log_df['Ins Paid Plan 2']
    )
    log_df['Winning PCN'] = np.where(
        log_df['Winning_BIN'] == log_df['Plan 1 BIN'],
        log_df['Plan 1 PCN'],
        log_df['Plan 2 PCN']
    )
    log_df['Winning Group'] = np.where(
        log_df['Winning_BIN'] == log_df['Plan 1 BIN'],
        log_df['Plan 1 Group #'],
        log_df['Plan 2 Group #']
    )
    # >>> NEW: determine row-level Processor now and FILTER if user selected any
    log_df['Processor'] = log_df['Winning_BIN'].map(bin_to_proc)
    rx_compare_source = log_df.copy()
    if selected_processors:
        # Normalize case and whitespace just to be safe
        allowed = {p.strip().casefold() for p in selected_processors}
        log_df = log_df[
            log_df['Processor'].fillna("").astype(
                str).str.strip().str.casefold().isin(allowed)
        ].copy()

    # ===== 1) Aggregate BY (NDC #, Winning_BIN) FIRST =====
    agg_bin = (log_df.groupby(['NDC #', 'Winning_BIN'], as_index=False)
               .agg({'Qty Filled': 'sum',
                     'Winning_Paid': 'sum',
                     'Drug Name': 'first',
                     'Drug Pkg Size': 'first'}))

    # ===== 2) Map BIN -> Processor AFTER aggregation =====
    bin_to_proc = dict(zip(bin_df['BIN'], bin_df['Processor']))
    agg_bin['Processor'] = agg_bin['Winning_BIN'].map(bin_to_proc)
    agg_bin = agg_bin[agg_bin['Processor'].notna()].copy()

    # >>> NEW: collapse to (NDC, Processor) both Qty and Paid
    grp_proc = (agg_bin.groupby(['NDC #', 'Processor'], as_index=False).agg(

        {'Qty Filled': 'sum', 'Winning_Paid': 'sum'}))

    # unique_procs = grp_proc['Processor'].dropna().astype(str).str.strip().unique().tolist()

    # Identify rows whose Winning_BIN is missing in the master map
    unmapped_mask = ~log_df['Winning_BIN'].isin(bin_to_proc.keys())
    unmapped = log_df.loc[unmapped_mask].copy()

    # If your custom log has 'Rx #' column, collect per BIN; else leave empty
    if 'Rx #' in unmapped.columns:
        # make RXs concise: comma-separated unique RX # per BIN
        rx_by_unmapped_bin = (
            unmapped.groupby('Winning_BIN')['Rx #']
                    .apply(lambda s: ', '.join(sorted(set(str(x).strip() for x in s if pd.notna(x) and str(x).strip()))))
                    .to_dict()
        )
    else:
        rx_by_unmapped_bin = {}  # no RX info available

    used_bins = (log_df['Winning_BIN'].astype(str)
                 .str.replace(r'\D', '', regex=True)
                 .str.zfill(6)
                 .dropna()
                 .unique()
                 )

    # ===== Build package size & name per NDC from the FULL custom log =====
    pkg_df = (
        log_df[['NDC #', 'Drug Pkg Size', 'Drug Name']]
        .rename(columns={'Drug Pkg Size': 'Package Size'})
        .copy()
    )

    pkg_df = (log_df[['NDC #', 'Drug Pkg Size', 'Drug Name']]
              .rename(columns={'Drug Pkg Size': 'Package Size'}))
    pkg_df['Package Size'] = pd.to_numeric(
        pkg_df['Package Size'], errors='coerce')
    pkg_df = (pkg_df
              .dropna(subset=['Package Size'])
              .drop_duplicates(subset=['NDC #']))

    # ===== 3) Pivots to Processor_Q and Processor_T =====
    # We already built grp_proc above with both Qty Filled and Winning_Paid.
    # >>> CHANGED: do NOT overwrite grp_proc again; keep both measures in it.
    grp_q = grp_proc.pivot(index='NDC #', columns='Processor',
                           values='Qty Filled').fillna(0).reset_index()
    pivot_q = grp_q.copy()
    pivot_q.columns = ['NDC #'] + \
        [f'{c}_Q' for c in pivot_q.columns if c != 'NDC #']

    # >>> NEW: Insurance dollars per processor → *_T
    grp_t = grp_proc.pivot(index='NDC #', columns='Processor',
                           values='Winning_Paid').fillna(0).reset_index()
    pivot_t = grp_t.copy()
    pivot_t.columns = ['NDC #'] + \
        [f'{c}_T' for c in pivot_t.columns if c != 'NDC #']

    # ===== 4) Process vendor files =====
    vendor_pivot, vendor_price_pivot, vendor_names, kinray_all, kinray_latest, all_vendor_df = \
        parse_vendor_files(vendor_paths)
    # ===== END VENDOR AGGREGATION =====

    # ===== Merge everything =====
    merged = (pivot_q
              .merge(pivot_t, on='NDC #', how='left')          # brings *_T
              # Vendor1, Vendor2, ...
              .merge(vendor_pivot, on='NDC #', how='left')
              # Package Size, Drug Name
              .merge(pkg_df, on='NDC #', how='left')
              .merge(vendor_price_pivot, on='NDC #', how='left')
              .merge(kinray_latest, on='NDC #', how='left'))   # Kinray_UPrice

    # Ensure vendor qty cols exist & numeric
    for vn in vendor_names:
        if vn not in merged.columns:
            merged[vn] = 0
    merged[vendor_names] = merged[vendor_names].apply(
        pd.to_numeric, errors='coerce').fillna(0)

    # Ensure price cols numeric too
    price_cols = [c for c in merged.columns if c.endswith('_PRICE')]
    if price_cols:
        merged[price_cols] = merged[price_cols].apply(
            pd.to_numeric, errors='coerce').fillna(0)

    # Total Purchased = sum of vendor shipped qty across all vendor columns
    merged['Total Purchased'] = merged[vendor_names].sum(axis=1)

    # Normalize left key and numerics we'll use
    merged['NDC #'] = (merged['NDC #'].astype(str)
                       .str.replace(r'\D', '', regex=True)
                       .str.zfill(11))
    merged['Kinray_UPrice'] = pd.to_numeric(merged.get(
        'Kinray_UPrice', 0), errors='coerce').fillna(0)
    pkg = pd.to_numeric(merged.get('Package Size', 0),
                        errors='coerce').fillna(0)

    # ===== Bring in ALL PBM (Quantity & Total $) =====
    if all_pbm_path:
        # should return NDC #, ALL_PBM_Q, ALL_PBM_T, ALL_PBM_DrugName
        all_pbm = _load_all_pbm_csv(all_pbm_path)

        # Back-compat: if the file uses 'Total' instead of ALL_PBM_T
        if 'Total' in all_pbm.columns and 'ALL_PBM_T' not in all_pbm.columns:
            all_pbm = all_pbm.rename(columns={'Total': 'ALL_PBM_T'})

        # one row per NDC (defensive)
        all_pbm = all_pbm.drop_duplicates(subset=['NDC #'], keep='last')

        # only map onto rows you already have
        merged = merged.merge(
            all_pbm[['NDC #', 'ALL_PBM_Q', 'ALL_PBM_T', 'ALL_PBM_DrugName']],
            on='NDC #',
            how='left'
        )
    else:
        merged['ALL_PBM_Q'] = 0
        merged['ALL_PBM_T'] = 0
        merged['ALL_PBM_DrugName'] = pd.NA

    # Fill Drug Name from ALL PBM if missing
    if 'Drug Name' not in merged.columns:
        merged['Drug Name'] = pd.NA
    merged['Drug Name'] = merged['Drug Name'].fillna(
        merged.get('ALL_PBM_DrugName'))

    def ensure_numeric_col(df, col, default=0):
        if col not in df.columns:
            df.loc[:, col] = default
        df.loc[:, col] = pd.to_numeric(
            df[col], errors='coerce').fillna(default)

    # Ensure numeric + compute PBM derived columns
    ensure_numeric_col(merged, 'ALL_PBM_Q', 0)
    ensure_numeric_col(merged, 'ALL_PBM_T', 0)

    pkg = pd.to_numeric(merged.get('Package Size', 0),
                        errors='coerce').fillna(0)
    merged['ALL_PBM_P'] = (merged['ALL_PBM_Q'] / pkg).where(pkg > 0, 0)
    merged['ALL_PBM_D'] = merged['Total Purchased'] - merged['ALL_PBM_P']
    merged['ALL_PBM_Pur'] = merged['ALL_PBM_P'] * \
        pd.to_numeric(merged.get('Kinray_UPrice', 0),
                      errors='coerce').fillna(0)

    # ===== Discover processors from either *_Q or *_T (robust union) =====
    procs_from_q = {c[:-2] for c in merged.columns if c.endswith('_Q')}
    procs_from_t = {c[:-2] for c in merged.columns if c.endswith('_T')}
    processors = sorted(procs_from_q.union(procs_from_t))

    # Ensure missing *_Q and *_T exist (so later loops never KeyError)
    for pr in processors:
        qcol, tcol = f'{pr}_Q', f'{pr}_T'
        if qcol not in merged.columns:
            merged[qcol] = 0
        if tcol not in merged.columns:
            merged[tcol] = 0

    # Convert *_Q/*_T numeric
    for pr in processors:
        merged[f'{pr}_Q'] = pd.to_numeric(
            merged[f'{pr}_Q'], errors='coerce').fillna(0)
        merged[f'{pr}_T'] = pd.to_numeric(
            merged[f'{pr}_T'], errors='coerce').fillna(0)

    # ===== Build per-processor derived bands for *every* processor
    for pr in processors:
        q = merged[f'{pr}_Q']
        p = (q / pkg).where(pkg > 0, 0)                 # packages billed
        merged[f'{pr}_P'] = p
        merged[f'{pr}_D'] = merged['Total Purchased'] - p
        merged[f'{pr}_Pur'] = p * merged['Kinray_UPrice']
        merged[f'{pr}_Net'] = merged[f'{pr}_T'] - merged[f'{pr}_Pur']

    # ===== Final column ordering (your "desired columns" spec) =====

    def have(cols): return [c for c in cols if c in merged.columns]

    # Discover processors FROM CURRENT COLUMNS (better than from desired_columns)
    def _discover_processors_from_columns(cols):
        procs_q = {c[:-2] for c in cols if c.endswith('_Q')}
        procs_t = {c[:-2] for c in cols if c.endswith('_T')}
        procs_p = {c[:-2] for c in cols if c.endswith('_P')}
        procs_d = {c[:-2] for c in cols if c.endswith('_D')}
        procs_pur = {c[:-4] for c in cols if c.endswith('_Pur')}
        procs_net = {c[:-4] for c in cols if c.endswith('_Net')}
        return sorted(procs_q | procs_t | procs_p | procs_d | procs_pur | procs_net)

    # ✅ define processors BEFORE using them in bands
    processors = _discover_processors_from_columns(merged.columns)
    # 🔥 Force ALL_PBM first, others alphabetical
    if 'ALL_PBM' in processors:
        processors = ['ALL_PBM'] + \
            sorted([p for p in processors if p != 'ALL_PBM'])
    else:
        processors = sorted(processors)

    base_cols = ['NDC #', 'Drug Name', 'Package Size']
    vendor_qty_cols = vendor_names
    qty_band = have([f'{pr}_Q' for pr in processors])
    pkg_band = have([f'{pr}_P' for pr in processors])
    diff_band = have([f'{pr}_D' for pr in processors])
    paid_band = have([f'{pr}_T' for pr in processors])  # dollars paid
    # $$ purchased (Kinray)
    pur_band = have([f'{pr}_Pur' for pr in processors])
    net_band = have([f'{pr}_Net' for pr in processors])  # paid − purchased
    other_cols = have(['Total Purchased', 'Kinray_UPrice'])

    # ✅ Round off paid_band, pur_band, and net_band columns (no decimals)
    for band in [paid_band, pur_band, net_band]:
        for col in band:
            if col in merged.columns:
                merged[col] = np.round(merged[col]).astype('Int64')

    desired_columns = (
        base_cols +
        vendor_qty_cols + ['Total Purchased'] +
        [c for c in other_cols if c not in ('Total Purchased',)] +
        qty_band + pkg_band + diff_band +
        paid_band + pur_band + net_band
    )

    def _proc_from_col(col: str):
        suffixes = ('_Q', '_P', '_D', '_T', '_Pur', '_Net')
        for sfx in suffixes:
            if col.endswith(sfx):
                return col[:-len(sfx)]
        return None  # not a processor metric column

    # ✅ NEW: remove deselected processors' columns dynamically (ALWAYS include ALL_PBM)
    if selected_processors:
        # normalize to a set for membership and add forced include
        selected_upper = {p.strip().upper() for p in selected_processors}
        selected_upper.add('ALL_PBM')

        keep_cols = []
        for col in desired_columns:
            proc = _proc_from_col(col)
            if proc is None:
                # base/vendor/other columns (not *_Q/_P/_D/_T/_Pur/_Net)
                keep_cols.append(col)
            else:
                if proc.strip().upper() in selected_upper:
                    keep_cols.append(col)

        desired_columns = keep_cols
        merged = merged.reindex(
            columns=[c for c in desired_columns if c in merged.columns])

    # Keep only existing (defensive) and sort by Drug Name
    desired_columns = [c for c in desired_columns if c in merged.columns]
    final = merged[desired_columns].sort_values(
        'Drug Name', na_position='last')

    # Use the SAME extractor to compute processors from the final df
    def processors_from_df(df):
        procs = set()
        for c in df.columns:
            p = _proc_from_col(c)
            if p:
                procs.add(p)
        return sorted(procs)

    processors = processors_from_df(final)
    # Ensure ALL_PBM is first if present
    if 'ALL_PBM' in processors:
        processors = ['ALL_PBM'] + \
            sorted([p for p in processors if p != 'ALL_PBM'])
    else:
        processors = sorted(processors)

    # Build NDC → Drug Type map from Kinray
    kinray_type_map = {}
    try:
        kinray_kpath = [path for path in vendor_paths if 'kinray' in path.lower()][0]
        kdf = pd.read_csv(kinray_kpath, dtype=str)
        kdf = kdf[
            kdf['Invoice Number'].notna() &
            (kdf['Invoice Number'].astype(str).str.strip().ne('')) &
            (kdf['Invoice Number'].astype(str).str.strip().ne('nan')) &
            (kdf['Invoice Number'].astype(str).str.strip().ne('Invoice Number'))
        ].copy()
        kdf['NDC_norm'] = (
            kdf['NDC/UPC'].astype(str)
            .str.replace(r'\D', '', regex=True)
            .str.lstrip('0')
        )
        kinray_type_map = dict(zip(kdf['NDC_norm'], kdf['Type']))
    except Exception as e:
        print(f'[DEBUG] Drug type map error: {e}')

    # Add Drug Type to final
    if 'NDC #' in final.columns:
        final['NDC_norm'] = (
            final['NDC #'].astype(str)
            .str.replace(r'\D', '', regex=True)
            .str.lstrip('0')
        )
        final['Drug Type'] = (
            final['NDC_norm']
            .map(kinray_type_map)
            .fillna('Unclassified')
        )
        final = final.drop(columns=['NDC_norm'])

    # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    # NEW: Split out rows that have NaN anywhere in the final report
    nan_mask = final.isna().any(axis=1)

    # build a small reasons frame (which columns were NaN)
    def _nan_cols(row):
        return ", ".join([col for col, is_nan in row.isna().items() if is_nan])

    nan_reason_df = final.loc[nan_mask].apply(_nan_cols, axis=1)
    final_with_reason = final.copy()
    final_with_reason.loc[nan_mask, 'Reason_NaN_Columns'] = nan_reason_df

    # main sheet (clean): drop NaN rows
    final_clean = final_with_reason.loc[~nan_mask].copy()

    # important sheet: original Custom Log rows for those NDCs that had NaN in final
    nan_ndcs = set(final.loc[nan_mask, 'NDC #']
                   ) if 'NDC #' in final.columns else set()
    important_rows = log_df[log_df['NDC #'].isin(nan_ndcs)].copy()

    # annotate why (optional)
    if not important_rows.empty:
        # map NDC -> reason string (from merged/final)
        ndc_to_reason = final_with_reason.loc[nan_mask, [
            'NDC #', 'Reason_NaN_Columns']].drop_duplicates()
        # merge to show reason next to original rows
        important_rows = important_rows.merge(
            ndc_to_reason, on='NDC #', how='left')
        # put an attention banner col
        important_rows.insert(
            0, '⚠️ Check', 'This NDC had NaN in merged report — investigate!')

    # Save into app's processed folder so /download can serve it
    safe_name = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_',
                       f'{pharmacy_name} ({date_range}).xlsx')
    output_dir = os.path.join(current_app.root_path, current_app.config.get(
        'PROCESSED_FOLDER', 'processed'))
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, safe_name)
    final.to_excel(output_file,  index=False, float_format="%.3f")

    #print(f"Processed file saved at: {output_file}")  # Debugging line

    # written_data = pd.read_excel(output_file)
    # written_data = final_df.copy()
    if not os.path.exists(output_file):
        raise FileNotFoundError(f"Processed file not found at {output_file}")

    wb = load_workbook(output_file)
    ws = wb.active
    build_processed_data_sheet(wb, ws, final, desired_columns, processors,
                               pharmacy_name, date_range,
                               rx_compare_source, bin_to_proc)
    # ===== Create/replace "Vendor Data" sheet =====
    vendor_dfs = []
    for p in vendor_paths:                      # you already build this list earlier
        try:
            vendor_dfs.append(pd.read_csv(p, dtype=str))
        except Exception as e:
            print(f"[warn] vendor read failed {p}: {e}")

    vendor_df_all = (pd.concat(vendor_dfs, ignore_index=True)
                     if vendor_dfs else pd.DataFrame())

    try:
        # If your functions are in the same file, make sure they are defined ABOVE this call.
        # `insurance_paths` is optional; pass None (or a list if you actually use it).
        add_max_difference_sheet(wb, final, insurance_paths=None)
        min_difference_sheet(wb, final, insurance_paths=None)
        create_never_ordered_check_sheet(wb, final)
        add_rx_unit_compare_sheet_exact(
            wb, log_df=rx_compare_source, kinray_df=kinray_all, sheet_name="RX Comparison - All")
        add_rx_unit_compare_sheet_exact_pos(
            wb, log_df=rx_compare_source, kinray_df=kinray_all, sheet_name="RX Comparison +ve")
        add_mfp_drugs_sheet(
            wb, log_df=rx_compare_source, kinray_df=kinray_all, sheet_name="MFP Drugs - RX")
        # add_zero_refills_sheet(
        #     wb, log_df=rx_compare_source, sheet_name="Refills 0 - Call Doctor")
        #add_missed_refill_revenue_sheet(
            #wb, log_df=rx_compare_source, sheet_name="Missed Refill - Revenue Recovery", grace_days=7)
        add_summary_sheet(wb, processed_source="Processed Data", needs_title="Needs to be ordered - All",
                          header_row=3, data_start_row=4, pharmacy_name=pharmacy_name, date_range=date_range)
        #add_alternate_ndc_sheet(wb, custom_log_df, all_vendor_df)
        # ALT_SHEET_NAME = "Alternate NDC - Purchased"
        # add_alternate_ndc_sheet(wb, log_df, all_vendor_df, sheet_name=ALT_SHEET_NAME)


        audit_df = final_clean if 'final_clean' in locals() else final
        audit_path = None
        audit_name = None
        try:
            audit_path = generate_master_audit_workbook(
                audit_df,
                pharmacy_name=pharmacy_name,
                date_range=date_range,
                output_dir=output_dir
            )
            if audit_path:
                audit_name = os.path.basename(audit_path)

                # If caller requested a copy to a user-specified folder, copy it there
                if user_audit_dir:
                    try:
                        os.makedirs(user_audit_dir, exist_ok=True)
                        dest = os.path.join(user_audit_dir, audit_name)
                        shutil.copy2(audit_path, dest)
                        #print(f"[info] Copied audit workbook to user folder: {dest}")
                    except Exception as _e:
                        print(f"[warn] Could not copy audit workbook to user folder {user_audit_dir}: {_e}")
        except Exception as e:
            print(f"Order helper sheets skipped (audit): {e}")
            # keep going — audit optional
            audit_path = None
            audit_name = None
    except Exception as e:
        print(f"Order helper sheets skipped: {e}")

        # Apply the shared titles/orientation/print settings
    processors = discover_processors_from_df(
        final)   # or pass a list you already have
    apply_common_sheet_settings(wb, pharmacy_name=pharmacy_name,
                                date_range=date_range, processors=processors, header_row_main=3)

    # Round numeric values in key sheets to 2 decimal places
    for sheet_name in ["Needs to be ordered - All", "Do Not Order - ALL", "Never Ordered - Check", "Refills 0 - Call Doctor", "RX Comparison - All", "RX Comparison +ve", "Missed Refill - Revenue Recovery"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=3):  # skip header rows
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.value = round(cell.value, 2)

    # if selected_sheets:
    #     keep = set(selected_sheets)
    #     # always ensure the main sheet stays
    #     keep.add("Processed Data")
    #     keep.add(ALT_SHEET_NAME)   # 👈 NEW
    # else:
    #     keep = None  # means keep everything

    # if keep is not None:
    #     for sheet in wb.sheetnames.copy():
    #         if sheet not in keep:
    #             wb.remove(wb[sheet])

    if "Processed Data" in wb.sheetnames:
        ws = wb["Processed Data"]
        header_row = 3  # your main headers are on row 3
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(output_file)
    unblock_file(output_file)

    return{
        "main" : safe_name,
        "audit": audit_name
    }
