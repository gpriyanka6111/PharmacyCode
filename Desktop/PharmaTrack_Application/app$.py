"""
==================================================
Script: app.py
Author: Priyanka Gulgari
Date: 06-18-2024
Description: This script is a Flask web application designed to process pharmacy data. It allows users to upload multiple Excel files, including BestRx data, vendor data, and conversion data.
The script processes these files to aggregate and merge data, calculate package sizes, and generate a comprehensive report in Excel format.
The report includes a summary of purchased quantities,billed quantities, package size differences, and highlights any missing items that need to be updated in the master file.
The application supports optional insurance files and ensures data integrity throughout the process.
==================================================

License:
This script is the intellectual property of Priyanka Gulgari.
Unauthorized copying, distribution, modification, or use of this code, via any medium, is strictly prohibited without prior written permission from the author.

Contact:
For permissions or inquiries, please contact priyankagulgari@gmail.com .

==================================================
"""
# Standard library imports
import csv
import json
import mimetypes
import os
import re
import shutil
import smtplib
import sys
import tempfile
import threading
import urllib.parse
import webbrowser
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from glob import glob
from urllib.parse import quote
from uuid import uuid4

# Third-party imports
import numpy as np
import pandas as pd
from flask import (Flask, current_app, jsonify, make_response, redirect,
                   render_template, request, send_file, send_from_directory,
                   url_for)
from flaskwebgui import FlaskUI
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break, PageBreak
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename
import tkinter as tk
from tkinter import filedialog

from config import APP_DISPLAY_NAME, BASE_DIR, UPLOAD_FOLDER, PROCESSED_FOLDER
from utils.helpers import resource_path, get_screen_dimensions, unblock_file
from processing.all_pbm_parser import _load_all_pbm_csv
from processing.log_parser import _filter_custom_log_transmitted_paid_ins
from processing.vendor_parser import parse_vendor_files
from excel.formatting import (get_column_index, discover_processors_from_df,
                               apply_common_sheet_settings, set_print_area_excluding_headers)
from excel.order_sheets import add_max_difference_sheet, min_difference_sheet
from excel.support_sheets import create_never_ordered_check_sheet, create_bin_to_processor_sheet
from excel.rx_comparison_sheets import (add_rx_unit_compare_sheet_exact,
                                        add_rx_unit_compare_sheet_exact_pos,
                                        add_mfp_drugs_sheet)
from excel.refill_sheets import add_zero_refills_sheet
from excel.summary_sheet import add_summary_sheet
from excel.audit_workbook import generate_master_audit_workbook
from excel.processed_data_sheet import build_processed_data_sheet
from processing.pipeline import process_custom_log_data

# Initialize Tkinter window for screen dimensions
screen_width, screen_height = get_screen_dimensions()


app = Flask(__name__, template_folder='templates')
app.static_folder = 'static'

# window = webview.create_window('Pharmacy Data Processing Application',app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(PROCESSED_FOLDER):
    os.makedirs(PROCESSED_FOLDER)


from routes.main import bp as main_bp
app.register_blueprint(main_bp)



'''
def build_alternate_ndc_df(log_df: pd.DataFrame,
                           all_vendor_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build a DataFrame describing alternate NDC (same NDC-9) purchase opportunities.

    Uses:
      - log_df    : custom log with 'NDC #', 'Drug Name', 'Drug Pkg Size', 'Qty Filled'
      - all_vendor_df : vendor rows with 'NDC #', 'Shipped', 'PRICE', 'DATE', 'Vendor', 'Drug Name' (optional)

    Each row ~ (Billed NDC, Alternate Purchased NDC)

    Final columns:
      Generic_Group, Billed_NDC, Billed Drug Name, Billed Pkg Size,
      Total Qty Billed (NDC), Billed Unit Cost,
      Alt_NDC, Alt Drug Name, Alt Last Purchase Date, Alt Unit Cost, Alt Vendor
    """

    # ---------- 1) Prepare billed side ----------
    df_log = log_df.copy()

    need_cols = ['NDC #', 'Drug Name', 'Drug Pkg Size', 'Qty Filled']
    missing = [c for c in need_cols if c not in df_log.columns]
    if missing:
        raise KeyError(f"Custom log missing required cols for alt NDC: {missing}")

    df_log["NDC11_BILLED"] = df_log["NDC #"].apply(_alt_norm_ndc11)
    df_log = df_log[~df_log["NDC11_BILLED"].isna()].copy()
    df_log["NDC9"] = df_log["NDC11_BILLED"].apply(_alt_ndc11_to_ndc9)

    df_log["Qty Filled"] = pd.to_numeric(df_log["Qty Filled"], errors="coerce").fillna(0)
    df_log["Drug Pkg Size"] = pd.to_numeric(df_log["Drug Pkg Size"], errors="coerce").fillna(0)

    billed_ndc = (
        df_log.groupby(["NDC9", "NDC11_BILLED"], as_index=False)
              .agg(
                  Total_Qty_Billed=("Qty Filled", "sum"),
                  Billed_Drug_Name=("Drug Name", "first"),
                  Billed_Pkg_Size=("Drug Pkg Size", "first"),
              )
    )

    if billed_ndc.empty:
        return pd.DataFrame(columns=[
            "Generic_Group",
            "Billed_NDC", "Billed Drug Name", "Billed Pkg Size",
            "Total Qty Billed (NDC)",
            "Billed Unit Cost",
            "Alt_NDC", "Alt Drug Name", "Alt Last Purchase Date",
            "Alt Unit Cost", "Alt Vendor",
        ])

    # ---------- 2) Prepare vendor side ----------
    df_v = all_vendor_df.copy()
    # required for pricing; Drug Name optional
    need_v_cols = ["NDC #", "Shipped", "PRICE", "DATE", "Vendor"]
    missing_v = [c for c in need_v_cols if c not in df_v.columns]
    if missing_v:
        raise KeyError(f"all_vendor_df missing required cols: {missing_v}")

    if "Drug Name" not in df_v.columns:
        df_v["Drug Name"] = ""  # optional, but we want Alt Drug Name

    df_v["NDC11"] = df_v["NDC #"].apply(_alt_norm_ndc11)
    df_v = df_v[~df_v["NDC11"].isna()].copy()
    df_v["NDC9"] = df_v["NDC11"].apply(_alt_ndc11_to_ndc9)

    df_v["Shipped"] = pd.to_numeric(df_v["Shipped"], errors="coerce").fillna(0)
    df_v["PRICE"] = pd.to_numeric(df_v["PRICE"], errors="coerce").fillna(0)
    df_v["DATE"] = pd.to_datetime(df_v["DATE"], errors="coerce")

    vend_ndc = (
        df_v.sort_values(["NDC9", "NDC11", "DATE"])
            .groupby(["NDC9", "NDC11", "Vendor"], as_index=False)
            .agg(
                Alt_Last_Purchase_Date=("DATE", "last"),
                Total_Shipped=("Shipped", "sum"),
                Total_Invoice=("PRICE", "sum"),
                Alt_Drug_Name=("Drug Name", "last"),
            )
    )

    vend_ndc["Alt_Unit_Cost"] = np.where(
        vend_ndc["Total_Shipped"] > 0,
        vend_ndc["Total_Invoice"] / vend_ndc["Total_Shipped"],
        np.nan,
    )

    # ---------- 3) Billed unit cost ----------
    billed_cost = vend_ndc[["NDC9", "NDC11", "Vendor", "Alt_Unit_Cost"]].rename(
        columns={"NDC11": "NDC11_BILLED", "Alt_Unit_Cost": "Billed_Unit_Cost"}
    )

    # ---------- 4) Merge billed side with alts ----------
    vend_ndc_alt = vend_ndc.rename(columns={"NDC11": "Alt_NDC"})

    merged = billed_ndc.merge(
        vend_ndc_alt,
        on="NDC9",
        how="left",
    )

    merged = merged[merged["Alt_NDC"] != merged["NDC11_BILLED"]].copy()
    if merged.empty:
        return pd.DataFrame(columns=[
            "Generic_Group",
            "Billed_NDC", "Billed Drug Name", "Billed Pkg Size",
            "Total Qty Billed (NDC)",
            "Billed Unit Cost",
            "Alt_NDC", "Alt Drug Name", "Alt Last Purchase Date",
            "Alt Unit Cost", "Alt Vendor",
        ])

    # billed unit cost collapsed per NDC9 + billed NDC (ignore vendor differences)
    billed_cost_collapsed = (
        billed_cost.groupby(["NDC9", "NDC11_BILLED"], as_index=False)
                   .agg(Billed_Unit_Cost=("Billed_Unit_Cost", "mean"))
    )

    merged = merged.merge(
        billed_cost_collapsed,
        on=["NDC9", "NDC11_BILLED"],
        how="left",
    )

    # ---------- 5) Final mapping (no diff/savings now) ----------
    alt_df = merged.assign(
        Generic_Group=lambda df: df["NDC9"],
        Billed_NDC=lambda df: df["NDC11_BILLED"],
    )

    alt_df = alt_df.rename(columns={
        "Billed_Drug_Name": "Billed Drug Name",
        "Billed_Pkg_Size": "Billed Pkg Size",
        "Alt_Unit_Cost": "Alt Unit Cost",
        "Billed_Unit_Cost": "Billed Unit Cost",
    })

    alt_df = alt_df[
        [
            "Generic_Group",
            "Billed_NDC", "Billed Drug Name", "Billed Pkg Size",
            "Total_Qty_Billed",  # will rename just below
            "Billed Unit Cost",
            "Alt_NDC", "Alt_Drug_Name", "Alt_Last_Purchase_Date",
            "Alt Unit Cost", "Vendor",
        ]
    ]

    alt_df = alt_df.rename(columns={
        "Total_Qty_Billed": "Total Qty Billed (NDC)",
        "Alt_Drug_Name": "Alt Drug Name",
        "Alt_Last_Purchase_Date": "Alt Last Purchase Date",
        "Vendor": "Alt Vendor",
    })

    # If you want some ordering on best alts even w/o explicit savings,
    # we can sort by Alt Unit Cost ascending for each generic group + billed NDC.
    alt_df = alt_df.sort_values(
        ["Generic_Group", "Billed_NDC", "Alt Unit Cost"],
        ascending=[True, True, True],
    ).reset_index(drop=True)

    return alt_df

'''


# def open_browser():
#     webbrowser.open_new("http://127.0.0.1:5000/")

# if __name__ == '__main__':
#     threading.Timer(1.5, open_browser).start()

#     app.run(
#         host="127.0.0.1",
#         port=5000,
#         debug=False,
#         use_reloader=False
#     )
