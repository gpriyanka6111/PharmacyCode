"""
==================================================
Script: app.py
Author: Priyanka Gulgari
Date: 06-18-2024
Description: This script is a Flask web application designed to process pharmacy data. It allows users to upload multiple Excel files, including BestRx data, vendor data, and conversion data.
The script processes these files to aggregate and merge data, calculate package sizes, and generate a comprehensive report in Excel format.
The report includes a summary of purchased quantities,billed quantities, package size differences, and highlights any missing items that need to be updated in the master file.
The application supports optional insurance files and ensures data integrity throughout the process.
==================================================

License:
This script is the intellectual property of Priyanka Gulgari.
Unauthorized copying, distribution, modification, or use of this code, via any medium, is strictly prohibited without prior written permission from the author.

Contact:
For permissions or inquiries, please contact priyankagulgari@gmail.com .

==================================================
"""
# Standard library imports
import csv
import json
import mimetypes
import os
import re
import shutil
import smtplib
import sys
import tempfile
import threading
import urllib.parse
import webbrowser
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from glob import glob
from urllib.parse import quote
from uuid import uuid4

# Third-party imports
import numpy as np
import pandas as pd
from flask import (Flask, current_app, jsonify, make_response, redirect,
                   render_template, request, send_file, send_from_directory,
                   url_for)
from flaskwebgui import FlaskUI
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break, PageBreak
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.utils import secure_filename
import tkinter as tk
from tkinter import filedialog

# Initialize Tkinter window for screen dimensions
root = tk.Tk()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
root.destroy()


def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


app = Flask(__name__, template_folder='templates')
app.static_folder = 'static'
app_display_name = "RxInsight"  # Use this for display purposes


# window = webview.create_window('Pharmacy Data Processing Application',app)
UPLOAD_FOLDER = 'uploads'
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
PROCESSED_FOLDER = os.path.join(BASE_DIR, 'processed')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(PROCESSED_FOLDER):
    os.makedirs(PROCESSED_FOLDER)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    pharmacy_name = (request.form.get('pharmacy_name') or '').strip()
    date_range = (request.form.get('date_range') or '').strip()

    try:
        vendor_count = int(request.form.get('vendor_count', 0))
    except ValueError:
        vendor_count = 0

    # ---- FILE OBJECTS ----
    custom_log_file = request.files.get('custom_log')      # CSV
    all_pbm_file = request.files.get('all_pbm')         # CSV (optional)
    kinray_file = request.files.get('kinray_file')     # CSV
    bin_master_file = request.files.get('bin_master')      # CSV (required)

    # ---- Required uploads ----
    missing = [k for k, f in {
        'custom_log':  custom_log_file,
        'kinray_file': kinray_file,
        'bin_master':  bin_master_file,
    }.items() if not f or not f.filename]
    if missing:
        return (f"Missing required file(s): {', '.join(missing)}. "
                f"Got files: {list(request.files.keys())}", 400)

    # ---- Save uploads to per-job dir ----
    updir = app.config['UPLOAD_FOLDER']
    os.makedirs(updir, exist_ok=True)

    job_id = uuid4().hex
    job_dir = os.path.join(updir, job_id)
    os.makedirs(job_dir, exist_ok=True)

    custom_log_path = os.path.join(job_dir, 'custom_log.csv')
    kinray_path = os.path.join(job_dir, 'kinray.csv')
    bin_master_path = os.path.join(job_dir, 'bin_master.csv')
    all_pbm_path = None

    if not kinray_file.filename.lower().endswith('.csv'):
        return "KINRAY file must be a CSV.", 400

    custom_log_file.save(custom_log_path)
    kinray_file.save(kinray_path)
    bin_master_file.save(bin_master_path)

    if all_pbm_file and all_pbm_file.filename:
        if not all_pbm_file.filename.lower().endswith('.csv'):
            return "ALL PBM file must be a CSV.", 400
        all_pbm_path = os.path.join(job_dir, 'all_pbm.csv')
        all_pbm_file.save(all_pbm_path)

    # ---- Optional vendor files ----
    vendor_paths = []
    if vendor_count and vendor_count > 0:
        for i in range(1, vendor_count + 1):
            vf = request.files.get(f'vendor{i}_file')
            vname = (request.form.get(
                f'vendor{i}_name') or f'Vendor{i}').strip()
            if vf and vf.filename:
                safe = re.sub(r'[^A-Za-z0-9_.-]+', '_', vname) or f'Vendor{i}'
                dest = os.path.join(job_dir, f'{safe}.csv')
                vf.save(dest)
                vendor_paths.append(dest)

    # ---- Build summary for review (processors, total Rx, $ by processor) ----
    log_df = pd.read_csv(custom_log_path, dtype=str)
    log_df, _status_col, _kept_rows, _dropped_rows = _filter_custom_log_transmitted_paid_ins(log_df)
    bin_df = pd.read_csv(bin_master_path, dtype=str)
    # ✅ Add this right here
    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Plan 1 BIN', 'Plan 2 BIN']:
        if c not in log_df.columns:
            log_df[c] = 0 if 'Paid' in c else ''  # safe default

    # normalize for summary
    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2']:
        if c in log_df.columns:
            log_df[c] = pd.to_numeric(log_df[c], errors='coerce').fillna(0)

    # choose winning bin/paid per row
    log_df['Winning_BIN'] = np.where(log_df.get('Ins Paid Plan 1', 0) >= log_df.get('Ins Paid Plan 2', 0),
                                     log_df.get('Plan 1 BIN', ''), log_df.get('Plan 2 BIN', ''))
    log_df['Winning_BIN'] = log_df['Winning_BIN'].astype(
        str).str.replace(r'\D', '', regex=True).str.zfill(6)
    log_df['Winning_Paid'] = np.where(
        log_df['Winning_BIN'] == log_df.get('Plan 1 BIN', ''),
        log_df.get('Ins Paid Plan 1', 0),
        log_df.get('Ins Paid Plan 2', 0)
    )
    log_df['Winning_Group'] = np.where(log_df.get('Ins Paid Plan 1', 0) >= log_df.get('Ins Paid Plan 2', 0),
                                       log_df.get('Plan 1 Group #', ''), log_df.get('Plan 2 Group #', ''))
    log_df['Winning_PCN'] = np.where(log_df.get('Ins Paid Plan 1', 0) >= log_df.get('Ins Paid Plan 2', 0),
                                     log_df.get('Plan 1 PCN', ''), log_df.get('Plan 2 PCN', ''))
    # map BIN -> Processor
    bin_df['BIN'] = bin_df['BIN'].astype(str).str.replace(
        r'\D', '', regex=True).str.zfill(6)
    bin_df['Processor'] = bin_df['Processor'].astype(str).str.strip()
    bin_to_proc = dict(zip(bin_df['BIN'], bin_df['Processor']))
    log_df['Processor'] = log_df['Winning_BIN'].map(bin_to_proc)

    # ---- Unmapped BINs (Winning BINs with no Processor mapping) ----
    mask_unmapped = (
        log_df['Processor'].isna()
        & log_df['Winning_BIN'].astype(str).str.strip().ne('')
    )

    if 'Rx #' in log_df.columns:
        # count unique RX # per unmapped BIN (safer)
        unmapped_grp = (
            log_df.loc[mask_unmapped]
            .groupby('Winning_BIN', as_index=False)
            .agg(rx_count=('Rx #', lambda s: s.astype(str).str.strip()
                           .replace('', np.nan).dropna().nunique()))
        )
    else:
        # fallback: count rows
        unmapped_grp = (
            log_df.loc[mask_unmapped]
            .groupby('Winning_BIN', as_index=False)
            .size().rename(columns={'size': 'rx_count'})
        )

    unmapped_grp = unmapped_grp.sort_values('rx_count', ascending=False)

    unmapped_bins = [
        {'bin': r['Winning_BIN'], 'rx_count': int(r['rx_count'])}
        for _, r in unmapped_grp.iterrows()
    ]
    unmapped_total_bins = len(unmapped_bins)
    unmapped_total_rx = int(
        unmapped_grp['rx_count'].sum()) if not unmapped_grp.empty else 0

    # total Rx (unique)
    if 'Rx #' in log_df.columns:
        total_rx = (log_df['Rx #'].astype(str).str.strip().replace(
            '', np.nan).dropna().nunique())
    else:
        total_rx = len(log_df)

    grp = (log_df.dropna(subset=['Processor'])
                 .groupby('Processor', as_index=False)
                 .agg(rx_count=('Winning_BIN', 'count'),
                      total_paid=('Winning_Paid', 'sum')))
    grp = grp.sort_values('total_paid', ascending=False)

    # ✅ Insert here
    if grp.empty:
        by_processor = []
        processors = []
    else:
        by_processor = [
            {"processor": r['Processor'], "rx_count": int(
                r['rx_count']), "total_paid": float(r['total_paid'])}
            for _, r in grp.iterrows()
        ]
        processors = [bp['processor'] for bp in by_processor]

    # ✅ expose sheet choices for the modal
    sheets_available = [
        "Processed Data",
        "Needs to be ordered - All",
        "Do Not Order - ALL",
        "Never Ordered - Check",
        "Refills 0 - Call Doctor",
        "RX Comparison - All",
        "RX Comparison +ve",
        "MFP Drugs - RX",
        "Missed Refill - Revenue Recovery",
        "BIN to Processor",
        "Summary"
    ]
    # sensible defaults (you can change in UI)
    preselected_sheets = [
        "Processed Data",
        "Needs to be ordered - All",
        "Do Not Order - ALL",
        "Never Ordered - Check",
        "Refills 0 - Call Doctor",
        "MFP Drugs - RX",
        "Missed Refill - Revenue Recovery",
        "BIN to Processor",
        "Summary"
    ]

    summary = {
        "total_rx": int(log_df.shape[0]),
        "processors": sorted(log_df["Processor"].dropna().astype(str).str.strip().unique().tolist()),
        "by_processor": (
            log_df.groupby("Processor", dropna=True)
            .agg(rx_count=("Rx #", "count"), total_paid=("Winning_Paid", "sum"))
            .reset_index()
            .rename(columns={"Processor": "processor"})
            .to_dict(orient="records")
        ),
        "unmapped_bins": unmapped_bins,                # [{bin, rx_count}, ...]
        "unmapped_total_bins": unmapped_total_bins,    # e.g., 7
        "unmapped_total_rx": unmapped_total_rx,        # e.g., 128
        "note_unmapped": "Update the BIN Master file to map these BINs to processors."

    }

    # cache minimal job context for finalize
    _JOB_CACHE[job_id] = {
        "paths": {
            "job_dir": job_dir,
            "custom_log": custom_log_path,
            "kinray": kinray_path,
            "bin_master": bin_master_path,
            "all_pbm": all_pbm_path,
            "vendors": vendor_paths,
        },
        "pharmacy_name": pharmacy_name,
        "date_range": date_range,
        "summary": summary
    }

    # respond with job + summary (front-end will open the modal)
    return jsonify({
        "ok": True,
        "job_id": job_id,
        "summary": summary
    })


@app.route('/email', methods=['POST'])
def email_report():
    from_email = request.form.get('from_email', '').strip()
    to_email = request.form.get('to_email', '').strip()
    message = request.form.get('message', '').strip()
    job_id = request.form.get('job_id', '').strip()

    if not from_email or not to_email:
        return jsonify({"ok": False, "error": "Sender and recipient emails are required."}), 400

    # Figure out which file to attach
    processed_dir = os.path.join(
        current_app.root_path, current_app.config.get('PROCESSED_FOLDER', 'processed'))
    attach_path = None

    # 1) Prefer the job’s file if we stored it
    ctx = _JOB_CACHE.get(job_id) if job_id else None
    if ctx and 'outfile' in ctx and os.path.exists(ctx['outfile']):
        attach_path = ctx['outfile']
    else:
        # 2) Fallback to newest .xlsx in processed folder
        candidates = glob.glob(os.path.join(processed_dir, '*.xlsx'))
        if not candidates:
            return jsonify({"ok": False, "error": "No report file found to attach."}), 404
        attach_path = max(candidates, key=os.path.getmtime)

    # Build email with attachment
    msg = MIMEMultipart()
    msg['Subject'] = "PharmaTrack Report"
    msg['From'] = from_email
    msg['To'] = to_email
    msg.attach(MIMEText(message or "Here’s your PharmaTrack report."))

    ctype, enc = mimetypes.guess_type(attach_path)
    if ctype is None:
        ctype = 'application/octet-stream'
    maintype, subtype = ctype.split('/', 1)

    with open(attach_path, 'rb') as f:
        part = MIMEBase(maintype, subtype)
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment',
                        filename=os.path.basename(attach_path))
        msg.attach(part)

    # Send (configure SMTP for your environment)
    try:
        # Example: Gmail SMTP (requires an app password)
        SMTP_USER = os.environ.get('SMTP_USER') or from_email
        # app password or your relay’s credential
        SMTP_PASS = os.environ.get('SMTP_PASS')

        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(from_email, [to_email], msg.as_string())
        return jsonify({"ok": True})
    except Exception as e:
        # Print full traceback to server logs to help debugging (will appear in the console)
        import traceback as _tb
        _tb.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/review', methods=['GET'])
def review_job():
    job_id = request.args.get('job_id', '')
    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return jsonify({"ok": False, "error": "Invalid job_id"}), 404
    return jsonify({"ok": True, "job_id": job_id, "summary": ctx["summary"]})


@app.route('/download')
def download_file():
    # filename = request.args.get('filename', '')
    # if not filename:
    #     return "Missing filename", 400
    # # prevent path traversal
    # safe = os.path.basename(filename)
    # fullpath = os.path.join(app.root_path, app.config.get('PROCESSED_FOLDER', 'processed'), safe)
    # if not os.path.exists(fullpath):
    #     return "File not found", 404
    # return send_file(fullpath, as_attachment=True, download_name=safe)
    filename = request.args.get("filename")
    directory = os.path.join(app.root_path, app.config.get(
        'PROCESSED_FOLDER', 'processed'))
    response = make_response(send_from_directory(
        directory, filename, as_attachment=True))
    # 🚀 Remove "Internet zone" mark by using generic content-type
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.headers["Content-Type"] = "application/octet-stream"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.errorhandler(500)
def _handle_500(e):
    # Ensure JSON instead of HTML debugger
    return jsonify({"error": str(e)}), 500


# keep small “job context” between /upload -> /review -> /finalize
_JOB_CACHE = {}  # { job_id: { "paths": {...}, "summary": {...}, "pharmacy_name":..., "date_range":... } }


def _normalize_status_value(v):
    s = '' if pd.isna(v) else str(v)
    return re.sub(r'[^a-z0-9]+', '', s.strip().lower())


def _filter_custom_log_transmitted_paid_ins(df):
    """
    Keep only rows whose status is Transmitted or Paid-Ins.
    Returns (filtered_df, status_col_name, kept_rows, dropped_rows).
    Raises ValueError if a usable status column cannot be identified.
    """
    if df is None or df.empty:
        return df.copy(), None, 0, 0

    col_lookup = {str(c).strip().lower(): c for c in df.columns}
    preferred = [
        'status', 'rx status', 'claim status', 'transaction status',
        'transmission status', 'payment status', 'rx state', 'state'
    ]

    status_col = None
    for key in preferred:
        if key in col_lookup:
            status_col = col_lookup[key]
            break

    allowed = {'transmitted', 'paidins'}

    # If no direct status column match, infer by values.
    if status_col is None:
        for c in df.columns:
            vals = pd.Series(df[c]).dropna().astype(str)
            if vals.empty:
                continue
            norm_vals = set(vals.map(_normalize_status_value).unique().tolist())
            if ('transmitted' in norm_vals) and ('paidins' in norm_vals):
                status_col = c
                break

    if status_col is None:
        raise ValueError(
            "Custom Log must include a status column containing 'Transmitted' and 'Paid-Ins' values."
        )

    status_norm = df[status_col].map(_normalize_status_value)
    mask = status_norm.isin(allowed)
    kept = int(mask.sum())
    dropped = int((~mask).sum())
    return df.loc[mask].copy(), status_col, kept, dropped


def _build_insurance_summary(log_df, bin_df):
    """
    Returns:
      {
        "total_rx": <int>,  # unique Rx # (fallback to row count if missing)
        "by_processor": [
            {"processor": "CAREMARK", "rx_count": 200, "total_paid": 200000.00},
            ...
        ],
        "processors": ["CAREMARK","OPTUMRX",...]
      }
    """
    df = log_df.copy()

    # normalize numerics
    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    # pick winner per row
    df['Winning_BIN'] = np.where(df.get('Ins Paid Plan 1', 0) >= df.get('Ins Paid Plan 2', 0),
                                 df.get('Plan 1 BIN', ''), df.get('Plan 2 BIN', ''))
    df['Winning_BIN'] = df['Winning_BIN'].astype(
        str).str.replace(r'\D', '', regex=True).str.zfill(6)

    df['Winning_Paid'] = np.where(
        df['Winning_BIN'] == df.get('Plan 1 BIN', ''),
        df.get('Ins Paid Plan 1', 0),
        df.get('Ins Paid Plan 2', 0)
    )

    # BIN -> Processor
    bin_df = bin_df.copy()
    bin_df['BIN'] = bin_df['BIN'].astype(str).str.replace(
        r'\D', '', regex=True).str.zfill(6)
    bin_df['Processor'] = bin_df['Processor'].astype(str).str.strip()
    bin_to_proc = dict(zip(bin_df['BIN'], bin_df['Processor']))

    df['Processor'] = df['Winning_BIN'].map(bin_to_proc)

    # RX count
    if 'Rx #' in df.columns:
        total_rx = df['Rx #'].astype(str).str.strip().replace(
            '', np.nan).dropna().nunique()
    else:
        total_rx = len(df)

    # group by processor
    grp = (df.dropna(subset=['Processor'])
             .groupby('Processor', as_index=False)
             .agg(rx_count=('Winning_BIN', 'count'),
                  total_paid=('Winning_Paid', 'sum')))

    # nice ordering by total $
    grp = grp.sort_values('total_paid', ascending=False)
    processors = grp['Processor'].tolist()

    by_processor = [
        {
            "processor": r['Processor'],
            "rx_count": int(r['rx_count']),
            "total_paid": float(r['total_paid'])
        }
        for _, r in grp.iterrows()
    ]

    return {
        "total_rx": int(total_rx),
        "by_processor": by_processor,
        "processors": processors
    }


@app.route('/finalize', methods=['POST'])
def finalize_job():
    data = request.get_json(force=True, silent=True) or {}
    job_id = (data.get('job_id') or '').strip()

    # Validate required folder paths
    main_save_dir = (data.get('main_save_dir') or '').strip()
    audit_save_dir = (data.get('audit_save_dir') or '').strip()
    
    if not main_save_dir:
        return jsonify({"ok": False, "error": "Main report destination folder is required"}), 400
    if not audit_save_dir:
        return jsonify({"ok": False, "error": "Audit workbook destination folder is required"}), 400
    
    if not os.path.isdir(main_save_dir):
        return jsonify({"ok": False, "error": f"Main report folder does not exist: {main_save_dir}"}), 400
    if not os.path.isdir(audit_save_dir):
        return jsonify({"ok": False, "error": f"Audit workbook folder does not exist: {audit_save_dir}"}), 400

    selected_processors = [str(p).strip().upper() for p in (
        data.get('selected_processors') or []) if p]
    selected_sheets = [str(s).strip()
                       for s in (data.get('selected_sheets') or []) if s]

    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return jsonify({"ok": False, "error": "Invalid job_id"}), 404

    paths = ctx["paths"]
    processed_dir = os.path.join(
        app.root_path, app.config.get('PROCESSED_FOLDER', 'processed'))
    os.makedirs(processed_dir, exist_ok=True)

    try:
        # include Kinray first, then the extras
        vendor_paths = [paths["kinray"], *paths["vendors"]]

        result = process_custom_log_data(
            custom_log_path=paths["custom_log"],
            all_pbm_path=paths["all_pbm"],
            bin_master_path=paths["bin_master"],
            vendor_paths=vendor_paths,
            pharmacy_name=ctx.get("pharmacy_name", ""),
            date_range=ctx.get("date_range", ""),
            selected_processors=selected_processors,
            selected_sheets=selected_sheets,
            user_audit_dir=audit_save_dir,
        )
        if not result or not isinstance(result, dict):
            return jsonify({"ok": False, "error": "Report generation returned no filename"}), 500
        
        main_name = result.get("main")
        audit_name = result.get("audit")
        #print(main_name, audit_name)
        if not main_name:
            return jsonify({"ok": False, "error": "Missing main filename"}), 500
        # IMPORTANT: do NOT change the name returned by the writer
        #fullpath = os.path.join(processed_dir, out_filename)
        processed_dir = os.path.join(
            app.root_path, app.config.get('PROCESSED_FOLDER', 'processed'))
        main_path = os.path.join(processed_dir, main_name)
        audit_path = os.path.join(processed_dir, audit_name)

        if not os.path.exists(main_path):
            # print("[finalize] processed dir listing:", os.listdir(processed_dir))
            return jsonify({"ok": False, "error": "Main output file not found"}), 500

        ctx["outfile_main"] = main_path
        
        # Copy main file to user-specified folder (now required)
        try:
            import shutil
            main_dest = os.path.join(main_save_dir, main_name)
            shutil.copy2(main_path, main_dest)
            #print(f"[finalize] copied main report to: {main_dest}")
        except Exception as e:
            #print(f"[finalize] error copying main file to {main_save_dir}: {e}")
            return jsonify({"ok": False, "error": f"Failed to copy main report: {str(e)}"}), 500
        
        audit_exists = False
        if audit_name:
            audit_path = os.path.join(processed_dir, audit_name)
            if os.path.exists(audit_path):
                ctx["outfile_audit"] = audit_path
                audit_exists = True
                # Copy audit file to user-specified folder (now required)
                try:
                    import shutil
                    audit_dest = os.path.join(audit_save_dir, audit_name)
                    shutil.copy2(audit_path, audit_dest)
                    #(f"[finalize] copied audit report to: {audit_dest}")
                except Exception as e:
                    #print(f"[finalize] error copying audit file to {audit_save_dir}: {e}")
                    return jsonify({"ok": False, "error": f"Failed to copy audit report: {str(e)}"}), 500
            else:
                print(f"[finalize] audit file expected but not found: {audit_path}")


        # if not os.path.exists(fullpath):
        #     # (Optional) one fallback try: a regex-sanitized variant
        #     alt = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', out_filename)
        #     altpath = os.path.join(processed_dir, alt)
        #     if os.path.exists(altpath):
        #         out_filename = alt
        #     else:
        #         print("[finalize] expected at:", os.path.abspath(fullpath))
        #         print("[finalize] processed dir listing:",
        #               os.listdir(processed_dir))
        #         return jsonify({"ok": False, "error": f"Output file not found: {out_filename}"}), 500

        # ctx['outfile'] = fullpath

        # return jsonify({
        #     "ok": True,
        #     "filename": out_filename,
        #     "download_url": f"/download?filename={quote(out_filename)}"
        # })
        # return jsonify({
        #     "ok": True,
        #     "filename": main_name,
        #     "download_url": f"/download?filename={quote(main_name)}",
        #     # new behaviour (audit report)
        #     # "audit_filename": audit_name,
        #     # "audit_download_url": f"/download?filename={quote(audit_name)}",
        # })
        resp = {
            "ok": True,
            "main_filename": main_name,
            "main_download_url": f"/download?filename={quote(main_name)}",
            "download_url": f"/download?filename={quote(main_name)}"  # backwards-compatible
        }
        if audit_name and ctx.get("outfile_audit"):
            resp.update({
                "audit_filename": audit_name,
                "audit_download_url": f"/download?filename={quote(audit_name)}"
            })
        return jsonify(resp)


    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/pick_folder', methods=['GET'])
def pick_folder():
    """Open a native folder picker on the server (local desktop) and return the chosen path.

    Note: This only works safely when the Flask app runs on the user's desktop (local machine).
    """
    try:
        # Use a transient Tk root to show the directory dialog
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askdirectory(title='Choose folder to save audit workbook')
        root.destroy()
        if not path:
            return jsonify({"ok": False, "error": "No folder selected"}), 400
        return jsonify({"ok": True, "path": path})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def get_column_index(ws, header_name, header_row=None):
    """Find a column by header text. Returns 1-based index or None."""
    rows = [header_row] if header_row else [2, 3]
    for r in rows:
        try:
            for cell in ws[r]:
                if cell.value == header_name:
                    return cell.col_idx
        except Exception:
            continue
    return None


def adjust_specific_columns(ws, col_letters, width=12):
    for col_letter in col_letters:
        if col_letter:
            ws.column_dimensions[col_letter].width = width


def discover_processors_from_df(final_df):
    """
    Derive processor prefixes from columns like <Processor>_{Q,P,D,T,Pur,Net,Diff$}.
    Use this if you don't already have a processors list.
    """
    suffixes = ('_Q', '_P', '_D', '_T', '_Pur', '_Net', '_Net')
    procs = set()
    for c in final_df.columns:
        for s in suffixes:
            if c.endswith(s):
                procs.add(c[:-len(s)])
                break
    return sorted(procs)


def add_autosum_by_processors(ws, processors, start_row, end_row, header_row=3):
    """
    Write SUM formulas for each processor's *_T, *_Pur, *_Net into the row after end_row,
    and format totals in currency.
    """
    from openpyxl.utils import get_column_letter

    if not processors:
        return

    total_row = end_row + 1
    # Write "Totals" label if blank
    if ws.cell(row=total_row, column=1).value in (None, ""):
        ws.cell(row=total_row, column=1, value="Totals").font = Font(bold=True)

    for pr in processors:
        for suf in ("_T", "_Pur", "_Net"):
            hdr = f"{pr}{suf}"
            col_idx = get_column_index(ws, hdr, header_row=header_row)
            if not col_idx:
                continue

            col_letter = get_column_letter(col_idx)
            cell = ws.cell(
                row=total_row,
                column=col_idx,
                value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
            )
            # Format in currency and bold
            cell.number_format = '"$"#,##0.00'
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Optional: highlight the total row visually
    for c in range(1, ws.max_column + 1):
        ws.cell(row=total_row, column=c).border = Border(
            top=Side(style="thick"))


def apply_common_sheet_settings(
    wb,
    pharmacy_name: str,
    date_range: str,
    processors: list[str] | None = None,
    header_row_main: int = 3,
):
    """
    Apply titles, orientations, print settings to known sheets.
    If processors are provided (or discoverable), run autosum & width tweaks on 'Processed Data'.
    """
    RULES = {
        "Processed Data": {
            "title": "{pharmacy} ({range})",
            "font_size": 35,
            "min_row_for_height": header_row_main + 1,  # data starts after header
            "orientation": "landscape",
            "header_row": header_row_main,
        },
        "Needs to be Ordered": {
            "title": "{pharmacy} ({range}) - NTO CVS",
            "font_size": 25,
            "min_row_for_height": 2,
            "orientation": "landscape",
        },
        "Missing Items": {
            "title": "{pharmacy} ({range}) - Missing items, To be updated in master file",
            "font_size": 15,
            "min_row_for_height": 2,
            "orientation": "landscape",
        },
        "Do Not Order CVS": {
            "title": "{pharmacy} ({range}) - DNO CVS",
            "font_size": 25,
            "min_row_for_height": 2,
            "orientation": "landscape",
        },
        "Needs to be ordered - All": {
            "title": "{pharmacy} ({range}) - Need to Order - ALL",
            "font_size": 25,
            "min_row_for_height": 3,
            "orientation": "landscape",
        },
        "Do Not Order - ALL": {
            "title": "{pharmacy} ({range}) - Do Not Order",
            "font_size": 25,
            "min_row_for_height": 3,
            "orientation": "portrait",
        },
        "Never Ordered - Check": {
            "title": "{pharmacy} ({range}) - Never Ordered Package - Check",
            "font_size": 25,
            "min_row_for_height": 3,
            "orientation": "landscape",
        },
        "Never Ordered  - Check": {  # tolerate double-space variant
            "title": "{pharmacy} ({range}) - Never Ordered Package - Check",
            "font_size": 25,
            "min_row_for_height": 3,
            "orientation": "landscape",
        },
    }

    # Apply per-sheet render + common print settings
    for sheet in wb.worksheets:
        name = sheet.title
        cfg = RULES.get(name)
        if cfg:
            # Title in A1
            t = cfg["title"].format(pharmacy=pharmacy_name, range=date_range)
            c = sheet.cell(row=1, column=1)
            c.value = t
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=cfg["font_size"], bold=True)

            # Row heights
            min_row = cfg.get("min_row_for_height", 2)
            for row in sheet.iter_rows(min_row=min_row, max_row=sheet.max_row):
                sheet.row_dimensions[row[0].row].height = 20

            # Orientation
            sheet.page_setup.orientation = cfg.get("orientation", "landscape")

        # Common print config
        sheet.print_title_rows = "1:2"
        sheet.oddFooter.left.text = "Page &P of &N"
        sheet.oddFooter.left.size = 8
        sheet.oddFooter.left.font = "Arial,Bold"
        sheet.page_margins = PageMargins(
            left=0, right=0, top=0, bottom=0, header=0, footer=0.1)
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_options.horizontalCentered = True
        sheet.print_options.verticalCentered = True
        sheet.print_options.gridLines = True

        # Extra on main sheet: autosum + width for processor T/Pur/Diff$
        if name == "Processed Data":
            header_row = RULES[name]["header_row"]
            start_row = header_row + 1
            end_row = sheet.max_row
            if end_row >= start_row:
                # find processors if not provided
                procs = processors
                if not procs:
                    # try to discover from headers on the sheet
                    # (if you still have the DataFrame, prefer discover_processors_from_df(final))
                    # Here, we scan row `header_row` for names that look like <Proc>_<suffix>
                    import re
                    suffixes = ("_T", "_Pur", "_Net")
                    procs = set()
                    for cell in sheet[header_row]:
                        val = str(cell.value or "")
                        for s in suffixes:
                            if val.endswith(s):
                                procs.add(val[:-len(s)])
                    procs = sorted(procs)

                # autosum
                add_autosum_by_processors(
                    sheet, procs, start_row, end_row, header_row=header_row)

                # width tweaks for *_T, *_Pur, *_Net
                cols_to_adjust = []
                for pr in procs:
                    for suf in ("_T", "_Pur", "_Net"):
                        idx = get_column_index(
                            sheet, f"{pr}{suf}", header_row=header_row)
                        if idx:
                            cols_to_adjust.append(get_column_letter(idx))
                if cols_to_adjust:
                    adjust_specific_columns(sheet, cols_to_adjust, width=12)


def min_difference_sheet(wb, final_data, insurance_paths=None):
    """
    Build 'Do Not Order - ALL' sheet:
      - Identify all *_D columns (excluding ALL_PBM_D).
      - Keep only rows where ALL *_D values are strictly > 0 (i.e., no deficits anywhere).
      - 'Min Positive' = row-wise minimum across the *_D columns (strictly positive).
      - Display ORIGINAL *_D values (positives remain visible; we don't zero anything for display).

    Parameters
    ----------
    wb : openpyxl.Workbook
    final_data : pandas.DataFrame
        Must contain columns: 'NDC #', 'Drug Name', 'Package Size', and multiple '*_D' columns.
    insurance_paths : unused; kept only for signature parity.
    """
    import pandas as pd

    df = final_data.copy()

    # --- 1) Identify difference columns (exclude ALL_PBM_D)
    difference_columns = [
        c for c in df.columns if c.endswith('_D') and c != 'ALL_PBM_D']

    # Create/replace target sheet
    sheet_title = "Do Not Order - ALL"
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    if not difference_columns:
        ws['A1'] = "No difference columns (*_D) found."
        return

    # Ensure required base columns exist
    for base in ['NDC #', 'Drug Name', 'Package Size']:
        if base not in df.columns:
            df[base] = 0 if base != 'Drug Name' else pd.NA

    # --- 2) Coerce *_D to numeric (for logic), but keep original values for display
    dnum = df[difference_columns].apply(
        pd.to_numeric, errors='coerce').fillna(0)

    # Logic: "Do Not Order" rows are those with ALL positives (>0) across *_D
    # (Mirror of your previous: negatives→0 then min>0; equivalently, (dnum > 0).all(axis=1))
    min_positive = dnum.min(axis=1)
    mask = (dnum > 0).all(axis=1) & (min_positive > 0)

    if not mask.any():
        ws['A1'] = "No rows qualify: no items with all positive differences."
        return

    # --- 3) Build display frame (use ORIGINAL *_D values)
    out = df.loc[mask, ['NDC #', 'Drug Name',
                        'Package Size'] + difference_columns].copy()
    out['Do Not Order'] = min_positive.loc[mask]
    out['Paper Work'] = " "
    out.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)

    display_columns = ['NDC #', 'Drug Name', 'Pkg Size'] + \
        difference_columns + ['Do Not Order', 'Paper Work']
    out = out[display_columns].sort_values('Drug Name')

    # --- 4) Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(display_columns))
    title_cell = ws.cell(row=1, column=1, value="Do not order")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)
    ws.row_dimensions[1].height = 30

    # --- 5) Write table (headers at row 2; data from row 3)
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                # header
                cell.font = Font(bold=True, size=12)
                align = Alignment(horizontal='center', vertical='center')
                if display_columns[c_idx - 1] in difference_columns + ['Pkg Size', 'Do Not Order']:
                    align = Alignment(horizontal='center',
                                      vertical='bottom', text_rotation=90)
                cell.alignment = align
            else:
                # body
                if display_columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=False)
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')
                cell.font = Font(size=12)

    # --- 6) Column widths
    ws.column_dimensions['A'].width = 15   # NDC #
    ws.column_dimensions['B'].width = 50   # Drug Name
    ws.column_dimensions['C'].width = 7    # Pkg Size

    for col_name in difference_columns:
        if col_name in display_columns:
            idx = display_columns.index(col_name) + 1
            ws.column_dimensions[get_column_letter(idx)].width = 7
    for col_name in ['Do Not Order', 'Pkg Size']:
        if col_name in display_columns:
            idx = display_columns.index(col_name) + 1
            ws.column_dimensions[get_column_letter(idx)].width = 8
    # wrap Paper work column text
    if 'Paper Work' in display_columns:
        idx = display_columns.index('Paper Work') + 1
        ws.column_dimensions[get_column_letter(idx)].width = 10
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    # --- 7) Borders and header height
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    thick = Border(left=Side(style='thick'), right=Side(style='thick'),
                   top=Side(style='thick'), bottom=Side(style='thick'))

    ws.row_dimensions[2].height = 80  # header row height

    # Thick border for header row
    for c in range(1, len(display_columns) + 1):
        ws.cell(row=2, column=c).border = thick

    # Thin borders for data cells
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(display_columns)):
        for cell in row:
            cell.border = thin

    # Thick edge borders for key columns
    def apply_column_border(ws_, col_idx):
        col_letter = get_column_letter(col_idx)
        for r in range(2, ws_.max_row + 1):
            c = ws_[f"{col_letter}{r}"]
            c.border = Border(left=thick.left, right=thick.right,
                              top=c.border.top, bottom=c.border.bottom)

    for key in ['NDC #', 'Drug Name', 'Pkg Size', 'Do Not Order', 'Paper Work']:
        if key in display_columns:
            apply_column_border(ws, display_columns.index(key) + 1)

    # --- 8) Freeze panes
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_columns))}{ws.max_row}"


def set_print_area_excluding(wb, sheet_name, headers_to_exclude):
    """
    Define the Excel print area for a sheet, excluding specific columns by header name.
    This does NOT hide them in the workbook — it only removes them from the print page.
    """
    ws = wb[sheet_name]
    header_row = 1  # change if your headers are on another row

    # Map each column index to its header text
    headers = {
        c: (ws.cell(row=header_row, column=c).value or '').strip()
        for c in range(1, ws.max_column + 1)
    }

    # Identify which columns to skip
    exclude_cols = {
        c for c, name in headers.items()
        if name.strip().lower() in {h.lower() for h in headers_to_exclude}
    }

    # Build contiguous column ranges to print
    ranges = []
    start = None
    for c in range(1, ws.max_column + 1):
        if c in exclude_cols:
            if start is not None:
                ranges.append((start, c - 1))
                start = None
        else:
            if start is None:
                start = c
    if start is not None:
        ranges.append((start, ws.max_column))

    # Convert to A1 ranges spanning all rows
    last_row = ws.max_row
    a1_ranges = [
        f"{get_column_letter(a)}1:{get_column_letter(b)}{last_row}"
        for a, b in ranges if a <= b
    ]

    # ✅ Assign the print area (multiple blocks allowed)
    ws.print_area = a1_ranges


def set_print_area_excluding_headers(ws, header_row=2, exclude_headers=()):
    # Build a case-insensitive set of headers to exclude
    bad = {str(h).strip().casefold() for h in exclude_headers}

    # Read headers on header_row
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        headers.append(((v if v is not None else "").strip(), c))

    # Keep columns not in exclude list
    include_cols = [c for (name, c) in headers if name.casefold() not in bad]
    if not include_cols:
        return  # nothing to include

    # Build contiguous column blocks -> "A2:C{max_row},E2:G{max_row},..."
    maxr = ws.max_row
    blocks = []
    run_start = run_prev = include_cols[0]
    for c in include_cols[1:] + [None]:
        if c is None or c != run_prev + 1:
            # close current run
            left = get_column_letter(run_start)
            right = get_column_letter(run_prev)
            blocks.append(f"{left}{header_row}:{right}{maxr}")
            # start new
            if c is not None:
                run_start = c
        run_prev = c if c is not None else run_prev

    # Assign comma-separated print area
    ws.print_area = ",".join(blocks)


def add_max_difference_sheet(wb, final_data, insurance_paths=None):
    """
    Create a 'Needs to be ordered - All' worksheet showing:
      - All *_D columns (package differences per processor) with original values (positives kept visible)
      - 'To Order' computed ONLY from negatives: max package deficit across all *_D per row
      - PRICE chosen as Kinray_UPrice (if > 0), else first non-zero vendor *_PRICE, else 0
      - 'Total Order Price' = To Order * PRICE

    Parameters
    ----------
    wb : openpyxl.Workbook
    final_data : pandas.DataFrame
        Must contain: 'NDC #', 'Drug Name', 'Package Size', optional 'Kinray_UPrice', vendor '*_PRICE' cols, and '*_D' cols.
    insurance_paths : any (unused; kept for signature compatibility)
    """
    import pandas as pd
    import numpy as np

    df = final_data.copy()

    # 1) Identify difference columns (exclude ALL_PBM_D)
    difference_columns = [
        c for c in df.columns if c.endswith('_D') and c != 'ALL_PBM_D']

    # Create/replace the target sheet early
    sheet_title = "Needs to be ordered - All"
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)

    if not difference_columns:
        ws['A1'] = "No difference columns (*_D) found."
        return

    # Make sure base columns exist
    for base in ['NDC #', 'Drug Name', 'Package Size', 'Kinray_UPrice']:
        if base not in df.columns:
            df[base] = 0 if base != 'Drug Name' else pd.NA

    # 2) Coerce numerics
    df[difference_columns] = df[difference_columns].apply(
        pd.to_numeric, errors='coerce').fillna(0)
    df['Kinray_UPrice'] = pd.to_numeric(
        df['Kinray_UPrice'], errors='coerce').fillna(0)

    # 3) Choose PRICE per row
    # Prefer Kinray_UPrice (>0), else first non-zero vendor *_PRICE, else 0
    vendor_price_cols = [c for c in df.columns if c.endswith('_PRICE')]
    if vendor_price_cols:
        vendor_prices = df[vendor_price_cols].apply(
            pd.to_numeric, errors='coerce').replace(0, pd.NA)
        # bring the first non-null from the row to the leftmost position, then pick first col
        first_nonzero_vendor_price = vendor_prices.bfill(
            axis=1).iloc[:, 0].fillna(0)
    else:
        first_nonzero_vendor_price = pd.Series(0, index=df.index)

    df['PRICE'] = np.where(df['Kinray_UPrice'] > 0,
                           df['Kinray_UPrice'], first_nonzero_vendor_price)
    df['PRICE'] = pd.to_numeric(df['PRICE'], errors='coerce').fillna(0)

    # 4) Compute "To Order" using ONLY negatives (positives remain visible in the table)
    neg_for_logic = df[difference_columns].clip(
        upper=0)  # keep negatives, zero out positives
    # any negative deficit across insurers
    needs_mask = neg_for_logic.lt(0).any(axis=1)

    if not needs_mask.any():
        # No ordering needed, but keep the sheet useful
        ws['A1'] = "No rows require ordering (no negative deficits in *_D columns)."
        return

    needs = df.loc[needs_mask].copy()
    needs['To Order'] = neg_for_logic.loc[needs_mask].min(
        axis=1).abs()  # most negative (largest deficit), abs->packages
    needs['Pkg Size'] = needs['Package Size']
    needs['Paper Work'] = " "
    needs['Total Order Price'] = needs['To Order'] * needs['PRICE']
    needs['Total Order Price'] = pd.to_numeric(
        needs['Total Order Price'], errors='coerce').fillna(0)

    # 5) Build display frame (keep ORIGINAL *_D values = positives visible)
    display_columns = (
        ['NDC #', 'Drug Name', 'Pkg Size'] +
        difference_columns +
        ['To Order', 'Paper Work', 'PRICE', 'Total Order Price']
    )
    needs = needs[display_columns].sort_values('Drug Name')

    # 6) Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(display_columns))
    title_cell = ws.cell(
        row=1, column=1, value="Needs to be Ordered (Max Package Deficit Across Insurances)")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=20, bold=True)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 80  # header
    # 7) Write the table (headers at row 2, data from row 3)
    for r_idx, row in enumerate(dataframe_to_rows(needs, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                # header style
                cell.font = Font(bold=True, size=12)
                align = Alignment(horizontal='center', vertical='center')
                if display_columns[c_idx - 1] in difference_columns:
                    align = Alignment(horizontal='center',
                                      vertical='center', text_rotation=90)
                cell.alignment = align
            else:
                # body style
                if display_columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=False)
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

    # 8) Column widths
    ws.column_dimensions['A'].width = 15   # NDC #
    ws.column_dimensions['B'].width = 50   # Drug Name
    ws.column_dimensions['C'].width = 7    # Pkg Size
    # Rotate text in 'Pkg Size' column (C)
    ws["C2"].alignment = Alignment(
        horizontal='center', vertical='center', text_rotation=90)

    for col_name in difference_columns:
        if col_name in display_columns:
            idx = display_columns.index(col_name) + 1
            ws.column_dimensions[get_column_letter(idx)].width = 8
    if 'To Order' in display_columns:
        ws.column_dimensions[get_column_letter(
            display_columns.index('To Order') + 1)].width = 12
    # PAPER WORK column
    if 'Paper Work' in display_columns:
        idx = display_columns.index('Paper Work') + 1
        ws.column_dimensions[get_column_letter(idx)].width = 12
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    if 'PRICE' in display_columns:
        ws.column_dimensions[get_column_letter(
            display_columns.index('PRICE') + 1)].width = 12
    # TOTAL ORDER PRICE column
    if 'Total Order Price' in display_columns:
        idx = display_columns.index('Total Order Price') + 1
        ws.column_dimensions[get_column_letter(idx)].width = 16
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)

    # 9) Borders and header height
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    thick = Border(left=Side(style='thick'), right=Side(style='thick'),
                   top=Side(style='thick'), bottom=Side(style='thick'))

    for c in range(1, len(display_columns) + 1):
        ws.cell(row=2, column=c).border = thick
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(display_columns)):
        for cell in row:
            cell.border = thin

    # Edge thick borders for key columns (visual groups)
    edge_cols = ['NDC #', 'Drug Name', 'Pkg Size', 'PRICE',
                 'To Order', 'Total Order Price', 'Paper Work']
    for name in edge_cols:
        if name in display_columns:
            idx = display_columns.index(name) + 1
            col_letter = get_column_letter(idx)
            for r in range(2, ws.max_row + 1):
                c = ws[f"{col_letter}{r}"]
                c.border = Border(
                    left=thick.left if c.column == idx else c.border.left,
                    right=thick.right if c.column == idx else c.border.right,
                    top=c.border.top, bottom=c.border.bottom
                )
    # --- page setup BEFORE summary is fine (heights, breaks, etc.) ---
    ws.print_title_rows = "2:2"
    ws.row_breaks = PageBreak()
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[2].height = 80
    ws.freeze_panes = "A3"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- numeric formats for data columns ---
    price_idx = display_columns.index('PRICE') + 1
    total_idx = display_columns.index('Total Order Price') + 1
    to_order_idx = display_columns.index('To Order') + 1
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=price_idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=total_idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=to_order_idx).number_format = '0'

    # --- build the two summary columns (label + amount) ---
    def _find_header_col(ws, header_text, header_row=2):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=c).value
            if (v if v is not None else "") == header_text:
                return c
        return None

    top_total_hdr_col = _find_header_col(ws, "Total Order Price", header_row=2)
    if top_total_hdr_col is None:
        top_total_hdr_col = display_columns.index("Total Order Price") + 1

    summary_label_col = top_total_hdr_col + 1
    summary_value_col = top_total_hdr_col + 2

    hdr_cell = ws.cell(row=2, column=summary_label_col,
                       value="Insurance-wise Order Estimate ($)")
    hdr_cell.font = Font(bold=True, size=12)
    hdr_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    val_hdr_cell = ws.cell(row=2, column=summary_value_col, value="Amount")
    val_hdr_cell.font = Font(bold=True, size=12)
    val_hdr_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    # Lock the last data row BEFORE we start writing summary data
    last_data_row = ws.max_row
    price_col_letter = get_column_letter(price_idx)

    r = 3
    for diff_col in difference_columns:
        ws.cell(row=r, column=summary_label_col, value=diff_col).alignment = Alignment(
            horizontal="left", vertical="center")
        diff_idx = display_columns.index(diff_col) + 1
        diff_letter = get_column_letter(diff_idx)
        formula = (
            f"=SUMPRODUCT((-{diff_letter}3:{diff_letter}{last_data_row})"
            f"*({diff_letter}3:{diff_letter}{last_data_row}<0),"
            f"{price_col_letter}3:{price_col_letter}{last_data_row})"
        )
        vcell = ws.cell(row=r, column=summary_value_col, value=formula)
        vcell.number_format = '"$"#,##0.00'
        vcell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    # === Formatting for summary columns ===
    for col_idx in (summary_label_col, summary_value_col):
        col_letter = get_column_letter(col_idx)

        # Set fixed width
        ws.column_dimensions[col_letter].width = 20

        # Center + wrap for ALL cells in these columns
        for r in range(2, ws.max_row + 1):      # row 2 = header, rest = values
            cell = ws.cell(row=r, column=col_idx)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

    for c in (summary_label_col, summary_value_col):
        ws.cell(row=2, column=c).border = thick
    for rr in range(3, ws.max_row + 1):
        ws.cell(row=rr, column=summary_label_col).border = thin
        ws.cell(row=rr, column=summary_value_col).border = thin

    # === Grand Total footer OUTSIDE sort/filter range ===
    total_col_idx = display_columns.index('Total Order Price') + 1
    total_col_letter = get_column_letter(total_col_idx)
    footer_row = last_data_row + 2

    ws.cell(row=footer_row, column=total_col_idx - 1,
            value="Grand Total").font = Font(bold=True)
    footer_total_cell = ws.cell(
        row=footer_row, column=total_col_idx,
        value=f"=SUBTOTAL(109,{total_col_letter}3:{total_col_letter}{last_data_row})"
    )
    footer_total_cell.font = Font(bold=True)
    footer_total_cell.number_format = '"$"#,##0.00'

    # strong top border across the table width (not across summary columns)
    for c in range(1, len(display_columns) + 1):
        ws.cell(row=footer_row, column=c).border = Border(
            top=Side(style='thick'))

    # === Add conditional formatting for negative values (light grey for B&W printing) ===
    from openpyxl.formatting.rule import CellIsRule
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light grey
    
    # Apply to all *_D columns
    for diff_col in difference_columns:
        diff_idx = display_columns.index(diff_col) + 1
        diff_letter = get_column_letter(diff_idx)
        data_range = f"{diff_letter}3:{diff_letter}{last_data_row}"
        
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=False, fill=grey_fill)
        )

    # === NOW set print area to exclude the two summary columns ===
    set_print_area_excluding_headers(
        ws, header_row=2,
        exclude_headers=["Insurance-wise Order Estimate ($)", "Amount"]
    )

    # === Filter range limited to data only (keeps footer fixed) ===
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_columns))}{last_data_row}"


def create_never_ordered_check_sheet(wb, final_data):
    """
    Create 'Never Ordered - Check' sheet:
      • Rows where Total Purchased == 0
      • AND billed to insurance (any *_Q > 0 OR *_P > 0 OR *_T > 0)
      • Shows base cols + all insurer Q/P/T columns (excluding ALL_PBM_*).
    """
    import pandas as pd

    df = final_data.copy()

    # Identify insurance bands
    # q_cols = [c for c in df.columns if c.endswith('_Q')]
    p_cols = [c for c in df.columns if c == 'ALL_PBM_P']
    # t_cols = [c for c in df.columns if c.endswith('_T') and c != 'ALL_PBM_T']

    # Ensure required base columns exist
    for base in ['Drug Name', 'NDC #', 'Package Size', 'Total Purchased']:
        if base not in df.columns:
            df[base] = 0 if base != 'Drug Name' else pd.NA

    # Coerce numeric for logic
    def _to_num(cols):
        if not cols:
            return
        df.loc[:, cols] = df[cols].apply(
            pd.to_numeric, errors='coerce').fillna(0)

    # _to_num(q_cols)
    _to_num(p_cols)
    # _to_num(t_cols)
    df['Total Purchased'] = pd.to_numeric(
        df['Total Purchased'], errors='coerce').fillna(0)

    # "Billed to insurance" mask (any positive in Q/P/T)
    billed_mask = pd.Series(False, index=df.index)
    # if q_cols: billed_mask |= df[q_cols].gt(0).any(axis=1)
    if p_cols:
        billed_mask |= df[p_cols].gt(0).any(axis=1)
    # if t_cols: billed_mask |= df[t_cols].gt(0).any(axis=1)

    mask = (df['Total Purchased'] == 0) & billed_mask

    # Build output
    display_columns = (['Drug Name', 'NDC #', 'Package Size', 'Total Purchased']
                       + p_cols)
    out = df.loc[mask, display_columns].copy()
    out.rename(columns={'Package Size': 'Pkg Size'}, inplace=True)

    # Create/replace sheet
    title = "Never Ordered - Check"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title=title)

    if out.empty:
        ws['A1'] = "No rows with Total Purchased = 0 that were billed to insurance."
        return

    out = out.sort_values('Drug Name')

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(out.columns))
    cell = ws.cell(row=1, column=1)
    cell.value = "Never Ordered - Check (Billed to Insurance)"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=20, bold=True)
    ws.row_dimensions[1].height = 30

    # Write table (headers at row 2, data from row 3)
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 2:
                # Header formatting
                cell.font = Font(bold=True, size=12)
                # Rotate insurance columns (Q/P/T) and Total Purchased
                hdr = out.columns[c_idx - 1]
                rotate = hdr in (p_cols + ['Total Purchased', 'Pkg Size'])
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           text_rotation=(90 if rotate else 0), wrap_text=True)
            else:
                # Body formatting
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=False)
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')
                cell.font = Font(size=12)

    # Header thick border
    thick = Border(left=Side(style='thick'), right=Side(style='thick'),
                   top=Side(style='thick'), bottom=Side(style='thick'))
    for col_idx in range(1, len(out.columns) + 1):
        ws.cell(row=2, column=col_idx).border = thick

    # Thin borders for body
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(out.columns)):
        for cell in row:
            cell.border = thin

    # Thick edge borders for key columns
    def apply_column_border(sheet, col_idx):
        col_letter = get_column_letter(col_idx)
        for r in range(2, sheet.max_row + 1):
            c = sheet[f"{col_letter}{r}"]
            c.border = Border(left=thick.left, right=thick.right,
                              top=c.border.top, bottom=c.border.bottom)

    edge_cols = ['Drug Name', 'NDC #', 'Pkg Size', 'Total Purchased']
    for name in edge_cols:
        if name in out.columns:
            apply_column_border(ws, out.columns.get_loc(name) + 1)

    # Column widths
    widths = {
        'Drug Name': 70,
        'NDC #': 15,
        'Pkg Size': 10,
        'Total Purchased': 12
    }
    for idx, col_name in enumerate(out.columns, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = widths.get(col_name, 8)

    # Rotate headers already done; set header row height
    ws.row_dimensions[2].height = 80

    # # Currency for *_T columns
    # for tcol in t_cols:
    #     if tcol in out.columns:
    #         cidx = out.columns.get_loc(tcol) + 1
    #         for r in range(3, ws.max_row + 1):
    #             ws.cell(row=r, column=cidx).number_format = '"$"#,##0.00'

    # Freeze panes
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:{get_column_letter(len(display_columns))}{ws.max_row}"


def _load_all_pbm_csv(path):
    """
    Read ALL PBM file (CSV/TSV). Auto-detect delimiter. 
    Returns: ['NDC #','ALL_PBM_Q','ALL_PBM_T','ALL_PBM_DrugName']
    (No aggregation: if an NDC appears multiple times, we keep the LAST row.)
    """
    import pandas as pd
    import re

    # Try auto delimiter; handle BOM
    df = pd.read_csv(path, dtype=str, encoding='utf-8-sig',
                     sep=None, engine='python')
    if df.shape[1] == 1:
        # Often actually TAB-delimited saved as .csv
        try:
            df = pd.read_csv(path, dtype=str, encoding='utf-8-sig', sep='\t')
        except Exception:
            pass

    # Normalize headers -> lowercase, collapse spaces, remove nbsp
    def norm(s):
        return re.sub(r'\s+', ' ', str(s).replace('\u00A0', ' ')).strip().lower()
    df.columns = [norm(c) for c in df.columns]

    # Column pickers (exact first, then contains)
    def pick(*cands):
        for k in cands:
            if k in df.columns:
                return k
        for k in cands:
            for col in df.columns:
                if k in col:
                    return col
        return None

    ndc_k = pick('ndc #', 'ndc#', 'ndc', 'ndc upc', 'ndc/upc')
    qty_k = pick('quantity', 'qty', 'total quantity', 'total qty', 'Quantity')
    total_k = pick('all_pbm_t', 'total $', 'total$',
                   'total amount', 'amount', 'total', 'Total')
    name_k = pick('drug name', 'drug', 'name')

    if ndc_k is None:
        # Return empty frame with expected columns so merge won’t break
        return pd.DataFrame(columns=['NDC #', 'ALL_PBM_Q', 'ALL_PBM_T', 'ALL_PBM_DrugName'])

    out = pd.DataFrame()
    out['NDC #'] = (df[ndc_k].astype(str)
                    .str.replace(r'\D', '', regex=True)
                    .str.zfill(11))

    # ALL_PBM_Q (optional; default 0)
    if qty_k:
        q = (df[qty_k].astype(str)
             .str.replace(',', '', regex=False)
             .str.replace(r'[^0-9.\-]', '', regex=True))
        out['ALL_PBM_Q'] = pd.to_numeric(q, errors='coerce').fillna(0)
    else:
        out['ALL_PBM_Q'] = 0

    # ALL_PBM_T (Total $) — verbatim per row, no aggregation
    if total_k:
        t = (df[total_k].astype(str)
             .str.replace(',', '', regex=False)
             .str.replace('$', '', regex=False)
             .str.replace('(', '-', regex=False)
             .str.replace(')', '', regex=False)
             .str.replace(r'[^0-9.\-]', '', regex=True))
        out['ALL_PBM_T'] = pd.to_numeric(t, errors='coerce').fillna(0)
    else:
        out['ALL_PBM_T'] = 0

    out['ALL_PBM_DrugName'] = df[name_k].astype(
        str).str.strip() if name_k else pd.NA

    # No aggregation: if duplicates exist, keep the LAST one from the file
    out = out.drop_duplicates(subset=['NDC #'], keep='last')

    return out[['NDC #', 'ALL_PBM_Q', 'ALL_PBM_T', 'ALL_PBM_DrugName']]


def find_kinray_price_by_month(ndc, fill_date, kinray_df):
    """
    Find Kinray price for NDC based on fill_date:
    1. Search same month/year as fill_date (latest purchase in that month)
    2. If not found, search backwards month by month
    3. If not found, search forwards month by month
    4. Return 0 if never found
    """
    if pd.isna(fill_date) or kinray_df.empty:
        return 0
    
    # Filter for this NDC
    ndc_purchases = kinray_df[kinray_df['NDC #'] == ndc].copy()
    if ndc_purchases.empty:
        return 0
    
    # Ensure DATE is datetime
    ndc_purchases['DATE'] = pd.to_datetime(ndc_purchases['DATE'], errors='coerce')
    ndc_purchases = ndc_purchases.dropna(subset=['DATE', '__UnitPrice__'])
    
    if ndc_purchases.empty:
        return 0
    
    fill_date = pd.to_datetime(fill_date)
    target_year = fill_date.year
    target_month = fill_date.month
    
    # Try same month first
    same_month = ndc_purchases[
        (ndc_purchases['DATE'].dt.year == target_year) &
        (ndc_purchases['DATE'].dt.month == target_month)
    ]
    if not same_month.empty:
        return same_month.sort_values('DATE').iloc[-1]['__UnitPrice__']
    
    # Get min and max dates available
    min_date = ndc_purchases['DATE'].min()
    max_date = ndc_purchases['DATE'].max()
    
    # Search backwards
    current_date = fill_date
    while current_date >= min_date:
        current_date = current_date - pd.DateOffset(months=1)
        month_data = ndc_purchases[
            (ndc_purchases['DATE'].dt.year == current_date.year) &
            (ndc_purchases['DATE'].dt.month == current_date.month)
        ]
        if not month_data.empty:
            return month_data.sort_values('DATE').iloc[-1]['__UnitPrice__']
    
    # Search forwards
    current_date = fill_date
    while current_date <= max_date:
        current_date = current_date + pd.DateOffset(months=1)
        month_data = ndc_purchases[
            (ndc_purchases['DATE'].dt.year == current_date.year) &
            (ndc_purchases['DATE'].dt.month == current_date.month)
        ]
        if not month_data.empty:
            return month_data.sort_values('DATE').iloc[-1]['__UnitPrice__']
    
    return 0


def add_rx_unit_compare_sheet_exact(
    wb,
    log_df,
    kinray_df,
    sheet_name: str = "RX Comparison - All"
):
    """
    Output columns (exact order):
      Rx, NDC, Drug Name, Fill date, Qty filled, Package billed,
      Kinray Unit Price, Ins paid, Unit Ins paid, Difference

    ✅ Shows ONLY rows where Difference < 0 (underpaid RXs)
    ✅ Sorted by Fill Date descending (latest first)
    """

    df = log_df.copy()
    #print(df.head())
    if '* SDRA Amt' in df.columns and 'SDRA Amt' not in df.columns:
        df.rename(columns={'* SDRA Amt': 'SDRA Amt'}, inplace=True)
    if 'Copay' in df.columns and 'COPAY' not in df.columns:
        df.rename(columns={'Copay': 'COPAY'}, inplace=True)

    # --- Normalize numeric columns ---
    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'Plan 1 BIN',
              'Plan 2 BIN', 'SDRA Amt', 'COPAY']:
        df[c] = pd.to_numeric(df.get(c, 0), errors='coerce').fillna(0)

    # Normalize NDC
    df['NDC #'] = (df['NDC #'].astype(str)
                   .str.replace('-', '', regex=False)
                   .str.replace(r'\D', '', regex=True)
                   .str.zfill(11))
    
    # Detect Fill Date column first (needed for price lookup)
    date_candidates = ['Fill Date', 'Date',
                       'Rx Date', 'Dispense Date', 'Service Date']
    fill_date_col = next((c for c in date_candidates if c in df.columns), None)
    if fill_date_col:
        df['Fill Date'] = pd.to_datetime(df[fill_date_col], errors='coerce')
    else:
        df['Fill Date'] = pd.NaT
    
    # Apply month-based Kinray price lookup
    df['Kinray Unit Price'] = df.apply(
        lambda row: find_kinray_price_by_month(row['NDC #'], row['Fill Date'], kinray_df),
        axis=1
    )

    # --- Winning insurance paid ---
    df['Ins paid'] = np.where(
        df['Ins Paid Plan 1'].fillna(0) >= df['Ins Paid Plan 2'].fillna(0),
        df['Ins Paid Plan 1'].fillna(0),
        df['Ins Paid Plan 2'].fillna(0)
    )

    # --- Package billed ---
    df['Package billed'] = np.where(
        df['Drug Pkg Size'] > 0,
        df['Qty Filled'] / df['Drug Pkg Size'],
        np.nan
    )

    df['Kinray final Price'] = np.where(
        (df['Drug Pkg Size'] > 0) & (df['Kinray Unit Price'] > 0),
        (df['Kinray Unit Price']/df['Drug Pkg Size']) * df['Qty Filled'],
        0.0
    )

    # --- Unit insurance paid (per package logic) ---
    df['Unit Ins paid'] = np.where(
        df['Package billed'] > 0,
        df['Ins paid'] / df['Package billed'],
        np.nan
    )

    # Total paid (Insurance + SDRA + Copay)
    df['Total Ins paid'] = df['Ins paid'] + df['SDRA Amt'] + df['COPAY']

    # Difference = Total Ins paid - Kinray final Price
    # If Kinray Unit Price is 0, force Difference = 0
    df['Difference'] = np.where(
        df['Kinray Unit Price'] > 0,
        df['Total Ins paid'] - df['Kinray final Price'],
        0.0
    )

    # Drop rows where Difference is positive or 0
    # df = df[df['Difference'] > 0]

    # Map Rx column
    rx_col = 'Rx #' if 'Rx #' in df.columns else (
        'Rx' if 'Rx' in df.columns else None)
    df['RX'] = df[rx_col] if rx_col else pd.NA
    df['NDC'] = df['NDC #']
    df['Drug Name'] = df['Drug Name']
    df['Pkg Size'] = df['Drug Pkg Size']
    df['Qty Filled'] = df['Qty Filled']
    df['BIN'] = df['Winning_BIN']
    df['Processor'] = df['Processor']
    df['PCN'] = df['Winning PCN']
    df['Group'] = df['Winning Group']
    df['Fill Date'] = df['Fill Date']
    df['Kinray Final Price'] = df['Kinray final Price']
    df['Ins Paid'] = df['Ins paid']
    df['SDRA Amt'] = df['SDRA Amt']
    df['COPAY'] = df['COPAY']
    df['Total = (Ins Paid + SDRA + COPAY)'] = df['Total Ins paid']
    df['Package Billed'] = df['Package billed']
    out_cols = [
        'RX', 'Fill Date', 'NDC', 'Drug Name', 'Pkg Size', 'Qty Filled',
        'Package Billed', 'Kinray Final Price', 'Ins Paid', 'SDRA Amt', 'COPAY', 'Total = (Ins Paid + SDRA + COPAY)', 'Difference', 'BIN', 'Processor',
        'PCN', 'Group'
    ]

    # Filter, then sort by Fill Date DESCENDING (latest first)
    out = df.loc[:, out_cols].copy()
    # latest first, then largest diff
    out = out.sort_values('Drug Name', ascending=True)

    # # If no underpaid rows, create placeholder sheet
    # if out.empty:
    #     if sheet_name in wb.sheetnames:
    #         del wb[sheet_name]
    #     ws = wb.create_sheet(title=sheet_name)
    #     ws['A1'] = "No underpaid RXs found (Difference ≥ 0)."
    #     return

    # --- Create Sheet ---
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(out_cols))
    t = ws.cell(row=1, column=1, value="RX Comparision Analysis (All RXs)")
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 26

    # Write table
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
            else:
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

    # ✅ Wrap specific headers
    for cell_ref in ["E2", "F2", "G2", "L2"]:
        ws[cell_ref].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

    # Borders & formatting
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(out_cols)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 9, 'NDC': 14, 'Drug Name': 45, 'Pkg Size': 8, 'Fill Date': 12,
        'Qty Filled': 8, 'Package Billed': 9, 'Kinray Final Price': 16,
        'Ins Paid': 14, 'SDRA Amt': 12, 'COPAY': 10, 'Total = (Ins Paid + SDRA + COPAY)': 24,
        'Difference': 14, 'BIN': 8,
        'PCN': 12, 'Group': 12, 'Processor': 15
    }
    ws.row_dimensions[2].height = 50

    for i, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    # Number formats
    for r in range(3, ws.max_row + 1):
        for name in ['Kinray Final Price', 'Ins Paid', 'SDRA Amt', 'COPAY', 'Total = (Ins Paid + SDRA + COPAY)', 'Difference']:
            idx = out_cols.index(name) + 1
            ws.cell(row=r, column=idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=out_cols.index(
            'Qty Filled') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index(
            'Package Billed') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index(
            'Fill Date') + 1).number_format = 'yyyy-mm-dd'

    diff_idx = out_cols.index('Difference') + 1
    last_data_row = ws.max_row
    total_row = last_data_row + 1

    # Label cell (optional)
    label_col = diff_idx - 1
    label_cell = ws.cell(row=total_row, column=label_col,
                         value="Total Difference")
    label_cell.font = Font(bold=True, size=12)
    label_cell.alignment = Alignment(horizontal='center', vertical='center')
    drug_idx = out_cols.index('Drug Name') + 1
    diff_idx = out_cols.index('Difference') + 1
    left_idx = min(drug_idx, diff_idx)
    right_idx = max(drug_idx, diff_idx)

    left_col = get_column_letter(left_idx)
    right_col = get_column_letter(right_idx)

    # Apply number format to Total Difference cell
    total_diff_cell = ws.cell(row=total_row, column=diff_idx)
    total_diff_cell.number_format = '"$"#,##0.00'
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{last_data_row}"

    # AutoSum cell
    sum_col_letter = get_column_letter(diff_idx)
    total_cell = ws.cell(row=total_row, column=diff_idx)
    # total_cell.value = f"=SUM({sum_col_letter}3:{sum_col_letter}{last_data_row})"
    total_cell.value = f"=SUBTOTAL(109,{sum_col_letter}3:{sum_col_letter}{last_data_row})"
    total_cell.number_format = '"$"#,##0.00'  # ✅ Currency format

    total_cell.font = Font(bold=True, size=12)
    total_cell.number_format = 'General'
    total_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ensure Excel recalculates when opening
    ws.parent.calculation.fullCalcOnLoad = True
    # Freeze panes
    ws.freeze_panes = "A3"
    from openpyxl.formatting.rule import CellIsRule
    diff_col_letter = get_column_letter(out_cols.index("Difference") + 1)
    data_range = f"{diff_col_letter}3:{diff_col_letter}{ws.max_row}"

    # 🔴 Negative values = RED FILL
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='lessThan', formula=['0'],
                fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )

    # 🟢 Positive values = GREEN FILL
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    # Set page orientation to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

def add_rx_unit_compare_sheet_exact_pos(
    wb,
    log_df,
    kinray_df,
    sheet_name: str = "RX Comparison +ve"
):
    """
    Output columns (exact order):
      Rx, NDC, Drug Name, Fill date, Qty filled, Package billed,
      Kinray Unit Price, Ins paid, Unit Ins paid, Difference

    ✅ Shows ONLY rows where Difference < 0 (underpaid RXs)
    ✅ Sorted by Fill Date descending (latest first)
    """
    df = log_df.copy()
    # print(df.head())
    if '* SDRA Amt' in df.columns and 'SDRA Amt' not in df.columns:
        df.rename(columns={'* SDRA Amt': 'SDRA Amt'}, inplace=True)
    if 'Copay' in df.columns and 'COPAY' not in df.columns:
        df.rename(columns={'Copay': 'COPAY'}, inplace=True)

    # --- Normalize numeric columns ---
    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'Plan 1 BIN',
              'Plan 2 BIN', 'SDRA Amt', 'COPAY']:
        df[c] = pd.to_numeric(df.get(c, 0), errors='coerce').fillna(0)

    # Normalize NDC
    df['NDC #'] = (df['NDC #'].astype(str)
                   .str.replace('-', '', regex=False)
                   .str.replace(r'\D', '', regex=True)
                   .str.zfill(11))
    
    # Detect Fill Date column first (needed for price lookup)
    date_candidates = ['Fill Date', 'Date',
                       'Rx Date', 'Dispense Date', 'Service Date']
    fill_date_col = next((c for c in date_candidates if c in df.columns), None)
    if fill_date_col:
        df['Fill Date'] = pd.to_datetime(df[fill_date_col], errors='coerce')
    else:
        df['Fill Date'] = pd.NaT
    
    # Apply month-based Kinray price lookup
    df['Kinray Unit Price'] = df.apply(
        lambda row: find_kinray_price_by_month(row['NDC #'], row['Fill Date'], kinray_df),
        axis=1
    )

    # --- Winning insurance paid ---
    df['Ins paid'] = np.where(
        df['Ins Paid Plan 1'].fillna(0) >= df['Ins Paid Plan 2'].fillna(0),
        df['Ins Paid Plan 1'].fillna(0),
        df['Ins Paid Plan 2'].fillna(0)
    )

    # --- Package billed ---
    df['Package billed'] = np.where(
        df['Drug Pkg Size'] > 0,
        df['Qty Filled'] / df['Drug Pkg Size'],
        np.nan
    )

    df['Kinray final Price'] = np.where(
        (df['Drug Pkg Size'] > 0) & (df['Kinray Unit Price'] > 0),
        (df['Kinray Unit Price']/df['Drug Pkg Size']) * df['Qty Filled'],
        0.0
    )

    # --- Unit insurance paid (per package logic) ---
    df['Unit Ins paid'] = np.where(
        df['Package billed'] > 0,
        df['Ins paid'] / df['Package billed'],
        np.nan
    )

    # Total paid (Insurance + SDRA + Copay)
    df['Total Ins paid'] = df['Ins paid'] + df['SDRA Amt'] + df['COPAY']

    # Difference = Total Ins paid - Kinray final Price
    # If Kinray Unit Price is 0, force Difference = 0
    df['Difference'] = np.where(
        df['Kinray Unit Price'] > 0,
        df['Total Ins paid'] - df['Kinray final Price'],
        0.0
    )

    # Drop rows where Difference is negative or 0
    df = df[df['Difference'] > 0]

    # Map Rx column
    rx_col = 'Rx #' if 'Rx #' in df.columns else (
        'Rx' if 'Rx' in df.columns else None)
    df['RX'] = df[rx_col] if rx_col else pd.NA
    df['NDC'] = df['NDC #']
    df['Drug Name'] = df['Drug Name']
    df['Pkg Size'] = df['Drug Pkg Size']
    df['Qty Filled'] = df['Qty Filled']
    df['BIN'] = df['Winning_BIN']
    df['Processor'] = df['Processor']
    df['PCN'] = df['Winning PCN']
    df['Group'] = df['Winning Group']
    df['Fill Date'] = df['Fill Date']
    df['Kinray Final Price'] = df['Kinray final Price']
    df['Ins Paid'] = df['Ins paid']
    df['SDRA Amt'] = df['SDRA Amt']
    df['COPAY'] = df['COPAY']
    df['Total = (Ins Paid + SDRA + COPAY)'] = df['Total Ins paid']
    df['Package Billed'] = df['Package billed']
    out_cols = [
        'RX', 'Fill Date', 'NDC', 'Drug Name', 'Pkg Size', 'Qty Filled',
        'Package Billed', 'Kinray Final Price', 'Ins Paid', 'SDRA Amt', 'COPAY', 'Total = (Ins Paid + SDRA + COPAY)', 'Difference', 'BIN', 'Processor',
        'PCN', 'Group'
    ]

    # Filter, then sort by Fill Date DESCENDING (latest first)
    out = df.loc[:, out_cols].copy()
    # latest first, then largest diff
    out = out.sort_values('Drug Name', ascending=True)

    # # If no underpaid rows, create placeholder sheet
    # if out.empty:
    #     if sheet_name in wb.sheetnames:
    #         del wb[sheet_name]
    #     ws = wb.create_sheet(title=sheet_name)
    #     ws['A1'] = "No underpaid RXs found (Difference ≥ 0)."
    #     return

    # --- Create Sheet ---
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(out_cols))
    t = ws.cell(row=1, column=1, value="RX Comparision +ve Analysis (All RXs)")
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 26

    # Write table
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
            else:
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

    # ✅ Wrap specific headers
    for cell_ref in ["E2", "F2", "G2", "L2"]:
        ws[cell_ref].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

    # Borders & formatting
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(out_cols)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 9, 'NDC': 14, 'Drug Name': 45, 'Pkg Size': 8, 'Fill Date': 12,
        'Qty Filled': 8, 'Package Billed': 9, 'Kinray Final Price': 16,
        'Ins Paid': 14, 'SDRA Amt': 12, 'COPAY': 10, 'Total = (Ins Paid + SDRA + COPAY)': 24,
        'Difference': 14, 'BIN': 8,
        'PCN': 12, 'Group': 12, 'Processor': 15
    }
    ws.row_dimensions[2].height = 50

    for i, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    # Number formats
    for r in range(3, ws.max_row + 1):
        for name in ['Kinray Final Price', 'Ins Paid', 'SDRA Amt', 'COPAY', 'Total = (Ins Paid + SDRA + COPAY)', 'Difference']:
            idx = out_cols.index(name) + 1
            ws.cell(row=r, column=idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=out_cols.index(
            'Qty Filled') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index(
            'Package Billed') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index(
            'Fill Date') + 1).number_format = 'yyyy-mm-dd'

    diff_idx = out_cols.index('Difference') + 1
    last_data_row = ws.max_row
    total_row = last_data_row + 1

    # Label cell (optional)
    label_col = diff_idx - 1
    label_cell = ws.cell(row=total_row, column=label_col,
                         value="Total Difference")
    label_cell.font = Font(bold=True, size=12)
    label_cell.alignment = Alignment(horizontal='center', vertical='center')
    drug_idx = out_cols.index('Drug Name') + 1
    diff_idx = out_cols.index('Difference') + 1
    left_idx = min(drug_idx, diff_idx)
    right_idx = max(drug_idx, diff_idx)

    left_col = get_column_letter(left_idx)
    right_col = get_column_letter(right_idx)

    # Apply number format to Total Difference cell
    total_diff_cell = ws.cell(row=total_row, column=diff_idx)
    total_diff_cell.number_format = '"$"#,##0.00'
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{last_data_row}"

    # AutoSum cell
    sum_col_letter = get_column_letter(diff_idx)
    total_cell = ws.cell(row=total_row, column=diff_idx)
    # total_cell.value = f"=SUM({sum_col_letter}3:{sum_col_letter}{last_data_row})"
    total_cell.value = f"=SUBTOTAL(109,{sum_col_letter}3:{sum_col_letter}{last_data_row})"
    total_cell.number_format = '"$"#,##0.00'  # ✅ Currency format

    total_cell.font = Font(bold=True, size=12)
    total_cell.number_format = 'General'
    total_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ensure Excel recalculates when opening
    ws.parent.calculation.fullCalcOnLoad = True
    # Freeze panes
    ws.freeze_panes = "A3"

    # Set page orientation to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_mfp_drugs_sheet(
    wb,
    log_df,
    kinray_df,
    sheet_name: str = "MFP Drugs - RX"
):
    """
    Per-RX MFP analysis sheet.

    Rule used:
    - Any RX row with SDRA Amt != 0 is considered an MFP RX.
    """
    df = log_df.copy()

    for c in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'SDRA Amt', 'COPAY']:
        df[c] = pd.to_numeric(df.get(c, 0), errors='coerce').fillna(0)

    df['NDC #'] = (df['NDC #'].astype(str)
                   .str.replace('-', '', regex=False)
                   .str.replace(r'\D', '', regex=True)
                   .str.zfill(11))

    date_candidates = ['Fill Date', 'Date', 'Rx Date', 'Dispense Date', 'Service Date']
    fill_date_col = next((c for c in date_candidates if c in df.columns), None)
    if fill_date_col:
        df['Fill Date'] = pd.to_datetime(df[fill_date_col], errors='coerce')
    else:
        df['Fill Date'] = pd.NaT

    df = df[df['SDRA Amt'].fillna(0) != 0].copy()

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    if df.empty:
        ws['A1'] = "No MFP RXs found (SDRA Amt is 0/blank for all rows)."
        ws['A1'].font = Font(size=14, bold=True)
        return

    df['Kinray Unit Price'] = df.apply(
        lambda row: find_kinray_price_by_month(row['NDC #'], row['Fill Date'], kinray_df),
        axis=1
    )

    df['Package billed'] = np.where(
        df['Drug Pkg Size'] > 0,
        df['Qty Filled'] / df['Drug Pkg Size'],
        np.nan
    )
    df['Kinray Final Price'] = np.where(
        (df['Drug Pkg Size'] > 0) & (df['Kinray Unit Price'] > 0),
        (df['Kinray Unit Price'] / df['Drug Pkg Size']) * df['Qty Filled'],
        0.0
    )

    df['Winning Ins Paid'] = np.where(
        df['Ins Paid Plan 1'].fillna(0) >= df['Ins Paid Plan 2'].fillna(0),
        df['Ins Paid Plan 1'].fillna(0),
        df['Ins Paid Plan 2'].fillna(0)
    )
    df['Total Collected'] = df['Winning Ins Paid'] + df['SDRA Amt'] + df['COPAY']
    df['Difference'] = df['Total Collected'] - df['Kinray Final Price']

    rx_col = 'Rx #' if 'Rx #' in df.columns else ('Rx' if 'Rx' in df.columns else None)
    df['RX'] = df[rx_col] if rx_col else pd.NA
    df['NDC'] = df['NDC #']
    df['Drug Name'] = df.get('Drug Name', '')
    df['Pkg Size'] = df.get('Drug Pkg Size', 0)
    df['Qty Filled'] = df.get('Qty Filled', 0)
    df['BIN'] = df.get('Winning_BIN', '')
    df['Processor'] = df.get('Processor', '')
    df['PCN'] = df.get('Winning PCN', '')
    df['Group'] = df.get('Winning Group', '')

    out_cols = [
        'RX', 'Fill Date', 'NDC', 'Drug Name', 'Pkg Size', 'Qty Filled',
        'Package billed', 'Kinray Unit Price', 'Kinray Final Price',
        'Ins Paid Plan 1', 'Ins Paid Plan 2', 'SDRA Amt', 'COPAY',
        'Winning Ins Paid', 'Total Collected', 'Difference',
        'BIN', 'Processor', 'PCN', 'Group'
    ]
    out = df.loc[:, out_cols].copy().sort_values(['Drug Name', 'Fill Date'], ascending=[True, False])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out_cols))
    title = ws.cell(row=1, column=1, value="MFP DRUGS")
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 26

    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                if out.columns[c_idx - 1] == 'Drug Name':
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(out_cols)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 9, 'Fill Date': 12, 'NDC': 14, 'Drug Name': 40, 'Pkg Size': 8, 'Qty Filled': 10,
        'Package billed': 12, 'Kinray Unit Price': 14, 'Kinray Final Price': 16,
        'Ins Paid Plan 1': 14, 'Ins Paid Plan 2': 14, 'SDRA Amt': 12, 'COPAY': 10,
        'Winning Ins Paid': 14, 'Total Collected': 15, 'Difference': 13,
        'BIN': 9, 'Processor': 15, 'PCN': 12, 'Group': 12
    }
    ws.row_dimensions[2].height = 45
    for i, name in enumerate(out_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    currency_cols = {
        'Kinray Unit Price', 'Kinray Final Price',
        'Ins Paid Plan 1', 'Ins Paid Plan 2', 'SDRA Amt', 'COPAY',
        'Winning Ins Paid', 'Total Collected', 'Difference'
    }
    for r in range(3, ws.max_row + 1):
        for name in out_cols:
            idx = out_cols.index(name) + 1
            if name in currency_cols:
                ws.cell(row=r, column=idx).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=out_cols.index('Qty Filled') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index('Package billed') + 1).number_format = '0.0'
        ws.cell(row=r, column=out_cols.index('Fill Date') + 1).number_format = 'yyyy-mm-dd'

    last_data_row = ws.max_row
    total_row = last_data_row + 1
    label_col = max(1, out_cols.index('Difference'))
    ws.cell(row=total_row, column=label_col, value='Totals').font = Font(bold=True)
    ws.cell(row=total_row, column=label_col).alignment = Alignment(horizontal='right', vertical='center')

    for name in ['Kinray Final Price', 'Ins Paid Plan 1', 'Ins Paid Plan 2', 'SDRA Amt', 'COPAY',
                 'Winning Ins Paid', 'Total Collected', 'Difference']:
        idx = out_cols.index(name) + 1
        col_letter = get_column_letter(idx)
        tcell = ws.cell(row=total_row, column=idx)
        tcell.value = f"=SUBTOTAL(109,{col_letter}3:{col_letter}{last_data_row})"
        tcell.font = Font(bold=True)
        tcell.number_format = '"$"#,##0.00'
        tcell.alignment = Alignment(horizontal='center', vertical='center')
        tcell.border = thin

    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{last_data_row}"
    ws.freeze_panes = "A3"

    diff_col_letter = get_column_letter(out_cols.index('Difference') + 1)
    data_range = f"{diff_col_letter}3:{diff_col_letter}{last_data_row}"
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
    )
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
    )

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_zero_refills_sheet(
    wb,
    log_df,
    sheet_name: str = "Refills 0 - Call Doctor"
):
    df = log_df.copy()

    def pick_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    rx_col = pick_col(['Rx #', 'Rx', 'RX'])
    refill_col = pick_col(['Refills Left', 'Refills Remaining'])
    fill_col = pick_col(['Fill Date', 'Dispense DateTime', 'Date', 'Rx Date', 'Service Date'])
    days_col = pick_col(['Days', 'Days Supply'])

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    if not rx_col or not refill_col:
        ws['A1'] = "Missing required columns for zero-refill list (Rx and Refills Left)."
        return

    df['__RX__'] = df[rx_col].astype(str).str.strip()
    df = df[df['__RX__'].ne('')].copy()

    if fill_col:
        df['__FillDate__'] = pd.to_datetime(df[fill_col], errors='coerce')
    else:
        df['__FillDate__'] = pd.NaT

    df['__RefillsLeft__'] = pd.to_numeric(df[refill_col], errors='coerce').fillna(0)
    if days_col:
        df['__Days__'] = pd.to_numeric(df[days_col], errors='coerce').fillna(0)
    else:
        df['__Days__'] = 0

    df = df.sort_values(['__RX__', '__FillDate__'])
    latest = df.drop_duplicates(subset=['__RX__'], keep='last').copy()

    out = latest[latest['__RefillsLeft__'] <= 0].copy()
    if out.empty:
        ws['A1'] = "No RX found with Refills Left = 0."
        return

    patient_col = pick_col(['Patient Name'])
    dob_col = pick_col(['Patient DOB'])
    phone_col = pick_col(['Patient Cell Phone', 'Patient Phone', 'Patient Work Phone'])
    drug_col = pick_col(['Drug Name'])
    ndc_col = pick_col(['Drug NDC', 'NDC #'])
    prescriber_col = pick_col(['Prescriber Name'])
    status_col = pick_col(['Rx Status', 'Status'])
    workflow_col = pick_col(['Workflow Status'])

    out['Expected Next Fill Date'] = out['__FillDate__'] + pd.to_timedelta(out['__Days__'], unit='D')

    display = pd.DataFrame({
        'RX': out['__RX__'],
        'Patient Name': out[patient_col] if patient_col else pd.NA,
        'Patient DOB': out[dob_col] if dob_col else pd.NA,
        'Phone': out[phone_col] if phone_col else pd.NA,
        'Drug Name': out[drug_col] if drug_col else pd.NA,
        'NDC': out[ndc_col] if ndc_col else pd.NA,
        'Last Fill Date': out['__FillDate__'],
        'Days Supply': out['__Days__'],
        'Refills Left': out['__RefillsLeft__'],
        'Expected Next Fill Date': out['Expected Next Fill Date'],
        'Prescriber': out[prescriber_col] if prescriber_col else pd.NA,
        'Rx Status': out[status_col] if status_col else pd.NA,
        'Workflow Status': out[workflow_col] if workflow_col else pd.NA,
        'Doctor Called?': '',
        'Call Date': pd.NaT,
        'Outcome': '',
        'New RX Received?': '',
    })

    display = display.sort_values(['Expected Next Fill Date', 'Last Fill Date'], ascending=[True, False])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display.columns))
    title = ws.cell(row=1, column=1, value="Refills Left = 0 (Call Doctor List)")
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42

    for r_idx, row in enumerate(dataframe_to_rows(display, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(display.columns)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 10,
        'Patient Name': 28,
        'Patient DOB': 12,
        'Phone': 14,
        'Drug Name': 40,
        'NDC': 14,
        'Last Fill Date': 13,
        'Days Supply': 10,
        'Refills Left': 10,
        'Expected Next Fill Date': 16,
        'Prescriber': 26,
        'Rx Status': 14,
        'Workflow Status': 18,
        'Doctor Called?': 12,
        'Call Date': 12,
        'Outcome': 22,
        'New RX Received?': 14,
    }
    for i, name in enumerate(display.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)

    date_cols = [
        display.columns.get_loc('Patient DOB') + 1,
        display.columns.get_loc('Last Fill Date') + 1,
        display.columns.get_loc('Expected Next Fill Date') + 1,
        display.columns.get_loc('Call Date') + 1,
    ]
    for r in range(3, ws.max_row + 1):
        for dc in date_cols:
            ws.cell(row=r, column=dc).number_format = 'mm-dd-yyyy'

    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = 'A3'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_missed_refill_revenue_sheet(
    wb,
    log_df,
    sheet_name: str = "Missed Refill - Revenue Recovery",
    grace_days: int = 7
):
    df = log_df.copy()

    def pick_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    rx_col = pick_col(['Rx #', 'Rx', 'RX'])
    fill_col = pick_col(['Fill Date', 'Dispense DateTime', 'Date', 'Rx Date', 'Service Date'])
    days_col = pick_col(['Days', 'Days Supply'])
    refills_left_col = pick_col(['Refills Left', 'Refills Remaining'])

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    if not rx_col or not fill_col or not days_col or not refills_left_col:
        ws['A1'] = "Missing required columns for missed refill analysis (Rx, Fill Date, Days, Refills Left)."
        return

    df['__RX__'] = df[rx_col].astype(str).str.strip()
    df = df[df['__RX__'].ne('')].copy()

    df['__FillDate__'] = pd.to_datetime(df[fill_col], errors='coerce')
    df['__Days__'] = pd.to_numeric(df[days_col], errors='coerce').fillna(0)
    df['__RefillsLeft__'] = pd.to_numeric(df[refills_left_col], errors='coerce').fillna(0)

    total_col = pick_col(['Total', 'Ins Paid Total'])
    if total_col:
        df['__TotalCollected__'] = pd.to_numeric(df[total_col], errors='coerce').fillna(0)
    else:
        ins1 = pd.to_numeric(df.get('Ins Paid Plan 1', 0), errors='coerce').fillna(0)
        ins2 = pd.to_numeric(df.get('Ins Paid Plan 2', 0), errors='coerce').fillna(0)
        copay = pd.to_numeric(df.get('Copay', df.get('COPAY', 0)), errors='coerce').fillna(0)
        sdra = pd.to_numeric(df.get('* SDRA Amt', df.get('SDRA Amt', 0)), errors='coerce').fillna(0)
        df['__TotalCollected__'] = np.maximum(ins1, ins2) + copay + sdra

    patient_col = pick_col(['Patient Name'])
    patient_id_col = pick_col(['Patient ID'])
    dob_col = pick_col(['Patient DOB'])
    phone_col = pick_col(['Patient Cell Phone', 'Patient Phone', 'Patient Work Phone'])
    drug_col = pick_col(['Drug Name'])
    ndc_col = pick_col(['Drug NDC', 'NDC #'])
    drug_group_col = pick_col(['Drug Group'])
    prescriber_col = pick_col(['Prescriber Name'])
    status_col = pick_col(['Rx Status', 'Status'])
    workflow_col = pick_col(['Workflow Status'])
    bin_col = pick_col(['Plan 1 BIN'])
    processor_col = pick_col(['Ins Group', 'Processor'])

    def norm_text(s):
        if pd.isna(s):
            return ''
        return re.sub(r'[^A-Z0-9]+', '', str(s).upper().strip())

    patient_base = (
        df[patient_id_col].map(norm_text)
        if patient_id_col else
        (df[patient_col].map(norm_text) if patient_col else pd.Series('', index=df.index))
    )
    if dob_col:
        dob_norm = pd.to_datetime(df[dob_col], errors='coerce').dt.strftime('%Y%m%d').fillna('')
    else:
        dob_norm = pd.Series('', index=df.index)

    drug_group_base = df[drug_group_col].map(norm_text) if drug_group_col else pd.Series('', index=df.index)
    drug_name_base = df[drug_col].map(norm_text) if drug_col else pd.Series('', index=df.index)
    ndc_base = (df[ndc_col].astype(str)
                .str.replace(r'\D', '', regex=True)
                .str.zfill(11)) if ndc_col else pd.Series('', index=df.index)

    uom_col = pick_col(['Unit of Measure'])
    uom_base = df[uom_col].map(norm_text) if uom_col else pd.Series('', index=df.index)

    def infer_form_token(name_val, uom_val):
        if uom_val:
            if uom_val in {'ML', 'EACH', 'TABLET', 'CAPSULE', 'GRAM'}:
                return uom_val
        n = '' if pd.isna(name_val) else str(name_val).upper()
        checks = [
            ('PEN', 'PEN'),
            ('INJECTION', 'INJECTION'),
            ('SYRINGE', 'SYRINGE'),
            ('TABLET', 'TABLET'),
            ('TAB', 'TABLET'),
            ('CAPSULE', 'CAPSULE'),
            ('CAP ', 'CAPSULE'),
            ('AEROSOL', 'INHALER'),
            ('INHAL', 'INHALER'),
            ('STRIP', 'STRIP'),
            ('SOLUTION', 'SOLUTION'),
            ('SUSPENSION', 'SUSPENSION'),
            ('PATCH', 'PATCH'),
            ('CREAM', 'CREAM'),
            ('OINTMENT', 'OINTMENT'),
            ('DROPS', 'DROPS'),
        ]
        for key, val in checks:
            if key in n:
                return val
        return ''

    THERAPY_ALIAS_MAP = {
        'APIXABAN': ['ELIQUIS', 'APIXABAN'],
        'RIVAROXABAN': ['XARELTO', 'RIVAROXABAN'],
        'DABIGATRAN': ['PRADAXA', 'DABIGATRAN'],
        'EDOXABAN': ['SAVAYSA', 'EDOXABAN'],
        'WARFARIN': ['COUMADIN', 'JANTOVEN', 'WARFARIN'],

        'ATORVASTATIN': ['LIPITOR', 'ATORVASTATIN'],
        'ROSUVASTATIN': ['CRESTOR', 'ROSUVASTATIN'],
        'SIMVASTATIN': ['ZOCOR', 'SIMVASTATIN'],
        'PRAVASTATIN': ['PRAVACHOL', 'PRAVASTATIN'],
        'PITAVASTATIN': ['LIVALO', 'PITAVASTATIN'],

        'LEVOTHYROXINE': ['SYNTHROID', 'LEVOXYL', 'UNITHROID', 'LEVOTHYROXINE'],
        'METFORMIN': ['GLUCOPHAGE', 'METFORMIN'],
        'EMPAGLIFLOZIN': ['JARDIANCE', 'EMPAGLIFLOZIN'],
        'DAPAGLIFLOZIN': ['FARXIGA', 'DAPAGLIFLOZIN'],
        'CANAGLIFLOZIN': ['INVOKANA', 'CANAGLIFLOZIN'],
        'ERTUGLIFLOZIN': ['STEGLATRO', 'ERTUGLIFLOZIN'],
        'SITAGLIPTIN': ['JANUVIA', 'SITAGLIPTIN'],
        'SITAGLIPTIN_METFORMIN': ['JANUMET', 'SITAGLIPTIN/METFORMIN'],
        'SAXAGLIPTIN': ['ONGLYZA', 'SAXAGLIPTIN'],
        'LINAGLIPTIN': ['TRADJENTA', 'LINAGLIPTIN'],
        'PIOGLITAZONE': ['ACTOS', 'PIOGLITAZONE'],
        'GLIMEPIRIDE': ['AMARYL', 'GLIMEPIRIDE'],
        'GLIPIZIDE': ['GLUCOTROL', 'GLIPIZIDE'],
        'GLYBURIDE': ['MICRONASE', 'DIABETA', 'GLYNASE', 'GLYBURIDE'],

        'SEMAGLUTIDE_INJ': ['OZEMPIC', 'WEGOVY', 'SEMAGLUTIDE'],
        'SEMAGLUTIDE_ORAL': ['RYBELSUS'],
        'DULAGLUTIDE': ['TRULICITY', 'DULAGLUTIDE'],
        'LIRAGLUTIDE': ['VICTOZA', 'SAXENDA', 'LIRAGLUTIDE'],
        'TIRZEPATIDE': ['MOUNJARO', 'ZEPBOUND', 'TIRZEPATIDE'],
        'EXENATIDE': ['BYETTA', 'EXENATIDE'],
        'EXENATIDE_ER': ['BYDUREON', 'BYDUREON BCISE', 'EXENATIDE ER'],

        'INSULIN_GLARGINE': ['LANTUS', 'BASAGLAR', 'INSULIN GLARGINE'],
        'INSULIN_GLARGINE_U300': ['TOUJEO'],
        'INSULIN_DETEMIR': ['LEVEMIR', 'INSULIN DETEMIR'],
        'INSULIN_DEGLUDEC': ['TRESIBA', 'INSULIN DEGLUDEC'],
        'INSULIN_LISPRO': ['HUMALOG', 'ADMELOG', 'INSULIN LISPRO'],
        'INSULIN_ASPART': ['NOVOLOG', 'FIASP', 'INSULIN ASPART'],
        'INSULIN_GLULISINE': ['APIDRA', 'INSULIN GLULISINE'],

        'BUDESONIDE_FORMOTEROL': ['SYMBICORT', 'BUDESONIDE/FORMOTEROL'],
        'FLUTICASONE_SALMETEROL': ['ADVAIR', 'WIXELA', 'FLUTICASONE/SALMETEROL'],
        'FLUTICASONE_VILANTEROL': ['BREO', 'FLUTICASONE/VILANTEROL'],
        'FLUTICASONE_UMECLIDINIUM_VILANTEROL': ['TRELEGY', 'FLUTICASONE/UMECLIDINIUM/VILANTEROL'],
        'TIOTROPIUM': ['SPIRIVA', 'TIOTROPIUM'],
        'UMECLIDINIUM': ['INCRUSE', 'UMECLIDINIUM'],
        'ALBUTEROL': ['VENTOLIN', 'PROAIR', 'PROVENTIL', 'ALBUTEROL'],

        'CYCLOSPORINE_OPHTHALMIC': ['RESTASIS', 'CYCLOSPORINE'],
        'LIFITEGRAST_OPHTHALMIC': ['XIIDRA', 'LIFITEGRAST'],
        'RIMEGEPANT': ['NURTEC', 'RIMEGEPANT'],

        'DARUNAVIR_COBICISTAT_EMTRICITABINE_TAF': ['SYMTUZA'],
    }

    def canonical_therapy(name_val, group_val):
        g = '' if pd.isna(group_val) else str(group_val).upper()
        n = '' if pd.isna(name_val) else str(name_val).upper()
        hay = f"{g} {n}".strip()
        for canon, aliases in THERAPY_ALIAS_MAP.items():
            for alias in aliases:
                if alias in hay:
                    return canon
        return ''

    canonical_base = pd.Series(
        [canonical_therapy(nv, gv) for nv, gv in zip(df[drug_col] if drug_col else pd.Series('', index=df.index),
                                                     df[drug_group_col] if drug_group_col else pd.Series('', index=df.index))],
        index=df.index,
        dtype='object'
    )

    therapy_token = canonical_base.where(canonical_base.ne(''), drug_group_base)
    therapy_token = therapy_token.where(therapy_token.ne(''), drug_name_base)
    therapy_token = therapy_token.where(therapy_token.ne(''), ndc_base)
    form_token = pd.Series(
        [infer_form_token(nv, uv) for nv, uv in zip(df[drug_col] if drug_col else pd.Series('', index=df.index), uom_base)],
        index=df.index,
        dtype='object'
    )

    df['__PatientKey__'] = patient_base + np.where(dob_norm.ne(''), '_' + dob_norm, '')
    df['__TherapyToken__'] = therapy_token
    df['__FormToken__'] = form_token
    df['__TherapyKey__'] = np.where(
        df['__TherapyToken__'].ne(''),
        df['__PatientKey__'] + '|' + df['__TherapyToken__'],
        ''
    )

    df = df.sort_values(['__RX__', '__FillDate__'])
    latest = df.drop_duplicates(subset=['__RX__'], keep='last').copy()

    latest['Expected Refill Date'] = latest['__FillDate__'] + pd.to_timedelta(latest['__Days__'], unit='D')
    latest['Expected Refill Date + Grace'] = latest['Expected Refill Date'] + pd.to_timedelta(grace_days, unit='D')
    today = pd.Timestamp.today().normalize()

    missed = latest[
        (latest['__RefillsLeft__'] > 0)
        & latest['__FillDate__'].notna()
        & (latest['__Days__'] > 0)
        & (latest['Expected Refill Date + Grace'] < today)
    ].copy()

    if missed.empty:
        ws['A1'] = "No missed refills found for the selected grace window."
        return

    missed['Days Overdue'] = (today - missed['Expected Refill Date']).dt.days.clip(lower=0)

    coverage_rows = []
    therapy_groups = latest.groupby('__TherapyKey__', dropna=False)
    for idx, row in missed.iterrows():
        tkey = row.get('__TherapyKey__', '')
        current_fill = row['__FillDate__']
        coverage_rx = ''
        coverage_fill = pd.NaT
        covered = False
        if tkey and tkey in therapy_groups.groups:
            group_rows = therapy_groups.get_group(tkey)
            cands = group_rows[
                (group_rows['__RX__'] != row['__RX__'])
                & group_rows['__FillDate__'].notna()
                & (group_rows['__FillDate__'] > current_fill)
                & (group_rows['__FillDate__'] <= today)
            ]
            row_form = row.get('__FormToken__', '')
            if row_form:
                cands = cands[(cands.get('__FormToken__', '') == row_form) | (cands.get('__FormToken__', '') == '')]
            if not cands.empty:
                hit = cands.sort_values('__FillDate__').iloc[-1]
                coverage_rx = hit['__RX__']
                coverage_fill = hit['__FillDate__']
                covered = True

        rx_status_val = str(row[status_col]).strip().lower() if status_col else ''
        workflow_val = str(row[workflow_col]).strip().lower() if workflow_col else ''
        excluded_terms = ['void', 'reversed', 'cancel', 'transferred', 'deleted']
        excluded = any(t in rx_status_val for t in excluded_terms) or any(t in workflow_val for t in excluded_terms)

        if excluded:
            final_action = 'Exclude'
        elif covered:
            final_action = 'Covered by Other Rx'
        else:
            final_action = 'Open - Missed'

        coverage_rows.append((idx, covered, coverage_rx, coverage_fill, final_action))

    coverage_df = pd.DataFrame(
        coverage_rows,
        columns=['__idx__', '__Covered__', '__CoverRx__', '__CoverFill__', '__FinalAction__']
    ).set_index('__idx__')
    missed = missed.join(coverage_df, how='left')
    missed['__Covered__'] = missed['__Covered__'].fillna(False)
    missed['Est Recoverable $'] = np.where(
        missed['__FinalAction__'].eq('Open - Missed'),
        missed['__TotalCollected__'],
        0
    )

    out = pd.DataFrame({
        'RX': missed['__RX__'],
        'Patient Name': missed[patient_col] if patient_col else pd.NA,
        'Patient DOB': missed[dob_col] if dob_col else pd.NA,
        'Phone': missed[phone_col] if phone_col else pd.NA,
        'Drug Name': missed[drug_col] if drug_col else pd.NA,
        'NDC': missed[ndc_col] if ndc_col else pd.NA,
        'Therapy Key': missed['__TherapyKey__'],
        'Last Fill Date': missed['__FillDate__'],
        'Days Supply': missed['__Days__'],
        'Refills Left': missed['__RefillsLeft__'],
        'Expected Refill Date': missed['Expected Refill Date'],
        'Days Overdue': missed['Days Overdue'],
        'Covered by Other Rx': np.where(missed['__Covered__'], 'Yes', 'No'),
        'Covering Rx #': missed['__CoverRx__'],
        'Covering Fill Date': missed['__CoverFill__'],
        'Final Action': missed['__FinalAction__'],
        'Est Recoverable $': missed['Est Recoverable $'],
        'Rx Status': missed[status_col] if status_col else pd.NA,
        'Workflow Status': missed[workflow_col] if workflow_col else pd.NA,
        'Processor': missed[processor_col] if processor_col else pd.NA,
        'BIN': missed[bin_col] if bin_col else pd.NA,
        'Prescriber': missed[prescriber_col] if prescriber_col else pd.NA,
    })

    out = out.sort_values(['Final Action', 'Days Overdue', 'Est Recoverable $'], ascending=[True, False, False])

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(out.columns))
    title = ws.cell(row=1, column=1, value=f"Missed Refill - Revenue Recovery (Grace: {grace_days} days)")
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.font = Font(size=16, bold=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 42

    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 2:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(out.columns)):
        for cell in row:
            cell.border = thin

    widths = {
        'RX': 10,
        'Patient Name': 28,
        'Patient DOB': 12,
        'Phone': 14,
        'Drug Name': 42,
        'NDC': 14,
        'Therapy Key': 28,
        'Last Fill Date': 13,
        'Days Supply': 10,
        'Refills Left': 10,
        'Expected Refill Date': 15,
        'Days Overdue': 11,
        'Covered by Other Rx': 15,
        'Covering Rx #': 12,
        'Covering Fill Date': 14,
        'Final Action': 18,
        'Est Recoverable $': 16,
        'Rx Status': 16,
        'Workflow Status': 18,
        'Processor': 12,
        'BIN': 9,
        'Prescriber': 28,
    }

    for i, name in enumerate(out.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 14)

    money_col = out.columns.get_loc('Est Recoverable $') + 1
    date_cols = [
        out.columns.get_loc('Patient DOB') + 1,
        out.columns.get_loc('Last Fill Date') + 1,
        out.columns.get_loc('Expected Refill Date') + 1,
        out.columns.get_loc('Covering Fill Date') + 1,
    ]
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=money_col).number_format = '"$"#,##0.00'
        for dc in date_cols:
            ws.cell(row=r, column=dc).number_format = 'mm-dd-yyyy'

    total_row = ws.max_row + 1
    label_col = out.columns.get_loc('Days Overdue') + 1
    ws.cell(row=total_row, column=label_col, value='Total Potential Recovery').font = Font(bold=True)
    total_cell = ws.cell(
        row=total_row,
        column=money_col,
        value=f"=SUBTOTAL(109,{get_column_letter(money_col)}3:{get_column_letter(money_col)}{ws.max_row})"
    )
    total_cell.font = Font(bold=True)
    total_cell.number_format = '"$"#,##0.00'

    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row - 1}"
    ws.freeze_panes = 'A3'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_summary_sheet(
    wb,
    processed_source="Processed Data",           # str sheet name OR a worksheet
    needs_title="Needs to be ordered - All",
    header_row=3,
    data_start_row=4,
    pharmacy_name=None,
    date_range=None,
):
    """
    Builds a 'Summary' sheet with columns = processors (+ ALL_PBM if present) and rows:
      - Insurance $ Paid
      - $$ Purchased (Kinray)
      - Net (Paid − Purchased)
      - Insurance-wise Order Estimate ($)  <-- pulled from Needs sheet

    Notes:
    - processed_source can be a sheet name (str) or an openpyxl Worksheet.
    - Formats money as currency.
    """

    # --- Resolve processed worksheet ---
    if isinstance(processed_source, str):
        if processed_source not in wb.sheetnames:
            return
        ws_pd = wb[processed_source]
        processed_title = processed_source
    else:
        # assume it's a worksheet
        ws_pd = processed_source
        processed_title = ws_pd.title

    # --- read headers from the header_row ---
    headers = [
        ws_pd.cell(row=header_row, column=c).value
        for c in range(1, ws_pd.max_column + 1)
    ]
    headers = [h for h in headers if h]

    def _procs_by_suffix(suffix: str):
        return sorted({
            h[:-len(suffix)]
            for h in headers
            if isinstance(h, str) and h.endswith(suffix)
        })

    procs_T = _procs_by_suffix("_T")
    procs_Pur = _procs_by_suffix("_Pur")
    procs_Net = _procs_by_suffix("_Net")

    processors = sorted(set(procs_T) | set(procs_Pur) | set(procs_Net))

    # include ALL_PBM if any ALL_PBM_* exists
    if "ALL_PBM" not in processors and any(
        isinstance(h, str) and h.startswith("ALL_PBM_") for h in headers
    ):
        processors.append("ALL_PBM")

    if not processors:
        # create a small Summary sheet indicating no processor columns
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
        ws = wb.create_sheet("Summary")
        ws["A1"] = "No processor metric columns (_T, _Pur, _Net) found in processed sheet."
        return

    # helper: find exact header col index
    def col_idx_for(header_text):
        for c in range(1, ws_pd.max_column + 1):
            if ws_pd.cell(row=header_row, column=c).value == header_text:
                return c
        return None

    # build processor -> column letter maps for each band
    def band_cols(suffix: str):
        out = {}
        for p in processors:
            hdr = f"{p}{suffix}"
            idx = col_idx_for(hdr)
            if idx:
                out[p] = get_column_letter(idx)
        return out

    cols_T = band_cols("_T")
    cols_Pur = band_cols("_Pur")
    cols_Net = band_cols("_Net")

    # find last data row (use max_row as a safe upper bound)
    last_data_row = ws_pd.max_row

    # --- Create/replace Summary at last index ---
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")

    # Title (row 1)
    end_col = len(processors) + 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    title_txt = "Summary"
    ws.cell(row=1, column=1, value=title_txt).font = Font(bold=True, size=16)
    ws.cell(row=1, column=1).alignment = Alignment(
        horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Optional subtitle with pharmacy/date (row 2) if provided
    if pharmacy_name or date_range:
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=end_col)
        if pharmacy_name and date_range:
            sub = f"Summary of {pharmacy_name} for the date range {date_range}"
        else:
            sub = " · ".join([t for t in [pharmacy_name, date_range] if t])
        ws.cell(row=2, column=1, value=sub).alignment = Alignment(
            horizontal="center")
        ws.row_dimensions[2].height = 22
        header_base_row = 3
    else:
        header_base_row = 2

    # Header row (Metric + processors)
    ws.cell(row=header_base_row, column=1,
            value="Metric").font = Font(bold=True)
    for j, p in enumerate(processors, start=2):
        cell = ws.cell(row=header_base_row, column=j, value=p)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row labels start below header
    start_data_row = header_base_row + 1
    metrics = [
        "Insurance $(BestRX)",
        "100% $$ Purchased (Kinray)",
        "Net (Paid − Purchased)",
        "Needs to Ordered Sheet, Insurance-wise Order Estimate ($)",
    ]
    
    for i, m in enumerate(metrics):
        cell = ws.cell(row=start_data_row + i, column=1, value=m)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # Increase row height for better visibility
        ws.row_dimensions[start_data_row + i].height = 28

    # Currency format
    money_fmt = '"$"#,##0.00'

    # Fill formulas for the first 3 metrics from Processed Data
    row_paid = start_data_row
    row_pur = start_data_row + 1
    row_net = start_data_row + 2
    row_est = start_data_row + 3

    for j, p in enumerate(processors, start=2):
        # Paid (T)
        if p in cols_T:
            col = cols_T[p]
            ws.cell(row=row_paid, column=j,
                    value=f"=SUM('{processed_title}'!{col}{data_start_row}:{col}{last_data_row})"
                    ).number_format = money_fmt
        else:
            ws.cell(row=row_paid, column=j, value=0).number_format = money_fmt

        # Purchased (Pur)
        if p in cols_Pur:
            col = cols_Pur[p]
            ws.cell(row=row_pur, column=j,
                    value=f"=SUM('{processed_title}'!{col}{data_start_row}:{col}{last_data_row})"
                    ).number_format = money_fmt
        else:
            ws.cell(row=row_pur, column=j, value=0).number_format = money_fmt

        # Net
        if p in cols_Net:
            col = cols_Net[p]
            ws.cell(row=row_net, column=j,
                    value=f"=SUM('{processed_title}'!{col}{data_start_row}:{col}{last_data_row})"
                    ).number_format = money_fmt
        else:
            ws.cell(row=row_net, column=j, value=0).number_format = money_fmt

    # --- Insurance-wise Order Estimate ($) from Needs sheet
    if needs_title in wb.sheetnames:
        ws_need = wb[needs_title]

        def find_header_col(ws0, text, hdr_row=2):
            for c in range(1, ws0.max_column + 1):
                if (ws0.cell(row=hdr_row, column=c).value or "") == text:
                    return c
            return None

        label_col_idx = find_header_col(
            ws_need, "Insurance-wise Order Estimate ($)")
        value_col_idx = find_header_col(ws_need, "Amount")

        # Fallback if headers differ: default to first two columns
        if not label_col_idx:
            label_col_idx = 1
        if not value_col_idx:
            value_col_idx = 2

        label_col_letter = get_column_letter(label_col_idx)
        value_col_letter = get_column_letter(value_col_idx)
        last_need_row = ws_need.max_row

        for j, p in enumerate(processors, start=2):
            # Looks up "<processor>_D" label in Needs sheet
            formula = (
                f"=IFERROR(INDEX('{needs_title}'!${value_col_letter}$3:${value_col_letter}${last_need_row},"
                f"MATCH(\"{p}_D\", '{needs_title}'!${label_col_letter}$3:${label_col_letter}${last_need_row}, 0)), 0)"
            )
            c = ws.cell(row=row_est, column=j, value=formula)
            c.number_format = money_fmt
    else:
        # No Needs sheet -> zeroes
        for j, _ in enumerate(processors, start=2):
            ws.cell(row=row_est, column=j, value=0).number_format = money_fmt

    # Freeze & filter
    freeze_row = header_base_row + 1
    ws.freeze_panes = f"B{freeze_row}"
    ws.auto_filter.ref = f"A{header_base_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Column widths
    ws.column_dimensions['A'].width = 32
    for j in range(2, len(processors) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 16

    # Borders & header fill
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for r in range(header_base_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = thin

    header_fill = PatternFill(start_color="D0CECE",
                              end_color="D0CECE", fill_type="solid")
    for c in range(1, ws.max_column + 1):
        ws.cell(row=header_base_row, column=c).fill = header_fill

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


# def generate_insurance_audit_files(final_df, pharmacy_name, date_range, output_dir):
#     """
#     For each insurance/processor in final_df (INCLUDING ALL_PBM), create ONE Excel file with 2 sheets:

#       Sheet 1: 'Top 100 - Packages'
#       Sheet 2: 'Top 100 - Paid'

#     Columns (both sheets):
#       Rank | Insurance | NDC # | Drug Name | Package Size |
#       Qty Billed | Packages Billed | Total Purchased | Difference | Actual $ Paid

#     Mapping for a processor PR:
#       Qty Billed      -> PR_Q
#       Packages Billed -> PR_P
#       Difference      -> PR_D
#       Actual $ Paid   -> PR_T
#       Total Purchased -> 'Total Purchased'
#     """

#     os.makedirs(output_dir, exist_ok=True)

#     def _proc_from_col(col: str):
#         """Extract processor/insurance name from *_Q/_P/_D/_T/_Pur/_Net columns."""
#         suffixes = ('_Q', '_P', '_D', '_T', '_Pur', '_Net')
#         for sfx in suffixes:
#             if col.endswith(sfx):
#                 return col[:-len(sfx)]
#         return None

#     # ---- Discover processors from columns (INCLUDING ALL_PBM) ----
#     processors = set()
#     for c in final_df.columns:
#         p = _proc_from_col(c)
#         if p:
#             processors.add(p)

#     processors = sorted(processors)
#     if not processors:
#         print("[audit] No processor metrics found. Skipping audit files.")
#         return []

#     # For safe filenames
#     safe_pharmacy = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', str(pharmacy_name)).strip()
#     safe_range    = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', str(date_range)).strip()

#     created_files = []

#     # Shared styling
#     header_font = Font(bold=True)
#     header_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
#     thin_border = Border(
#         left=Side(style='thin', color="A9A9A9"),
#         right=Side(style='thin', color="A9A9A9"),
#         top=Side(style='thin', color="A9A9A9"),
#         bottom=Side(style='thin', color="A9A9A9")
#     )

#     def _apply_header(ws, insurance_name, last_col_index):
#         """Row 1: Pharmacy — Date Range — INSURANCE."""
#         ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_index)
#         cell = ws.cell(row=1, column=1)
#         cell.value = f"{pharmacy_name} — {date_range} — {insurance_name}"
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         cell.font = Font(size=14, bold=True)
#         ws.row_dimensions[1].height = 30

#     def _build_top_df(df_proc, pr, sort_col, top_n=100):
#         """Build the Top N dataframe for one processor, one metric (P or T)."""
#         col_q = f"{pr}_Q"
#         col_p = f"{pr}_P"
#         col_d = f"{pr}_D"
#         col_t = f"{pr}_T"

#         df = df_proc.copy()

#         # Ensure numeric
#         for c in [col_q, col_p, col_d, col_t, 'Total Purchased']:
#             if c in df.columns:
#                 df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

#         if sort_col not in df.columns:
#             return pd.DataFrame(columns=[
#                 "Rank", "Insurance", "NDC #", "Drug Name", "Package Size",
#                 "Qty Billed", "Packages Billed", "Total Purchased",
#                 "Difference", "Actual $ Paid"
#             ])

#         df = df[df[sort_col] > 0].copy()
#         if df.empty:
#             return pd.DataFrame(columns=[
#                 "Rank", "Insurance", "NDC #", "Drug Name", "Package Size",
#                 "Qty Billed", "Packages Billed", "Total Purchased",
#                 "Difference", "Actual $ Paid"
#             ])

#         df = df.sort_values(sort_col, ascending=False).head(top_n).reset_index(drop=True)

#         out = pd.DataFrame()
#         out["Rank"]            = range(1, len(df) + 1)
#         out["Insurance"]       = pr
#         out["NDC #"]           = df.get("NDC #", "")
#         out["Drug Name"]       = df.get("Drug Name", "")
#         out["Package Size"]    = df.get("Package Size", "")
#         out["Qty Billed"]      = df.get(col_q, 0)
#         out["Packages Billed"] = df.get(col_p, 0)
#         out["Total Purchased"] = df.get("Total Purchased", 0)
#         out["Difference"]      = df.get(col_d, 0)
#         out["Actual $ Paid"]   = df.get(col_t, 0)

#         return out

#     def _write_table(ws, df, subtitle=None):
#         """Write df to ws starting row 3, with optional subtitle on row 2."""
#         if subtitle:
#             ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, size=10)

#         if df.empty:
#             ws.cell(row=3, column=1, value="No data for this criteria.")
#             return

#         cols = list(df.columns)

#         # Row 3: headers
#         for c_idx, col_name in enumerate(cols, start=1):
#             cell = ws.cell(row=3, column=c_idx, value=col_name)
#             cell.font = header_font
#             cell.fill = header_fill
#             cell.alignment = Alignment(horizontal='center', vertical='center')
#             cell.border = thin_border

#         # Data from row 4
#         for r_idx, row in enumerate(df.itertuples(index=False), start=4):
#             for c_idx, col_name in enumerate(cols, start=1):
#                 val = getattr(row, col_name.replace(" ", "_"), None)
#                 if val is None:
#                     val = df.iloc[r_idx - 4][col_name]

#                 cell = ws.cell(row=r_idx, column=c_idx, value=val)
#                 if isinstance(val, (int, float)):
#                     cell.alignment = Alignment(horizontal='right', vertical='center')
#                 else:
#                     cell.alignment = Alignment(horizontal='left', vertical='center')
#                 cell.border = thin_border

#         # Auto width
#         for col_idx in range(1, len(cols) + 1):
#             col_letter = get_column_letter(col_idx)
#             max_len = 0
#             for row in ws.iter_rows(min_row=3, max_row=ws.max_row,
#                                     min_col=col_idx, max_col=col_idx):
#                 for cell in row:
#                     try:
#                         txt = str(cell.value) if cell.value is not None else ""
#                         max_len = max(max_len, len(txt))
#                     except Exception:
#                         pass
#             ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

#         ws.freeze_panes = "A4"

#     # ---- Loop over each processor and create files ----
#     for pr in processors:
#         col_p = f"{pr}_P"
#         col_t = f"{pr}_T"

#         if col_p not in final_df.columns and col_t not in final_df.columns:
#             continue

#         df_proc = final_df.copy()

#         top_packages = _build_top_df(df_proc, pr, col_p, top_n=100)
#         top_paid     = _build_top_df(df_proc, pr, col_t, top_n=100)

#         if top_packages.empty and top_paid.empty:
#             continue

#         safe_proc = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', str(pr)).strip()
#         filename  = f"{safe_pharmacy} - {safe_proc}_Audit_{safe_range}.xlsx"
#         filepath  = os.path.join(output_dir, filename)

#         wb = Workbook()

#         # Sheet 1: Top 100 - Packages
#         ws1 = wb.active
#         ws1.title = "Top 100 Drugs - Packages Billed {pr}"
#         last_col_index_1 = len(top_packages.columns) if not top_packages.empty else 10
#         _apply_header(ws1, pr, last_col_index_1)
#         _write_table(ws1, top_packages,
#                      subtitle=f"Top 100 drugs by Packages Billed for {pr}")

#         # Sheet 2: Top 100 - Paid
#         ws2 = wb.create_sheet(title="Top 100 Drugs - Ins Paid {pr}")
#         last_col_index_2 = len(top_paid.columns) if not top_paid.empty else 10
#         _apply_header(ws2, pr, last_col_index_2)
#         _write_table(ws2, top_paid,
#                      subtitle=f"Top 100 drugs by Actual $ Paid for {pr}")

#         wb.save(filepath)
#         created_files.append(filepath)
#         print(f"[audit] Created audit file: {filepath}")

#     return created_files
def generate_master_audit_workbook(final_df, pharmacy_name, date_range, output_dir):
    """
    ONE workbook with ALL insurances (including ALL_PBM).

    For each insurance/processor PR, creates:
      - Sheet: "{PR} - Top 100 Packages"
      - Sheet: "{PR} - Top 100 Ins Paid"

        Columns (both sheets):
            Rank | Insurance | NDC # | Drug Name | Package Size |
            Qty Billed | Packages Billed | Total Purchased | Difference | Actual $ Paid | Amount to be Paid to Insurance (If Audit)
    """
    import os
    import re
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(output_dir, exist_ok=True)

    # ---------- Helper: extract processor from column name ----------
    def _proc_from_col(col: str):
        """Extract processor/insurance name from *_Q/_P/_D/_T/_Pur/_Net columns."""
        suffixes = ('_Q', '_P', '_D', '_T', '_Pur', '_Net')
        for sfx in suffixes:
            if col.endswith(sfx):
                return col[:-len(sfx)]
        return None

    # ---- Discover processors from columns (INCLUDING ALL_PBM) ----
    processors = set()
    for c in final_df.columns:
        p = _proc_from_col(c)
        if p:
            processors.add(p)

    processors = sorted(processors)
    if not processors:
        print("[audit] No processor metrics found. Skipping audit sheets.")
        return None

    # Safe filename & sheet names
    safe_pharmacy = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', str(pharmacy_name)).strip()
    safe_range    = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', str(date_range)).strip()
    filename      = f"{safe_pharmacy}_Audit_{safe_range}.xlsx"
    filepath      = os.path.join(output_dir, filename)

    def safe_sheet_name(name: str) -> str:
        r"""
        Clean a string to be a valid Excel sheet name:
        - remove invalid characters: \ / ? * [ ] :
        - truncate to max 31 characters
        """
        cleaned = re.sub(r'[\\/*?:\[\]]', '_', str(name))
        return cleaned[:31]

    # ---------- Shared styling ----------
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color="A9A9A9"),
        right=Side(style='thin', color="A9A9A9"),
        top=Side(style='thin', color="A9A9A9"),
        bottom=Side(style='thin', color="A9A9A9")
    )

    def _apply_header(ws, insurance_name, last_col_index):
        """Row 1: Pharmacy — Date Range — INSURANCE."""
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col_index)
        cell = ws.cell(row=1, column=1)
        cell.value = f"{pharmacy_name} — {date_range} — {insurance_name}"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(size=20, bold=True)
        ws.row_dimensions[1].height = 30

    def _build_top_df(df_proc, pr, sort_col, top_n=None):
        """Build the Top N dataframe for one processor, one metric (P or T)."""
        col_q = f"{pr}_Q"
        col_p = f"{pr}_P"
        col_d = f"{pr}_D"
        col_t = f"{pr}_T"

        df = df_proc.copy()

        # Ensure numeric
        for c in [col_q, col_p, col_d, col_t, 'Total Purchased']:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

        if sort_col not in df.columns:
            return pd.DataFrame(columns=[
                "Rank", "Insurance", "NDC #", "Drug Name", "Package Size",
                "Qty Billed", "Packages Billed", "Total Purchased",
                "Difference", "Actual $ Paid", "Amount to be Paid to Insurance (If Audit)",
            ])

        df = df[df[sort_col] > 0].copy()
        if df.empty:
            return pd.DataFrame(columns=[
                "Rank", "Insurance", "NDC #", "Drug Name", "Package Size",
                "Qty Billed", "Packages Billed", "Total Purchased",
                "Difference", "Actual $ Paid", "Amount to be Paid to Insurance (If Audit)",
            ])

        # 🔑 reset index so values line up row-by-row
        df = df.sort_values(sort_col, ascending=False)
        if top_n is not None:
            df = df.head(top_n)
        df = df.reset_index(drop=True)
        out = pd.DataFrame()
        out["Rank"]            = range(1, len(df) + 1)
        out["Insurance"]       = pr
        out["NDC #"]           = df.get("NDC #", "")
        out["Drug Name"]       = df.get("Drug Name", "")
        out["Package Size"]    = df.get("Package Size", "")
        out["Qty Billed"]      = df.get(col_q, 0)
        out["Packages Billed"] = df.get(col_p, 0)
        out["Total Purchased"] = df.get("Total Purchased", 0)
        out["Difference"]      = df.get(col_d, 0)
        out["Actual $ Paid"]   = df.get(col_t, 0)
        out["Amount to be Paid to Insurance (If Audit)"] = 0

        # Clean <NA> → proper numbers / blanks
        num_cols  = ["Rank", "Qty Billed", "Packages Billed",
                 "Total Purchased", "Difference", "Actual $ Paid", "Amount to be Paid to Insurance (If Audit)"]
        text_cols = ["Insurance", "NDC #", "Drug Name", "Package Size"]

        #Rename Qty Billed to Qty Billed to {pr}
        out.rename(columns={"Qty Billed": f"Qty Billed to {pr}"}, inplace=True)
        #Rename Packages Billed to Packages Billed to {pr}
        out.rename(columns={"Packages Billed": f"Packages Billed to {pr}"}, inplace=True)
        #Rename Total Purchased to Total Qty Purchased
        out.rename(columns={"Total Purchased": "Total Qty Purchased"}, inplace=True) 
        #Rename Difference to Qty Difference for {pr}
        out.rename(columns={"Difference": f"Qty Difference for {pr}"}, inplace=True)
        #Rename Actual $ Paid to Actual $ Paid by {pr}
        out.rename(columns={"Actual $ Paid": f"Actual $ Paid by {pr} (BestRX)"}, inplace=True)
        for c in num_cols:
            if c in out.columns:
                out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)

        for c in text_cols:
            if c in out.columns:
                out[c] = out[c].astype(str).where(out[c].notna(), "")

        return out

    #def _write_table(ws, df, subtitle=None):
    def _write_table(ws, df, subtitle=None):
        """Write df to ws starting row 3, with optional subtitle on row 2."""
        import pandas as pd

        if df is None or df.empty:
            # Even if empty, give a subtitle (merged row 2) if requested
            if subtitle:
                empty_last_col = len(df.columns) if (df is not None and hasattr(df, 'columns') and len(df.columns) > 0) else 11
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=empty_last_col)
                c2 = ws.cell(row=2, column=1, value=subtitle)
                c2.font = Font(size=15, bold=True)
                c2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=3, column=1, value="No data for this criteria.")
            return

        # Make sure there is no pandas <NA> for openpyxl
        df = df.copy()
        df = df.astype(object).where(pd.notna(df), None)

        cols = list(df.columns)
        last_col = len(cols)

        # ---- Row 2: merged subtitle (Rank .. Actual $ Paid) ----
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
            c2 = ws.cell(row=2, column=1, value=subtitle)
            c2.font = Font(size=15, bold=True)   # 🔹 font size 15 & bold
            c2.alignment = Alignment(horizontal='center',
                                     vertical='center',
                                     wrap_text=True)
        ws.row_dimensions[2].height = 24

        # ---- Row 3: headers ----
        for c_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=3, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True   # 🔹 wrap text for header row
            )
            cell.border = thin_border

        # Row 3 height = 35
        ws.row_dimensions[3].height = 65

        # ---- Data from row 4 (positionally) ----
        for r_idx, row_vals in enumerate(df.itertuples(index=False, name=None), start=4):
            for c_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                # Default alignment: left for text, right for numbers
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border

        # ---- Column K formula: per-row audit estimate (negative only) ----
        # =IFERROR(IF((ROUND((J4/G4)*I4,1))<0,(ROUND((J4/G4)*I4,1)),0),0)
        if last_col >= 11:
            for r_idx in range(4, ws.max_row + 1):
                k_cell = ws.cell(row=r_idx, column=11)
                k_cell.value = f"=IFERROR(IF((ROUND((J{r_idx}/G{r_idx})*I{r_idx},1))<0,(ROUND((J{r_idx}/G{r_idx})*I{r_idx},1)),0),0)"
                k_cell.border = thin_border

        # ---- Column-specific formatting: E..K ----
        # E: Package Size
        # F: Qty Billed
        # G: Packages Billed
        # H: Total Purchased
        # I: Difference
        # J: Actual $ Paid
        # K: Amount to be Paid to Insurance (If Audit)
        first_data_row = 4
        last_data_row = ws.max_row

        accounting_fmt = '$#,##0.00;[Red]-$#,##0.00'

        for row in ws.iter_rows(min_row=first_data_row, max_row=last_data_row,
                                min_col=5, max_col=min(11, last_col)):
            for cell in row:
                # Center alignment & 2 decimal format
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=(cell.column == 11)
                )
                if isinstance(cell.value, (int, float)):
                    if cell.column in (10, 11):
                        cell.number_format = accounting_fmt
                    else:
                        cell.number_format = "0.00"

        if last_col >= 11:
            for r_idx in range(first_data_row, last_data_row + 1):
                ws.cell(row=r_idx, column=11).number_format = accounting_fmt

        if last_col >= 10:
            for r_idx in range(first_data_row, last_data_row + 1):
                ws.cell(row=r_idx, column=10).number_format = accounting_fmt

        # ---- Conditional formatting: highlight negatives in E..K ----
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        rng_neg = f"E{first_data_row}:{get_column_letter(min(11, last_col))}{last_data_row}"
        ws.conditional_formatting.add(
            rng_neg,
            CellIsRule(operator="lessThan", formula=["0"], stopIfTrue=False, fill=red_fill, font=red_font)
        )

        # ---- Auto-sum row at bottom (F..K) ----
        total_row = last_data_row + 1
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=total_row, column=5).border = thin_border
        ws.cell(row=total_row, column=5).fill = total_fill

        for c_idx in range(6, min(11, last_col) + 1):
            col_letter = get_column_letter(c_idx)
            tcell = ws.cell(row=total_row, column=c_idx)
            tcell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            tcell.font = Font(bold=True)
            tcell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=(c_idx == 11))
            if c_idx in (10, 11):
                tcell.number_format = accounting_fmt
            else:
                tcell.number_format = "0.00"
            tcell.border = thin_border
            tcell.fill = total_fill

        # ---- Bottom quick metrics ----
        if last_col >= 11:
            audited_count_row = total_row + 1
            exposure_row = total_row + 2

            ws.cell(row=audited_count_row, column=10, value="Audited Drugs Count").font = Font(bold=True)
            ws.cell(row=audited_count_row, column=10).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=audited_count_row, column=10).border = thin_border
            ws.cell(row=audited_count_row, column=11, value=f"=COUNTIF(K{first_data_row}:K{last_data_row},\"<0\")")
            ws.cell(row=audited_count_row, column=11).font = Font(bold=True)
            ws.cell(row=audited_count_row, column=11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=audited_count_row, column=11).border = thin_border

            ws.cell(row=exposure_row, column=10, value="Total Audit Exposure").font = Font(bold=True)
            ws.cell(row=exposure_row, column=10).alignment = Alignment(horizontal='right', vertical='center')
            ws.cell(row=exposure_row, column=10).border = thin_border
            ws.cell(row=exposure_row, column=11, value=f"=SUMIF(K{first_data_row}:K{last_data_row},\"<0\",K{first_data_row}:K{last_data_row})")
            ws.cell(row=exposure_row, column=11).font = Font(bold=True)
            ws.cell(row=exposure_row, column=11).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=exposure_row, column=11).number_format = accounting_fmt
            ws.cell(row=exposure_row, column=11).border = thin_border

        # ---- Auto column widths ----
        for col_idx in range(1, last_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=col_idx, max_col=col_idx):
                val = row[0].value
                if val is None:
                    continue
                txt = str(val)
                if len(txt) > max_len:
                    max_len = len(txt)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        ws.column_dimensions['A'].width = 5   # Rank
        ws.column_dimensions['B'].width = 12  # Insurance
        ws.column_dimensions['C'].width = 12  # NDC #
        ws.column_dimensions['D'].width = 35  # Drug Name
        ws.column_dimensions['E'].width = 8  # Package Size
        ws.column_dimensions['F'].width = 12  # Qty Billed
        ws.column_dimensions['G'].width = 12  # Packages Billed
        ws.column_dimensions['H'].width = 12  # Total Purchased
        ws.column_dimensions['I'].width = 12  # Difference
        ws.column_dimensions['J'].width = 15  # Actual $ Paid
        if last_col >= 11:
            ws.column_dimensions['K'].width = 15  # Amount to be Paid to Insurance (If Audit)
        # ---- Freeze panes at row 4 (row 1–3 fixed) ----
        ws.freeze_panes = "E4"

        # ---- Enable filter on header row (row 3) ----
        if last_col >= 11:
            ws.auto_filter.ref = f"A3:K{last_data_row}"
        else:
            ws.auto_filter.ref = f"A3:{get_column_letter(last_col)}{last_data_row}"
         # ---- 📄 Page setup: Landscape + fit all columns on one page width ----
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # or "landscape"
        ws.page_setup.fitToWidth = 1   # fit all columns to one page wide
        ws.page_setup.fitToHeight = 0  # allow multiple pages down
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # Optional: tiny margins for better fit
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        # ✅ Print debug info
        #print(f"[audit] Sheet '{ws.title}' written with {len(df)} rows and {last_col} columns (landscape, fit-to-width).")
        

    # ---------- ONE workbook for all processors ----------
    wb = Workbook()
    wb.remove(wb.active)   # we'll create sheets ourselves

    for pr in processors:
        col_q = f"{pr}_Q"
        col_p = f"{pr}_P"
        col_t = f"{pr}_T"

        # skip if we have neither Q nor T
        if col_q not in final_df.columns and col_t not in final_df.columns:
            continue

        df_proc = final_df.copy()

        top_packages = _build_top_df(df_proc, pr, col_q)
        top_paid     = _build_top_df(df_proc, pr, col_t)

        # If both empty, skip
        if top_packages.empty and top_paid.empty:
            continue

        # # ---- Sheet 1: Top 100 by Packages ----
        # sheet_name_pkg = safe_sheet_name(f"{pr} - Top 100 Qty Billed")
        # ws1 = wb.create_sheet(title=sheet_name_pkg)
        # last_col_index_1 = len(top_packages.columns) if not top_packages.empty else 10
        # _apply_header(ws1, pr, last_col_index_1)
        # _write_table(ws1, top_packages,
        #              subtitle=f"Top 100 drugs for {pr} by Qty Billed")

        # Top 100 by Ins Paid ----
        sheet_name_paid = safe_sheet_name(f"{pr}")
        ws2 = wb.create_sheet(title=sheet_name_paid)
        last_col_index_2 = len(top_paid.columns) if not top_paid.empty else 10
        _apply_header(ws2, pr, last_col_index_2)
        _write_table(ws2, top_paid,
                     subtitle=f"Overall {pr} Overview")

    # Save single master workbook
    wb.save(filepath)
    try:
        import os
        # 🧹 Remove Windows "Zone.Identifier" metadata (Protected View trigger)
        if os.name == "nt":  # only on Windows
            import subprocess
            subprocess.run(
                [
                    "powershell",
                    "-Command",
                    f'Unblock-File -Path "{filepath}"'
                ],
                shell=True
            )
    except Exception as e:
        print(f"[audit warn] Could not unblock audit file: {e}")

    #print(f"[audit] Master audit workbook created: {filepath}")
    return filepath

import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule


def _alt_norm_ndc11(ndc):
    """Normalize any NDC-like value to an 11-digit numeric string or None."""
    if pd.isna(ndc):
        return None
    s = ''.join(ch for ch in str(ndc) if ch.isdigit())
    if not s:
        return None
    return s.zfill(11)


def _alt_ndc11_to_ndc9(ndc11: str):
    if not ndc11:
        return None
    return ndc11[:9]


def build_alternate_ndc_df(log_df: pd.DataFrame,
                           all_vendor_df: pd.DataFrame) -> pd.DataFrame:
    """
    Build a DataFrame describing alternate NDC (same NDC-9) purchase opportunities.

    Uses:
      - log_df    : custom log with 'NDC #', 'Drug Name', 'Drug Pkg Size', 'Qty Filled'
      - all_vendor_df : vendor rows with 'NDC #', 'Shipped', 'PRICE', 'DATE', 'Vendor', 'Drug Name' (optional)

    Each row ~ (Billed NDC, Alternate Purchased NDC)

    Final columns:
      Generic_Group, Billed_NDC, Billed Drug Name, Billed Pkg Size,
      Total Qty Billed (NDC), Billed Unit Cost,
      Alt_NDC, Alt Drug Name, Alt Last Purchase Date, Alt Unit Cost, Alt Vendor
    """

    # ---------- 1) Prepare billed side ----------
    df_log = log_df.copy()

    need_cols = ['NDC #', 'Drug Name', 'Drug Pkg Size', 'Qty Filled']
    missing = [c for c in need_cols if c not in df_log.columns]
    if missing:
        raise KeyError(f"Custom log missing required cols for alt NDC: {missing}")

    df_log["NDC11_BILLED"] = df_log["NDC #"].apply(_alt_norm_ndc11)
    df_log = df_log[~df_log["NDC11_BILLED"].isna()].copy()
    df_log["NDC9"] = df_log["NDC11_BILLED"].apply(_alt_ndc11_to_ndc9)

    df_log["Qty Filled"] = pd.to_numeric(df_log["Qty Filled"], errors="coerce").fillna(0)
    df_log["Drug Pkg Size"] = pd.to_numeric(df_log["Drug Pkg Size"], errors="coerce").fillna(0)

    billed_ndc = (
        df_log.groupby(["NDC9", "NDC11_BILLED"], as_index=False)
              .agg(
                  Total_Qty_Billed=("Qty Filled", "sum"),
                  Billed_Drug_Name=("Drug Name", "first"),
                  Billed_Pkg_Size=("Drug Pkg Size", "first"),
              )
    )

    if billed_ndc.empty:
        return pd.DataFrame(columns=[
            "Generic_Group",
            "Billed_NDC", "Billed Drug Name", "Billed Pkg Size",
            "Total Qty Billed (NDC)",
            "Billed Unit Cost",
            "Alt_NDC", "Alt Drug Name", "Alt Last Purchase Date",
            "Alt Unit Cost", "Alt Vendor",
        ])

    # ---------- 2) Prepare vendor side ----------
    df_v = all_vendor_df.copy()
    # required for pricing; Drug Name optional
    need_v_cols = ["NDC #", "Shipped", "PRICE", "DATE", "Vendor"]
    missing_v = [c for c in need_v_cols if c not in df_v.columns]
    if missing_v:
        raise KeyError(f"all_vendor_df missing required cols: {missing_v}")

    if "Drug Name" not in df_v.columns:
        df_v["Drug Name"] = ""  # optional, but we want Alt Drug Name

    df_v["NDC11"] = df_v["NDC #"].apply(_alt_norm_ndc11)
    df_v = df_v[~df_v["NDC11"].isna()].copy()
    df_v["NDC9"] = df_v["NDC11"].apply(_alt_ndc11_to_ndc9)

    df_v["Shipped"] = pd.to_numeric(df_v["Shipped"], errors="coerce").fillna(0)
    df_v["PRICE"] = pd.to_numeric(df_v["PRICE"], errors="coerce").fillna(0)
    df_v["DATE"] = pd.to_datetime(df_v["DATE"], errors="coerce")

    vend_ndc = (
        df_v.sort_values(["NDC9", "NDC11", "DATE"])
            .groupby(["NDC9", "NDC11", "Vendor"], as_index=False)
            .agg(
                Alt_Last_Purchase_Date=("DATE", "last"),
                Total_Shipped=("Shipped", "sum"),
                Total_Invoice=("PRICE", "sum"),
                Alt_Drug_Name=("Drug Name", "last"),
            )
    )

    vend_ndc["Alt_Unit_Cost"] = np.where(
        vend_ndc["Total_Shipped"] > 0,
        vend_ndc["Total_Invoice"] / vend_ndc["Total_Shipped"],
        np.nan,
    )

    # ---------- 3) Billed unit cost ----------
    billed_cost = vend_ndc[["NDC9", "NDC11", "Vendor", "Alt_Unit_Cost"]].rename(
        columns={"NDC11": "NDC11_BILLED", "Alt_Unit_Cost": "Billed_Unit_Cost"}
    )

    # ---------- 4) Merge billed side with alts ----------
    vend_ndc_alt = vend_ndc.rename(columns={"NDC11": "Alt_NDC"})

    merged = billed_ndc.merge(
        vend_ndc_alt,
        on="NDC9",
        how="left",
    )

    merged = merged[merged["Alt_NDC"] != merged["NDC11_BILLED"]].copy()
    if merged.empty:
        return pd.DataFrame(columns=[
            "Generic_Group",
            "Billed_NDC", "Billed Drug Name", "Billed Pkg Size",
            "Total Qty Billed (NDC)",
            "Billed Unit Cost",
            "Alt_NDC", "Alt Drug Name", "Alt Last Purchase Date",
            "Alt Unit Cost", "Alt Vendor",
        ])

    # billed unit cost collapsed per NDC9 + billed NDC (ignore vendor differences)
    billed_cost_collapsed = (
        billed_cost.groupby(["NDC9", "NDC11_BILLED"], as_index=False)
                   .agg(Billed_Unit_Cost=("Billed_Unit_Cost", "mean"))
    )

    merged = merged.merge(
        billed_cost_collapsed,
        on=["NDC9", "NDC11_BILLED"],
        how="left",
    )

    # ---------- 5) Final mapping (no diff/savings now) ----------
    alt_df = merged.assign(
        Generic_Group=lambda df: df["NDC9"],
        Billed_NDC=lambda df: df["NDC11_BILLED"],
    )

    alt_df = alt_df.rename(columns={
        "Billed_Drug_Name": "Billed Drug Name",
        "Billed_Pkg_Size": "Billed Pkg Size",
        "Alt_Unit_Cost": "Alt Unit Cost",
        "Billed_Unit_Cost": "Billed Unit Cost",
    })

    alt_df = alt_df[
        [
            "Generic_Group",
            "Billed_NDC", "Billed Drug Name", "Billed Pkg Size",
            "Total_Qty_Billed",  # will rename just below
            "Billed Unit Cost",
            "Alt_NDC", "Alt_Drug_Name", "Alt_Last_Purchase_Date",
            "Alt Unit Cost", "Vendor",
        ]
    ]

    alt_df = alt_df.rename(columns={
        "Total_Qty_Billed": "Total Qty Billed (NDC)",
        "Alt_Drug_Name": "Alt Drug Name",
        "Alt_Last_Purchase_Date": "Alt Last Purchase Date",
        "Vendor": "Alt Vendor",
    })

    # If you want some ordering on best alts even w/o explicit savings,
    # we can sort by Alt Unit Cost ascending for each generic group + billed NDC.
    alt_df = alt_df.sort_values(
        ["Generic_Group", "Billed_NDC", "Alt Unit Cost"],
        ascending=[True, True, True],
    ).reset_index(drop=True)

    return alt_df

'''
def add_alternate_ndc_sheet(
    wb,
    log_df: pd.DataFrame,
    all_vendor_df: pd.DataFrame,
    sheet_name: str = "Alternate NDC - Purchased",
):
    """
    Compute alternate NDC opportunities and write them into a new sheet in `wb`.

    Now includes Alt Drug Name and no longer shows diff/savings.
    """

    alt_df = build_alternate_ndc_df(log_df, all_vendor_df)
    print(f"[ALT NDC] rows={len(alt_df)}")

    # Create / replace sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(title=sheet_name)

    if alt_df.empty:
        ws["A1"] = "No alternate NDC purchase opportunities found for this period."
        ws["A1"].font = Font(size=14, bold=True)
        return

    # Title row
    num_cols = alt_df.shape[1]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Alternate NDC Purchase Opportunities"
    title_cell.font = Font(size=18, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    header_row_idx = 2
    ws.row_dimensions[header_row_idx].height = 30
    header_font = Font(size=12, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(alt_df.columns, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = header_alignment

    # Column widths
    widths = {
        "Generic_Group": 14,
        "Billed_NDC": 16,
        "Billed Drug Name": 40,
        "Billed Pkg Size": 14,
        "Total Qty Billed (NDC)": 18,
        "Billed Unit Cost": 16,
        "Alt_NDC": 16,
        "Alt Drug Name": 40,
        "Alt Last Purchase Date": 18,
        "Alt Unit Cost": 16,
        "Alt Vendor": 18,
    }
    for col_idx, col_name in enumerate(alt_df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(col_name, 14)

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    data_start_row = 3
    for row_offset, (_, row) in enumerate(alt_df.iterrows(), start=0):
        excel_row = data_start_row + row_offset
        for col_idx, col_name in enumerate(alt_df.columns, start=1):
            value = row[col_name]
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border

            # Date formatting for Alt Last Purchase Date
            if col_name == "Alt Last Purchase Date" and pd.notna(value):
                cell.number_format = "yyyy-mm-dd"   # no timestamp
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(value, (int, float, np.integer, np.floating)):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A3"

    last_row = data_start_row + len(alt_df) - 1
    last_col_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A2:{last_col_letter}{last_row}"

    # No more diff/savings CF for now; keep it simple
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
'''

def process_custom_log_data(custom_log_path, bin_master_path, vendor_paths, pharmacy_name, date_range, all_pbm_path,
                            selected_processors=None, selected_sheets=None, vendor_count=None,
                            job_dir=None, user_audit_dir=None):
    import pandas as pd
    import os
    import re
    import numpy as np
    # ===== Load =====
    bin_df = pd.read_csv(bin_master_path, dtype=str)
    log_df = pd.read_csv(custom_log_path, dtype=str)

    # Normalize incoming headers
    log_df.columns = [str(c).strip() for c in log_df.columns]

    # Keep only insurance-adjudicated rows early (ignore cash/other states)
    log_df, _status_col, _kept_rows, _dropped_rows = _filter_custom_log_transmitted_paid_ins(log_df)

    # Normalize "Drug NDC" -> "NDC #", if needed
    for c in list(log_df.columns):
        if c.strip().lower() in ('drug ndc', 'drug ndc#', 'drug ndc #'):
            log_df.rename(columns={c: 'NDC #'}, inplace=True)
            break

    # Normalize SDRA naming to one canonical column
    if '* SDRA Amt' in log_df.columns and 'SDRA Amt' not in log_df.columns:
        log_df.rename(columns={'* SDRA Amt': 'SDRA Amt'}, inplace=True)

    # Normalize Copay naming/case to one canonical column
    if 'Copay' in log_df.columns and 'COPAY' not in log_df.columns:
        log_df.rename(columns={'Copay': 'COPAY'}, inplace=True)

    # ===== Validate required columns =====
    need_log = [
        'Rx #', 'NDC #', 'Drug Name',
        'Plan 1 BIN', 'Plan 1 PCN', 'Plan 1 Group #',
        'Plan 2 BIN', 'Plan 2 PCN', 'Plan 2 Group #',
        'Ins Paid Plan 1', 'Ins Paid Plan 2',
        'Qty Filled', 'Drug Pkg Size',
        'SDRA Amt', 'COPAY'
    ]
    missing = [c for c in need_log if c not in log_df.columns]
    if missing:
        raise ValueError(
            f"Custom Log missing required column(s): {', '.join(missing)}")

    # User requirement: treat the custom log as a 20-column mandatory input
    if len(log_df.columns) < 20:
        raise ValueError(
            f"Custom Log must contain at least 20 columns. Found {len(log_df.columns)} columns.")

    if 'BIN' not in bin_df.columns or 'Processor' not in bin_df.columns:
        raise ValueError(
            "BIN master must contain 'BIN' and 'Processor' columns.")

    # ===== Normalize =====
    # BIN master
    bin_df['BIN'] = (bin_df['BIN'].astype(str)
                     .str.replace(r'\D', '', regex=True)
                     .str.zfill(6))

    bin_df['Processor'] = bin_df['Processor'].astype(str).str.strip()
    # >>> Build the BIN -> Processor map ONCE (used below for row-level filter)
    bin_to_proc = dict(zip(bin_df['BIN'], bin_df['Processor']))

    # Custom log
    for col in ['Plan 1 BIN', 'Plan 2 BIN']:
        log_df[col] = (log_df[col].astype(str)
                                  .str.replace(r'\D', '', regex=True)
                                  .str.zfill(6))

    for col in ['Ins Paid Plan 1', 'Ins Paid Plan 2', 'Qty Filled', 'Drug Pkg Size', 'SDRA Amt', 'COPAY']:
        log_df[col] = pd.to_numeric(log_df[col], errors='coerce').fillna(0)

    log_df['NDC #'] = (log_df['NDC #'].astype(str)
                       .str.replace('-', '', regex=False)
                       .str.strip()
                       .str.zfill(11))

    log_df['Drug Name'] = (
        log_df['Drug Name']
        .astype(str)
        .str.strip()
        .str.replace(r'\*+$', '', regex=True)  # remove only trailing ****
    )

    # ===== Choose winning BIN per row =====
    log_df['Winning_BIN'] = log_df.apply(
        lambda r: r['Plan 1 BIN'] if r['Ins Paid Plan 1'] >= r['Ins Paid Plan 2'] else r['Plan 2 BIN'],
        axis=1
    ).str.zfill(6)

    # capture the winning insurance dollars only (this becomes Processor_T later)
    log_df['Winning_Paid'] = np.where(
        log_df['Winning_BIN'] == log_df['Plan 1 BIN'],
        log_df['Ins Paid Plan 1'],
        log_df['Ins Paid Plan 2']
    )
    log_df['Winning PCN'] = np.where(
        log_df['Winning_BIN'] == log_df['Plan 1 BIN'],
        log_df['Plan 1 PCN'],
        log_df['Plan 2 PCN']
    )
    log_df['Winning Group'] = np.where(
        log_df['Winning_BIN'] == log_df['Plan 1 BIN'],
        log_df['Plan 1 Group #'],
        log_df['Plan 2 Group #']
    )
    # >>> NEW: determine row-level Processor now and FILTER if user selected any
    log_df['Processor'] = log_df['Winning_BIN'].map(bin_to_proc)
    rx_compare_source = log_df.copy()
    if selected_processors:
        # Normalize case and whitespace just to be safe
        allowed = {p.strip().casefold() for p in selected_processors}
        log_df = log_df[
            log_df['Processor'].fillna("").astype(
                str).str.strip().str.casefold().isin(allowed)
        ].copy()

    # ===== 1) Aggregate BY (NDC #, Winning_BIN) FIRST =====
    agg_bin = (log_df.groupby(['NDC #', 'Winning_BIN'], as_index=False)
               .agg({'Qty Filled': 'sum',
                     'Winning_Paid': 'sum',
                     'Drug Name': 'first',
                     'Drug Pkg Size': 'first'}))

    # ===== 2) Map BIN -> Processor AFTER aggregation =====
    bin_to_proc = dict(zip(bin_df['BIN'], bin_df['Processor']))
    agg_bin['Processor'] = agg_bin['Winning_BIN'].map(bin_to_proc)
    agg_bin = agg_bin[agg_bin['Processor'].notna()].copy()

    # >>> NEW: collapse to (NDC, Processor) both Qty and Paid
    grp_proc = (agg_bin.groupby(['NDC #', 'Processor'], as_index=False).agg(
        {'Qty Filled': 'sum', 'Winning_Paid': 'sum'}))

    # unique_procs = grp_proc['Processor'].dropna().astype(str).str.strip().unique().tolist()

    # Identify rows whose Winning_BIN is missing in the master map
    unmapped_mask = ~log_df['Winning_BIN'].isin(bin_to_proc.keys())
    unmapped = log_df.loc[unmapped_mask].copy()

    # If your custom log has 'Rx #' column, collect per BIN; else leave empty
    if 'Rx #' in unmapped.columns:
        # make RXs concise: comma-separated unique RX # per BIN
        rx_by_unmapped_bin = (
            unmapped.groupby('Winning_BIN')['Rx #']
                    .apply(lambda s: ', '.join(sorted(set(str(x).strip() for x in s if pd.notna(x) and str(x).strip()))))
                    .to_dict()
        )
    else:
        rx_by_unmapped_bin = {}  # no RX info available

    used_bins = (log_df['Winning_BIN'].astype(str)
                 .str.replace(r'\D', '', regex=True)
                 .str.zfill(6)
                 .dropna()
                 .unique()
                 )

    # ===== Build package size & name per NDC from the FULL custom log =====
    pkg_df = (
        log_df[['NDC #', 'Drug Pkg Size', 'Drug Name']]
        .rename(columns={'Drug Pkg Size': 'Package Size'})
        .copy()
    )

    pkg_df = (log_df[['NDC #', 'Drug Pkg Size', 'Drug Name']]
              .rename(columns={'Drug Pkg Size': 'Package Size'}))
    pkg_df['Package Size'] = pd.to_numeric(
        pkg_df['Package Size'], errors='coerce')
    pkg_df = (pkg_df
              .dropna(subset=['Package Size'])
              .drop_duplicates(subset=['NDC #']))

    # ===== 3) Pivots to Processor_Q and Processor_T =====
    # We already built grp_proc above with both Qty Filled and Winning_Paid.
    # >>> CHANGED: do NOT overwrite grp_proc again; keep both measures in it.
    grp_q = grp_proc.pivot(index='NDC #', columns='Processor',
                           values='Qty Filled').fillna(0).reset_index()
    pivot_q = grp_q.copy()
    pivot_q.columns = ['NDC #'] + \
        [f'{c}_Q' for c in pivot_q.columns if c != 'NDC #']

    # >>> NEW: Insurance dollars per processor → *_T
    grp_t = grp_proc.pivot(index='NDC #', columns='Processor',
                           values='Winning_Paid').fillna(0).reset_index()
    pivot_t = grp_t.copy()
    pivot_t.columns = ['NDC #'] + \
        [f'{c}_T' for c in pivot_t.columns if c != 'NDC #']

    # ===== 4) Process vendor files =====
    all_vendor_rows = []   # for Alternate NDC logic

    vendor_frames_qty, vendor_frames_price, vendor_names = [], [], []
    kinray_rows = []   # stash rows (NDC, DATE, UnitPrice) for Kinray only

    def _norm_headers(cols):
        # collapse whitespace and convert NBSP to space
        return [re.sub(r'\s+', ' ', str(c).replace('\xa0', ' ')).strip() for c in cols]

    def _pick(lower_to_orig, cands):
        # exact case-insensitive first
        for c in cands:
            if c in lower_to_orig:
                return lower_to_orig[c]
        # partial contains fallback
        for c in cands:
            for low, orig in lower_to_orig.items():
                if c in low:
                    return orig
        return None

    for i, vp in enumerate(vendor_paths, start=1):
        raw = pd.read_csv(vp, dtype=str)

        # Normalize headers
        raw.columns = _norm_headers(raw.columns)
        lower_to_orig = {c.lower(): c for c in raw.columns}

        # Required columns (robust to vendor variations)
        ndc_col = _pick(lower_to_orig, ['ndc/upc', 'ndc #', 'ndc', 'ndc#', 'ndc number',
                        'ndc no', 'ndc upc', 'ndcupc', 'NDC/UPC', 'Item NDC/UPC (Current)'])
        ship_col = _pick(lower_to_orig, ['ship qty', 'shipped', 'shipped qty', 'qty shipped',
                         'quantity shipped', 'ship quantity', 'qty', 'Ship Qty', 'Purchase History Ordered Quantity'])
        price_col = _pick(lower_to_orig, [
                          'invoice $', 'invoice$', 'invoice amount', 'invoice', 'price', 'Invoice $'])

        if not ndc_col or not ship_col or not price_col:
            raise ValueError(
                f"Vendor file '{os.path.basename(vp)}' must contain NDC/UPC, Ship Qty, and Invoice $ (or equivalents). "
                f"Found: {list(raw.columns)}"
            )

        # Optional date column (for Kinray "latest" logic)
        date_col = _pick(lower_to_orig, [
                         'invoice date', 'ship date', 'shipping date', 'order date', 'date'])
        ourcase_col = _pick(lower_to_orig, ['ourcase', 'our case', 'case',
                                            'invoice #', 'invoice number', 'inv #',
                                            'document number', 'doc #', 'doc no'])
        # # 🔹 NEW: optional drug-name column
        # drug_col = _pick(lower_to_orig, [
        #     'drug name', 'item description', 'description',
        #     'product name', 'item name','Description'
        # ])


        keep_cols = [ndc_col, ship_col, price_col] \
            + ([date_col] if date_col else []) \
            + ([ourcase_col] if ourcase_col else [])

        v = raw[keep_cols].copy()

        rename_map = {ndc_col: 'NDC #',
                      ship_col: 'Shipped', price_col: 'PRICE'}
        if date_col:
            rename_map[date_col] = 'DATE'
        if ourcase_col:
            rename_map[ourcase_col] = 'OURCASE'
        v.rename(columns=rename_map, inplace=True)
        # if drug_col:
        #     rename_map[drug_col] = 'Drug Name'   # 🔹 NEW
        # DATE
        if 'DATE' in v.columns:
            v['DATE'] = pd.to_datetime(v['DATE'], errors='coerce')
        else:
            v['DATE'] = pd.NaT

        # OURCASE: keep as string for stable sorting
        if 'OURCASE' in v.columns:
            v['OURCASE'] = v['OURCASE'].astype(str).str.strip()
        else:
            v['OURCASE'] = ''

        # Normalize NDC -> 11 digits
        v['NDC #'] = (v['NDC #'].astype(str)
                                .str.replace(r'\D', '', regex=True)
                                .str.strip()
                                .str.zfill(11))

        # Normalize Shipped -> numeric
        v['Shipped'] = (v['Shipped'].astype(str)
                        .str.replace(',', '', regex=False)
                        .str.replace('(', '-', regex=False)
                        .str.replace(')', '', regex=False)
                        .str.replace(r'[^0-9.\-]', '', regex=True)
                        .str.strip())
        v['Shipped'] = pd.to_numeric(v['Shipped'], errors='coerce').fillna(0)

        # Normalize PRICE -> numeric
        v['PRICE'] = (v['PRICE'].astype(str)
                                .str.replace(',', '', regex=False)
                                .str.replace('$', '', regex=False)
                                .str.replace('(', '-', regex=False)
                                .str.replace(')', '', regex=False)
                                .str.replace(r'[^0-9.\-]', '', regex=True)
                                .str.strip())
        v['PRICE'] = pd.to_numeric(v['PRICE'], errors='coerce').fillna(0)

        # Parse date (if present)
        if 'DATE' in v.columns:
            v['DATE'] = pd.to_datetime(v['DATE'], errors='coerce')
        else:
            v['DATE'] = pd.NaT

        # Vendor label (keep your convention)
        # ---------- ADD THIS BLOCK ----------
        # Derive a nice label from the file name (e.g. 'Kinray.xlsx' -> 'Kinray')
        vendor_label = os.path.splitext(os.path.basename(vp))[0]
        # If you want to force uppercase like 'MCK', you can do:
        # vendor_label = vendor_label.upper()

        v['Vendor'] = vendor_label
        # name = f'Vendor{i}'
        # v['Vendor'] = name
        # vendor_names.append(name)
        # Row-level unit price (only where Shipped > 0)
        v['__UnitPrice__'] = np.where(
            v['Shipped'] > 0,
            np.round(v['PRICE'] / v['Shipped'], 2),  # keep as float first
            np.nan
        )

        # remove infinities (e.g., division by 0)
        v['__UnitPrice__'] = np.where(np.isfinite(
            v['__UnitPrice__']), v['__UnitPrice__'], np.nan)

        # if you want whole numbers only, round and cast safely
        v['__UnitPrice__'] = v['__UnitPrice__'].round(0)  # no decimals
        v['__UnitPrice__'] = v['__UnitPrice__'].astype(
            'Int64')  # keeps NaN as <NA>

        is_kinray = 'kinray' in os.path.basename(vp).lower()
        if is_kinray:
            # keep all selectors we need to decide "latest"
            kinray_rows.append(
                v[['NDC #', 'DATE', 'OURCASE', 'Shipped',
                    'PRICE', '__UnitPrice__']].copy()
            )

        # Collect for aggregation/pivots
        vendor_frames_qty.append(v[['NDC #', 'Vendor', 'Shipped']])
        vendor_frames_price.append(v[['NDC #', 'Vendor', 'PRICE']])
        # --------- add this to feed Alternate NDC logic ----------
        cols_for_alt = ['NDC #', 'Shipped', 'PRICE', 'DATE', 'Vendor']
        if 'Drug Name' in v.columns:
            cols_for_alt.append('Drug Name')
        else:
            v['Drug Name'] = ""   # ensure column exists
            cols_for_alt.append('Drug Name')

        all_vendor_rows.append(v[cols_for_alt].copy())

    if all_vendor_rows:
        all_vendor_df = pd.concat(all_vendor_rows, ignore_index=True)
    else:
        all_vendor_df = pd.DataFrame(
            columns=['NDC #', 'Shipped', 'PRICE', 'DATE', 'Vendor'])
        
    #print("\n[DEBUG Vendor Preview]")
    #print(all_vendor_df[['NDC #', 'Drug Name']].head(), "\n")

    # ===== Qty pivot (sum of shipped by NDC×Vendor)
    vendor_combined = (pd.concat(vendor_frames_qty, ignore_index=True)
                       if vendor_frames_qty else pd.DataFrame(columns=['NDC #', 'Vendor', 'Shipped']))
    vendor_agg = vendor_combined.groupby(
        ['NDC #', 'Vendor'], as_index=False)['Shipped'].sum()
    vendor_pivot = (vendor_agg.pivot(index='NDC #', columns='Vendor', values='Shipped')
                    .fillna(0)
                    .reset_index())
    vendor_names = [c for c in vendor_pivot.columns if c != 'NDC #']
    # Capitalize vendor names
    vendor_rename_map = {vn: vn.upper() for vn in vendor_names}
    # Apply rename to pivots
    vendor_pivot = vendor_pivot.rename(columns=vendor_rename_map)
    # Update vendor_names list
    vendor_names = [vn.upper() for vn in vendor_names]

    # ===== Price pivot (last seen price by NDC×Vendor)
    vendor_price_combined = (pd.concat(vendor_frames_price, ignore_index=True)
                             if vendor_frames_price else pd.DataFrame(columns=['NDC #', 'Vendor', 'PRICE']))
    vendor_price_agg = (vendor_price_combined
                        .groupby(['NDC #', 'Vendor'], as_index=False)['PRICE']
                        .last())
    vendor_price_pivot = (vendor_price_agg.pivot(index='NDC #', columns='Vendor', values='PRICE')
                          .fillna(0)
                          .reset_index())
    vendor_price_pivot = vendor_price_pivot.rename(columns={c: (c.upper() + "_PRICE")
                                                            if c not in ("NDC #") else c for c in vendor_price_pivot.columns})
    vendor_price_pivot.columns = [
        'NDC #'] + [f'{c}_PRICE' for c in vendor_price_pivot.columns if c != 'NDC #']

    if kinray_rows:
        kinray_all = pd.concat(kinray_rows, ignore_index=True)

        # Treat missing DATE as very old, so real dates win
        min_ts = pd.Timestamp(1970, 1, 1)
        kinray_all['__DATE__'] = kinray_all['DATE'].fillna(min_ts)

        # If OURCASE is purely numeric in many files, try to rank; otherwise string compare works
        # Sort by NDC, then DATE asc, then OURCASE asc; keep last = latest
        kinray_latest = (
            kinray_all
            # must have calculable unit price
            .dropna(subset=['__UnitPrice__'])
            .sort_values(['NDC #', '__DATE__', 'OURCASE'])
            .drop_duplicates(subset=['NDC #'], keep='last')  # latest per NDC
            .loc[:, ['NDC #', '__UnitPrice__']]
            .rename(columns={'__UnitPrice__': 'Kinray_UPrice'})
        )
    else:
        kinray_all = pd.DataFrame(columns=['NDC #', 'DATE', 'OURCASE', 'Shipped', 'PRICE', '__UnitPrice__'])
        kinray_latest = pd.DataFrame(columns=['NDC #', 'Kinray_UPrice'])

    kinray_price_map = dict(
        zip(kinray_latest['NDC #'], kinray_latest['Kinray_UPrice']))

    # ===== END VENDOR AGGREGATION =====

    # ===== Merge everything =====
    merged = (pivot_q
              .merge(pivot_t, on='NDC #', how='left')          # brings *_T
              # Vendor1, Vendor2, ...
              .merge(vendor_pivot, on='NDC #', how='left')
              # Package Size, Drug Name
              .merge(pkg_df, on='NDC #', how='left')
              .merge(vendor_price_pivot, on='NDC #', how='left')
              .merge(kinray_latest, on='NDC #', how='left'))   # Kinray_UPrice

    # Ensure vendor qty cols exist & numeric
    for vn in vendor_names:
        if vn not in merged.columns:
            merged[vn] = 0
    merged[vendor_names] = merged[vendor_names].apply(
        pd.to_numeric, errors='coerce').fillna(0)

    # Ensure price cols numeric too
    price_cols = [c for c in merged.columns if c.endswith('_PRICE')]
    if price_cols:
        merged[price_cols] = merged[price_cols].apply(
            pd.to_numeric, errors='coerce').fillna(0)

    # Total Purchased = sum of vendor shipped qty across all vendor columns
    merged['Total Purchased'] = merged[vendor_names].sum(axis=1)

    # Normalize left key and numerics we’ll use
    merged['NDC #'] = (merged['NDC #'].astype(str)
                       .str.replace(r'\D', '', regex=True)
                       .str.zfill(11))
    merged['Kinray_UPrice'] = pd.to_numeric(merged.get(
        'Kinray_UPrice', 0), errors='coerce').fillna(0)
    pkg = pd.to_numeric(merged.get('Package Size', 0),
                        errors='coerce').fillna(0)

    # ===== Bring in ALL PBM (Quantity & Total $) =====
    if all_pbm_path:
        # should return NDC #, ALL_PBM_Q, ALL_PBM_T, ALL_PBM_DrugName
        all_pbm = _load_all_pbm_csv(all_pbm_path)

        # Back-compat: if the file uses 'Total' instead of ALL_PBM_T
        if 'Total' in all_pbm.columns and 'ALL_PBM_T' not in all_pbm.columns:
            all_pbm = all_pbm.rename(columns={'Total': 'ALL_PBM_T'})

        # one row per NDC (defensive)
        all_pbm = all_pbm.drop_duplicates(subset=['NDC #'], keep='last')

        # only map onto rows you already have
        merged = merged.merge(
            all_pbm[['NDC #', 'ALL_PBM_Q', 'ALL_PBM_T', 'ALL_PBM_DrugName']],
            on='NDC #',
            how='left'
        )
    else:
        merged['ALL_PBM_Q'] = 0
        merged['ALL_PBM_T'] = 0
        merged['ALL_PBM_DrugName'] = pd.NA

    # Fill Drug Name from ALL PBM if missing
    if 'Drug Name' not in merged.columns:
        merged['Drug Name'] = pd.NA
    merged['Drug Name'] = merged['Drug Name'].fillna(
        merged.get('ALL_PBM_DrugName'))

    def ensure_numeric_col(df, col, default=0):
        if col not in df.columns:
            df.loc[:, col] = default
        df.loc[:, col] = pd.to_numeric(
            df[col], errors='coerce').fillna(default)

    # Ensure numeric + compute PBM derived columns
    ensure_numeric_col(merged, 'ALL_PBM_Q', 0)
    ensure_numeric_col(merged, 'ALL_PBM_T', 0)

    pkg = pd.to_numeric(merged.get('Package Size', 0),
                        errors='coerce').fillna(0)
    merged['ALL_PBM_P'] = (merged['ALL_PBM_Q'] / pkg).where(pkg > 0, 0)
    merged['ALL_PBM_D'] = merged['Total Purchased'] - merged['ALL_PBM_P']
    merged['ALL_PBM_Pur'] = merged['ALL_PBM_P'] * \
        pd.to_numeric(merged.get('Kinray_UPrice', 0),
                      errors='coerce').fillna(0)

    # ===== Discover processors from either *_Q or *_T (robust union) =====
    procs_from_q = {c[:-2] for c in merged.columns if c.endswith('_Q')}
    procs_from_t = {c[:-2] for c in merged.columns if c.endswith('_T')}
    processors = sorted(procs_from_q.union(procs_from_t))

    # Ensure missing *_Q and *_T exist (so later loops never KeyError)
    for pr in processors:
        qcol, tcol = f'{pr}_Q', f'{pr}_T'
        if qcol not in merged.columns:
            merged[qcol] = 0
        if tcol not in merged.columns:
            merged[tcol] = 0

    # Convert *_Q/*_T numeric
    for pr in processors:
        merged[f'{pr}_Q'] = pd.to_numeric(
            merged[f'{pr}_Q'], errors='coerce').fillna(0)
        merged[f'{pr}_T'] = pd.to_numeric(
            merged[f'{pr}_T'], errors='coerce').fillna(0)

    # ===== Build per-processor derived bands for *every* processor
    for pr in processors:
        q = merged[f'{pr}_Q']
        p = (q / pkg).where(pkg > 0, 0)                 # packages billed
        merged[f'{pr}_P'] = p
        merged[f'{pr}_D'] = merged['Total Purchased'] - p
        merged[f'{pr}_Pur'] = p * merged['Kinray_UPrice']
        merged[f'{pr}_Net'] = merged[f'{pr}_T'] - merged[f'{pr}_Pur']

    # ===== Final column ordering (your “desired columns” spec) =====

    def have(cols): return [c for c in cols if c in merged.columns]

    # Discover processors FROM CURRENT COLUMNS (better than from desired_columns)
    def _discover_processors_from_columns(cols):
        procs_q = {c[:-2] for c in cols if c.endswith('_Q')}
        procs_t = {c[:-2] for c in cols if c.endswith('_T')}
        procs_p = {c[:-2] for c in cols if c.endswith('_P')}
        procs_d = {c[:-2] for c in cols if c.endswith('_D')}
        procs_pur = {c[:-4] for c in cols if c.endswith('_Pur')}
        procs_net = {c[:-4] for c in cols if c.endswith('_Net')}
        return sorted(procs_q | procs_t | procs_p | procs_d | procs_pur | procs_net)

    # ✅ define processors BEFORE using them in bands
    processors = _discover_processors_from_columns(merged.columns)
    # 🔥 Force ALL_PBM first, others alphabetical
    if 'ALL_PBM' in processors:
        processors = ['ALL_PBM'] + \
            sorted([p for p in processors if p != 'ALL_PBM'])
    else:
        processors = sorted(processors)

    base_cols = ['NDC #', 'Drug Name', 'Package Size']
    vendor_qty_cols = vendor_names
    qty_band = have([f'{pr}_Q' for pr in processors])
    pkg_band = have([f'{pr}_P' for pr in processors])
    diff_band = have([f'{pr}_D' for pr in processors])
    paid_band = have([f'{pr}_T' for pr in processors])  # dollars paid
    # $$ purchased (Kinray)
    pur_band = have([f'{pr}_Pur' for pr in processors])
    net_band = have([f'{pr}_Net' for pr in processors])  # paid − purchased
    other_cols = have(['Total Purchased', 'Kinray_UPrice'])

    # ✅ Round off paid_band, pur_band, and net_band columns (no decimals)
    for band in [paid_band, pur_band, net_band]:
        for col in band:
            if col in merged.columns:
                merged[col] = np.round(merged[col]).astype('Int64')

    desired_columns = (
        base_cols +
        vendor_qty_cols + ['Total Purchased'] +
        [c for c in other_cols if c not in ('Total Purchased',)] +
        qty_band + pkg_band + diff_band +
        paid_band + pur_band + net_band
    )

    def _proc_from_col(col: str):
        suffixes = ('_Q', '_P', '_D', '_T', '_Pur', '_Net')
        for sfx in suffixes:
            if col.endswith(sfx):
                return col[:-len(sfx)]
        return None  # not a processor metric column

    # ✅ NEW: remove deselected processors’ columns dynamically (ALWAYS include ALL_PBM)
    if selected_processors:
        # normalize to a set for membership and add forced include
        selected_upper = {p.strip().upper() for p in selected_processors}
        selected_upper.add('ALL_PBM')

        keep_cols = []
        for col in desired_columns:
            proc = _proc_from_col(col)
            if proc is None:
                # base/vendor/other columns (not *_Q/_P/_D/_T/_Pur/_Net)
                keep_cols.append(col)
            else:
                if proc.strip().upper() in selected_upper:
                    keep_cols.append(col)

        desired_columns = keep_cols
        merged = merged.reindex(
            columns=[c for c in desired_columns if c in merged.columns])

    # Keep only existing (defensive) and sort by Drug Name
    desired_columns = [c for c in desired_columns if c in merged.columns]
    final = merged[desired_columns].sort_values(
        'Drug Name', na_position='last')

    # Use the SAME extractor to compute processors from the final df
    def processors_from_df(df):
        procs = set()
        for c in df.columns:
            p = _proc_from_col(c)
            if p:
                procs.add(p)
        return sorted(procs)

    processors = processors_from_df(final)
    # Ensure ALL_PBM is first if present
    if 'ALL_PBM' in processors:
        processors = ['ALL_PBM'] + \
            sorted([p for p in processors if p != 'ALL_PBM'])
    else:
        processors = sorted(processors)

    # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    # NEW: Split out rows that have NaN anywhere in the final report
    nan_mask = final.isna().any(axis=1)

    # build a small reasons frame (which columns were NaN)
    def _nan_cols(row):
        return ", ".join([col for col, is_nan in row.isna().items() if is_nan])

    nan_reason_df = final.loc[nan_mask].apply(_nan_cols, axis=1)
    final_with_reason = final.copy()
    final_with_reason.loc[nan_mask, 'Reason_NaN_Columns'] = nan_reason_df

    # main sheet (clean): drop NaN rows
    final_clean = final_with_reason.loc[~nan_mask].copy()

    # important sheet: original Custom Log rows for those NDCs that had NaN in final
    nan_ndcs = set(final.loc[nan_mask, 'NDC #']
                   ) if 'NDC #' in final.columns else set()
    important_rows = log_df[log_df['NDC #'].isin(nan_ndcs)].copy()

    # annotate why (optional)
    if not important_rows.empty:
        # map NDC -> reason string (from merged/final)
        ndc_to_reason = final_with_reason.loc[nan_mask, [
            'NDC #', 'Reason_NaN_Columns']].drop_duplicates()
        # merge to show reason next to original rows
        important_rows = important_rows.merge(
            ndc_to_reason, on='NDC #', how='left')
        # put an attention banner col
        important_rows.insert(
            0, '⚠️ Check', 'This NDC had NaN in merged report — investigate!')

    # Save into app's processed folder so /download can serve it
    safe_name = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_',
                       f'{pharmacy_name} ({date_range}).xlsx')
    output_dir = os.path.join(current_app.root_path, current_app.config.get(
        'PROCESSED_FOLDER', 'processed'))
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, safe_name)
    final.to_excel(output_file,  index=False, float_format="%.3f")

    #print(f"Processed file saved at: {output_file}")  # Debugging line

    # written_data = pd.read_excel(output_file)
    # written_data = final_df.copy()
    if not os.path.exists(output_file):
        raise FileNotFoundError(f"Processed file not found at {output_file}")

    wb = load_workbook(output_file)
    ws = wb.active
    header_row = 3

    # Merge the first row and set the pharmacy name and date range in the center
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(desired_columns))
    cell = ws.cell(row=1, column=1)
    cell.value = f"{pharmacy_name} ({date_range})"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=35, bold=True)
    ws.row_dimensions[1].height = 60

    # Move the data down by one row
    ws.insert_rows(2)
    # Move the data down by one row
    ws.insert_rows(3)

    # Explicitly set the headers in the second row
    for col_num, header in enumerate(desired_columns, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=False)
        cell.font = Font(bold=False, size=15)

    # Dynamically calculate the start and end columns for each merged cell

    def get_column_index(ws, header_name, header_row=None):
        """Find a column by header text. Returns 1-based index or None."""
        rows = [header_row] if header_row else [2, 3]
        for r in rows:
            try:
                for cell in ws[r]:
                    if cell.value == header_name:
                        return cell.col_idx
            except Exception:
                continue
        return None

    for col_name in base_cols:
        idx = get_column_index(ws, col_name)
        if not idx:
            continue

        header_cell = ws.cell(row=header_row, column=idx)

        if col_name == 'Package Size':
            # Header rotated 90°, centered
            header_cell.alignment = Alignment(
                horizontal='center',
                vertical='bottom',
                text_rotation=90,
                wrap_text=False
            )
        else:
            # Normal center/center for NDC # and Drug Name
            header_cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=False
            )

    if "Total Purchased" in desired_columns:
        total_purchased_col = desired_columns.index("Total Purchased") + 1
        total_purchased_indices = [total_purchased_col]
    else:
        raise ValueError("'Total Purchased' not found in desired_columns")

    # >>> NEW: group headers for Insurance $ Paid, $$ Purchased (Kinray), Net
    def _band_bounds_from_suffix(sfx):
        cols = [get_column_index(ws, f"{pr}_{sfx}") for pr in processors]
        cols = [c for c in cols if c]  # drop None
        if not cols:
            return None, None
        return min(cols), max(cols)

    def _ranges_intersect(a, b):
        a_min_col, a_min_row, a_max_col, a_max_row = range_boundaries(str(a))
        b_min_col, b_min_row, b_max_col, b_max_row = range_boundaries(str(b))
        return not (a_max_col < b_min_col or b_max_col < a_min_col or
                    a_max_row < b_min_row or b_max_row < a_min_row)

    def _merge_band(row, start_col, end_col, title):
        """
        Safely put `title` across start_col..end_col on `row`.
        - write value BEFORE merge (avoids MergedCell read-only)
        - unmerge any overlapping prior merges
        - style the anchor (top-left) cell
        """
        if not (start_col and end_col and start_col <= end_col):
            return

        # If it's a single column, just set value & style—no merge needed.
        if start_col == end_col:
            anchor = ws.cell(row=row, column=start_col)
            anchor.value = title
            anchor.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            anchor.font = Font(bold=True)
            return

        target_ref = f"{ws.cell(row=row, column=start_col).coordinate}:{ws.cell(row=row, column=end_col).coordinate}"

        # Unmerge any existing merged ranges that overlap our target span
        for mr in list(ws.merged_cells.ranges):
            if _ranges_intersect(mr, target_ref):
                ws.unmerge_cells(str(mr))

        # Write BEFORE merging
        anchor = ws.cell(row=row, column=start_col)
        anchor.value = title

        # Merge & style the anchor
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=end_col)
        anchor.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        anchor.font = Font(bold=True)

    paid_s, paid_e = _band_bounds_from_suffix('T')
    pur_s,  pur_e = _band_bounds_from_suffix('Pur')
    net_s,  net_e = _band_bounds_from_suffix('Net')
    qt_s, qt_e = _band_bounds_from_suffix('Q')
    pk_b_s, pk_b_e = _band_bounds_from_suffix('P')
    pk_d_s, pk_d_e = _band_bounds_from_suffix('D')

    _merge_band(2, qt_s,  qt_e,  "Quantity Billed = BestRX")
    _merge_band(2, pk_b_s, pk_b_e,
                "Package Size Billed = Quantity Billed(BestRx) ÷ Package Size")
    _merge_band(2, pk_d_s, pk_d_e,
                "Package Size Difference = Total Packages Purchased(Vendors) − Package Size Billed(BestRx)")
    _merge_band(2, paid_s, paid_e, "Actual $ Paid by Insurance = BestRX")
    _merge_band(2, pur_s,  pur_e,
                "Actual $ Purchased (Kinray Unit Price × Packages Billed To Ins)")
    _merge_band(2, net_s,  net_e,
                "Net(Profit/Loss)$ = Actual $ Paid(BestRx) − Actual $ Purchased(Kinray)")

    # ✅ Enable wrap text for Drug Name column (column B)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(
                horizontal='left', vertical='top', wrap_text=False)

    # Set the desired column widths
    column_widths = {
        'A': 15,  # NDC
        'B': 45,  # Drug Name
    }

    # Setting up width for the other columns
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    # Set widths for dynamic columns
    for col_num in range(4, len(desired_columns) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 7

    # Set the height for the first row
    ws.row_dimensions[1].height = 35

    # Set header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D0CECE",
                              end_color="D0CECE", fill_type="solid")
    # Set border style
    thin_border = Border(left=Side(style='thin'), right=Side(
        style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(
        style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
    # Enable text wrapping for the second row
    for cell in ws[3]:
        if cell.col_idx > 3:
            cell.alignment = Alignment(
                text_rotation=90, horizontal='center', wrap_text=False)
            cell.font = Font(bold=False, size=14, name='Calibri')
            cell.fill = header_fill
            cell.border = thin_border
        else:
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=False)
            cell.fill = header_fill
            cell.border = thin_border

    # Text rotation for Package Size column header
    # text rotation is not happening for Package Size column

    pkg_size_col_idx = get_column_index(ws, 'Package Size')
    if pkg_size_col_idx:
        pkg_cell = ws.cell(row=3, column=pkg_size_col_idx)
        pkg_cell.alignment = Alignment(
            text_rotation=90,
            horizontal='center',
            vertical='center',
            wrap_text=False
        )
        pkg_cell.font = Font(bold=False, size=14, name='Calibri')

    ws.row_dimensions[3].height = 100
    # Freeze the first row
    # ws.freeze_panes = 'A4'
    # freeze panes till kinray unit price column
    ws.freeze_panes = get_column_letter(total_purchased_col + 2) + '4'

    red_fill = PatternFill(start_color="FFC7CE",
                           end_color="FFC7CE", fill_type="solid")   # soft red
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # soft green
    row_fill_soft = PatternFill(
        start_color="E8E6FF", end_color="E8E6FF", fill_type="solid")  # subtle lavender-gray

    data_first_row = header_row + 1
    data_last_row = ws.max_row

    # Map header text -> column index from the worksheet
    header_map = {ws.cell(row=header_row, column=c).value: c for c in range(
        1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value}

    # All *_D columns present in the sheet
    diff_cols = [h for h in header_map.keys() if isinstance(h, str)
                 and h.endswith('_D')]

    # Apply per-column red/green CF to *_D columns
    for h in diff_cols:
        col_idx = header_map[h]
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}{data_first_row}:{col_letter}{data_last_row}"

        # red < 0
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="lessThan", formula=[
                       "0"], stopIfTrue=False, fill=red_fill)
        )
        # green > 0
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=[
                       "0"], stopIfTrue=False, fill=green_fill)
        )

    # Whole-row soft highlight if ANY *_D in that row is negative
    if diff_cols:
        first_idx = min(header_map[h] for h in diff_cols)
        last_idx = max(header_map[h] for h in diff_cols)
        first_letter = get_column_letter(first_idx)
        last_letter = get_column_letter(last_idx)

        # IMPORTANT: the formula must be relative to the TOP row of the CF range
        row_range = f"A{data_first_row}:{get_column_letter(ws.max_column)}{data_last_row}"
        formula = f'COUNTIF(${first_letter}{data_first_row}:${last_letter}{data_first_row},"<0")>0'
        ws.conditional_formatting.add(
            row_range,
            FormulaRule(formula=[formula],
                        stopIfTrue=False, fill=row_fill_soft)
        )

    # AutoFilter over the full data region (row 3 headers)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{data_last_row}"

    # Center align all data
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set the first two columns to left alignment
    for row in ws.iter_rows(min_row=4):
        row[0].alignment = Alignment(horizontal='left',  vertical='center')
        row[1].alignment = Alignment(horizontal='left',  vertical='center')
        row[2].alignment = Alignment(horizontal='center', vertical='center')

    def apply_thick_border(ws, start_col, end_col, start_row, end_row):
        # Apply the thick border to the top row
        for col_num in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.border = Border(
                top=thick_border.top,
                left=cell.border.left,
                right=cell.border.right,
                bottom=cell.border.bottom
            )

        # Apply the thick border to the bottom row
        for col_num in range(start_col, end_col + 1):
            cell = ws.cell(row=end_row, column=col_num)
            cell.border = Border(
                bottom=thick_border.bottom,
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top
            )

        # Apply the thick border to the left column
        for row_num in range(start_row, end_row + 1):
            cell = ws.cell(row=row_num, column=start_col)
            cell.border = Border(
                left=thick_border.left,
                top=cell.border.top,
                right=cell.border.right,
                bottom=cell.border.bottom
            )

        # Apply the thick border to the right column
        for row_num in range(start_row, end_row + 1):
            cell = ws.cell(row=row_num, column=end_col)
            cell.border = Border(
                right=thick_border.right,
                top=cell.border.top,
                left=cell.border.left,
                bottom=cell.border.bottom
            )
    start_row = 1
    end_row = ws.max_row

    def style_sheet(ws):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    def get_column_indices(ws, column_names):
        indices = []
        for col in ws[3]:
            if col.value in column_names:
                indices.append(col.col_idx)
        return indices
    total_purchased_col = get_column_index(ws, 'Total Purchased')
    if total_purchased_col is None:
        raise ValueError("Header 'Total Purchased' not found in the worksheet")

    # Apply thick border for specific column ranges
    quantity_billed_indices = get_column_indices(
        ws, [f'{pr}_Q' for pr in processors])
    package_size_billed_indices = get_column_indices(
        ws, [f'{pr}_P' for pr in processors])
    package_size_difference_indices = get_column_indices(
        ws, [f'{pr}_D' for pr in processors])

    total_purchased_indices = [total_purchased_col]

    def apply_thick_border_to_groups(ws, column_groups, start_row, end_row):
        for group in column_groups:
            if group:
                start_col = group[0]
                end_col = group[-1]
                apply_thick_border(ws, start_col, end_col, start_row, end_row)

    thin_border = Border(left=Side(style='thin', color="A9A9A9"), right=Side(style='thin', color="A9A9A9"), top=Side(
        style='thin', color="A9A9A9"), bottom=Side(style='thin', color="A9A9A9"))

    for cell in ws[3]:
        cell.border = thin_border

    # Set up styles
    cell_fill_red = PatternFill(
        start_color="F88379", end_color="F88379", fill_type="solid")
    row_fill_blue = PatternFill(
        start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Grouping column indices
    paid_indices = get_column_indices(ws, [f'{pr}_T' for pr in processors])
    pur_indices = get_column_indices(ws, [f'{pr}_Pur' for pr in processors])
    net_indices = get_column_indices(ws, [f'{pr}_Net' for pr in processors])
    Kinray_UPrice_index = get_column_index(ws, 'Kinray_UPrice')
    all_pbmm = get_column_index(ws, 'ALL_PBM_Pur')
    all_pbmd = get_column_index(ws, 'ALL_PBM_D')
    all_pbmt = get_column_index(ws, 'ALL_PBM_T')
    all_pbmq = get_column_index(ws, 'ALL_PBM_Q')
    all_pbmn = get_column_index(ws, 'ALL_PBM_Net')
    all_pbmp = get_column_index(ws, 'ALL_PBM_P')
    raw_groups = [
        quantity_billed_indices,
        package_size_billed_indices,
        package_size_difference_indices,
        paid_indices,
        pur_indices,
        net_indices,
        total_purchased_indices,
        Kinray_UPrice_index,
        all_pbmm,
        all_pbmd,
        all_pbmt,
        all_pbmq,
        all_pbmn,
        all_pbmp
    ]

    # Normalize groups so each entry is a sequence (list). Some items like
    # Kinray_UPrice_index are single ints; the helper expects indexable groups
    # (group[0], group[-1]). Convert ints -> [int], None -> [].
    column_groups = []
    for g in raw_groups:
        if isinstance(g, (list, tuple)):
            column_groups.append(list(g))
        elif isinstance(g, int):
            column_groups.append([g])
        else:
            column_groups.append([])

    apply_thick_border_to_groups(ws, column_groups, start_row, end_row)
    apply_thick_border(ws, start_col=1, end_col=1,
                       start_row=start_row, end_row=end_row)
    apply_thick_border(ws, start_col=2, end_col=2,
                       start_row=start_row, end_row=end_row)
    apply_thick_border(ws, start_col=3, end_col=3,
                       start_row=start_row, end_row=end_row)
    apply_thick_border(ws, start_col=4, end_col=4,
                       start_row=start_row, end_row=end_row)

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0, right=0, top=0, bottom=0, header=0, footer=0)

    # Set the title of the active worksheet
    ws.title = "Processed Data"
    # ws.protection.sheet = True
    # ===== Create/replace "BIN to Processor" sheet =====
    title_sheet = "BIN to Processor"
    if title_sheet in wb.sheetnames:
        del wb[title_sheet]
    ws2 = wb.create_sheet(title_sheet)

    # Title
    ws2.insert_rows(1)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    tcell = ws2.cell(
        row=1, column=1, value="BIN Numbers Billed (from Custom Log)")
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal='center', vertical='center')

    def find_fill_date_column(df):
        """Return the first column name that looks like a Fill Date."""
        candidates = ['Fill Date']
        # exact match first
        for c in candidates:
            if c in df.columns:
                return c
        # loose match (case-insensitive contains "fill" & "date")
        for c in df.columns:
            cl = str(c).strip().lower()
            if "date" in cl and ("fill" in cl or "filled" in cl):
                return c
        return None

    # --- Build BIN → Processor counts from the UNFILTERED custom log ---
    src_df = rx_compare_source.copy()  # unfiltered copy created earlier
    # Use the UNFILTERED log for totals so processor filters don't shrink the counts

    # <- "rows" (count rows), "qty" (sum Qty Filled), or "unique_rx" (distinct Rx #)
    COUNT_MODE = "rows"

    def build_rx_counts(src_df, mode="rows"):
        # Normalize BIN; include NaN/blank → '000000'
        bins = (src_df['Winning_BIN']
                .astype('string')
                .fillna('')                       # keep empties
                .str.replace(r'\D', '', regex=True)
                .str.zfill(6))                    # '' -> '000000'

        df = src_df.copy().assign(__BIN=bins)

        if mode == "rows":
            out = (df.groupby('__BIN', as_index=False)
                   .size()
                   .rename(columns={'__BIN': 'BIN', 'size': 'Total Rx'}))
            label = 'Total Rx'
        elif mode == "qty":
            out = (df.groupby('__BIN', as_index=False)['Qty Filled']
                   .sum()
                   .rename(columns={'__BIN': 'BIN', 'Qty Filled': 'Total Qty'}))
            label = 'Total Qty'
        else:  # unique_rx
            out = (df.groupby('__BIN', as_index=False)['Rx #']
                   .nunique()
                   .rename(columns={'__BIN': 'BIN', 'Rx #': 'Total Rx'}))
            label = 'Total Rx'
        return out, label

    rx_counts_df, RX_LABEL = build_rx_counts(src_df, COUNT_MODE)

    bin_proc_df = (rx_counts_df[['BIN']].copy()
                   .assign(Processor=lambda d: d['BIN'].map(bin_to_proc))
                   # keep 000000
                   .assign(Processor=lambda d: d['Processor'].fillna('Unmapped'))
                   .merge(rx_counts_df, on='BIN', how='left')
                   .sort_values(['Processor', 'BIN'])
                   .reset_index(drop=True))

    # Write headers with the dynamic label in C
    headers = ["BIN", "Processor", RX_LABEL]
    for cidx, h in enumerate(headers, start=1):
        cell = ws2.cell(row=2, column=cidx, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
    # header_row = 2
    # start_data_row = header_row + 1

    # Write A:C
    for r, (bin_, proc, total) in enumerate(
            bin_proc_df[['BIN', 'Processor', RX_LABEL]].itertuples(index=False, name=None), start=3):
        ws2.cell(row=r, column=1, value=str(bin_))
        ws2.cell(row=r, column=2, value=str(proc))
        ws2.cell(row=r, column=3, value=int(total))

    # Optional grand total row to sanity-check equals src_df.shape[0] when COUNT_MODE=="rows"
    gt_row = ws2.max_row + 1
    ws2.cell(row=gt_row, column=2, value="Grand Total").font = Font(bold=True)
    ws2.cell(row=gt_row, column=3,
             value=f"=SUM(C3:C{gt_row-1})").font = Font(bold=True)

    # Widths / filter
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 28
    ws2.column_dimensions['C'].width = 10
    ws2.auto_filter.ref = f"A2:C{ws2.max_row}"
    ws2.freeze_panes = "A3"

    # # Optional: bottom TOTAL row (helps you QA against expected 7,100 etc.)
    # end_row = ws2.max_row
    # total_row = end_row + 1
    # ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    # # Sum of column C
    # ws2.cell(row=total_row, column=3,
    #         value=f"=SUM(C{start_data_row}:C{end_row})").font = Font(bold=True)

    src_norm = src_df.copy()
    src_norm['__BIN'] = (src_norm['Winning_BIN'].astype('string')
                         .fillna('')
                         .str.replace(r'\D', '', regex=True)
                         .str.zfill(6))

    unmapped_rows = src_norm[src_norm['__BIN'] == '000000'].copy()
    fill_col = find_fill_date_column(unmapped_rows)

    ws2.merge_cells('F1:H1')
    title = ws2.cell(row=1, column=6, value="Unmapped BIN Numbers (000000)")
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.font = Font(bold=True, size=14)
    ws2['F2'] = "BIN"
    ws2['G2'] = "RX #"
    if fill_col:
        ws2['H2'] = "Fill Date"

    # Style + widths
    for col in ['F', 'G'] + (['H'] if fill_col else []):
        head = ws2[f'{col}2']
        head.font = Font(bold=True, color="000000")
        head.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[col].width = 18
    if fill_col:
        ws2.column_dimensions['H'].width = 14

    # Coerce date (for pretty output); safe even if mixed types
    if fill_col:
        try:
            unmapped_rows[fill_col] = pd.to_datetime(
                unmapped_rows[fill_col], errors='coerce')
        except Exception:
            pass

   # Write ALL rows (no set()/groupby dedupe): F=BIN, G=RX #, H=Fill Date
    start_row_unmapped = 3
    cols = ['__BIN', 'Rx #'] + ([fill_col] if fill_col else [])
    for r_idx, row in enumerate(unmapped_rows[cols].itertuples(index=False, name=None),
                                start=start_row_unmapped):
        # F -> BIN (000000)
        ws2.cell(row=r_idx, column=6, value=row[0])
        ws2.cell(row=r_idx, column=7, value=str(row[1]))          # G -> RX #
        if fill_col:
            v = row[2]
            # format Timestamp nicely
            if hasattr(v, "strftime"):
                v = v.strftime('%Y-%m-%d')
            # H -> Fill Date
            ws2.cell(row=r_idx, column=8, value=v)

    # Filter across A..H if H exists; else A..G
    last_col_letter = 'H' if fill_col else 'G'
    ws2.auto_filter.ref = f"A2:{last_col_letter}{ws2.max_row}"
    # ===== Create/replace "Vendor Data" sheet =====
    vendor_dfs = []
    for p in vendor_paths:                      # you already build this list earlier
        try:
            vendor_dfs.append(pd.read_csv(p, dtype=str))
        except Exception as e:
            print(f"[warn] vendor read failed {p}: {e}")

    vendor_df_all = (pd.concat(vendor_dfs, ignore_index=True)
                     if vendor_dfs else pd.DataFrame())

    # ---- Formatting
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 10
    ws2.freeze_panes = 'A3'  # keep title+headers fixed
    ws2.auto_filter.ref = f"A2:B{ws2.max_row}"  # filter on BIN/Processor only

    try:
        # If your functions are in the same file, make sure they are defined ABOVE this call.
        # `insurance_paths` is optional; pass None (or a list if you actually use it).
        add_max_difference_sheet(wb, final, insurance_paths=None)
        min_difference_sheet(wb, final, insurance_paths=None)
        create_never_ordered_check_sheet(wb, final)
        add_rx_unit_compare_sheet_exact(
            wb, log_df=rx_compare_source, kinray_df=kinray_all, sheet_name="RX Comparison - All")
        add_rx_unit_compare_sheet_exact_pos(
            wb, log_df=rx_compare_source, kinray_df=kinray_all, sheet_name="RX Comparison +ve")
        add_mfp_drugs_sheet(
            wb, log_df=rx_compare_source, kinray_df=kinray_all, sheet_name="MFP Drugs - RX")
        add_zero_refills_sheet(
            wb, log_df=rx_compare_source, sheet_name="Refills 0 - Call Doctor")
        add_missed_refill_revenue_sheet(
            wb, log_df=rx_compare_source, sheet_name="Missed Refill - Revenue Recovery", grace_days=7)
        add_summary_sheet(wb, processed_source="Processed Data", needs_title="Needs to be ordered - All",
                          header_row=3, data_start_row=4, pharmacy_name=pharmacy_name, date_range=date_range)
        #add_alternate_ndc_sheet(wb, custom_log_df, all_vendor_df)
        # ALT_SHEET_NAME = "Alternate NDC - Purchased"
        # add_alternate_ndc_sheet(wb, log_df, all_vendor_df, sheet_name=ALT_SHEET_NAME)

        #audit_source_df = final_clean if 'final_clean' in locals() else final
        # generate_insurance_audit_files(
        #     audit_source_df,
        #     pharmacy_name=pharmacy_name,
        #     date_range=date_range,
        #     output_dir=output_dir,   # same processed folder
        # )
        audit_df = final_clean if 'final_clean' in locals() else final
        audit_path = None
        audit_name = None
        try:
            audit_path = generate_master_audit_workbook(
                audit_df,
                pharmacy_name=pharmacy_name,
                date_range=date_range,
                output_dir=output_dir
            )
            if audit_path:
                audit_name = os.path.basename(audit_path)

                # If caller requested a copy to a user-specified folder, copy it there
                if user_audit_dir:
                    try:
                        os.makedirs(user_audit_dir, exist_ok=True)
                        dest = os.path.join(user_audit_dir, audit_name)
                        shutil.copy2(audit_path, dest)
                        #print(f"[info] Copied audit workbook to user folder: {dest}")
                    except Exception as _e:
                        print(f"[warn] Could not copy audit workbook to user folder {user_audit_dir}: {_e}")
        except Exception as e:
            print(f"Order helper sheets skipped (audit): {e}")
            # keep going — audit optional
            audit_path = None
            audit_name = None
    except Exception as e:
        print(f"Order helper sheets skipped: {e}")

        # Apply the shared titles/orientation/print settings
    processors = discover_processors_from_df(
        final)   # or pass a list you already have
    apply_common_sheet_settings(wb, pharmacy_name=pharmacy_name,
                                date_range=date_range, processors=processors, header_row_main=3)

    # Round numeric values in key sheets to 2 decimal places
    for sheet_name in ["Needs to be ordered - All", "Do Not Order - ALL", "Never Ordered - Check", "Refills 0 - Call Doctor", "RX Comparison - All", "RX Comparison +ve", "Missed Refill - Revenue Recovery"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=3):  # skip header rows
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.value = round(cell.value, 2)

    # if selected_sheets:
    #     keep = set(selected_sheets)
    #     # always ensure the main sheet stays
    #     keep.add("Processed Data")
    #     keep.add(ALT_SHEET_NAME)   # 👈 NEW
    # else:
    #     keep = None  # means keep everything

    # if keep is not None:
    #     for sheet in wb.sheetnames.copy():
    #         if sheet not in keep:
    #             wb.remove(wb[sheet])

    if "Processed Data" in wb.sheetnames:
        ws = wb["Processed Data"]
        header_row = 3  # your main headers are on row 3
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(output_file)
    try:
        import os
        # 🧹 Remove Windows "Zone.Identifier" metadata (Protected View trigger)
        if os.name == "nt":  # only on Windows
            import subprocess
            subprocess.run(["powershell", "-Command",
                           f'Unblock-File -Path "{output_file}"'], shell=True)
    except Exception as e:
        print(f"[warn] Could not unblock file: {e}")

    return{
        "main" : safe_name,
        "audit": audit_name
    }

# def open_browser():
#     webbrowser.open_new("http://127.0.0.1:5000/")

# if __name__ == '__main__':
#     threading.Timer(1.5, open_browser).start()

#     app.run(
#         host="127.0.0.1",
#         port=5000,
#         debug=False,
#         use_reloader=False
#     )
if __name__ == '__main__':
    # Use FlaskWebGUI to create a desktop window
    FlaskUI(
        app=app,
        server="flask",
        width=1200,
        height=800,
        port=5000
    ).run()