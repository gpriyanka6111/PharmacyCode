# All Flask routes: /, /upload, /finalize, /email, /review, /download, /pick_folder; holds _JOB_CACHE.

import mimetypes
import os
import re
import shutil
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import json
from datetime import datetime
from glob import glob
from urllib.parse import quote
from uuid import uuid4

import numpy as np
import pandas as pd
import tkinter as tk
from tkinter import filedialog

from flask import (Blueprint, current_app, jsonify, make_response, redirect,
                   render_template, request, send_file, send_from_directory,
                   url_for)
from werkzeug.utils import secure_filename

from processing.log_parser import _filter_custom_log_transmitted_paid_ins
from processing.pipeline import process_custom_log_data
from processing.vendor_parser import read_vendor_file

bp = Blueprint('main', __name__)

# keep small "job context" between /upload -> /review -> /finalize
_JOB_CACHE = {}  # { job_id: { "paths": {...}, "summary": {...}, "pharmacy_name":..., "date_range":... } }


def _file_ext(upload):
    return os.path.splitext(secure_filename(upload.filename))[1].lower()


def _is_csv(upload):
    return _file_ext(upload) == '.csv'


def _is_vendor_file(upload):
    return _file_ext(upload) in ('.csv', '.xlsx', '.xls')


@bp.route('/') # Landing page with upload form
def index():
    past_reports = []
    reports_dir = os.path.join(current_app.root_path, 'reports')
    if os.path.isdir(reports_dir):
        for i, fpath in enumerate(sorted(glob(os.path.join(reports_dir, '*_report.json')), reverse=True)):
            try:
                with open(fpath) as f:
                    meta = json.load(f)
                past_reports.append({
                    'filename': os.path.basename(fpath),
                    'generated_at': meta.get('generated_at', ''),
                    'total_rx': meta.get('summary', {}).get('total_rx', 0),
                    'all_pbm_total': meta.get('summary', {}).get('all_pbm_total', 0.0),
                    'pharmacy_name': meta.get('pharmacy_name', ''),
                    'date_range': meta.get('date_range', ''),
                    'is_latest': i == 0,
                    'csv_backup_path': meta.get('csv_backup_path'),
                })
            except Exception:
                pass
    return render_template('index.html', past_reports=past_reports)


@bp.route('/upload', methods=['POST'])
def upload_file():
    print("\n===== UPLOAD DEBUG =====")
    print("FORM KEYS:", list(request.form.keys()))
    print("FILES KEYS:", list(request.files.keys()))
    print("FORM DATA:", request.form.to_dict())
    print("========================\n")

    pharmacy_name = (request.form.get('pharmacy_name') or '').strip()
    date_range = (request.form.get('date_range') or '').strip()

    try:
        vendor_count = int(request.form.get('vendor_count', 0))
    except ValueError:
        vendor_count = 0

    # ---- FILE OBJECTS ----
    custom_log_file = request.files.get('custom_log')      # CSV
    all_pbm_file = request.files.get('all_pbm')         # CSV (optional)
    kinray_file = request.files.get('kinray_file')     # CSV/Excel
    bin_master_file = request.files.get('bin_master')      # CSV (required)

    # ---- Required uploads ----
    missing = [k for k, f in {
        'custom_log':  custom_log_file,
        'kinray_file': kinray_file,
        'bin_master':  bin_master_file,
    }.items() if not f or not f.filename]
    if missing:
        return (f"Missing required file(s): {', '.join(missing)}. "
                f"Got files: {list(request.files.keys())}", 400)

    if not _is_csv(custom_log_file) or not _is_csv(bin_master_file):
        return "Custom Log, ALL PBM, and BIN Master must be CSV", 400
    if all_pbm_file and all_pbm_file.filename and not _is_csv(all_pbm_file):
        return "Custom Log, ALL PBM, and BIN Master must be CSV", 400
    if not _is_vendor_file(kinray_file):
        return "Kinray file must be CSV or Excel", 400

    # ---- Save uploads to per-job dir ----
    updir = current_app.config['UPLOAD_FOLDER']
    os.makedirs(updir, exist_ok=True)

    job_id = uuid4().hex
    job_dir = os.path.join(updir, job_id)
    os.makedirs(job_dir, exist_ok=True)

    custom_log_path = os.path.join(job_dir, 'custom_log.csv')
    kinray_path = os.path.join(job_dir, f'kinray{_file_ext(kinray_file)}')
    bin_master_path = os.path.join(job_dir, 'bin_master.csv')
    all_pbm_path = None

    custom_log_file.save(custom_log_path)
    kinray_file.save(kinray_path)
    bin_master_file.save(bin_master_path)

    if all_pbm_file and all_pbm_file.filename:
        all_pbm_path = os.path.join(job_dir, 'all_pbm.csv')
        all_pbm_file.save(all_pbm_path)

    # ---- Optional vendor files ----
    vendor_paths = []
    if vendor_count and vendor_count > 0:
        for i in range(1, vendor_count + 1):
            vf = request.files.get(f'vendor{i}_file')
            vname = (request.form.get(
                f'vendor{i}_name') or f'Vendor{i}').strip()
            if vf and vf.filename:
                if not _is_vendor_file(vf):
                    return "Vendor file must be CSV or Excel", 400
                safe = re.sub(r'[^A-Za-z0-9_.-]+', '_', vname) or f'Vendor{i}'
                dest = os.path.join(job_dir, f'{safe}{_file_ext(vf)}')
                vf.save(dest)
                vendor_paths.append(dest)

    try:
        # ---- Build summary for review (processors, total Rx, $ by processor) ----
        log_df = pd.read_csv(custom_log_path, dtype=str)
        filter_result = _filter_custom_log_transmitted_paid_ins(log_df)
        log_df = filter_result[0]
        _status_col = filter_result[1] if len(filter_result) > 1 else None
        _kept_rows = filter_result[2] if len(filter_result) > 2 else None
        _dropped_rows = filter_result[3] if len(filter_result) > 3 else None
        bin_df = pd.read_csv(bin_master_path, dtype=str)
        # Normalize BIN columns
        for col in ['Plan 1 BIN', 'Plan 2 BIN']:
            if col in log_df.columns:
                log_df[col] = (log_df[col].astype(str)
                    .str.replace(r'\D', '', regex=True)
                    .str.zfill(6))

        # Normalize payment columns
        for col in ['Ins Paid Plan 1', 'Ins Paid Plan 2']:
            if col in log_df.columns:
                log_df[col] = pd.to_numeric(
                    log_df[col], errors='coerce'
                ).fillna(0)

        # Choose winning BIN and paid per row
        log_df['Winning_BIN'] = log_df.apply(
            lambda r: r['Plan 1 BIN']
            if r['Ins Paid Plan 1'] >= r['Ins Paid Plan 2']
            else r['Plan 2 BIN'], axis=1
        ).str.zfill(6)

        log_df['Winning_Paid'] = np.where(
            log_df['Winning_BIN'] == log_df['Plan 1 BIN'],
            log_df['Ins Paid Plan 1'],
            log_df['Ins Paid Plan 2']
        )

        # Map BIN to Processor
        bin_df['BIN'] = (bin_df['BIN'].astype(str)
            .str.replace(r'\D', '', regex=True)
            .str.zfill(6))
        bin_df['Processor'] = (bin_df['Processor']
            .astype(str).str.strip())
        bin_to_proc = dict(zip(
            bin_df['BIN'], bin_df['Processor']
        ))

        log_df['Processor'] = log_df['Winning_BIN'].map(bin_to_proc)

        # ---- Unmapped BINs (Winning BINs with no Processor mapping) ----
        mask_unmapped = (
            log_df['Processor'].isna()
            & log_df['Winning_BIN'].astype(str).str.strip().ne('')
        )

        if 'Rx #' in log_df.columns:
            unmapped_grp = (
                log_df.loc[mask_unmapped]
                .groupby('Winning_BIN', as_index=False)
                .agg(rx_count=('Rx #', lambda s: s.astype(str).str.strip()
                               .replace('', np.nan).dropna().nunique()))
            )
        else:
            unmapped_grp = (
                log_df.loc[mask_unmapped]
                .groupby('Winning_BIN', as_index=False)
                .size().rename(columns={'size': 'rx_count'})
            )

        unmapped_grp = unmapped_grp.sort_values('rx_count', ascending=False)

        unmapped_bins = [
            {'bin': r['Winning_BIN'], 'rx_count': int(r['rx_count'])}
            for _, r in unmapped_grp.iterrows()
        ]
        unmapped_total_bins = len(unmapped_bins)
        unmapped_total_rx = int(
            unmapped_grp['rx_count'].sum()) if not unmapped_grp.empty else 0

        # total Rx (unique)
        if 'Rx #' in log_df.columns:
            total_rx = (log_df['Rx #'].astype(str).str.strip().replace(
                '', np.nan).dropna().nunique())
        else:
            total_rx = len(log_df)

        # Insurance only rows
        ins_df = log_df[
            log_df['Rx Status'].astype(str)
            .str.strip().str.lower()
            .str.replace(r'[^a-z]', '', regex=True)
            .isin(['paidins', 'transmitted'])
        ].copy()

        # Group by processor
        grp = (ins_df
            .dropna(subset=['Processor'])
            .groupby('Processor', as_index=False)
            .agg(
                rx_count=('Winning_BIN', 'count'),
                total_paid=('Winning_Paid', 'sum')
            )
            .sort_values('total_paid', ascending=False)
        )

        processors = grp['Processor'].tolist()
        by_processor = [
            {
                'processor': str(r['Processor']),
                'rx_count': int(r['rx_count']),
                'total_paid': float(r['total_paid'])
            }
            for _, r in grp.iterrows()
        ]

        # Add CASH separately
        cash_df = log_df[
            log_df['Rx Status'].astype(str)
            .str.strip().str.lower()
            .str.replace(r'[^a-z]', '', regex=True)
            == 'paidcash'
        ].copy()

        if len(cash_df) > 0 and 'Total' in cash_df.columns:
            cash_df['__total__'] = pd.to_numeric(
                cash_df['Total'].astype(str)
                .str.replace(',', '', regex=False)
                .str.replace('$', '', regex=False)
                .str.replace(r'[^0-9.\-]', '', regex=True),
                errors='coerce'
            ).fillna(0)

            by_processor.append({
                'processor': 'CASH',
                'rx_count': int(len(cash_df)),
                'total_paid': float(cash_df['__total__'].sum())
            })
            processors.append('CASH')

        # ✅ expose sheet choices for the modal
        sheets_available = [
            "Processed Data",
            "Needs to be ordered - All",
            "Do Not Order - ALL",
            "Refills 0 - Call Doctor",
            "RX Comparison - All",
            "RX Comparison +ve",
            "MFP Drugs - RX",
            "Missed Refill - Revenue Recovery",
            "BIN to Processor",
            "Summary"
        ]
        # sensible defaults (you can change in UI)
        preselected_sheets = [
            "Processed Data",
            "Needs to be ordered - All",
            "Do Not Order - ALL",
            "Refills 0 - Call Doctor",
            "MFP Drugs - RX",
            "Missed Refill - Revenue Recovery",
            "BIN to Processor",
            "Summary"
        ]

        # ---- Insurance Paid Total (from BestRx custom log) ----
        if 'Ins Paid Total' in log_df.columns:
            all_pbm_total = pd.to_numeric(
                log_df['Ins Paid Total'].astype(str)
                .str.replace(',', '', regex=False)
                .str.replace('$', '', regex=False)
                .str.replace(r'[^0-9.\-]', '', regex=True),
                errors='coerce'
            ).fillna(0).sum()
        else:
            all_pbm_total = pd.to_numeric(
                log_df.get('Winning_Paid',
                pd.Series([0])), errors='coerce'
            ).fillna(0).sum()

        # ---- Kinray Total (Real invoices only) ----
        kinray_raw = read_vendor_file(kinray_path)

        kinray_clean = kinray_raw[
            kinray_raw['Invoice Number'].notna() &
            (kinray_raw['Invoice Number'].astype(str)
             .str.strip().ne('')) &
            (kinray_raw['Invoice Number'].astype(str)
             .str.strip().ne('nan')) &
            (kinray_raw['Invoice Number'].astype(str)
             .str.strip().ne('Invoice Number'))
        ].copy()

        kinray_clean['__price__'] = pd.to_numeric(
            kinray_clean['Invoice $'].astype(str)
            .str.replace(',', '', regex=False)
            .str.replace('$', '', regex=False)
            .str.replace('(', '-', regex=False)
            .str.replace(')', '', regex=False)
            .str.replace(r'[^0-9.\-]', '', regex=True),
            errors='coerce'
        ).fillna(0)

        total_kinray = float(kinray_clean['__price__'].sum())
        print(f'[DEBUG] Kinray clean rows: {len(kinray_clean)}')
        print(f'[DEBUG] Kinray total: {total_kinray}')
        print(f'[DEBUG] Service total: {kinray_clean[kinray_clean["Type"] == "Service"]["__price__"].sum()}')
        kinray_rows = len(kinray_clean)
        kinray_ndcs = int(kinray_clean['NDC/UPC'].nunique())

        # === Branded drugs purchased but not billed ===
        branded_not_billed = []

        try:
            branded_kinray = kinray_clean[
                kinray_clean['Type'] == 'Branded Drug'
            ].copy()

            branded_kinray['NDC_norm'] = (
                branded_kinray['NDC/UPC'].astype(str)
                .str.replace(r'\D', '', regex=True)
                .str.lstrip('0')
            )

            log_df['NDC_norm'] = (
                log_df['Drug NDC'].astype(str)
                .str.replace(r'\D', '', regex=True)
                .str.lstrip('0')
            )

            billed_ndcs = set(log_df['NDC_norm'].unique())

            branded_grp = (
                branded_kinray
                .groupby(['NDC_norm', 'Description'])
                .agg(total_cost=('__price__', 'sum'))
                .reset_index()
                .sort_values('total_cost', ascending=False)
            )

            never_billed = branded_grp[
                ~branded_grp['NDC_norm'].isin(billed_ndcs)
            ]

            VACCINE_KEYWORDS = [
                'SHINGRIX', 'PREVNAR', 'FLUZONE',
                'FLUARIX', 'FLUCELVAX', 'PNEUMOVAX',
                'BEXSERO', 'TRUMENBA', 'GARDASIL',
                'VARIVAX', 'PROQUAD', 'ZOSTAVAX',
                'RECOMBIVAX', 'ENGERIX', 'TWINRIX',
                'HAVRIX', 'VAQTA', 'TDVAX', 'BOOSTRIX',
                'DAPTACEL', 'INFANRIX', 'PEDIARIX'
            ]

            for _, row in never_billed.iterrows():
                desc = str(row['Description']).upper()
                is_vaccine = any(v in desc for v in VACCINE_KEYWORDS)
                branded_not_billed.append({
                    'drug': str(row['Description']),
                    'cost': float(row['total_cost']),
                    'is_vaccine': is_vaccine,
                    'status': 'Vaccine — billed separately' if is_vaccine else 'Investigate'
                })

        except Exception as e:
            print(f'[DEBUG] Branded not billed error: {e}')

        # Breakdown by Type
        kinray_by_type = []
        if 'Type' in kinray_clean.columns:
            type_grp = (
                kinray_clean.groupby('Type')['__price__']
                .sum()
                .sort_values(ascending=False)
                .reset_index()
            )
            kinray_by_type = [
                {
                    'type': str(r['Type']),
                    'total': float(r['__price__'])
                }
                for _, r in type_grp.iterrows()
                if str(r['Type']).strip() not in ('', 'nan', 'Type')
            ]

        summary = {
            "total_rx": int(log_df.shape[0]),
            "processors": processors,
            "by_processor": by_processor,
            "unmapped_bins": unmapped_bins,
            "unmapped_total_bins": unmapped_total_bins,
            "unmapped_total_rx": unmapped_total_rx,
            "note_unmapped": "Update the BIN Master file to map these BINs to processors.",
            "all_pbm_total": float(all_pbm_total),
            "total_kinray": total_kinray,
            "kinray_rows": kinray_rows,
            "kinray_ndcs": kinray_ndcs,
            "kinray_by_type": kinray_by_type,
            "cash_rx_count": int(len(cash_df)) if len(cash_df) > 0 else 0,
            "cash_rx_total": float(cash_df['__total__'].sum()) if len(cash_df) > 0 and '__total__' in cash_df.columns else 0.0,
            "branded_not_billed": branded_not_billed,
            "branded_not_billed_count": len(branded_not_billed),
            "branded_not_billed_total": sum(d['cost'] for d in branded_not_billed),
            "branded_not_billed_investigate_total": sum(d['cost'] for d in branded_not_billed if not d['is_vaccine']),
        }

        _JOB_CACHE[job_id] = {
            "paths": {
                "job_dir": job_dir,
                "custom_log": custom_log_path,
                "kinray": kinray_path,
                "bin_master": bin_master_path,
                "all_pbm": all_pbm_path,
                "vendors": vendor_paths,
            },
            "pharmacy_name": pharmacy_name,
            "date_range": date_range,
            "summary": summary
        }

        return jsonify({
            "ok": True,
            "job_id": job_id,
            "summary": summary
        })
    except Exception as e:
        import traceback
        print("\n" + "=" * 80)
        print("UPLOAD ERROR")
        print("=" * 80)
        traceback.print_exc()
        print("=" * 80 + "\n")

        return jsonify({
            'success': False,
            'error': 'Upload failed',
            'message': str(e),
            'hint': 'Please check your CSV files and try again.'
        }), 400


@bp.route('/email', methods=['POST'])
def email_report():
    from_email = request.form.get('from_email', '').strip()
    to_email = request.form.get('to_email', '').strip()
    message = request.form.get('message', '').strip()
    job_id = request.form.get('job_id', '').strip()

    if not from_email or not to_email:
        return jsonify({"ok": False, "error": "Sender and recipient emails are required."}), 400

    # Figure out which file to attach
    processed_dir = os.path.join(
        current_app.root_path, current_app.config.get('PROCESSED_FOLDER', 'processed'))
    attach_path = None

    # 1) Prefer the job's file if we stored it
    ctx = _JOB_CACHE.get(job_id) if job_id else None
    if ctx and 'outfile' in ctx and os.path.exists(ctx['outfile']):
        attach_path = ctx['outfile']
    else:
        # 2) Fallback to newest .xlsx in processed folder
        candidates = glob.glob(os.path.join(processed_dir, '*.xlsx'))
        if not candidates:
            return jsonify({"ok": False, "error": "No report file found to attach."}), 404
        attach_path = max(candidates, key=os.path.getmtime)

    # Build email with attachment
    msg = MIMEMultipart()
    msg['Subject'] = "RxInsight Report"
    msg['From'] = from_email
    msg['To'] = to_email
    msg.attach(MIMEText(message or "Here's your RxInsight report."))

    ctype, enc = mimetypes.guess_type(attach_path)
    if ctype is None:
        ctype = 'application/octet-stream'
    maintype, subtype = ctype.split('/', 1)

    with open(attach_path, 'rb') as f:
        part = MIMEBase(maintype, subtype)
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment',
                        filename=os.path.basename(attach_path))
        msg.attach(part)

    # Send (configure SMTP for your environment)
    try:
        # Example: Gmail SMTP (requires an app password)
        SMTP_USER = os.environ.get('SMTP_USER') or from_email
        # app password or your relay's credential
        SMTP_PASS = os.environ.get('SMTP_PASS')

        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(from_email, [to_email], msg.as_string())
        return jsonify({"ok": True})
    except Exception as e:
        # Print full traceback to server logs to help debugging (will appear in the console)
        import traceback as _tb
        _tb.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route('/review', methods=['GET'])
def review_job():
    job_id = request.args.get('job_id', '')
    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return jsonify({"ok": False, "error": "Invalid job_id"}), 404
    return jsonify({"ok": True, "job_id": job_id, "summary": ctx["summary"]})


@bp.route('/download')
def download_file():
    # filename = request.args.get('filename', '')
    # if not filename:
    #     return "Missing filename", 400
    # # prevent path traversal
    # safe = os.path.basename(filename)
    # fullpath = os.path.join(app.root_path, app.config.get('PROCESSED_FOLDER', 'processed'), safe)
    # if not os.path.exists(fullpath):
    #     return "File not found", 404
    # return send_file(fullpath, as_attachment=True, download_name=safe)
    filename = request.args.get("filename")
    directory = os.path.join(current_app.root_path, current_app.config.get(
        'PROCESSED_FOLDER', 'processed'))
    response = make_response(send_from_directory(
        directory, filename, as_attachment=True))
    # 🚀 Remove "Internet zone" mark by using generic content-type
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.headers["Content-Type"] = "application/octet-stream"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@bp.errorhandler(500)
def _handle_500(e):
    # Ensure JSON instead of HTML debugger
    return jsonify({"error": str(e)}), 500


@bp.route('/finalize', methods=['POST'])
def finalize_job():
    data = request.get_json(force=True, silent=True) or {}
    job_id = (data.get('job_id') or '').strip()

    # Validate required folder paths
    main_save_dir = (data.get('main_save_dir') or '').strip()
    audit_save_dir = (data.get('audit_save_dir') or '').strip()

    if not main_save_dir:
        return jsonify({"ok": False, "error": "Main report destination folder is required"}), 400
    if not audit_save_dir:
        return jsonify({"ok": False, "error": "Audit workbook destination folder is required"}), 400

    if not os.path.isdir(main_save_dir):
        return jsonify({"ok": False, "error": f"Main report folder does not exist: {main_save_dir}"}), 400
    if not os.path.isdir(audit_save_dir):
        return jsonify({"ok": False, "error": f"Audit workbook folder does not exist: {audit_save_dir}"}), 400

    selected_processors = [str(p).strip().upper() for p in (
        data.get('selected_processors') or []) if p]
    selected_sheets = [str(s).strip()
                       for s in (data.get('selected_sheets') or []) if s]

    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return jsonify({"ok": False, "error": "Invalid job_id"}), 404

    paths = ctx["paths"]
    processed_dir = os.path.join(
        current_app.root_path, current_app.config.get('PROCESSED_FOLDER', 'processed'))
    os.makedirs(processed_dir, exist_ok=True)

    try:
        # include Kinray first, then the extras
        vendor_paths = [paths["kinray"], *paths["vendors"]]

        result = process_custom_log_data(
            custom_log_path=paths["custom_log"],
            all_pbm_path=paths["all_pbm"],
            bin_master_path=paths["bin_master"],
            vendor_paths=vendor_paths,
            pharmacy_name=ctx.get("pharmacy_name", ""),
            date_range=ctx.get("date_range", ""),
            selected_processors=selected_processors,
            selected_sheets=selected_sheets,
            user_audit_dir=audit_save_dir,
        )
        if not result or not isinstance(result, dict):
            return jsonify({"ok": False, "error": "Report generation returned no filename"}), 500

        main_name = result.get("main")
        audit_name = result.get("audit")
        #print(main_name, audit_name)
        if not main_name:
            return jsonify({"ok": False, "error": "Missing main filename"}), 500
        # IMPORTANT: do NOT change the name returned by the writer
        #fullpath = os.path.join(processed_dir, out_filename)
        processed_dir = os.path.join(
            current_app.root_path, current_app.config.get('PROCESSED_FOLDER', 'processed'))
        main_path = os.path.join(processed_dir, main_name)
        audit_path = os.path.join(processed_dir, audit_name) if audit_name else None

        if not os.path.exists(main_path):
            # print("[finalize] processed dir listing:", os.listdir(processed_dir))
            return jsonify({"ok": False, "error": "Main output file not found"}), 500


        ctx["outfile_main"] = main_path

        # Copy main file to user-specified folder (now required)
        try:
            import shutil
            main_dest = os.path.join(main_save_dir, main_name)
            shutil.copy2(main_path, main_dest)
            #print(f"[finalize] copied main report to: {main_dest}")
        except Exception as e:
            #print(f"[finalize] error copying main file to {main_save_dir}: {e}")
            return jsonify({"ok": False, "error": f"Failed to copy main report: {str(e)}"}), 500

        audit_exists = False
        if audit_name:
            audit_path = os.path.join(processed_dir, audit_name)
            if os.path.exists(audit_path):
                ctx["outfile_audit"] = audit_path
                audit_exists = True
                # Copy audit file to user-specified folder (now required)
                try:
                    import shutil
                    audit_dest = os.path.join(audit_save_dir, audit_name)
                    shutil.copy2(audit_path, audit_dest)
                    #(f"[finalize] copied audit report to: {audit_dest}")
                except Exception as e:
                    #print(f"[finalize] error copying audit file to {audit_save_dir}: {e}")
                    return jsonify({"ok": False, "error": f"Failed to copy audit report: {str(e)}"}), 500
            else:
                print(f"[finalize] audit file expected but not found: {audit_path}")


        # if not os.path.exists(fullpath):
        #     # (Optional) one fallback try: a regex-sanitized variant
        #     alt = re.sub(r'[^A-Za-z0-9()._\-\s]+', '_', out_filename)
        #     altpath = os.path.join(processed_dir, alt)
        #     if os.path.exists(altpath):
        #         out_filename = alt
        #     else:
        #         print("[finalize] expected at:", os.path.abspath(fullpath))
        #         print("[finalize] processed dir listing:",
        #               os.listdir(processed_dir))
        #         return jsonify({"ok": False, "error": f"Output file not found: {out_filename}"}), 500

        # ctx['outfile'] = fullpath

        # return jsonify({
        #     "ok": True,
        #     "filename": out_filename,
        #     "download_url": f"/download?filename={quote(out_filename)}"
        # })
        # return jsonify({
        #     "ok": True,
        #     "filename": main_name,
        #     "download_url": f"/download?filename={quote(main_name)}",
        #     # new behaviour (audit report)
        #     # "audit_filename": audit_name,
        #     # "audit_download_url": f"/download?filename={quote(audit_name)}",
        # })
        # Persist report metadata to disk so it survives restarts
        try:
            reports_dir = os.path.join(current_app.root_path, 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            csv_backup_path = None
            try:
                backup_dir = os.path.join(reports_dir, f'{ts}_csvs')
                os.makedirs(backup_dir, exist_ok=True)
                for src in [paths.get('custom_log'), paths.get('kinray'),
                            paths.get('bin_master'), paths.get('all_pbm')]:
                    if src and os.path.exists(src):
                        shutil.copy2(src, backup_dir)
                csv_backup_path = backup_dir
            except Exception as _e:
                print(f'[finalize] csv backup error: {_e}')
            report_meta = {
                'generated_at': datetime.now().isoformat(),
                'excel_path': main_path,
                'pharmacy_name': ctx.get('pharmacy_name', ''),
                'date_range': ctx.get('date_range', ''),
                'summary': ctx.get('summary', {}),
                'paths': ctx.get('paths', {}),
                'csv_backup_path': csv_backup_path,
            }
            with open(os.path.join(reports_dir, f'{ts}_report.json'), 'w') as f:
                json.dump(report_meta, f)
        except Exception as e:
            print(f'[finalize] report save error: {e}')

        resp = {
            "ok": True,
            "main_filename": main_name,
            "main_download_url": f"/download?filename={quote(main_name)}",
            "download_url": f"/download?filename={quote(main_name)}",  # backwards-compatible
        }
        if audit_name and ctx.get("outfile_audit"):
            resp.update({
                "audit_filename": audit_name,
                "audit_download_url": f"/download?filename={quote(audit_name)}"
            })
        return jsonify(resp)


    except Exception as e:
        import traceback
        traceback.print_exc()   # prints full traceback to terminal
        return jsonify({
            "ok": False,
            "error": f"Finalize failed: {str(e)}"
        }), 500


@bp.route('/dashboard')
def dashboard():
    job_id = request.args.get('job_id', '')
    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return redirect('/')

    if not ctx.get('summary'):
        return render_template('error.html',
            title='No Report Loaded',
            message='No report data found in memory.',
            hint='Please select a past report or upload new CSV files to continue.',
            back_url='/',
            back_label='Go to Home'
        ), 404

    try:
        log_df = pd.read_csv(ctx['paths']['custom_log'], dtype=str)
    except Exception:
        log_df = pd.DataFrame()
    total_rx = len(log_df) or ctx.get('summary', {}).get('total_rx', 0)

    top_doctors = []
    if 'Prescriber NPI #' in log_df.columns:

        def most_common(x):
            return x.value_counts().index[0] if len(x) > 0 else ''

        npi_grp = (
            log_df.groupby('Prescriber NPI #', as_index=False)
            .agg(
                rx_count=('Rx #', 'count'),
                ins_rx=('Rx Status', lambda x: int(x.isin(['Paid-Ins', 'Transmitted']).sum())),
                cash_rx=('Rx Status', lambda x: int(
                    (x.str.strip().str.lower().str.replace(r'[^a-z]', '', regex=True) == 'paidcash').sum()
                )),
                prescriber_name=('Prescriber Name', most_common),
                prescriber_phone=('Prescriber Phone #', 'first'),
                prescriber_fax=('Prescriber Fax #', 'first'),
                prescriber_city=('Prescriber City', 'first'),
                prescriber_state=('Prescriber State', 'first'),
                prescriber_address=('Prescriber Address 1', 'first'),
                prescriber_zip=('Prescriber Zip', 'first'),
                unique_addresses=('Prescriber Address', 'nunique'),
                npi=('Prescriber NPI #', 'first')
            )
            .sort_values('rx_count', ascending=False)
            .head(20)
        )

        for _, row in npi_grp.iterrows():
            name = str(row['prescriber_name']).strip()
            parts = [p.strip().capitalize() for p in name.split() if p.strip()]
            display_name = ' '.join(parts)
            initials = ''.join([p[0].upper() for p in parts[:2]])
            pct = round((int(row['rx_count']) / total_rx) * 100, 1)

            top_doctors.append({
                'npi': str(row['npi']),
                'name': display_name,
                'initials': initials,
                'phone': str(row['prescriber_phone']).strip(),
                'fax': str(row['prescriber_fax']).strip(),
                'city': str(row['prescriber_city']).strip(),
                'state': str(row['prescriber_state']).strip(),
                'address': str(row['prescriber_address']).strip(),
                'zip': str(row['prescriber_zip']).strip(),
                'rx_count': int(row['rx_count']),
                'ins_rx': int(row['ins_rx']),
                'cash_rx': int(row['cash_rx']),
                'pct': pct,
                'locations': int(row['unique_addresses'])
            })

    summary = ctx.get('summary', {})

    main_file = ctx.get('outfile_main')
    excel_data = {}

    if main_file and os.path.exists(main_file):
        for key, sheet_name in {
            'needs_ordering': 'Needs to be ordered - All',
            'do_not_order': 'Do Not Order - ALL',
            'rx_comparison': 'RX Comparison - All',
        }.items():
            try:
                df = pd.read_excel(
                    main_file,
                    sheet_name=sheet_name,
                    dtype=str,
                    header=1
                )
                df = df.dropna(how='all').fillna('')
                excel_data[key] = {
                    'count': len(df),
                    'columns': df.columns.tolist(),
                    'rows': df.head(100).values.tolist()
                }
            except Exception as e:
                print(f'[DEBUG] {key}: {e}')
                excel_data[key] = {
                    'count': 0,
                    'columns': [],
                    'rows': []
                }

    # Top discrepancies (real data)
    top_discrepancies = []
    try:
        log = pd.read_csv(ctx['paths']['custom_log'], dtype=str)

        kinray_raw = read_vendor_file(ctx['paths']['kinray'])

        kinray_clean = kinray_raw[
            kinray_raw['Invoice Number'].notna() &
            (kinray_raw['Invoice Number'].astype(str).str.strip().ne('')) &
            (kinray_raw['Invoice Number'].astype(str).str.strip().ne('nan')) &
            (kinray_raw['Invoice Number'].astype(str).str.strip().ne('Invoice Number'))
        ].copy()

        kinray_clean['NDC_norm'] = (
            kinray_clean['NDC/UPC'].astype(str)
            .str.replace(r'\D', '', regex=True)
            .str.lstrip('0')
        )
        kinray_clean['Ship Qty'] = pd.to_numeric(
            kinray_clean['Ship Qty'], errors='coerce'
        ).fillna(0)

        kinray_qty = (
            kinray_clean
            .groupby('NDC_norm')
            .agg(kinray_qty=('Ship Qty', 'sum'), drug_name=('Description', 'first'))
            .reset_index()
        )

        log['NDC_norm'] = (
            log['Drug NDC'].astype(str)
            .str.replace(r'\D', '', regex=True)
            .str.lstrip('0')
        )
        log['Qty Filled'] = pd.to_numeric(log['Qty Filled'], errors='coerce').fillna(0)
        log['Drug Pkg Size'] = pd.to_numeric(log['Drug Pkg Size'], errors='coerce').fillna(1)
        log['pkgs_billed'] = log['Qty Filled'] / log['Drug Pkg Size']

        billing_qty = (
            log.groupby('NDC_norm')
            .agg(billed_pkgs=('pkgs_billed', 'sum'), drug_name=('Drug Name', 'first'))
            .reset_index()
        )

        merged = billing_qty.merge(kinray_qty, on='NDC_norm', how='inner')
        merged['diff'] = (merged['kinray_qty'] - merged['billed_pkgs']).round(1)

        understocked = merged[merged['diff'] < 0].sort_values('diff').head(3)
        overstocked = merged[merged['diff'] > 0].sort_values('diff', ascending=False).head(2)
        combined = pd.concat([understocked, overstocked]).sort_values('diff')

        for _, row in combined.iterrows():
            diff = float(row['diff'])
            top_discrepancies.append({
                'drug': str(row['drug_name_x']).strip()[:40],
                'billed': round(float(row['billed_pkgs']), 1),
                'purchased': round(float(row['kinray_qty']), 1),
                'diff': diff,
                'status': 'understocked' if diff < 0 else 'overstocked'
            })

    except Exception as e:
        print(f'[DEBUG] Discrepancy error: {e}')

    # Build real key insights from actual data
    key_insights = []

    # 1. Check CVS concentration risk
    try:
        by_proc = summary.get('by_processor', [])
        total_rev = sum(p['total_paid'] for p in by_proc)
        if total_rev > 0:
            for p in by_proc:
                pct = (p['total_paid'] / total_rev) * 100
                if pct > 50 and p['processor'] != 'CASH':
                    key_insights.append({
                        'type': 'red',
                        'title': f"{p['processor']} pays {pct:.1f}% of all revenue",
                        'sub': 'High dependency — risk if this payer changes rates'
                    })
    except Exception as e:
        print(f'[DEBUG] Insight 1: {e}')

    # 2. Branded drugs not billed
    try:
        count = summary.get('branded_not_billed_count', 0)
        total = summary.get('branded_not_billed_total', 0)
        investigate = summary.get('branded_not_billed_investigate_total', 0)
        if count > 0:
            key_insights.append({
                'type': 'red',
                'title': f"{count} branded drugs purchased but not billed",
                'sub': f"${investigate:,.0f} needs investigation (excl. vaccines)"
            })
    except Exception as e:
        print(f'[DEBUG] Insight 2: {e}')

    # 3. Drugs need ordering
    try:
        order_count = excel_data.get('needs_ordering', {}).get('count', 0)
        if order_count > 0:
            key_insights.append({
                'type': 'amber',
                'title': f"{order_count} drugs need to be ordered",
                'sub': 'View order list for estimated restock cost'
            })
    except Exception as e:
        print(f'[DEBUG] Insight 3: {e}')

    # 4. Underpayments from Rx comparison
    try:
        rx_count = excel_data.get('rx_comparison', {}).get('count', 0)
        if rx_count > 0:
            key_insights.append({
                'type': 'amber',
                'title': f"{rx_count} Rx comparisons analyzed",
                'sub': 'Check Rx comparison sheet for underpayments eligible for dispute'
            })
    except Exception as e:
        print(f'[DEBUG] Insight 4: {e}')


    # 6. Overstocked drugs
    try:
        over_count = excel_data.get('do_not_order', {}).get('count', 0)
        if over_count > 0:
            key_insights.append({
                'type': 'green',
                'title': f"{over_count} drugs are overstocked",
                'sub': 'Do not order these — cash tied up in inventory'
            })
    except Exception as e:
        print(f'[DEBUG] Insight 6: {e}')

    # 7. Total Rx processed
    try:
        total_rx = summary.get('total_rx', 0)
        all_pbm = summary.get('all_pbm_total', 0)
        if total_rx > 0:
            avg = all_pbm / total_rx if total_rx else 0
            key_insights.append({
                'type': 'green',
                'title': f"{total_rx:,} total Rx processed this period",
                'sub': f"Avg insurance paid: ${avg:.2f} per Rx"
            })
    except Exception as e:
        print(f'[DEBUG] Insight 7: {e}')

    # 8. Top processor dominance warning
    try:
        by_proc = summary.get('by_processor', [])
        if len(by_proc) >= 2:
            top2 = sorted(by_proc, key=lambda x: x['total_paid'], reverse=True)[:2]
            total_rev = sum(p['total_paid'] for p in by_proc)
            top2_pct = sum(p['total_paid'] for p in top2) / total_rev * 100 if total_rev else 0
            if top2_pct > 80:
                key_insights.append({
                    'type': 'amber',
                    'title': f"Top 2 insurances = {top2_pct:.0f}% of revenue",
                    'sub': f"{top2[0]['processor']} + {top2[1]['processor']} dominate your revenue"
                })
    except Exception as e:
        print(f'[DEBUG] Insight 8: {e}')

    return render_template('dashboard.html',
        pharmacy_name=ctx.get('pharmacy_name', 'Pharmacy'),
        date_range=ctx.get('date_range', ''),
        summary=summary,
        top_doctors=top_doctors,
        excel_data=excel_data,
        top_discrepancies=top_discrepancies,
        job_id=job_id,
        key_insights=key_insights
    )


def _apply_sheet_filters(rows, columns, sheet_key, search, severity, drug_type, processor):
    SEV_THRESHOLDS = {
        'needs_ordering': {'high': 5,  'med_min': 2, 'med_max': 4},
        'do_not_order':   {'high': 10, 'med_min': 5, 'med_max': 10},
    }
    name_idx = next((i for i, c in enumerate(columns) if 'drug' in c.lower() and 'name' in c.lower()), -1)
    ndc_idx  = next((i for i, c in enumerate(columns) if 'ndc' in c.lower()), -1)
    type_idx = next((i for i, c in enumerate(columns) if c == 'Drug Type'), -1)
    sev_col_name = {'needs_ordering': 'To Order', 'do_not_order': 'Do Not Order'}.get(sheet_key, '')
    sev_col_idx  = next((i for i, c in enumerate(columns) if c == sev_col_name), -1)

    if search:
        rows = [r for r in rows if
            (name_idx >= 0 and search in str(r[name_idx]).lower()) or
            (ndc_idx  >= 0 and search in str(r[ndc_idx]).lower())]

    if severity and sheet_key in SEV_THRESHOLDS:
        thresh = SEV_THRESHOLDS[sheet_key]
        def _get_sev(row):
            if sev_col_idx < 0:
                return 'low'
            try:
                v = abs(float(str(row[sev_col_idx]).replace(',', '') or 0))
            except (ValueError, TypeError):
                v = 0
            if v > thresh['high']:
                return 'high'
            if thresh['med_min'] <= v <= thresh['med_max']:
                return 'med'
            return 'low'
        rows = [r for r in rows if _get_sev(r) == severity]

    if drug_type and type_idx >= 0:
        if drug_type == 'Unclassified':
            rows = [r for r in rows if str(r[type_idx]).strip() == 'Unclassified']
        else:
            rows = [r for r in rows if str(r[type_idx]).strip() in (drug_type, 'Unclassified')]

    if processor:
        include_procs = {p.strip() for p in processor.split(',') if p.strip()}
        proc_indices = [i for i, c in enumerate(columns)
                        if (c.endswith('_D') or c.endswith('_d')) and c != 'ALL_PBM_D'
                        and c in include_procs]
        if proc_indices:
            def _has_proc(row):
                for i in proc_indices:
                    try:
                        if float(str(row[i]).replace(',', '')) > 0:
                            return True
                    except (ValueError, TypeError):
                        pass
                return False
            rows = [r for r in rows if _has_proc(r)]

    return rows


@bp.route('/api/sheet_data')
def sheet_data():
    job_id    = request.args.get('job_id', '')
    sheet     = request.args.get('sheet', '')
    page      = request.args.get('page', 1, type=int)
    per_page  = 50
    search    = (request.args.get('search')    or '').strip().lower()
    severity  = (request.args.get('severity')  or '').strip()
    drug_type = (request.args.get('drug_type') or '').strip()
    processor = (request.args.get('processor') or '').strip()

    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return jsonify({'ok': False}), 404

    main_file = ctx.get('outfile_main')
    if not main_file or not os.path.exists(main_file):
        return jsonify({'ok': False}), 404

    sheet_map = {
        'do_not_order':    'Do Not Order - ALL',
        'needs_ordering':  'Needs to be ordered - All',
        'rx_comparison':   'RX Comparison - All',
    }
    sheet_name = sheet_map.get(sheet)
    if not sheet_name:
        return jsonify({'ok': False}), 400

    try:
        df = pd.read_excel(main_file, sheet_name=sheet_name, dtype=str, header=1)
        df = df.dropna(how='all').fillna('')
        columns = df.columns.tolist()
        rows = df.values.tolist()

        rows = _apply_sheet_filters(rows, columns, sheet, search, severity, drug_type, processor)

        total_count = len(rows)
        start = (page - 1) * per_page
        end   = start + per_page
        page_rows = rows[start:end]

        result_rows = [
            {'row_number': start + idx + 1, 'data': row}
            for idx, row in enumerate(page_rows)
        ]

        return jsonify({
            'ok':          True,
            'total_count': total_count,
            'page':        page,
            'rows':        result_rows,
            'columns':     columns,
            'has_more':    end < total_count,
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/api/export_excel/<sheet_key>')
def export_excel(sheet_key):
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    job_id    = request.args.get('job_id', '')
    search    = (request.args.get('search')    or '').strip().lower()
    severity  = (request.args.get('severity')  or '').strip()
    drug_type = (request.args.get('drug_type') or '').strip()
    processor = (request.args.get('processor') or '').strip()

    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return jsonify({'ok': False}), 404
    main_file = ctx.get('outfile_main')
    if not main_file or not os.path.exists(main_file):
        return jsonify({'ok': False}), 404

    sheet_map = {
        'do_not_order':    'Do Not Order - ALL',
        'needs_ordering':  'Needs to be ordered - All',
        'rx_comparison':   'RX Comparison - All',
    }
    sheet_name = sheet_map.get(sheet_key)
    if not sheet_name:
        return jsonify({'ok': False}), 400

    EXPORT_HIDDEN = {'Drug Type', 'Insurance-wise Order Estimate ($)', 'Amount'}

    try:
        df = pd.read_excel(main_file, sheet_name=sheet_name, dtype=str, header=1)
        df = df.dropna(how='all').fillna('')
        columns = df.columns.tolist()
        rows = df.values.tolist()

        rows = _apply_sheet_filters(rows, columns, sheet_key, search, severity, drug_type, processor)

        visible = [(i, col) for i, col in enumerate(columns)
                   if col not in EXPORT_HIDDEN and not col.lower().startswith('insurance-wise')]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        hdr_font = Font(bold=True)
        hdr_fill = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
        for col_num, (_, col_name) in enumerate(visible, start=1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        for row_num, row in enumerate(rows, start=2):
            for col_num, (col_idx, _) in enumerate(visible, start=1):
                val = row[col_idx] if col_idx < len(row) else ''
                try:
                    fval = float(str(val).replace(',', ''))
                    val = int(fval) if fval == int(fval) else fval
                except (ValueError, TypeError):
                    pass
                ws.cell(row=row_num, column=col_num, value=val)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = re.sub(r'[^A-Za-z0-9_-]', '_', sheet_key)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f'{safe_name}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/sheet/do_not_order')
def sheet_do_not_order():
    job_id = request.args.get('job_id', '')
    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return redirect('/')

    main_file = ctx.get('outfile_main')
    rows = []
    columns = []
    error = None

    if main_file and os.path.exists(main_file):
        try:
            df = pd.read_excel(
                main_file,
                sheet_name='Do Not Order - ALL',
                dtype=str,
                header=1
            )
            df = df.dropna(how='all').fillna('')
            columns = df.columns.tolist()
            rows = df.values.tolist()
        except Exception as e:
            error = str(e)

    drug_types = []
    if main_file and os.path.exists(main_file):
        try:
            dft = pd.read_excel(
                main_file,
                sheet_name='Do Not Order - ALL',
                dtype=str, header=1
            )
            dft = dft.dropna(how='all')
            if 'Drug Type' in dft.columns:
                type_order = [
                    'Branded Drug', 'Generic Drug', 'OTC Drug',
                    'Home Health Care', 'Health and Beauty Aid', 'Unclassified'
                ]
                all_types = dft['Drug Type'].dropna().unique().tolist()
                drug_types = sorted(
                    [t for t in all_types if t and t != 'nan'],
                    key=lambda x: type_order.index(x) if x in type_order else 99
                )
        except Exception as e:
            print(f'[DEBUG] drug types: {e}')

    # Fallback if Drug Type not in sheet yet (pre-regeneration)
    if not drug_types:
        drug_types = [
            'Branded Drug', 'Generic Drug', 'OTC Drug',
            'Home Health Care', 'Health and Beauty Aid', 'Unclassified'
        ]

    return render_template(
        'sheet_view.html',
        pharmacy_name=ctx.get('pharmacy_name', 'Pharmacy'),
        date_range=ctx.get('date_range', ''),
        job_id=job_id,
        columns=columns,
        rows=rows,
        error=error,
        sheet_key='do_not_order',
        sheet_title='Drugs overstocked — do not order',
        sheet_count=len(rows),
        sheet_badge_color='#EAF3DE',
        sheet_badge_text_color='#3B6D11',
        sheet_accent_color='#639922',
        sheet_export_bg='#3b7c0f',
        severity_col='Do Not Order',
        high_label='High >10',
        med_label='Medium 5-10',
        low_label='Low <5',
        drug_types=drug_types,
        stat_labels={
            'total_sub': 'overstocked',
            'high_label': 'High overstock',
            'high_sub': 'do not order',
            'high_threshold': '>10',
            'med_label': 'Medium overstock',
            'med_sub': 'monitor',
            'med_threshold': '5-10',
            'low_label': 'Low overstock',
            'low_sub': 'acceptable',
            'low_threshold': '<5',
            'high_min': 10,
            'med_min': 5,
            'med_max': 10,
        },
    )


@bp.route('/sheet/needs_ordering')
def sheet_needs_ordering():
    job_id = request.args.get('job_id', '')
    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return redirect('/')

    main_file = ctx.get('outfile_main')
    sheet_data = {
        'count': 0,
        'columns': [],
        'rows': []
    }

    if main_file and os.path.exists(main_file):
        try:
            df = pd.read_excel(
                main_file,
                sheet_name='Needs to be ordered - All',
                dtype=str,
                header=1
            )
            df = df.dropna(how='all').fillna('')
            sheet_data = {
                'count': len(df),
                'columns': df.columns.tolist(),
                'rows': df.head(100).values.tolist()
            }
        except Exception as e:
            print(f'[DEBUG] Needs ordering sheet error: {e}')

    drug_types = []
    if main_file and os.path.exists(main_file):
        try:
            dft = pd.read_excel(
                main_file,
                sheet_name='Needs to be ordered - All',
                dtype=str, header=1
            )
            dft = dft.dropna(how='all')
            if 'Drug Type' in dft.columns:
                type_order = [
                    'Branded Drug', 'Generic Drug', 'OTC Drug',
                    'Home Health Care', 'Health and Beauty Aid', 'Unclassified'
                ]
                all_types = dft['Drug Type'].dropna().unique().tolist()
                drug_types = sorted(
                    [t for t in all_types if t and t != 'nan'],
                    key=lambda x: type_order.index(x) if x in type_order else 99
                )
        except Exception as e:
            print(f'[DEBUG] drug types: {e}')

    # Fallback if Drug Type not in sheet yet (pre-regeneration)
    if not drug_types:
        drug_types = [
            'Branded Drug', 'Generic Drug', 'OTC Drug',
            'Home Health Care', 'Health and Beauty Aid', 'Unclassified'
        ]

    return render_template(
        'sheet_view.html',
        job_id=job_id,
        pharmacy_name=ctx.get('pharmacy_name', ''),
        date_range=ctx.get('date_range', ''),
        sheet_key='needs_ordering',
        sheet_title='Drugs that need ordering',
        sheet_count=sheet_data['count'],
        sheet_badge_color='#FCEBEB',
        sheet_badge_text_color='#A32D2D',
        sheet_accent_color='#E24B4A',
        sheet_export_bg='#E24B4A',
        columns=sheet_data['columns'],
        rows=sheet_data['rows'],
        severity_col='To Order',
        high_label='Critical >5',
        med_label='Medium 2-4',
        low_label='Low ≤1',
        high_color='#E24B4A',
        med_color='#EF9F27',
        low_color='#639922',
        drug_types=drug_types,
        stat_labels={
            'total_sub': 'need ordering',
            'high_label': 'High understock',
            'high_sub': 'urgent',
            'high_threshold': '>5',
            'med_label': 'Medium understock',
            'med_sub': 'monitor',
            'med_threshold': '2-4',
            'low_label': 'Low understock',
            'low_sub': 'low priority',
            'low_threshold': '≤1',
            'high_min': 5,
            'med_min': 2,
            'med_max': 4,
        },
    )


@bp.route('/sheet/rx_comparison')
def sheet_rx_comparison():
    job_id = request.args.get('job_id', '')
    ctx = _JOB_CACHE.get(job_id)
    if not ctx:
        return redirect('/')

    main_file = ctx.get('outfile_main')
    sheet_data = {
        'count': 0,
        'columns': [],
        'rows': []
    }

    if main_file and os.path.exists(main_file):
        try:
            df = pd.read_excel(
                main_file,
                sheet_name='RX Comparison - All',
                dtype=str,
                header=1
            )
            df = df.dropna(how='all').fillna('')
            sheet_data = {
                'count': len(df),
                'columns': df.columns.tolist(),
                'rows': df.head(100).values.tolist()
            }
        except Exception as e:
            print(f'[DEBUG] RX Comparison sheet error: {e}')

    return render_template(
        'sheet_view.html',
        job_id=job_id,
        pharmacy_name=ctx.get('pharmacy_name', ''),
        date_range=ctx.get('date_range', ''),
        sheet_key='rx_comparison',
        sheet_title='Rx comparison — underpayment analysis',
        sheet_count=sheet_data['count'],
        sheet_badge_color='#E6F1FB',
        sheet_badge_text_color='#185FA5',
        sheet_accent_color='#185FA5',
        sheet_export_bg='#185FA5',
        columns=sheet_data['columns'],
        rows=sheet_data['rows'],
        severity_col='Difference',
        high_label='Underpaid >$10',
        med_label='Underpaid $5-$10',
        low_label='Underpaid <$5',
        high_color='#E24B4A',
        med_color='#EF9F27',
        low_color='#639922',
        drug_types=[],
        stat_labels=None,
    )


@bp.route('/load_report/<filename>')
def load_report(filename):
    try:
        safe = os.path.basename(filename)
        if not safe.endswith('_report.json'):
            return redirect('/')
        reports_dir = os.path.join(current_app.root_path, 'reports')
        path = os.path.join(reports_dir, safe)
        if not os.path.exists(path):
            return redirect('/')
        with open(path) as f:
            meta = json.load(f)
        job_id = uuid4().hex
        _JOB_CACHE[job_id] = {
            'paths': meta.get('paths', {}),
            'pharmacy_name': meta.get('pharmacy_name', ''),
            'date_range': meta.get('date_range', ''),
            'summary': meta.get('summary', {}),
            'outfile_main': meta.get('excel_path', ''),
            'csv_backup_path': meta.get('csv_backup_path'),
        }
        return redirect(url_for('main.dashboard', job_id=job_id))
    except Exception:
        return render_template('error.html',
            title='Could Not Load Report',
            message='The report file may be corrupted or missing.',
            hint='Please try another report or upload new CSV files.',
            back_url='/',
            back_label='Go to Home'
        ), 500


@bp.route('/pick_folder', methods=['GET'])
def pick_folder():
    """Open a native folder picker on the server (local desktop) and return the chosen path.

    Note: This only works safely when the Flask app runs on the user's desktop (local machine).
    """
    try:
        # Use a transient Tk root to show the directory dialog
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askdirectory(title='Choose folder to save audit workbook')
        root.destroy()
        if not path:
            return jsonify({"ok": False, "error": "No folder selected"}), 400
        return jsonify({"ok": True, "path": path})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
